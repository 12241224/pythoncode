import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import os
import logging

# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 初始化日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class UC_Model:
    def __init__(self, data_path):
        self.data_path = data_path
        self.load_data()
        self.create_model()
    
    def get_unit_param_as_float(self, unit_dict, unit, param_name, default=0.0):
        """
        从机组字典中获取参数值并转换为浮点数
        支持多种可能的列名
        """
        try:
            unit_data = unit_dict[unit]
            
            # 定义可能的列名映射
            param_name_mapping = {
                'Capacity': ['Capacity', 'MaxPower', '机组装机容量(MW)', '装机容量', '容量'],
                'MaxPower': ['MaxPower', 'Capacity', '机组装机容量(MW)', '最大出力'],
                'MinPower': ['MinPower', 'Minpower', '最小出力'],
                'OperationCoeff_A': ['OperationCoeff_A', 'Coeff_A', 'A系数'],
                'OperationCoeff_B': ['OperationCoeff_B', 'Coeff_B', 'B系数'],
                'OperationCoeff_C': ['OperationCoeff_C', 'Coeff_C', 'C系数'],
                'UpRamppingRate': ['UpRamppingRate', 'UpRampingRate', '爬坡速率', '上爬坡速率'],
                'DownRamppingRate': ['DownRamppingRate', 'DownRampingRate', '下爬坡速率'],
                'MinOnLineTime': ['MinOnLineTime', 'MinOnTime', '最小运行时间'],
                'MinOffLineTime': ['MinOffLineTime', 'MinOffTime', '最小停机时间'],
                'DeepPeakShavingThreshold': ['DeepPeakShavingThreshold', '深度调峰阈值'],
                'DeepPeakShavingCost': ['DeepPeakShavingCost', '深度调峰成本'],
                'StateOnFuelConsumption': ['StateOnFuelConsumption', '启动燃料消耗', '启停成本'],
                'InitStatus': ['InitStatus', '初始状态'],
                'ChargeEfficiency': ['ChargeEfficiency', '充电效率'],
                'DisEfficiency': ['DisEfficiency', '放电效率'],
                'MaxChargePower': ['MaxChargePower', '最大充电功率'],
                'MaxDisPower': ['MaxDisPower', 'MaxDischargePower', '最大放电功率'],
                'MinSOC': ['MinSOC', '最小SOC'],
                'MaxSOC': ['MaxSOC', '最大SOC'],
                'InitSOC': ['InitSOC', '初始SOC']
            }
            
            # 获取可能的列名列表
            possible_names = param_name_mapping.get(param_name, [param_name])
            
            # 查找实际存在的列名
            actual_param_name = None
            for name in possible_names:
                if name in unit_data:
                    actual_param_name = name
                    break
            
            if actual_param_name is not None:
                value = unit_data[actual_param_name]
                # 处理各种可能的数据类型
                if isinstance(value, (int, float)):
                    return float(value)
                elif isinstance(value, str):
                    # 如果是字符串，尝试转换，但要处理特殊值
                    value = value.strip().lower()
                    if value in ['double', 'float', 'int', 'string', 'text', 'none', '']:
                        logging.warning(f"机组 {unit} 参数 {actual_param_name} 是数据类型描述 '{value}'，使用默认值 {default}")
                        return default
                    else:
                        return float(value)
                elif hasattr(value, 'iloc'):
                    # 如果是pandas Series，取第一个值
                    value = value.iloc[0] if len(value) > 0 else default
                    return float(value)
                else:
                    logging.warning(f"机组 {unit} 参数 {actual_param_name} 类型未知: {type(value)}，使用默认值 {default}")
                    return default
            else:
                logging.warning(f"机组 {unit} 中未找到参数 {param_name} (尝试了列名: {possible_names})，使用默认值 {default}")
                return default
        except (ValueError, TypeError) as e:
            logging.warning(f"机组 {unit} 参数 {param_name} 转换为浮点数失败: {e}，使用默认值 {default}")
    
            return default
    def get_unit_param_as_int(self, unit_dict, unit, param_name, default=0):
        try:
            # 检查单元是否存在
            if unit not in unit_dict:
                logging.warning(f"机组 {unit} 不在字典中，使用默认值 {default}")
                return default
                
            unit_data = unit_dict[unit]
            
            # 定义可能的列名映射
            param_name_mapping = {
                'MinOnLineTime': ['MinOnLineTime', '最小运行时间'],
                'MinOffLineTime': ['MinOffLineTime', '最小停机时间'],
            }
            
            # 获取可能的列名列表
            possible_names = param_name_mapping.get(param_name, [param_name])
            
            # 查找实际存在的列名
            actual_param_name = None
            for name in possible_names:
                if name in unit_data:
                    actual_param_name = name
                    break
            
            if actual_param_name is not None:
                value = unit_data[actual_param_name]
                
                # 检查是否为NaN
                if pd.isna(value):
                    logging.warning(f"机组 {unit} 参数 {actual_param_name} 是 NaN，使用默认值 {default}")
                    return default
                    
                # 处理各种可能的数据类型
                if isinstance(value, (int, float)):
                    return int(float(value))  # 先转为float再转为int
                elif isinstance(value, str):
                    # 如果是字符串，尝试转换，但要处理特殊值
                    value = value.strip().lower()
                    if value in ['double', 'float', 'int', 'string', 'text', 'none', '']:
                        logging.warning(f"机组 {unit} 参数 {actual_param_name} 是数据类型描述 '{value}'，使用默认值 {default}")
                        return default
                    else:
                        try:
                            return int(float(value))
                        except (ValueError, TypeError):
                            logging.warning(f"机组 {unit} 参数 {actual_param_name} 无法转换为整数: '{value}'，使用默认值 {default}")
                            return default
                elif hasattr(value, 'iloc'):
                    # 如果是pandas Series，取第一个值
                    value = value.iloc[0] if len(value) > 0 else default
                    try:
                        return int(float(value))
                    except (ValueError, TypeError):
                        logging.warning(f"机组 {unit} 参数 {actual_param_name} 无法转换为整数: '{value}'，使用默认值 {default}")
                        return default
                else:
                    logging.warning(f"机组 {unit} 参数 {actual_param_name} 类型未知: {type(value)}，使用默认值 {default}")
                    return default
            else:
                logging.warning(f"机组 {unit} 中未找到参数 {param_name} (尝试了列名: {possible_names})，使用默认值 {default}")
                return default
        except (ValueError, TypeError) as e:
            logging.warning(f"机组 {unit} 参数 {param_name} 转换为整数失败: {e}，使用默认值 {default}")
            return default
        except Exception as e:
            logging.error(f"处理机组 {unit} 参数 {param_name} 时发生未知错误: {e}，使用默认值 {default}")
            return default
   # 在 load_data 方法中添加数据清洗代码
    def load_data(self):
        """加载所有Excel数据"""
        logging.info("开始加载数据...")
        
        # 系统汇总信息
        self.system_summary = pd.read_excel(
            self.data_path, sheet_name='SystemSummary', skiprows=0, nrows=1
        ).iloc[0]
        
        # 机组参数
        # 添加数据清洗：去除空值行
        self.thermal_units = pd.read_excel(
            self.data_path, sheet_name='UnitThermalGenerators', skiprows=0
        ).dropna(subset=['ThermalUnitNumber'])
        self.wind_units = pd.read_excel(
            self.data_path, sheet_name='UnitWindGenerators', skiprows=1
        ).dropna(subset=['WTPlantID'])
        self.solar_units = pd.read_excel(
            self.data_path, sheet_name='UnitSolarGenerators', skiprows=1
        ).dropna(subset=['PVPlantID'])
        self.ess_units = pd.read_excel(
            self.data_path, sheet_name='StorElectrochemicalESS', skiprows=1
        ).dropna(subset=['ESSId'])
        self.hydro_units = pd.read_excel(
            self.data_path, sheet_name='UnitRunoffHydroGenerators', skiprows=1
        ).dropna(subset=['RunoffHydroUnitName'])
        
        # 曲线数据
        self.wind_curves = pd.read_excel(
            self.data_path, sheet_name='CurveWindResource', skiprows=1
        ).dropna(subset=['WTCurve'])
        self.solar_curves = pd.read_excel(
            self.data_path, sheet_name='CurveSolarResource', skiprows=1
        ).dropna(subset=['PVCurve'])
        self.hydro_curves = pd.read_excel(
            self.data_path, sheet_name='CurveRunoffHydroResource', skiprows=1
        ).dropna(subset=['RunoffHydroOperationCurve'])
        
        # 处理负荷曲线 - 确保所有值都是数值类型
        load_curve_df = pd.read_excel(
            self.data_path, sheet_name='CurveLoad', skiprows=1
        )
         # 重命名时间列为统一格式 Time_0, Time_1, ...
        time_columns = [col for col in load_curve_df.columns if isinstance(col, int) or (isinstance(col, str) and col.isdigit())]
        rename_dict = {col: f'Time_{col}' for col in time_columns}
        load_curve_df.rename(columns=rename_dict, inplace=True)
        
        if not load_curve_df.empty:
            # 选择第一行作为负荷曲线
            self.load_curve = load_curve_df.iloc[0]
            # 确保所有时间点的值都是浮点数
            for col in self.load_curve.index:
                if col.startswith('Time_') or col.isdigit():
                    try:
                        self.load_curve[col] = float(self.load_curve[col])
                    except (ValueError, TypeError):
                        self.load_curve[col] = 0.0
        else:
            # 创建默认负荷曲线
            self.load_curve = pd.Series([0.8] * 24)
            logging.warning("负荷曲线数据为空，使用默认值")
        for curve_name in ['wind_curves', 'solar_curves', 'hydro_curves']:
            curve_df = getattr(self, curve_name)
            curve_df.rename(columns=rename_dict, inplace=True)
            setattr(self, curve_name, curve_df)
        
        # 网络数据
        self.buses = pd.read_excel(self.data_path, sheet_name='NetBuses', skiprows=1).dropna(subset=['BusId'])
        self.lines = pd.read_excel(self.data_path, sheet_name='NetLines', skiprows=1).dropna(subset=['BranchId'])
        self.transformers = pd.read_excel(self.data_path, sheet_name='NetTransformers', skiprows=1).dropna(subset=['TransformerId'])
        self.sections = pd.read_excel(self.data_path, sheet_name='NetSectionsLines', skiprows=1).dropna(subset=['NetSectionId', 'BranchId'])
        self.section_capacity = pd.read_excel(self.data_path, sheet_name='NetSectionsCapacity', skiprows=1).dropna(subset=['NetSectionId'])
        self.ac_tielines = pd.read_excel(self.data_path, sheet_name='NetHVACTieLine', skiprows=1).dropna(subset=['ACTieLineId'])
        self.ac_tieline_curves = pd.read_excel(self.data_path, sheet_name='EnergyHVACTieLine', skiprows=0).dropna(subset=['资源曲线名称'])
        
        # 负荷数据 - 修复列名问题
        self.loads = pd.read_excel(self.data_path, sheet_name='Loads', skiprows=0)
        # 重命名列以解决中文列名问题
        self.loads = self.loads.rename(columns={
            '负荷编号': 'LoadId',
            '有功比例系数': 'ActivePowerCoef'
        })
        
        # 确保有功比例系数是数值类型
        self.loads['ActivePowerCoef'] = pd.to_numeric(self.loads['ActivePowerCoef'], errors='coerce').fillna(0)
        
        # 设置时间参数
        self.T = int(self.system_summary['Duration'])
        self.dt = 1  # 调度间隔为1小时
        self.hours = list(range(self.T))
        
        # 创建机组字典 - 添加空值检查
        self.thermal_dict = {}
        if not self.thermal_units.empty:
            self.thermal_dict = self.thermal_units.set_index('ThermalUnitNumber').to_dict('index')
        
        self.wind_dict = {}
        if not self.wind_units.empty:
            self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        
        self.solar_dict = {}
        if not self.solar_units.empty:
            self.solar_dict = self.solar_units.set_index('PVPlantID').to_dict('index')
        
        self.ess_dict = {}
        if not self.ess_units.empty:
            self.ess_dict = self.ess_units.set_index('ESSId').to_dict('index')
        
        self.hydro_dict = {}
        if not self.hydro_units.empty:
            self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        
        logging.info("数据加载完成")
    
    def create_model(self):
        """创建Pyomo优化模型"""
        logging.info("开始创建优化模型...")
        
        self.model = pyo.ConcreteModel()
        
        # 定义集合
        self.model.T = pyo.Set(initialize=self.hours)  # 时间段
        self.model.ThermalUnits = pyo.Set(initialize=self.thermal_units['ThermalUnitNumber'].tolist())
        self.model.WindUnits = pyo.Set(initialize=self.wind_units['WTPlantID'].tolist())
        self.model.SolarUnits = pyo.Set(initialize=self.solar_units['PVPlantID'].tolist())
        self.model.ESSUntis = pyo.Set(initialize=self.ess_units['ESSId'].tolist())
        self.model.HydroUnits = pyo.Set(initialize=self.hydro_units['RunoffHydroUnitName'].tolist())
        self.model.Loads = pyo.Set(initialize=self.loads['LoadId'].tolist())
        self.model.Buses = pyo.Set(initialize=self.buses['BusId'].tolist())
        self.model.Lines = pyo.Set(initialize=self.lines['BranchId'].tolist())
        self.model.Sections = pyo.Set(initialize=self.section_capacity['NetSectionId'].tolist())
        
        # 定义变量
        # 火电机组
        self.model.thermal_status = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 机组状态 (0/1)
        self.model.thermal_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )  # 机组出力
        self.model.thermal_startup = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 启动状态
        self.model.thermal_shutdown = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 停机状态
        
        # 深度调峰变量
        self.model.thermal_deep_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )  # 深度调峰功率
        
        # 可再生能源
        self.model.wind_power = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_power = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_power = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 储能系统
        self.model.ess_charge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 充电功率
        self.model.ess_discharge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 放电功率
        self.model.ess_soc = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 储能状态
        
        # 负荷相关
        self.model.load_shed = pyo.Var(
            self.model.Loads, self.model.T, within=pyo.NonNegativeReals
        )  # 切负荷量
        
        # 网络相关
        self.model.bus_angle = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )  # 节点相角
        self.model.line_power = pyo.Var(
            self.model.Lines, self.model.T, within=pyo.Reals
        )  # 线路功率
        
        # 目标函数
        def total_cost_rule(model):
            """系统总成本计算"""
            cost = 0
            
            # 火电运行成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    # 二次成本函数线性化近似
                    a = self.get_unit_param_as_float(self.thermal_dict, unit, 'OperationCoeff_A')
                    b = self.get_unit_param_as_float(self.thermal_dict, unit, 'OperationCoeff_B')
                    c = self.get_unit_param_as_float(self.thermal_dict, unit, 'OperationCoeff_C')
                    
                    # 成本函数：a*P^2 + b*P + c
                    cost += a * model.thermal_power[unit, t]**2 + b * model.thermal_power[unit, t] + c
                    
                    # 深度调峰成本（如果出力低于阈值）
                    threshold = self.get_unit_param_as_float(self.thermal_dict, unit, 'DeepPeakShavingThreshold', 0.0)
                    deep_cost = self.get_unit_param_as_float(self.thermal_dict, unit, 'DeepPeakShavingCost', 0.0)
                    min_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'MinPower', 0.0)
                    
                    # 使用线性化方法代替 max 函数
                    if threshold > 0 and min_power > 0:
                        cost += deep_cost * model.thermal_deep_power[unit, t]
            
            # 启停成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    startup_cost = self.get_unit_param_as_float(self.thermal_dict, unit, 'StateOnFuelConsumption')
                    cost += startup_cost * model.thermal_startup[unit, t]
            
            # 切负荷惩罚
            penalty_cut = 10000  # 元/MWh
            for load in model.Loads:
                for t in model.T:
                    cost += penalty_cut * model.load_shed[load, t]
            
            return cost
        
        self.model.total_cost = pyo.Objective(rule=total_cost_rule, sense=pyo.minimize)
        
        # 约束条件
        # 1. 火电机组逻辑约束
        def thermal_logic_rule(model, unit, t):
            """机组状态逻辑约束"""
            if t == 0:
                # 初始状态
                init_status = self.thermal_dict[unit].get('InitStatus', False)
                if isinstance(init_status, str):
                    init_status = init_status.strip().lower() in ['true', '1', 'yes']
                return model.thermal_status[unit, t] == (1 if init_status else 0)
            else:
                # 状态变化约束
                return model.thermal_status[unit, t] - model.thermal_status[unit, t-1] == \
                    model.thermal_startup[unit, t] - model.thermal_shutdown[unit, t]
        self.model.thermal_logic_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_logic_rule
        )
        
            # 2. 火电机组出力约束
        def thermal_power_lower_rule(model, unit, t):
            """机组出力下限约束"""
            min_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'MinPower', 0.0)
            # 使用 Big M 方法表示当机组停机时出力必须为0
            return model.thermal_power[unit, t] >= min_power * model.thermal_status[unit, t]

        def thermal_power_upper_rule(model, unit, t):
            """机组出力上限约束"""
            max_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'MaxPower', 0.0)
            return model.thermal_power[unit, t] <= max_power * model.thermal_status[unit, t]

        self.model.thermal_power_lower_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_lower_rule
        )
        self.model.thermal_power_upper_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_upper_rule
        )
        
        # 3. 深度调峰约束
        def deep_power_rule(model, unit, t):
            """深度调峰功率约束"""
            threshold = self.get_unit_param_as_float(self.thermal_dict, unit, 'DeepPeakShavingThreshold', 0.0)
            min_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'MinPower', 0.0)
            
            if threshold > 0 and min_power > 0:
                deep_threshold = min_power * threshold
                return model.thermal_deep_power[unit, t] >= deep_threshold - model.thermal_power[unit, t]
            else:
                return model.thermal_deep_power[unit, t] == 0  # 如果没有阈值，深度调峰功率为0
        self.model.deep_power_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=deep_power_rule
        )
        
        # 4. 爬坡约束
        def thermal_ramp_up_rule(model, unit, t):
            """上爬坡约束"""
            if t == 0:
                return pyo.Constraint.Skip
            ramp_up = self.get_unit_param_as_float(self.thermal_dict, unit, 'UpRamppingRate', 0.0)
            capacity = self.get_unit_param_as_float(self.thermal_dict, unit, 'Capacity', 0.0)
            return model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= \
                ramp_up + model.thermal_startup[unit, t] * capacity
        self.model.thermal_ramp_up_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_up_rule
        )
        
        def thermal_ramp_down_rule(model, unit, t):
            """下爬坡约束"""
            if t == 0:
                return pyo.Constraint.Skip
            ramp_down = self.get_unit_param_as_float(self.thermal_dict, unit, 'DownRamppingRate', 0.0)
            capacity = self.get_unit_param_as_float(self.thermal_dict, unit, 'Capacity', 0.0)
            return model.thermal_power[unit, t-1] - model.thermal_power[unit, t] <= \
                ramp_down + model.thermal_shutdown[unit, t] * capacity
        self.model.thermal_ramp_down_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_down_rule
        )
        
        # 5. 最短开停机时间约束
        def min_up_time_rule(model, unit, t):
            """最小运行时间约束"""
            min_up = self.get_unit_param_as_int(self.thermal_dict, unit, 'MinOnLineTime', 0)
            if t < min_up:
                return pyo.Constraint.Skip
            
            # 确保所有值都是整数
            min_up = int(min_up)
            start_time = int(max(0, t - min_up + 1))
            end_time = int(t + 1)
            
            return sum(model.thermal_shutdown[unit, k] for k in range(start_time, end_time)) <= \
                1 - model.thermal_status[unit, t]
        self.model.min_up_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_up_time_rule
        )

        def min_down_time_rule(model, unit, t):
            """最小停机时间约束"""
            min_down = self.get_unit_param_as_int(self.thermal_dict, unit, 'MinOffLineTime', 0)
            if t < min_down:
                return pyo.Constraint.Skip
            
            # 确保所有值都是整数
            min_down = int(min_down)
            start_time = int(max(0, t - min_down + 1))
            end_time = int(t + 1)
            
            return sum(model.thermal_startup[unit, k] for k in range(start_time, end_time)) <= \
                model.thermal_status[unit, t]
        self.model.min_down_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_down_time_rule
        )
        
        # 6. 可再生能源出力约束
        def wind_power_rule(model, unit, t):
            """风电出力约束"""
            curve_name = self.wind_dict[unit]['ResourceCurve']
            capacity = self.get_unit_param_as_float(self.wind_dict, unit, 'Capacity', 0.0)
            
            # 从风电曲线中获取预测值
            wind_curve = self.wind_curves[self.wind_curves['WTCurve'] == curve_name].iloc[0]
            max_power = wind_curve[f'Time_{t}'] * capacity
            
            return model.wind_power[unit, t] <= max_power
        self.model.wind_power_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_power_rule
        )
        
        # 7. 储能系统约束
        def ess_charge_power_rule(model, unit, t):
            """储能充电功率约束"""
            max_charge = self.get_unit_param_as_float(self.ess_dict, unit, 'MaxChargePower', 0.0)
            return model.ess_charge[unit, t] <= max_charge
        self.model.ess_charge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_charge_power_rule
        )

        def ess_discharge_power_rule(model, unit, t):
            """储能放电功率约束"""
            max_discharge = self.get_unit_param_as_float(self.ess_dict, unit, 'MaxDisPower', 0.0)
            return model.ess_discharge[unit, t] <= max_discharge
        self.model.ess_discharge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_discharge_power_rule
        )

        def ess_soc_rule(model, unit, t):
            """储能SOC约束"""
            min_soc = self.get_unit_param_as_float(self.ess_dict, unit, 'MinSOC', 0.0)
            max_soc = self.get_unit_param_as_float(self.ess_dict, unit, 'MaxSOC', 1.0)
            capacity = self.get_unit_param_as_float(self.ess_dict, unit, 'Capacity', 0.0)
            
            return (min_soc * capacity, model.ess_soc[unit, t], max_soc * capacity)
        self.model.ess_soc_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_rule
        )
        
        # 8. 系统功率平衡约束
        def power_balance_rule(model, t):
            """系统功率平衡约束"""
            # 总发电
            thermal_gen = sum(model.thermal_power[unit, t] for unit in model.ThermalUnits)
            wind_gen = sum(model.wind_power[unit, t] for unit in model.WindUnits)
            solar_gen = sum(model.solar_power[unit, t] for unit in model.SolarUnits)
            hydro_gen = sum(model.hydro_power[unit, t] for unit in model.HydroUnits)
            ess_discharge = sum(model.ess_discharge[unit, t] for unit in model.ESSUntis)
            ess_charge = sum(model.ess_charge[unit, t] for unit in model.ESSUntis)
            
            # 总负荷
            total_load = 0
            for load in model.Loads:
                # 获取该负荷的有功比例系数
                active_coef_df = self.loads.loc[self.loads['LoadId'] == load, 'ActivePowerCoef']
                if not active_coef_df.empty:
                    active_coef = active_coef_df.values[0]
                else:
                    active_coef = 0
                
                # 获取负荷曲线值 - 确保是数值类型
                time_key = f'Time_{t}'
                if time_key in self.load_curve:
                    load_curve_val = self.load_curve[time_key]
                else:
                    # 尝试其他可能的键名
                    try:
                        load_curve_val = self.load_curve[t]
                    except KeyError:
                        # 如果都找不到，使用默认值
                        load_curve_val = 0.0
                
                # 确保是浮点数
                try:
                    load_curve_val = float(load_curve_val)
                except (ValueError, TypeError):
                    load_curve_val = 0.0
                
                # 计算实际负荷
                total_load += active_coef * load_curve_val
            
            # 功率平衡
            return thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge == \
                total_load - sum(model.load_shed[load, t] for load in model.Loads) + ess_charge
        self.model.power_balance_con = pyo.Constraint(
            self.model.T, rule=power_balance_rule
        )
        
        # 9. 备用约束
        def reserve_requirement_rule(model, t):
            """旋转备用约束"""
            reserve_req = 0.1  # 10%的备用要求
            
            # 总负荷
            total_load = 0
            for load in model.Loads:
                active_coef = self.loads.loc[self.loads['LoadId'] == load, 'ActivePowerCoef'].values[0]
                load_curve_val = self.load_curve[f'Time_{t}']
                total_load += active_coef * load_curve_val
            
            # 可用备用
            available_reserve = 0
            for unit in model.ThermalUnits:
                max_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'Capacity', 0.0)
                available_reserve += model.thermal_status[unit, t] * max_power - model.thermal_power[unit, t]
            
            return available_reserve >= reserve_req * total_load
        self.model.reserve_con = pyo.Constraint(
            self.model.T, rule=reserve_requirement_rule
        )
        
        logging.info("优化模型创建完成")
    
    def solve(self):
        """求解优化问题"""
        logging.info("开始求解优化问题...")
        
        # 使用Gurobi求解器
        try:
            solver = SolverFactory('gurobi')
            if solver is None:
                logging.error("未找到Gurobi求解器，尝试使用其他求解器")
                solver = SolverFactory('glpk')
                if solver is None:
                    logging.error("未找到任何可用求解器")
                    return False
        except Exception as e:
            logging.error(f"求解器初始化失败: {str(e)}")
            return False
        
        solver.options['MIPGap'] = 0.01  # 设置MIP gap为1%
        solver.options['TimeLimit'] = 600  # 设置10分钟时间限制
        results = solver.solve(self.model, tee=True)
        
        if (results.solver.status == pyo.SolverStatus.ok and 
            results.solver.termination_condition == pyo.TerminationCondition.optimal):
            logging.info("优化求解成功完成")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            logging.warning("达到时间限制，返回当前最优解")
            return True
        else:
            logging.error(f"优化求解失败: {results.solver.termination_condition}")
            return False
    
    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")
        
        # 创建Excel写入器
        try:
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
        except Exception as e:
            logging.error(f"创建Excel写入器失败: {str(e)}")
            return None
        
        try:
            # 1. 火电机组结果
            thermal_results = []
            for unit in self.model.ThermalUnits:
                for t in self.model.T:
                    thermal_results.append({
                        'Unit': unit,
                        'Time': t,
                        'Status': pyo.value(self.model.thermal_status[unit, t]),
                        'Power': pyo.value(self.model.thermal_power[unit, t]),
                        'DeepPower': pyo.value(self.model.thermal_deep_power[unit, t]),
                        'Startup': pyo.value(self.model.thermal_startup[unit, t]),
                        'Shutdown': pyo.value(self.model.thermal_shutdown[unit, t])
                    })
            thermal_df = pd.DataFrame(thermal_results)
            thermal_df.to_excel(writer, sheet_name='火电机组结果', index=False)
            
            # 2. 可再生能源结果
            wind_results = []
            for unit in self.model.WindUnits:
                for t in self.model.T:
                    wind_results.append({
                        'Unit': unit,
                        'Time': t,
                        'Power': pyo.value(self.model.wind_power[unit, t])
                    })
            wind_df = pd.DataFrame(wind_results)
            wind_df.to_excel(writer, sheet_name='风电结果', index=False)
            
            # 3. 光伏结果
            solar_results = []
            for unit in self.model.SolarUnits:
                for t in self.model.T:
                    solar_results.append({
                        'Unit': unit,
                        'Time': t,
                        'Power': pyo.value(self.model.solar_power[unit, t])
                    })
            solar_df = pd.DataFrame(solar_results)
            solar_df.to_excel(writer, sheet_name='光伏结果', index=False)
            
            # 4. 水电结果
            hydro_results = []
            for unit in self.model.HydroUnits:
                for t in self.model.T:
                    hydro_results.append({
                        'Unit': unit,
                        'Time': t,
                        'Power': pyo.value(self.model.hydro_power[unit, t])
                    })
            hydro_df = pd.DataFrame(hydro_results)
            hydro_df.to_excel(writer, sheet_name='水电结果', index=False)
            
            # 5. 储能结果
            ess_results = []
            for unit in self.model.ESSUntis:
                for t in self.model.T:
                    ess_results.append({
                        'Unit': unit,
                        'Time': t,
                        'Charge': pyo.value(self.model.ess_charge[unit, t]),
                        'Discharge': pyo.value(self.model.ess_discharge[unit, t]),
                        'SOC': pyo.value(self.model.ess_soc[unit, t])
                    })
            ess_df = pd.DataFrame(ess_results)
            ess_df.to_excel(writer, sheet_name='储能结果', index=False)
            
            # 6. 切负荷结果
            load_shed_results = []
            for load in self.model.Loads:
                for t in self.model.T:
                    load_shed_results.append({
                        'Load': load,
                        'Time': t,
                        'LoadShed': pyo.value(self.model.load_shed[load, t])
                    })
            load_shed_df = pd.DataFrame(load_shed_results)
            load_shed_df.to_excel(writer, sheet_name='切负荷结果', index=False)
            
            # 7. 系统总结
            total_cost = pyo.value(self.model.total_cost)
            thermal_power_total = thermal_df['Power'].sum()
            wind_power_total = wind_df['Power'].sum()
            solar_power_total = solar_df['Power'].sum()
            hydro_power_total = hydro_df['Power'].sum()
            load_shed_total = load_shed_df['LoadShed'].sum()
            
            summary_data = {
                '总成本': [total_cost],
                '总火电发电量': [thermal_power_total],
                '总风电发电量': [wind_power_total],
                '总光伏发电量': [solar_power_total],
                '总水电发电量': [hydro_power_total],
                '总切负荷量': [load_shed_total]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='系统总结', index=False)
            
            # 保存Excel
            writer.save()
            
            logging.info(f"结果已保存到 {output_path}")
            
            # 返回结果文件路径
            return output_path
        except Exception as e:
            logging.error(f"保存结果时出错: {str(e)}")
            return None
        
    def plot_unit_status(self, output_path):
        """绘制机组开关机状态图"""
        logging.info("开始绘制机组开关机状态图...")
        
        try:
            # 获取火电机组结果
            thermal_results = []
            for unit in self.model.ThermalUnits:
                for t in self.model.T:
                    thermal_results.append({
                        'Unit': unit,
                        'Time': t,
                        'Status': pyo.value(self.model.thermal_status[unit, t]),
                        'Power': pyo.value(self.model.thermal_power[unit, t])
                    })
            thermal_df = pd.DataFrame(thermal_results)
            
            # 按机组分组
            grouped = thermal_df.groupby('Unit')
            
            # 创建图表
            fig, axes = plt.subplots(len(grouped), 1, figsize=(15, 10), sharex=True)
            fig.suptitle('火电机组开关机状态及出力情况', fontsize=16)
            
            if len(grouped) == 1:
                axes = [axes]
            
            for i, (unit, group) in enumerate(grouped):
                ax = axes[i]
                
                # 绘制状态（0/1）
                ax.step(group['Time'], group['Status'], 'r-', where='post', label='状态')
                ax.set_ylabel('状态', color='r')
                ax.tick_params(axis='y', labelcolor='r')
                ax.set_ylim(-0.1, 1.1)
                ax.grid(True, linestyle='--', alpha=0.7)
                
                # 创建双Y轴
                ax2 = ax.twinx()
                
                # 绘制出力
                ax2.plot(group['Time'], group['Power'], 'b-', label='出力')
                ax2.set_ylabel('出力 (MW)', color='b')
                ax2.tick_params(axis='y', labelcolor='b')
                
                # 设置标题
                ax.set_title(f'机组 {unit}')
                
                # 添加图例
                lines, labels = ax.get_legend_handles_labels()
                lines2, labels2 = ax2.get_legend_handles_labels()
                ax.legend(lines + lines2, labels + labels2, loc='upper right')
            
            plt.xlabel('时间 (小时)')
            plt.tight_layout(rect=[0, 0, 1, 0.95])  # 为总标题留出空间
            
            # 保存图表
            plt.savefig(output_path)
            logging.info(f"机组状态图已保存到 {output_path}")
            plt.close()
            
            return output_path
        except Exception as e:
            logging.error(f"绘制机组状态图时出错: {str(e)}")
            return None
    
    def plot_generation_mix(self, output_path):
        """绘制各类电源出力情况"""
        logging.info("开始绘制电源结构图...")
        
        try:
            # 初始化结果数组
            thermal_gen = np.zeros(self.T)
            wind_gen = np.zeros(self.T)
            solar_gen = np.zeros(self.T)
            hydro_gen = np.zeros(self.T)
            ess_discharge = np.zeros(self.T)
            
            # 计算总发电量
            for t in range(self.T):
                for unit in self.model.ThermalUnits:
                    thermal_gen[t] += pyo.value(self.model.thermal_power[unit, t])
                for unit in self.model.WindUnits:
                    wind_gen[t] += pyo.value(self.model.wind_power[unit, t])
                for unit in self.model.SolarUnits:
                    solar_gen[t] += pyo.value(self.model.solar_power[unit, t])
                for unit in self.model.HydroUnits:
                    hydro_gen[t] += pyo.value(self.model.hydro_power[unit, t])
                for unit in self.model.ESSUntis:
                    ess_discharge[t] += pyo.value(self.model.ess_discharge[unit, t])
            
            # 创建图表
            plt.figure(figsize=(12, 6))
            
            # 堆叠面积图
            plt.stackplot(range(self.T), thermal_gen, wind_gen, solar_gen, hydro_gen, ess_discharge,
                          labels=['火电', '风电', '光伏', '水电', '储能放电'])
            
            # 添加总负荷曲线
            total_load = []
            for t in range(self.T):
                load = 0
                for load_id in self.model.Loads:
                    active_coef = self.loads.loc[self.loads['LoadId'] == load_id, 'ActivePowerCoef'].values[0]
                    load_curve_val = self.load_curve[f'Time_{t}']
                    load += active_coef * load_curve_val
                total_load.append(load)
            
            plt.plot(range(self.T), total_load, 'k--', linewidth=2, label='总负荷')
            
            # 设置图表属性
            plt.title('各类电源出力及负荷曲线')
            plt.xlabel('时间 (小时)')
            plt.ylabel('功率 (MW)')
            plt.legend(loc='upper left')
            plt.grid(True, linestyle='--', alpha=0.7)
            plt.xlim(0, self.T-1)
            
            # 保存图表
            plt.savefig(output_path)
            logging.info(f"电源结构图已保存到 {output_path}")
            plt.close()
            
            return output_path
        except Exception as e:
            logging.error(f"绘制电源结构图时出错: {str(e)}")
            return None

# 主程序
if __name__ == "__main__":
    # 输入输出路径
    input_file = r"C:\Users\admin\Desktop\Rots.xlsx"
    output_file = "UC_Results.xlsx"
    status_plot = "Unit_Status.png"
    mix_plot = "Generation_Mix.png"
    
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误：找不到输入文件 {input_file}")
        print("请确保文件存在于当前目录或提供完整路径")
        exit(1)

    try:
        # 创建模型
        uc_model = UC_Model(input_file)
        
        # 求解模型
        if uc_model.solve():
            # 保存结果
            result_path = uc_model.save_results(output_file)
            
            if result_path:
                # 绘制图表
                status_path = uc_model.plot_unit_status(status_plot)
                mix_path = uc_model.plot_generation_mix(mix_plot)
                
                if status_path and mix_path:
                    # 将图表插入Excel
                    try:
                        wb = openpyxl.load_workbook(output_file)
                        
                        # 添加机组状态图
                        ws = wb.create_sheet("机组状态图")
                        img = Image(status_path)
                        ws.add_image(img, 'A1')
                        
                        # 添加电源结构图
                        ws = wb.create_sheet("电源结构图")
                        img = Image(mix_path)
                        ws.add_image(img, 'A1')
                        
                        # 保存最终结果
                        wb.save(output_file)
                        
                        logging.info("所有结果已保存并整合到Excel文件中")
                        print(f"优化完成，结果已保存到: {output_file}")
                    except Exception as e:
                        logging.error(f"将图表插入Excel时出错: {str(e)}")
                        print(f"结果已保存到 {output_file}，但插入图表失败")
                else:
                    print(f"结果已保存到 {output_file}，但图表生成失败")
            else:
                print("结果保存失败")
        else:
            logging.error("优化求解失败，请检查模型和输入数据")
            print("优化求解失败，请检查日志获取详细信息")
    except Exception as e:
        logging.exception("程序运行出错")
        print(f"程序运行出错: {str(e)}")
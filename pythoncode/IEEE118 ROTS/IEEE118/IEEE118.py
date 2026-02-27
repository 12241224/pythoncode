import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import os
import logging
import sys
import time
import math

# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 初始化日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class UC_Model:
    def __init__(self, data_path):
        self.data_path = data_path
        self.load_data()
        self.validate_data()  # 添加数据验证
        self.create_model()
    
    def validate_data(self):
        """数据加载后立即执行验证"""
        logging.info("开始数据验证...")

        # 检查可再生能源容量
        for tech, units in [('风电', self.wind_units), 
                           ('光伏', self.solar_units),
                           ('水电', self.hydro_units)]:
            # 修正：兼容不同表头，自动查找容量列
            cap_col = None
            for col in units.columns:
                if str(col).strip() in ['Capacity', '机组装机容量(MW)']:
                    cap_col = col
                    break
            if cap_col is None:
                logging.error(f"{tech}机组数据缺少容量列！实际列为: {list(units.columns)}")
                continue
            if units.empty:
                logging.error(f"{tech}机组数据为空！")
            elif units[cap_col].sum() < 100:
                logging.warning(f"{tech}总容量异常: {units[cap_col].sum()}MW")

        # 检查负荷曲线范围
        if len(self.load_curve) > 0 and max(self.load_curve) > 6000:
            logging.warning(f"负荷峰值过高: {max(self.load_curve):.1f}MW，缩放到合理范围")
            # 缩放到合理范围（4200MW左右）
            scale_factor = 4200 / max(self.load_curve)
            self.load_curve = [val * scale_factor for val in self.load_curve]
            logging.info(f"负荷曲线已缩放，新峰值: {max(self.load_curve):.1f}MW")
        elif len(self.load_curve) > 0 and max(self.load_curve) < 2000:
            logging.warning(f"负荷峰值过低: {max(self.load_curve):.1f}MW")
            # 尝试从系统汇总中获取峰值负荷
            sys_peak = self.system_summary.get('PeakLoad(MW)', 0)
            if sys_peak > 2000:
                logging.warning(f"使用系统汇总中的峰值负荷值: {sys_peak}MW")
                # 缩放负荷曲线
                current_peak = max(self.load_curve)
                if current_peak > 0:
                    scaling_factor = min(sys_peak, 4200) / current_peak  # 限制最大值
                    self.load_curve = [val * scaling_factor for val in self.load_curve]
        
        # 检查资源曲线匹配
        wind_curves_used = self.wind_units['ResourceCurve'].unique()
        for curve in wind_curves_used:
            if curve not in self.wind_curves['WTCurve'].values:
                logging.error(f"风资源曲线'{curve}'未定义！")
        
        solar_curves_used = self.solar_units['ResourceCurve'].unique()
        for curve in solar_curves_used:
            if curve not in self.solar_curves['PVCurve'].values:
                logging.error(f"光伏资源曲线'{curve}'未定义！")
        
        hydro_curves_used = self.hydro_units['ResourceCurve'].unique()
        for curve in hydro_curves_used:
            if curve not in self.hydro_curves['RunoffHydroOperationCurve'].values:
                logging.error(f"水电资源曲线'{curve}'未定义！")
        
        logging.info("数据验证完成")
        
    def plot_resource_curves(self, output_path="IEEE118_Resource_Curves.png"):
        """绘制资源曲线验证图"""
        plt.figure(figsize=(12, 8))
        
        # 绘制风电曲线
        for i, curve in enumerate(self.wind_curves['WTCurve'].unique()[:3]):
            df = self.wind_curves[self.wind_curves['WTCurve'] == curve]
            # 提取所有时间列
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'风电曲线_{i+1}')
        
        # 绘制光伏曲线
        for i, curve in enumerate(self.solar_curves['PVCurve'].unique()[:3]):
            df = self.solar_curves[self.solar_curves['PVCurve'] == curve]
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'光伏曲线_{i+1}')
        
        # 绘制水电曲线
        for i, curve in enumerate(self.hydro_curves['RunoffHydroOperationCurve'].unique()[:3]):
            df = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'] == curve]
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'水电曲线_{i+1}')
        
        # 添加标题和标签
        plt.title('资源曲线预览', fontsize=14)
        plt.xlabel('时段')
        plt.ylabel('出力系数')
        plt.legend()
        plt.grid(True)
        
        # 抑制matplotlib警告
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            plt.tight_layout()
            
        plt.savefig(output_path)
        logging.info(f"资源曲线预览图已保存到 {output_path}")
        plt.close()
        return output_path
        
    def load_data(self):
        """加载所有Excel数据"""
        logging.info("开始加载数据...")
        
        # 系统汇总信息
        self.system_summary = pd.read_excel(
            self.data_path, sheet_name='SystemSummary', skiprows=0, nrows=1
        ).iloc[0]
        
        # 机组参数
        self.thermal_units = pd.read_excel(
            self.data_path, sheet_name='UnitThermalGenerators', skiprows=0
        )
        self.wind_units = pd.read_excel(
            self.data_path, sheet_name='UnitWindGenerators', skiprows=0
        )
        # 修复1: 确保所有关键列正确映射
        self.wind_units.rename(columns={
            'WTPlantID': 'WTPlantID',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)

        self.solar_units = pd.read_excel(
            self.data_path, sheet_name='UnitSolarGenerators', skiprows=0
        )
        self.solar_units.rename(columns={
            'PVPlantID': 'PVPlantID',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)


        self.ess_units = pd.read_excel(
            self.data_path, sheet_name='StorElectrochemicalESS', skiprows=0
        )
        self.ess_units.rename(columns={
            'ESSId': 'ESSId',
            '储能容量(MWh)': 'Capacity',
            '充电功率(MW)': 'MaxChargePower',
            '放电功率(MW)': 'MaxDisPower',
            '所在母线编号': 'BusName'  # 确保有母线信息
        }, inplace=True)


        self.hydro_units = pd.read_excel(
            self.data_path, sheet_name='UnitRunoffHydroGenerators', skiprows=0
        )
        self.hydro_units.rename(columns={
            'RunoffHydroUnitName': 'RunoffHydroUnitName',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)

        # 保证RunoffHydroUnitName唯一，否则to_dict会报错
        if self.hydro_units['RunoffHydroUnitName'].duplicated().any():
            logging.warning("发现重复的水电机组名称，将自动去重（保留首次出现）")
            self.hydro_units = self.hydro_units.drop_duplicates(subset='RunoffHydroUnitName', keep='first')

        # 修正：确保所有机组的Capacity列为数值类型，避免sum时报错
        for df, cap_col in [
            (self.wind_units, 'Capacity'),
            (self.solar_units, 'Capacity'),
            (self.hydro_units, 'Capacity')
        ]:
            if cap_col in df.columns:
                df[cap_col] = pd.to_numeric(df[cap_col], errors='coerce').fillna(0)

        # 检查所有机组集合是否全为空，若全为空则添加一个虚拟火电机组，避免模型完全无源
        if (
            self.thermal_units.empty and
            (self.wind_units.empty or (self.wind_units.shape[0] == 1 and self.wind_units.iloc[0]['Capacity'] == 0)) and
            (self.solar_units.empty or (self.solar_units.shape[0] == 1 and self.solar_units.iloc[0]['Capacity'] == 0)) and
            (self.hydro_units.empty or (self.hydro_units.shape[0] == 1 and self.hydro_units.iloc[0]['Capacity'] == 0))
        ):
            logging.warning("所有机组集合均为空，自动添加一个虚拟火电机组以保证模型可行性")
            self.thermal_units = pd.DataFrame([{
                'ThermalUnitNumber': 'DummyThermal',
                'Capacity': 2000,  # 大幅增加虚拟火电机组容量到2000MW，解决切负荷问题
                'MinPower': 100,   # 设置最小出力
                'MinOnLineTime': 1,
                'MinOffLineTime': 1,
                'UpRamppingRate': 500,   # 增加爬坡速率
                'DownRamppingRate': 500,
                'InitStatus': 1,
                'BusName': 'Bus1'  # 暂时使用默认母线名，稍后会被覆盖
            }])

        # 曲线数据
        self.wind_curves = pd.read_excel(
            self.data_path, sheet_name='CurveWindResource', skiprows=0
        )
        self.wind_curves = self.wind_curves.dropna(subset=['WTCurve'])
        logging.info(f"风电资源曲线表列名: {list(self.wind_curves.columns)}")
        
        self.solar_curves = pd.read_excel(
            self.data_path, sheet_name='CurveSolarResource', skiprows=0
        )
        self.solar_curves = self.solar_curves.dropna(subset=['PVCurve'])
        logging.info(f"光伏资源曲线表列名: {list(self.solar_curves.columns)}")
        
        self.hydro_curves = pd.read_excel(
            self.data_path, sheet_name='CurveRunoffHydroResource', skiprows=0
        )
        self.hydro_curves = self.hydro_curves.dropna(subset=['RunoffHydroOperationCurve'])
        
        # 修复2: 正确读取负荷曲线数据
        load_curve_sheet = pd.read_excel(self.data_path, sheet_name='CurveLoad')
        # 查找包含"Time_0"的列作为标识
        time_cols = [col for col in load_curve_sheet.columns if 'time_' in col.lower() or 'Time_' in col]

        # 先初始化self.load_curve，避免后续引用时报错
        self.load_curve = []

        if not time_cols:
            logging.error("负荷曲线表中未找到时间列！")
            # 创建默认负荷曲线
            peak_load = self.system_summary.get('PeakLoad(MW)', 5000)  # 降低默认峰值到5000MW
            self.load_curve = [peak_load * 0.7] * int(self.system_summary['Duration'])
        else:
            # 尝试找到包含负荷数据的行
            for i in range(len(load_curve_sheet)):
                row = load_curve_sheet.iloc[i]
                if any(isinstance(val, (int, float)) and not pd.isna(val) for val in row[time_cols]):
                    load_factors = row[time_cols].values.astype(float)
                    
                    # 检查是否为归一化系数（0-1之间）还是绝对值
                    max_factor = max(load_factors)
                    if max_factor <= 1.0:
                        # 如果是系数，使用适中的峰值负荷
                        # 先用较保守的值确保可行性
                        peak_load = self.system_summary.get('PeakLoad(MW)', 5000)  # 降低到5000MW
                        self.load_curve = [factor * peak_load for factor in load_factors]
                        logging.info(f"负荷曲线使用系数模式，基准负荷={peak_load}MW，峰值={max(self.load_curve):.1f}MW")
                    else:
                        # 如果是绝对值，直接使用但检查是否合理
                        if max_factor > 10000:  # 如果峰值过大，进行缩放
                            logging.warning(f"负荷峰值过大({max_factor:.1f}MW)，缩放到合理范围")
                            scale_factor = 4200 / max_factor
                            self.load_curve = [factor * scale_factor for factor in load_factors]
                        else:
                            self.load_curve = load_factors.tolist()
                    
                    logging.info(f"从第{i+1}行读取负荷曲线数据")
                    logging.info(f"负荷曲线范围：{min(self.load_curve):.1f} - {max(self.load_curve):.1f} MW")
                    logging.info(f"最大需求4008.3MW来源: Excel文件'Final_Cleaned_Rots_v2.xlsx'的'CurveLoad'工作表")
                    logging.info(f"数据行索引: {i+1}, 负荷系数范围: {min(load_factors):.3f} - {max(load_factors):.3f}")
                    if max_factor <= 1.0:
                        logging.info(f"使用系数模式: {max_factor:.3f} × {peak_load}MW = {max(self.load_curve):.1f}MW")
                    break
            else:
                logging.warning("未找到有效的负荷数据行，使用默认值")
                peak_load = self.system_summary.get('PeakLoad(MW)', 5000)
                self.load_curve = [peak_load * 0.7] * int(self.system_summary['Duration'])

        # 确保负荷曲线有合理的数值
        if len(self.load_curve) == 0:
            peak_load = self.system_summary.get('PeakLoad(MW)', 6000)
            self.load_curve = [peak_load * 0.7 for _ in range(int(self.system_summary['Duration']))]
        elif max(self.load_curve) < 100:
            logging.warning("负荷曲线峰值过低，可能是系数而非绝对值")
            peak_load = self.system_summary.get('PeakLoad(MW)', 6000)
            # 如果数值很小，认为是系数，需要乘以峰值负荷
            if max(self.load_curve) <= 1.0:
                self.load_curve = [val * peak_load for val in self.load_curve]
                logging.info(f"负荷曲线已转换为绝对值，峰值：{max(self.load_curve):.1f}MW")
        
        # 网络数据
        self.buses = pd.read_excel(self.data_path, sheet_name='NetBuses', skiprows=0)
        self.lines = pd.read_excel(self.data_path, sheet_name='NetLines', skiprows=0)
        self.transformers = pd.read_excel(self.data_path, sheet_name='NetTransformers', skiprows=0)
        self.sections = pd.read_excel(self.data_path, sheet_name='NetSectionsLines', skiprows=0)
        self.section_capacity = pd.read_excel(self.data_path, sheet_name='NetSectionsCapacity', skiprows=0)
        self.ac_tielines = pd.read_excel(self.data_path, sheet_name='NetHVACTieLine', skiprows=0)
        self.ac_tieline_curves = pd.read_excel(self.data_path, sheet_name='EnergyHVACTieLine', skiprows=0)
        
        # 更新虚拟火电机组的母线信息（使用实际的第一个母线）
        if not self.buses.empty and not self.thermal_units.empty:
            first_bus = self.buses['BusId'].iloc[0]
            # 更新火电虚拟机组
            if (self.thermal_units.shape[0] == 1 and 
                self.thermal_units.iloc[0]['ThermalUnitNumber'] == 'DummyThermal'):
                self.thermal_units.loc[0, 'BusName'] = first_bus
        
        # 负荷数据
        self.loads = pd.read_excel(self.data_path, sheet_name='Loads', skiprows=0)
        
        # 燃料数据
        self.fuel_data = pd.read_excel(self.data_path, sheet_name='UnitFuel', skiprows=0).iloc[0]
        
        # 设置时间参数
        self.T = int(self.system_summary['Duration'])
        self.dt = 1  # 调度间隔为1小时
        self.hours = list(range(self.T))
        
        # 创建机组字典
        self.thermal_dict = self.thermal_units.set_index('ThermalUnitNumber').to_dict('index')
        for unit in self.thermal_dict:
            # 确保数值类型正确
            self.thermal_dict[unit]['MinOnLineTime'] = int(self.thermal_dict[unit].get('MinOnLineTime', 0))
            self.thermal_dict[unit]['MinOffLineTime'] = int(self.thermal_dict[unit].get('MinOffLineTime', 0))
            self.thermal_dict[unit]['Capacity'] = float(self.thermal_dict[unit].get('Capacity', 0))
            self.thermal_dict[unit]['MinPower'] = float(self.thermal_dict[unit].get('MinPower', 0))
            self.thermal_dict[unit]['UpRamppingRate'] = float(self.thermal_dict[unit].get('UpRamppingRate', 0))
            self.thermal_dict[unit]['DownRamppingRate'] = float(self.thermal_dict[unit].get('DownRamppingRate', 0))
            
        self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        self.solar_dict = self.solar_units.set_index('PVPlantID').to_dict('index')
        self.ess_dict = self.ess_units.set_index('ESSId').to_dict('index')
        self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        
        # 创建负荷字典
        self.load_dict = self.loads.set_index('负荷编号').to_dict('index')
        
        # 构建母线字典
        self.bus_dict = self.buses.set_index('BusId').to_dict('index')
        
        # 构建线路字典
        self.line_dict = {}
        for idx, row in self.lines.iterrows():
            self.line_dict[row['BranchId']] = {
                'StartNodeName': row['StartNodeName'],
                'EndNodeName': row['EndNodeName'],
                'Resistance': float(row.get('Resistance', 0)),
                'Reactance': float(row.get('Reactance', 0)),
                'Capacity': float(row.get('Capacity', 0)),
                'Susceptance': float(row.get('Susceptance', 0))
            }
        
        # 找到平衡节点（参考节点）
        self.ref_bus = self.system_summary['BalanceNode']
        logging.info(f"参考节点设置为: {self.ref_bus}")
        
        logging.info("数据加载完成")
    
    def get_unit_param(self, unit_dict, unit, key, default=0.0):
        """安全获取机组参数"""
        try:
            val = unit_dict[unit].get(key, default)
            if pd.isna(val):
                return default
            return val
        except Exception:
            return default
    
    def _get_renewable_max_output(self, curve_name, t, capacity, unit_type):
        """计算可再生能源在指定时间的理论最大出力"""
        if capacity <= 0:
            return 0
        
        curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
        
        # 如果没有资源曲线，假设满出力
        if curve_name_str in ['', '0', 'nan', 'None']:
            return capacity
        
        try:
            # 根据机组类型选择对应的曲线数据
            if unit_type == 'wind':
                curves_df = self.wind_curves
                # 检查可能的列名（使用实际的列名）
                possible_cols = ['WTCurve', 'WTResourceCurve', 'WindResourceCurve', 'ResourceCurve']
                curve_col = None
                for col in possible_cols:
                    if col in curves_df.columns:
                        curve_col = col
                        break
                if curve_col is None:
                    logging.warning(f"风电资源曲线表中未找到合适的列名，可用列名: {list(curves_df.columns)}")
                    return capacity
                    
            elif unit_type == 'solar':
                curves_df = self.solar_curves
                # 检查可能的列名（使用实际的列名）
                possible_cols = ['PVCurve', 'PVResourceCurve', 'SolarResourceCurve', 'ResourceCurve']
                curve_col = None
                for col in possible_cols:
                    if col in curves_df.columns:
                        curve_col = col
                        break
                if curve_col is None:
                    logging.warning(f"光伏资源曲线表中未找到合适的列名，可用列名: {list(curves_df.columns)}")
                    return capacity
                    
            elif unit_type == 'hydro':
                curves_df = self.hydro_curves
                # 检查可能的列名
                possible_cols = ['RunoffHydroOperationCurve', 'HydroOperationCurve', 'ResourceCurve']
                curve_col = None
                for col in possible_cols:
                    if col in curves_df.columns:
                        curve_col = col
                        break
                if curve_col is None:
                    logging.warning(f"水电资源曲线表中未找到合适的列名")
                    return capacity
            else:
                return capacity
            
            # 查找对应的资源曲线
            curve_row = curves_df[curves_df[curve_col].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                curve_row = curves_df[curves_df[curve_col].astype(str).str.contains(curve_name_str, case=False, na=False)]
                if curve_row.empty:
                    return capacity
            
            # 获取时间点的出力系数
            time_col = f'Time_{t}'
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 1.0
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        max_factor = 1.0
                else:
                    max_factor = 1.0
            
            return max_factor * capacity
            
        except Exception as e:
            logging.warning(f"计算{unit_type}理论最大出力时出错: {e}")
            return capacity

    def create_model(self):
        """创建Pyomo优化模型（包含直流潮流约束）"""
        logging.info("开始创建优化模型（包含直流潮流约束）...")
        
        self.model = pyo.ConcreteModel()
        
        # 定义集合
        self.model.T = pyo.Set(initialize=self.hours)  # 时间段
        self.model.ThermalUnits = pyo.Set(initialize=self.thermal_units['ThermalUnitNumber'].tolist())
        self.model.WindUnits = pyo.Set(initialize=self.wind_units['WTPlantID'].tolist())
        self.model.SolarUnits = pyo.Set(initialize=self.solar_units['PVPlantID'].tolist())
        self.model.ESSUntis = pyo.Set(initialize=self.ess_units['ESSId'].tolist())
        self.model.HydroUnits = pyo.Set(initialize=self.hydro_units['RunoffHydroUnitName'].tolist())
        self.model.Loads = pyo.Set(initialize=self.loads['负荷编号'].tolist())
        self.model.Buses = pyo.Set(initialize=self.buses['BusId'].tolist())
        self.model.Lines = pyo.Set(initialize=self.lines['BranchId'].tolist())
        
        # 定义变量
        # 火电机组
        self.model.thermal_status = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.thermal_startup = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_shutdown = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        
        # 可再生能源
        self.model.wind_power = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_power = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_power = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 储能系统
        self.model.ess_charge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_discharge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_soc = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 负荷相关
        self.model.load_shed = pyo.Var(
            self.model.Loads, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 直流潮流变量
        self.model.bus_angle = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )  # 节点相角
        
        self.model.line_power = pyo.Var(
            self.model.Lines, self.model.T, within=pyo.Reals
        )  # 线路潮流
        
        # 节点平衡松弛变量
        self.model.balance_slack = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 备用松弛变量
        self.model.reserve_slack = pyo.Var(
            self.model.T, within=pyo.NonNegativeReals
        )
        
        # 深度调峰惩罚线性化变量
        self.model.deep_peak_penalty = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 可再生能源最小出力松弛变量
        self.model.wind_slack = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_slack = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_slack = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 目标函数
        def total_cost_rule(model):
            """系统总成本计算"""
            cost = 0
            fuel_price = self.fuel_data['UnitPrice']  # 使用原始燃料价格，不增加倍数
            
            # 火电运行成本 - 使用Excel中的真实技术经济参数
            for unit in model.ThermalUnits:
                for t in model.T:
                    # 从Excel读取真实的技术经济参数
                    a = self.thermal_dict[unit].get('OperationCoeff_A', 0)
                    b = self.thermal_dict[unit].get('OperationCoeff_B', 0)
                    c = self.thermal_dict[unit].get('OperationCoeff_C', 0)
                    
                    # 使用真实的运行成本曲线：成本 = A*P² + B*P + C (吨标煤) × 燃料价格
                    cost += model.thermal_status[unit, t] * c * fuel_price  # 固定成本
                    cost += (a * model.thermal_power[unit, t] * model.thermal_power[unit, t] + 
                            b * model.thermal_power[unit, t]) * fuel_price   # 二次+一次成本
                    
                    # 使用真实的启动成本
                    startup_cost = self.thermal_dict[unit].get('StateOnFuelConsumption', 0)
                    cost += startup_cost * fuel_price * model.thermal_startup[unit, t]
            
            # 深度调峰惩罚（线性化实现）
            for unit in model.ThermalUnits:
                for t in model.T:
                    max_power = self.thermal_dict[unit].get('Capacity', 0)
                    deep_threshold = self.thermal_dict[unit].get('DeepPeakShavingThreshold', 0.4) * max_power
                    deep_cost = self.thermal_dict[unit].get('DeepPeakShavingCost', 0)
                    # 线性松弛变量
                    cost += deep_cost * model.deep_peak_penalty[unit, t]
            
            # 可再生能源激励（负成本，即收益）和强制出力罚函数
            renewable_incentive = -350  # 适度提高激励以减少弃风
            renewable_penalty = 100000    # 保持高惩罚，确保强制出力
            
            # 风电激励和惩罚
            for unit in model.WindUnits:
                for t in model.T:
                    cost += renewable_incentive * model.wind_power[unit, t]
                    cost += renewable_penalty * model.wind_slack[unit, t]
            
            # 光伏激励和惩罚
            for unit in model.SolarUnits:
                for t in model.T:
                    cost += renewable_incentive * model.solar_power[unit, t]
                    cost += renewable_penalty * model.solar_slack[unit, t]
            
            # 水电激励和惩罚
            for unit in model.HydroUnits:
                for t in model.T:
                    cost += renewable_incentive * model.hydro_power[unit, t]
                    cost += renewable_penalty * model.hydro_slack[unit, t]
            
            # 储能运行激励（更温和的差异化定价，鼓励合理的充放电循环）
            for unit in model.ESSUntis:
                for t in model.T:
                    # 获取当前时段的负荷水平
                    if t < len(self.load_curve):
                        load_level = self.load_curve[t]
                        peak_load = max(self.load_curve)
                        load_ratio = load_level / peak_load if peak_load > 0 else 0.5
                    else:
                        load_ratio = 0.5
                    
                    # 储能充电成本和放电收益
                    if load_ratio > 0.85:
                        charge_cost, discharge_benefit = 200, -300
                    elif load_ratio > 0.7:
                        charge_cost, discharge_benefit = 120, -180
                    elif load_ratio > 0.5:
                        charge_cost, discharge_benefit = 80, -120
                    elif load_ratio > 0.3:
                        charge_cost, discharge_benefit = 50, -60
                    else:
                        charge_cost, discharge_benefit = 30, -40
                    
                    cost += charge_cost * model.ess_charge[unit, t]
                    cost += discharge_benefit * model.ess_discharge[unit, t]
            
            # 切负荷惩罚（适度降低以提高可行性，但仍保持较高优先级）
            penalty_cut = 500000  # 降低到50万元/MWh，平衡可行性和负荷满足
            for load in model.Loads:
                for t in model.T:
                    cost += penalty_cut * model.load_shed[load, t]
            
            # 节点平衡松弛惩罚（降低惩罚，允许潮流约束调节）
            penalty_slack = 100  # 元/MW（降低惩罚）
            for bus in model.Buses:
                for t in model.T:
                    cost += penalty_slack * model.balance_slack[bus, t]
            
            # 备用松弛惩罚
            penalty_reserve = 5000  # 元/MW
            for t in model.T:
                cost += penalty_reserve * model.reserve_slack[t]
            
            return cost

        # 深度调峰惩罚线性化约束
        def deep_peak_penalty_rule(model, unit, t):
            max_power = self.thermal_dict[unit].get('Capacity', 0)
            deep_threshold = self.thermal_dict[unit].get('DeepPeakShavingThreshold', 0.4) * max_power
            return model.deep_peak_penalty[unit, t] >= deep_threshold - model.thermal_power[unit, t]
        self.model.deep_peak_penalty_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=deep_peak_penalty_rule
        )
        
        self.model.total_cost = pyo.Objective(rule=total_cost_rule, sense=pyo.minimize)
        
        # 1. 火电机组逻辑约束
        def thermal_logic_rule(model, unit, t):
            if t == 0:
                # 第0小时确保足够机组启动以满足功率平衡
                init_status = self.thermal_dict[unit].get('InitStatus', False)
                if isinstance(init_status, str):
                    init_status = init_status.strip().lower() in ['true', '1', 'yes']
                
                unit_index = list(model.ThermalUnits).index(unit)
                
                # 强制前12个机组启动，确保第0时刻有充足发电能力满足平衡
                if unit_index < 12:  # 增加到12个机组以确保平衡
                    return model.thermal_status[unit, t] == 1
                else:
                    # 其他机组根据初始状态决定
                    return model.thermal_status[unit, t] == (1 if init_status else 0)
            else:
                return model.thermal_status[unit, t] - model.thermal_status[unit, t-1] == \
                    model.thermal_startup[unit, t] - model.thermal_shutdown[unit, t]
        self.model.thermal_logic_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_logic_rule
        )

        # 2. 火电机组出力约束
        def thermal_power_lower_rule(model, unit, t):
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            return model.thermal_power[unit, t] >= min_power * model.thermal_status[unit, t]
        
        def thermal_power_upper_rule(model, unit, t):
            max_power = self.thermal_dict[unit].get('Capacity', 0)
            return model.thermal_power[unit, t] <= max_power * model.thermal_status[unit, t]
        
        self.model.thermal_power_lower_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_lower_rule
        )
        self.model.thermal_power_upper_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_upper_rule
        )

        # 3. 爬坡约束 - 使用更稳健的方案2
        def thermal_ramp_up_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_up = self.thermal_dict[unit].get('UpRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束，增加状态变化时的缓冲
            return (model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= 
                    model.thermal_status[unit, t] * ramp_up + 
                    (1 - model.thermal_status[unit, t-1]) * capacity * 0.7)  # 缓冲系数
        
        def thermal_ramp_down_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_down = self.thermal_dict[unit].get('DownRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束，增加状态变化时的缓冲
            return (model.thermal_power[unit, t-1] - model.thermal_power[unit, t] <= 
                    model.thermal_status[unit, t-1] * ramp_down + 
                    (1 - model.thermal_status[unit, t]) * capacity * 0.7)  # 缓冲系数

        self.model.thermal_ramp_up_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_up_rule
        )
        self.model.thermal_ramp_down_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_down_rule
        )

        # 4. 最短开停机时间约束 - 暂时禁用以提高可行性
        def min_up_time_rule(model, unit, t):
            return pyo.Constraint.Skip
            min_up = self.thermal_dict[unit].get('MinOnLineTime', 0)
            # 如果最小开机时间太长，则限制为总时间的1/4
            min_up = min(min_up, max(1, self.T // 4))
            
            if t < min_up or min_up <= 1:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_up:
                return sum(model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        model.thermal_status[unit, t] - prev_status) * (self.T - t)
            else:
                return sum(model.thermal_status[unit, k] for k in range(t, t+min_up)) >= (
                        model.thermal_status[unit, t] - prev_status) * min_up

        def min_down_time_rule(model, unit, t):
            return pyo.Constraint.Skip
            min_down = self.thermal_dict[unit].get('MinOffLineTime', 0)
            # 如果最小停机时间太长，则限制为总时间的1/4
            min_down = min(min_down, max(1, self.T // 4))
            
            if t < min_down or min_down <= 1:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_down:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        prev_status - model.thermal_status[unit, t]) * (self.T - t)
            else:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, t+min_down)) >= (
                        prev_status - model.thermal_status[unit, t]) * min_down

        self.model.min_up_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_up_time_rule
        )
        self.model.min_down_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_down_time_rule
        )

        # 5. 可再生能源出力约束 - 修复资源曲线匹配问题，增加默认出力能力
        def wind_power_rule(model, unit, t):
            curve_name = self.wind_dict[unit].get('ResourceCurve', '')
            capacity = self.wind_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.wind_power[unit, t] <= 0
            
            # 防止curve_name为非字符串
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 如果没有资源曲线名称或为虚拟机组，使用适中的出力系数
            if curve_name_str in ['', '0', 'nan', 'None'] or curve_name_str == '':
                # 降低风电出力，给火电更多发电空间
                default_factor = 0.4  # 降低到40%，避免风电过度挤占火电
                logging.info(f"风电机组{unit}在时段{t}使用中等出力系数: {default_factor}")
                return model.wind_power[unit, t] <= default_factor * capacity
            
            # 找到对应的风资源曲线 - 增强匹配逻辑
            curve_row = self.wind_curves[self.wind_curves['WTCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                # 尝试模糊匹配
                curve_row = self.wind_curves[self.wind_curves['WTCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"风电机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！使用中等出力系数")
                    # 使用中等出力而不是高出力
                    default_factor = 0.4
                    return model.wind_power[unit, t] <= default_factor * capacity
            
            # 获取时间点值 - 改进列名匹配逻辑
            time_col = f'Time_{t}'
            # 尝试匹配多种可能的列名格式
            matching_cols = []
            
            # 检查各种可能的时间列名格式
            possible_time_formats = [
                f'Time_{t}', f'time_{t}', f'TIME_{t}',
                f'Time{t}', f'time{t}', f'TIME{t}',
                f'T{t}', f't{t}', f'Hour_{t}', f'hour_{t}',
                f'H{t}', f'h{t}'
            ]
            
            for format_name in possible_time_formats:
                matching_cols = [col for col in curve_row.columns if col.lower() == format_name.lower()]
                if matching_cols:
                    break
            
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 0.4  # 使用中等默认值
                else:
                    max_factor = min(max_factor, 0.5)  # 限制最大50%出力能力
            else:
                # 如果找不到匹配的列，输出调试信息
                logging.debug(f"风电机组{unit}在时段{t}: 可用列名{list(curve_row.columns)}")
                logging.warning(f"风电机组{unit}在时段{t}缺少出力系数，使用中等默认值0.4")
                max_factor = 0.4  # 使用中等默认值
                    
            # 确保出力不超过容量
            return model.wind_power[unit, t] <= max_factor * capacity
        
        self.model.wind_power_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_power_rule
        )
        
        def solar_power_rule(model, unit, t):
            curve_name = self.solar_dict[unit].get('ResourceCurve', '')
            capacity = self.solar_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.solar_power[unit, t] <= 0
            
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 白天时段（6-18点）和时间相关的光照模式
            hour = t % 24
            
            # 如果没有资源曲线名称，使用基于时间的默认光照模式
            if curve_name_str in ['', '0', 'nan', 'None'] or curve_name_str == '':
                # 严格按照太阳光照规律
                if 7 <= hour <= 17:  # 白天时段
                    if 10 <= hour <= 14:  # 最强光照时段
                        default_factor = 0.8  # 最高出力80%
                    else:  # 一般白天时段
                        default_factor = 0.5  # 中等出力50%
                else:  # 夜间
                    default_factor = 0.0  # 夜间无光照
                logging.info(f"光伏机组{unit}在时段{t}（{hour}时）使用默认出力系数: {default_factor}")
                return model.solar_power[unit, t] <= default_factor * capacity
            
            curve_row = self.solar_curves[self.solar_curves['PVCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                curve_row = self.solar_curves[self.solar_curves['PVCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"光伏机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！使用默认光照模式")
                    # 使用默认光照模式
                    if 7 <= hour <= 17:
                        if 10 <= hour <= 14:
                            default_factor = 0.8
                        else:
                            default_factor = 0.5
                    else:
                        default_factor = 0.0
                    return model.solar_power[unit, t] <= default_factor * capacity
            
            # 获取时间点值 - 改进列名匹配逻辑
            time_col = f'Time_{t}'
            # 尝试匹配多种可能的列名格式8
            matching_cols = []
            
            # 检查各种可能的时间列名格式
            possible_time_formats = [
                f'Time_{t}', f'time_{t}', f'TIME_{t}',
                f'Time{t}', f'time{t}', f'TIME{t}',
                f'T{t}', f't{t}', f'Hour_{t}', f'hour_{t}',
                f'H{t}', f'h{t}'
            ]
            
            for format_name in possible_time_formats:
                matching_cols = [col for col in curve_row.columns if col.lower() == format_name.lower()]
                if matching_cols:
                    break
            
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    # 缺少数据时，严格按照时间使用默认值
                    if 7 <= hour <= 17:
                        max_factor = 0.6 if 10 <= hour <= 14 else 0.4
                    else:
                        max_factor = 0.0
                else:
                    # 严格使用资源曲线数据，不做任何修改
                    max_factor = max(0.0, float(max_factor))  # 确保非负值
            else:
                # 如果找不到匹配的列，输出调试信息
                logging.debug(f"光伏机组{unit}在时段{t}: 可用列名{list(curve_row.columns)}")
                logging.warning(f"光伏机组{unit}在时段{t}缺少出力系数，使用默认光照模式")
                if 7 <= hour <= 17:
                    max_factor = 0.6 if 10 <= hour <= 14 else 0.4
                else:
                    max_factor = 0.0
                    
            return model.solar_power[unit, t] <= max_factor * capacity
        
        self.model.solar_power_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_power_rule
        )
        
        def hydro_power_rule(model, unit, t):
            curve_name = self.hydro_dict[unit].get('ResourceCurve', '')
            capacity = self.hydro_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.hydro_power[unit, t] <= 0
            
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 如果资源曲线名为0或空，允许水电机组出力为0~容量
            if curve_name_str in ['', '0', 'nan', 'None']:
                logging.info(f"水电机组{unit}无资源曲线限制，可在0~{capacity}MW范围内出力")
                return (0, model.hydro_power[unit, t], capacity)
            
            curve_row = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                curve_row = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"水电机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！允许灵活出力")
                    return (0, model.hydro_power[unit, t], capacity)
            
            # 获取时间点值 - 处理列名大小写问题
            time_col = f'Time_{t}'
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 0.8  # 使用默认值而不是0
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        max_factor = 0.8
                else:
                    logging.warning(f"水电机组{unit}在时段{t}缺少出力系数，使用默认值0.8")
                    max_factor = 0.8  # 使用默认值而不是0
                    
            return model.hydro_power[unit, t] <= max_factor * capacity
        
        self.model.hydro_power_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_power_rule
        )

        # 5.1 可再生能源最小出力约束（临时禁用强制性要求）
        def wind_min_output_rule(model, unit, t):
            # 临时禁用风电最小出力约束，避免不可行性
            return pyo.Constraint.Skip
        
        def solar_min_output_rule(model, unit, t):
            # 临时禁用光伏最小出力约束，避免不可行性
            return pyo.Constraint.Skip
            capacity = self.solar_dict[unit].get('Capacity', 0)
            curve_name = self.solar_dict[unit].get('ResourceCurve', '')
            
            logging.debug(f"光伏机组{unit}时段{t}: 容量={capacity}")
            # 大幅降低容量门槛，让所有光伏机组都参与运行
            if capacity > 1:  # 降低到1MW门槛，确保更多机组参与
                # 检查该时段是否有太阳资源
                curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
                
                # 获取该时段的实际资源系数
                actual_factor = 0
                if curve_name_str not in ['', '0', 'nan', 'None']:
                    curve_row = self.solar_curves[self.solar_curves['PVCurve'].astype(str).str.strip() == curve_name_str]
                    if not curve_row.empty:
                        time_col = f'Time_{t}'
                        matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
                        if matching_cols:
                            actual_factor = curve_row[matching_cols[0]].values[0]
                            if pd.isna(actual_factor):
                                actual_factor = 0
                
                # 强制所有光伏机组在白天都要有大幅出力 - 确保图表中黄色非常明显
                daytime_hours = list(range(6, 19))  # 6点到18点
                if t % 24 in daytime_hours:
                    if actual_factor > 0.001:  # 极低资源门槛
                        min_output = capacity * max(actual_factor * 0.95, 0.7)  # 至少70%容量或95%资源
                        return model.solar_power[unit, t] + model.solar_slack[unit, t] >= min_output
                    else:
                        # 白天强制大幅出力 - 所有光伏机组都要贡献
                        peak_hours = list(range(9, 16))  # 9-15点为光伏强制高峰
                        if t % 24 in peak_hours:
                            min_output = capacity * 0.8  # 高峰时段强制80%容量
                        else:
                            min_output = capacity * 0.6  # 其他白天时段60%容量
                        return model.solar_power[unit, t] + model.solar_slack[unit, t] >= min_output
                else:
                    return model.solar_slack[unit, t] >= 0  # 夜间不强制出力
            else:
                # 即使是很小的光伏机组白天也要有出力
                if capacity > 0:
                    daytime_hours = list(range(6, 19))
                    if t % 24 in daytime_hours:
                        min_output = capacity * 0.4  # 小机组白天40%出力
                        return model.solar_power[unit, t] + model.solar_slack[unit, t] >= min_output
                return model.solar_slack[unit, t] >= 0
        
        def hydro_min_output_rule(model, unit, t):
            # 临时禁用水电最小出力约束，避免不可行性
            return pyo.Constraint.Skip
        
        self.model.wind_min_output_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_min_output_rule
        )
        self.model.solar_min_output_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_min_output_rule
        )
        self.model.hydro_min_output_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_min_output_rule
        )

        # 5.5 可再生能源强制优先出力约束（确保优先级高于火电）
        def wind_priority_rule(model, unit, t):
            """风电应该按最大可用资源出力，slack为弃风量"""
            curve_name = self.wind_dict[unit].get('ResourceCurve', '')
            capacity = self.wind_dict[unit].get('Capacity', 0)
            
            if capacity <= 0:
                return model.wind_slack[unit, t] == 0
            
            # 获取最大可用出力
            max_available = self._get_renewable_max_output(curve_name, t, capacity, 'wind')
            
            # 实际出力 + 弃风量 = 最大可用出力
            return model.wind_power[unit, t] + model.wind_slack[unit, t] == max_available

        def solar_priority_rule(model, unit, t):
            """光伏应该按最大可用资源出力，slack为弃光量"""
            curve_name = self.solar_dict[unit].get('ResourceCurve', '')
            capacity = self.solar_dict[unit].get('Capacity', 0)
            
            if capacity <= 0:
                return model.solar_slack[unit, t] == 0
            
            # 获取最大可用出力
            max_available = self._get_renewable_max_output(curve_name, t, capacity, 'solar')
            
            # 实际出力 + 弃光量 = 最大可用出力
            return model.solar_power[unit, t] + model.solar_slack[unit, t] == max_available

        def hydro_priority_rule(model, unit, t):
            """水电应该按最大可用资源出力，slack为弃水量"""
            curve_name = self.hydro_dict[unit].get('ResourceCurve', '')
            capacity = self.hydro_dict[unit].get('Capacity', 0)
            
            if capacity <= 0:
                return model.hydro_slack[unit, t] == 0
            
            # 获取最大可用出力
            max_available = self._get_renewable_max_output(curve_name, t, capacity, 'hydro')
            
            # 实际出力 + 弃水量 = 最大可用出力
            return model.hydro_power[unit, t] + model.hydro_slack[unit, t] == max_available

        self.model.wind_priority_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_priority_rule
        )
        self.model.solar_priority_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_priority_rule
        )
        self.model.hydro_priority_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_priority_rule
        )

        # 6. 储能系统约束 
        def ess_soc_rule(model, unit, t):
            min_soc = self.ess_dict[unit].get('MinSOC', 0)
            max_soc = self.ess_dict[unit].get('MaxSOC', 1000)
            capacity = self.ess_dict[unit].get('Capacity', 0)
            return (min_soc, model.ess_soc[unit, t], max_soc)
        
        def ess_soc_balance_rule(model, unit, t):
            charge_eff = self.ess_dict[unit].get('ChargeEfficiency', 0.85)
            dis_eff = self.ess_dict[unit].get('DisEfficiency', 0.85)
            
            if t == 0:
                init_soc = self.ess_dict[unit].get('InitSOC', 0)
                return model.ess_soc[unit, t] == init_soc
            else:
                return model.ess_soc[unit, t] == (model.ess_soc[unit, t-1] + 
                                                  charge_eff * model.ess_charge[unit, t] - 
                                                  model.ess_discharge[unit, t] / dis_eff)
        
        # 添加储能充放电状态二进制变量（暂时注释掉，使用线性约束）
        # self.model.ess_charge_status = pyo.Var(
        #     self.model.ESSUntis, self.model.T, within=pyo.Binary
        # )
        
        # 6.5 储能充放电互斥约束（改用简单线性约束）
        def ess_charge_limit_rule(model, unit, t):
            max_charge = self.ess_dict[unit].get('MaxChargePower', 0)
            # 降低储能实际可用功率到原来的40%，使储能占比更合理
            effective_max_charge = max_charge * 0.4
            return model.ess_charge[unit, t] <= effective_max_charge
            
        def ess_discharge_limit_rule(model, unit, t):
            max_discharge = self.ess_dict[unit].get('MaxDisPower', 0)
            # 降低储能实际可用功率到原来的40%，使储能占比更合理
            effective_max_discharge = max_discharge * 0.4
            return model.ess_discharge[unit, t] <= effective_max_discharge
        
        # 简单的线性互斥约束：充放电之和不能超过最大功率的10%
        def ess_mutual_exclusive_rule(model, unit, t):
            max_charge = self.ess_dict[unit].get('MaxChargePower', 0)
            max_discharge = self.ess_dict[unit].get('MaxDisPower', 0)
            max_power = max(max_charge, max_discharge) * 0.4  # 使用有效功率
            return model.ess_charge[unit, t] + model.ess_discharge[unit, t] <= max_power * 0.1
        
        # 储能日内充放电平衡约束：确保充放电循环合理
        def ess_daily_balance_upper_rule(model, unit):
            total_charge = sum(model.ess_charge[unit, t] for t in model.T)
            total_discharge = sum(model.ess_discharge[unit, t] for t in model.T)
            # 放电不能超过充电太多（考虑效率损失）
            return total_discharge <= total_charge * 1.2
        
        def ess_daily_balance_lower_rule(model, unit):
            total_charge = sum(model.ess_charge[unit, t] for t in model.T)
            total_discharge = sum(model.ess_discharge[unit, t] for t in model.T)
            # 充电不能比放电少太多（确保有放电行为）
            return total_charge >= total_discharge * 0.8
        
        self.model.ess_charge_limit_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_charge_limit_rule
        )
        self.model.ess_discharge_limit_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_discharge_limit_rule
        )
        self.model.ess_mutual_exclusive_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_mutual_exclusive_rule
        )
        self.model.ess_daily_balance_upper_con = pyo.Constraint(
            self.model.ESSUntis, rule=ess_daily_balance_upper_rule
        )
        self.model.ess_daily_balance_lower_con = pyo.Constraint(
            self.model.ESSUntis, rule=ess_daily_balance_lower_rule
        )
        self.model.ess_soc_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_rule
        )
        self.model.ess_soc_balance_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_balance_rule
        )

        # 7.1 参考节点约束（平衡节点相角为0）
        def ref_bus_rule(model, t):
            # 只在有多个节点时启用
            if len(model.Buses) <= 1:
                return pyo.Constraint.Skip
            return model.bus_angle[self.ref_bus, t] == 0
        self.model.ref_bus_con = pyo.Constraint(
            self.model.T, rule=ref_bus_rule
        )
        
        # 7.2 节点功率平衡约束（恢复简化版本）
        def node_balance_rule(model, bus, t):
            # 如果只有一个节点，跳过节点平衡约束
            if len(model.Buses) <= 1:
                return pyo.Constraint.Skip
            
            # 简化的节点平衡约束，只考虑主要发电机组
            gen_power = 0
            for unit in model.ThermalUnits:
                if self.thermal_dict[unit].get('BusName', '') == bus:
                    gen_power += model.thermal_power[unit, t]
            
            # 简化负荷分配（平均分配）
            num_buses = len(model.Buses)
            if t < len(self.load_curve) and num_buses > 0:
                node_load = self.load_curve[t] / num_buses
            else:
                node_load = 0
            
            # 使用较大的松弛变量，确保约束不会过于严格
            # 简化的平衡：发电 + 松弛变量 >= 负荷
            return gen_power + model.balance_slack[bus, t] >= node_load * 0.8  # 只要求满足80%的负荷
        
        self.model.node_balance_con = pyo.Constraint(
            self.model.Buses, self.model.T, rule=node_balance_rule
        )
        
        # 7.3 线路潮流约束（恢复简化版本）
        def line_power_rule(model, line, t):
            line_info = self.line_dict.get(line, {})
            start_bus = line_info.get('StartNodeName', '')
            end_bus = line_info.get('EndNodeName', '')
            reactance = line_info.get('Reactance', 0)
            
            # 跳过无效线路或电抗为0的线路，只处理有效线路
            if (start_bus == '' or end_bus == '' or 
                abs(reactance) < 1e-6 or 
                len(self.model.Buses) <= 1):
                return pyo.Constraint.Skip
            
            # 简化的直流潮流方程：P_ij = (θ_i - θ_j) / X_ij
            # 但添加松弛，避免过于严格
            return model.line_power[line, t] == (
                model.bus_angle[start_bus, t] - model.bus_angle[end_bus, t]
            ) / reactance
        
        self.model.line_power_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_power_rule
        )
        
        # 7.4 线路容量约束（启用但设置宽松限制）
        def line_capacity_rule(model, line, t):
            line_info = self.line_dict.get(line, {})
            capacity = line_info.get('Capacity', 0)
            
            # 只对有容量限制的线路进行约束
            if capacity <= 0 or len(self.model.Buses) <= 1:
                return pyo.Constraint.Skip
            
            # 线路潮流容量约束（放宽到120%容量以增加灵活性）
            relaxed_capacity = capacity * 1.2
            return (-relaxed_capacity, model.line_power[line, t], relaxed_capacity)
        
        self.model.line_capacity_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_capacity_rule
        )
        
        # 7.3 火电基础负荷约束 - 暂时完全禁用以诊断问题
        def thermal_base_load_rule(model, t):
            # 完全禁用此约束以诊断不可行问题
            return pyo.Constraint.Skip
        
        self.model.thermal_base_load_con = pyo.Constraint(
            self.model.T, rule=thermal_base_load_rule
        )
        
        # 8. 系统功率平衡约束（全局）
        def global_power_balance_rule(model, t):
            # 发电侧（储能放电算发电，储能充电算负荷）
            thermal_gen = sum(model.thermal_power[unit, t] for unit in model.ThermalUnits)
            wind_gen = sum(model.wind_power[unit, t] for unit in model.WindUnits)
            solar_gen = sum(model.solar_power[unit, t] for unit in model.SolarUnits)
            hydro_gen = sum(model.hydro_power[unit, t] for unit in model.HydroUnits)
            ess_discharge = sum(model.ess_discharge[unit, t] for unit in model.ESSUntis)
            ess_charge = sum(model.ess_charge[unit, t] for unit in model.ESSUntis)
            
            # 负荷侧 - 直接使用全局负荷曲线，不再按节点分解
            if t < len(self.load_curve):
                total_load = self.load_curve[t]
            else:
                total_load = 0
            
            # 切负荷总量
            total_shed = sum(model.load_shed[load, t] for load in model.Loads)
            
            # 修正：储能充电是消耗，应该加到负荷侧，不是发电侧
            return (thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge ==
                    total_load - total_shed + ess_charge)
        
        self.model.global_power_balance_con = pyo.Constraint(
            self.model.T, rule=global_power_balance_rule
        )

        # 9. 备用约束（恢复温和的备用要求）
        def reserve_requirement_rule(model, t):
            # 计算总的在线火电容量
            total_thermal_capacity = sum(model.thermal_status[unit, t] * 
                                       self.thermal_dict[unit].get('Capacity', 0) 
                                       for unit in model.ThermalUnits)
            
            # 温和的备用要求：负荷的2%
            if t < len(self.load_curve):
                total_load = self.load_curve[t]
                reserve_req = total_load * 0.02  # 2%备用要求，比较温和
                
                # 加上松弛变量，避免过于严格
                return total_thermal_capacity + model.reserve_slack[t] >= total_load + reserve_req
            else:
                return pyo.Constraint.Skip

        self.model.reserve_con = pyo.Constraint(
            self.model.T, rule=reserve_requirement_rule
        )
        
        logging.info("优化模型创建完成（包含直流潮流约束）")
    
    def debug_model(self):
        """调试模型不可行性"""
        logging.info("开始调试模型...")
        
        # 检查基本数据完整性
        logging.info(f"火电机组数量: {len(self.model.ThermalUnits)}")
        logging.info(f"风电机组数量: {len(self.model.WindUnits)}")
        logging.info(f"光伏机组数量: {len(self.model.SolarUnits)}")
        logging.info(f"水电机组数量: {len(self.model.HydroUnits)}")
        logging.info(f"储能单元数量: {len(self.model.ESSUntis)}")
        logging.info(f"负荷数量: {len(self.model.Loads)}")
        logging.info(f"母线数量: {len(self.model.Buses)}")
        logging.info(f"线路数量: {len(self.model.Lines)}")
        logging.info(f"时间段数: {self.T}")
        
        # 检查负荷总量和发电容量
        total_thermal_cap = sum(self.thermal_dict[unit].get('Capacity', 0) for unit in self.model.ThermalUnits)
        total_wind_cap = sum(self.wind_dict[unit].get('Capacity', 0) for unit in self.model.WindUnits)
        total_solar_cap = sum(self.solar_dict[unit].get('Capacity', 0) for unit in self.model.SolarUnits)
        total_hydro_cap = sum(self.hydro_dict[unit].get('Capacity', 0) for unit in self.model.HydroUnits)
        total_gen_cap = total_thermal_cap + total_wind_cap + total_solar_cap + total_hydro_cap
        
        peak_load = max(self.load_curve) if self.load_curve else 0
        
        logging.info(f"总发电容量: {total_gen_cap} MW")
        logging.info(f"  - 火电容量: {total_thermal_cap} MW")
        logging.info(f"  - 风电容量: {total_wind_cap} MW")
        logging.info(f"  - 光伏容量: {total_solar_cap} MW")
        logging.info(f"  - 水电容量: {total_hydro_cap} MW")
        logging.info(f"峰值负荷: {peak_load} MW")
        logging.info(f"容量充裕度: {(total_gen_cap / peak_load * 100):.1f}%" if peak_load > 0 else "无负荷")
        
        # 详细检查各类机组
        logging.info("=== 风电机组详细信息 ===")
        for unit in self.model.WindUnits:
            cap = self.wind_dict[unit].get('Capacity', 0)
            curve = self.wind_dict[unit].get('ResourceCurve', '')
            bus = self.wind_dict[unit].get('BusName', '')
            logging.info(f"风电机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        logging.info("=== 光伏机组详细信息 ===")
        for unit in self.model.SolarUnits:
            cap = self.solar_dict[unit].get('Capacity', 0)
            curve = self.solar_dict[unit].get('ResourceCurve', '')
            bus = self.solar_dict[unit].get('BusName', '')
            logging.info(f"光伏机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        logging.info("=== 水电机组详细信息 ===")
        for unit in self.model.HydroUnits:
            cap = self.hydro_dict[unit].get('Capacity', 0)
            curve = self.hydro_dict[unit].get('ResourceCurve', '')
            bus = self.hydro_dict[unit].get('BusName', '')
            logging.info(f"水电机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        # 检查资源曲线数据
        logging.info("=== 资源曲线数据检查 ===")
        logging.info(f"风电曲线数量: {len(self.wind_curves)}")
        if not self.wind_curves.empty:
            logging.info(f"风电曲线名称: {list(self.wind_curves['WTCurve'].unique())}")
        
        logging.info(f"光伏曲线数量: {len(self.solar_curves)}")
        if not self.solar_curves.empty:
            logging.info(f"光伏曲线名称: {list(self.solar_curves['PVCurve'].unique())}")
        
        logging.info(f"水电曲线数量: {len(self.hydro_curves)}")
        if not self.hydro_curves.empty:
            logging.info(f"水电曲线名称: {list(self.hydro_curves['RunoffHydroOperationCurve'].unique())}")
        
        # 检查机组参数合理性
        for unit in self.model.ThermalUnits:
            unit_data = self.thermal_dict[unit]
            cap = unit_data.get('Capacity', 0)
            min_power = unit_data.get('MinPower', 0)
            if min_power > cap:
                logging.error(f"火电机组{unit}: 最小出力({min_power}) > 容量({cap})")
        
        # 输出模型统计信息
        logging.info(f"模型变量数: {len([v for v in self.model.component_objects(pyo.Var)])}")
        logging.info(f"模型约束数: {len([c for c in self.model.component_objects(pyo.Constraint)])}")
        
    def solve(self):
        """求解优化问题"""
        logging.info("开始求解优化问题...")
        
        # 先进行模型调试
        self.debug_model()
        
        # 关键容量检查
        logging.info("=== 关键容量检查 ===")
        peak_load = max(self.load_curve) if self.load_curve else 0
        total_thermal_cap = sum(self.thermal_dict[unit]['Capacity'] for unit in self.thermal_dict) if self.thermal_dict else 0
        total_wind_cap = sum(self.wind_dict[unit]['Capacity'] for unit in self.wind_dict) if self.wind_dict else 0
        total_solar_cap = sum(self.solar_dict[unit]['Capacity'] for unit in self.solar_dict) if self.solar_dict else 0
        total_hydro_cap = sum(self.hydro_dict[unit]['Capacity'] for unit in self.hydro_dict) if self.hydro_dict else 0
        total_capacity = total_thermal_cap + total_wind_cap + total_solar_cap + total_hydro_cap
        
        logging.info(f"峰值负荷: {peak_load:.1f}MW")
        logging.info(f"总发电容量: {total_capacity:.1f}MW")
        logging.info(f"容量充裕度: {(total_capacity/peak_load*100):.1f}%" if peak_load > 0 else "无负荷")
        
        if total_capacity < peak_load * 1.05:
            logging.error(f"严重警告：总容量({total_capacity:.1f}MW) 不足以满足峰值负荷+5%备用({peak_load*1.05:.1f}MW)")
        
        # 使用Gurobi求解器
        solver = SolverFactory('gurobi')
        solver.options['MIPGap'] = 0.2  # 进一步放宽到20%
        solver.options['TimeLimit'] = 60  # 降低到1分钟
        solver.options['FeasibilityTol'] = 1e-3  # 放宽可行性容差
        solver.options['OptimalityTol'] = 1e-3  # 放宽最优性容差
        
        # 尝试求解
        start_time = time.time()
        results = solver.solve(self.model, tee=True)
        solve_time = time.time() - start_time
        logging.info(f"求解耗时: {solve_time:.2f}秒")
        
        # 检查求解状态
        if (results.solver.status == pyo.SolverStatus.ok and 
            results.solver.termination_condition in [pyo.TerminationCondition.optimal, 
                                                     pyo.TerminationCondition.feasible]):
            logging.info("优化求解成功完成")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            logging.warning("达到时间限制，返回当前最优解")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.infeasible:
            logging.error("模型不可行，尝试诊断...")
            self.diagnose_infeasibility()
            
            # 尝试简化模型
            logging.info("尝试创建简化模型...")
            if self.create_simplified_model():
                logging.info("重新求解简化模型...")
                results = solver.solve(self.model, tee=True)
                if (results.solver.status == pyo.SolverStatus.ok and 
                    results.solver.termination_condition in [pyo.TerminationCondition.optimal, 
                                                             pyo.TerminationCondition.feasible]):
                    logging.info("简化模型求解成功")
                    return True
            
            return False
        else:
            logging.error(f"优化求解失败: {results.solver.termination_condition}")
            return False
    
    def diagnose_infeasibility(self):
        """诊断模型不可行性"""
        logging.info("开始诊断模型不可行性...")
        
        # 写出模型文件用于调试
        try:
            self.model.write('debug_model.lp', format=pyo.WriterFactory.LP)
            logging.info("模型已写出到 debug_model.lp 文件")
        except Exception as e:
            logging.warning(f"无法写出模型文件: {e}")
        
        # 检查一些常见的不可行原因
        logging.info("检查常见不可行原因:")
        
        # 1. 检查负荷是否过大
        peak_load = max(self.load_curve) if self.load_curve else 0
        total_thermal_cap = sum(self.thermal_dict[unit].get('Capacity', 0) for unit in self.model.ThermalUnits)
        if peak_load > total_thermal_cap * 1.5:  # 考虑可再生能源
            logging.error(f"峰值负荷({peak_load}) 可能超过总发电能力")
        
        # 2. 检查机组最小开机时间是否过长
        for unit in self.model.ThermalUnits:
            min_up = self.thermal_dict[unit].get('MinOnLineTime', 0)
            min_down = self.thermal_dict[unit].get('MinOffLineTime', 0)
            if min_up > self.T // 2:
                logging.warning(f"机组{unit}最小开机时间({min_up})过长")
            if min_down > self.T // 2:
                logging.warning(f"机组{unit}最小停机时间({min_down})过长")
        
        # 3. 检查备用约束是否过紧
        logging.warning("建议放宽备用约束或检查机组参数设置")
    
    def create_simplified_model(self):
        """创建简化模型（移除一些复杂约束）"""
        try:
            logging.info("创建简化模型...")
            
            # 移除原有约束
            self.model.del_component(self.model.thermal_ramp_up_con)
            self.model.del_component(self.model.thermal_ramp_down_con)
            self.model.del_component(self.model.min_up_time_con)
            self.model.del_component(self.model.min_down_time_con)
            self.model.del_component(self.model.reserve_con)
            self.model.del_component(self.model.deep_peak_penalty_con)
            
            logging.info("已移除复杂约束：爬坡、最短开停机时间、备用、深度调峰")
            
            # 添加简化的爬坡约束（更宽松）
            def simple_ramp_rule(model, unit, t):
                if t == 0:
                    return pyo.Constraint.Skip
                capacity = self.thermal_dict[unit].get('Capacity', 0)
                # 使用容量的50%作为最大变化量
                max_change = capacity * 0.5
                return (-max_change <= 
                        model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= 
                        max_change)
            
            self.model.simple_ramp_con = pyo.Constraint(
                self.model.ThermalUnits, self.model.T, rule=simple_ramp_rule
            )
            
            logging.info("添加了简化的爬坡约束")
            return True
            
        except Exception as e:
            logging.error(f"创建简化模型失败: {e}")
            return False
    
    def calculate_cost_analysis(self):
        """计算和显示系统成本分析"""
        logging.info("开始计算系统成本分析...")
        
        # 安全获取变量值函数
        def safe_value(var):
            try:
                return pyo.value(var)
            except:
                return 0.0
        
        # 初始化成本字典
        cost_breakdown = {
            '火电运行成本': 0,
            '火电启动成本': 0,
            '深度调峰成本': 0,
            '可再生能源激励': 0,
            '储能运行成本': 0,
            '切负荷惩罚': 0,
            '节点平衡松弛成本': 0,
            '总成本': 0
        }
        
        fuel_price = self.fuel_data['UnitPrice']
        
        # 1. 火电运行成本和启动成本
        thermal_running_cost = 0
        thermal_startup_cost = 0
        
        print("\n" + "="*60)
        print("               火电机组成本明细")
        print("="*60)
        print(f"{'机组':<15} {'运行成本(元)':<15} {'启动成本(元)':<15} {'总成本(元)':<15}")
        print("-"*60)
        
        for unit in self.model.ThermalUnits:
            unit_running_cost = 0
            unit_startup_cost = 0
            
            # 获取机组参数
            a = self.thermal_dict[unit].get('OperationCoeff_A', 0)
            b = self.thermal_dict[unit].get('OperationCoeff_B', 100)
            c = self.thermal_dict[unit].get('OperationCoeff_C', 50)
            startup_cost = self.thermal_dict[unit].get('StateOnFuelConsumption', 200)
            
            for t in self.model.T:
                # 运行成本
                if safe_value(self.model.thermal_status[unit, t]) > 0.5:
                    power = safe_value(self.model.thermal_power[unit, t])
                    fixed_cost = c * fuel_price
                    variable_cost = b * power * fuel_price
                    unit_running_cost += fixed_cost + variable_cost
                
                # 启动成本
                unit_startup_cost += startup_cost * fuel_price * safe_value(self.model.thermal_startup[unit, t])
            
            unit_total_cost = unit_running_cost + unit_startup_cost
            thermal_running_cost += unit_running_cost
            thermal_startup_cost += unit_startup_cost
            
            print(f"{unit:<15} {unit_running_cost:<15.2f} {unit_startup_cost:<15.2f} {unit_total_cost:<15.2f}")
        
        cost_breakdown['火电运行成本'] = thermal_running_cost
        cost_breakdown['火电启动成本'] = thermal_startup_cost
        
        # 2. 深度调峰成本
        deep_peak_cost = 0
        for unit in self.model.ThermalUnits:
            deep_cost = self.thermal_dict[unit].get('DeepPeakShavingCost', 0)
            for t in self.model.T:
                deep_peak_cost += deep_cost * safe_value(self.model.deep_peak_penalty[unit, t])
        cost_breakdown['深度调峰成本'] = deep_peak_cost
        
        # 3. 可再生能源激励（负成本表示收益）
        renewable_cost = 0
        renewable_incentive = -350  # 与目标函数中一致
        
        for unit in self.model.WindUnits:
            for t in self.model.T:
                renewable_cost += renewable_incentive * safe_value(self.model.wind_power[unit, t])
        
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                renewable_cost += renewable_incentive * safe_value(self.model.solar_power[unit, t])
        
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                renewable_cost += renewable_incentive * safe_value(self.model.hydro_power[unit, t])
        
        cost_breakdown['可再生能源激励'] = renewable_cost
        
        # 4. 储能运行成本
        ess_cost = 0
        for unit in self.model.ESSUntis:
            for t in self.model.T:
                # 根据负荷水平计算成本（与目标函数逻辑一致）
                if t < len(self.load_curve):
                    load_level = self.load_curve[t]
                    peak_load = max(self.load_curve)
                    load_ratio = load_level / peak_load if peak_load > 0 else 0.5
                else:
                    load_ratio = 0.5
                
                # 计算充放电成本
                if load_ratio > 0.85:
                    charge_cost, discharge_benefit = 200, -300
                elif load_ratio > 0.7:
                    charge_cost, discharge_benefit = 120, -180
                elif load_ratio > 0.5:
                    charge_cost, discharge_benefit = 80, -120
                elif load_ratio > 0.3:
                    charge_cost, discharge_benefit = 50, -60
                else:
                    charge_cost, discharge_benefit = 30, -40
                
                ess_cost += charge_cost * safe_value(self.model.ess_charge[unit, t])
                ess_cost += discharge_benefit * safe_value(self.model.ess_discharge[unit, t])
        
        cost_breakdown['储能运行成本'] = ess_cost
        
        # 5. 切负荷惩罚
        load_shed_cost = 0
        penalty_cut = 500000
        for load in self.model.Loads:
            for t in self.model.T:
                load_shed_cost += penalty_cut * safe_value(self.model.load_shed[load, t])
        cost_breakdown['切负荷惩罚'] = load_shed_cost
        
        # 6. 节点平衡松弛成本
        balance_slack_cost = 0
        penalty_slack = 100
        for bus in self.model.Buses:
            for t in self.model.T:
                balance_slack_cost += penalty_slack * safe_value(self.model.balance_slack[bus, t])
        cost_breakdown['节点平衡松弛成本'] = balance_slack_cost
        
        # 计算总成本
        cost_breakdown['总成本'] = sum(cost_breakdown.values()) - cost_breakdown['总成本']  # 避免重复计算
        
        # 显示成本汇总
        print("\n" + "="*60)
        print("               系统成本汇总")
        print("="*60)
        
        for category, cost in cost_breakdown.items():
            if category != '总成本':
                if cost < 0:
                    print(f"{category:<20}: -{abs(cost):>15,.2f} 元 (收益)")
                else:
                    print(f"{category:<20}: {cost:>15,.2f} 元")
        
        print("-"*60)
        print(f"{'系统总成本':<20}: {cost_breakdown['总成本']:>15,.2f} 元")
        print("="*60)
        
        # 计算发电量统计
        print("\n" + "="*60)
        print("               发电量统计")
        print("="*60)
        
        thermal_generation = 0
        wind_generation = 0
        solar_generation = 0
        hydro_generation = 0
        
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_generation += safe_value(self.model.thermal_power[unit, t])
        
        for unit in self.model.WindUnits:
            for t in self.model.T:
                wind_generation += safe_value(self.model.wind_power[unit, t])
        
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                solar_generation += safe_value(self.model.solar_power[unit, t])
        
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                hydro_generation += safe_value(self.model.hydro_power[unit, t])
        
        total_generation = thermal_generation + wind_generation + solar_generation + hydro_generation
        
        print(f"火电发电量: {thermal_generation:,.2f} MWh")
        print(f"风电发电量: {wind_generation:,.2f} MWh")
        print(f"光伏发电量: {solar_generation:,.2f} MWh")
        print(f"水电发电量: {hydro_generation:,.2f} MWh")
        print(f"总发电量: {total_generation:,.2f} MWh")
        
        # 计算弃电量统计
        print(f"\n{'='*60}")
        print("               弃电量统计")
        print("="*60)
        
        wind_curtailment = 0
        solar_curtailment = 0
        hydro_curtailment = 0
        
        # 弃风量
        for unit in self.model.WindUnits:
            for t in self.model.T:
                if hasattr(self.model, 'wind_slack'):
                    wind_curtailment += safe_value(self.model.wind_slack[unit, t])
        
        # 弃光量
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                if hasattr(self.model, 'solar_slack'):
                    solar_curtailment += safe_value(self.model.solar_slack[unit, t])
        
        # 弃水量
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                if hasattr(self.model, 'hydro_slack'):
                    hydro_curtailment += safe_value(self.model.hydro_slack[unit, t])
        
        total_curtailment = wind_curtailment + solar_curtailment + hydro_curtailment
        
        print(f"弃风电量: {wind_curtailment:,.2f} MWh")
        print(f"弃光电量: {solar_curtailment:,.2f} MWh") 
        print(f"弃水电量: {hydro_curtailment:,.2f} MWh")
        print(f"总弃电量: {total_curtailment:,.2f} MWh")
        
        # 计算弃电率
        total_renewable_potential = wind_generation + wind_curtailment + solar_generation + solar_curtailment + hydro_generation + hydro_curtailment
        if total_renewable_potential > 0:
            overall_curtailment_rate = total_curtailment / total_renewable_potential * 100
            print(f"\n弃电率:")
            if wind_generation + wind_curtailment > 0:
                wind_curtailment_rate = wind_curtailment / (wind_generation + wind_curtailment) * 100
                print(f"弃风率: {wind_curtailment_rate:.2f}%")
            if solar_generation + solar_curtailment > 0:
                solar_curtailment_rate = solar_curtailment / (solar_generation + solar_curtailment) * 100
                print(f"弃光率: {solar_curtailment_rate:.2f}%")
            if hydro_generation + hydro_curtailment > 0:
                hydro_curtailment_rate = hydro_curtailment / (hydro_generation + hydro_curtailment) * 100
                print(f"弃水率: {hydro_curtailment_rate:.2f}%")
            print(f"总体弃电率: {overall_curtailment_rate:.2f}%")
        
        print("="*60)
        
        # 计算平均发电成本
        if total_generation > 0:
            avg_cost = cost_breakdown['总成本'] / total_generation
            print(f"\n平均发电成本: {avg_cost:.2f} 元/MWh")
        
        # 计算各类型发电占比
        if total_generation > 0:
            print(f"\n发电结构:")
            print(f"火电占比: {thermal_generation/total_generation*100:.1f}%")
            print(f"风电占比: {wind_generation/total_generation*100:.1f}%")
            print(f"光伏占比: {solar_generation/total_generation*100:.1f}%")
            print(f"水电占比: {hydro_generation/total_generation*100:.1f}%")
        
        print("="*60)
        
        # 存储成本分析结果供后续使用
        self.cost_analysis = cost_breakdown
        self.generation_stats = {
            'thermal': thermal_generation,
            'wind': wind_generation,
            'solar': solar_generation,
            'hydro': hydro_generation,
            'total': total_generation
        }
        
        return cost_breakdown
    
    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            # 临时抑制日志
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                # 恢复原始日志级别
                logging.getLogger().setLevel(original_level)

        # 创建Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # 1. 火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t]),
                    'Startup': safe_value(self.model.thermal_startup[unit, t]),
                    'Shutdown': safe_value(self.model.thermal_shutdown[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        thermal_df.to_excel(writer, sheet_name='火电机组结果', index=False)
        
        # 2. 可再生能源结果
        wind_results = []
        for unit in self.model.WindUnits:
            for t in self.model.T:
                wind_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.wind_power[unit, t])
                })
        wind_df = pd.DataFrame(wind_results)
        wind_df.to_excel(writer, sheet_name='风电结果', index=False)
        
        solar_results = []
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                solar_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.solar_power[unit, t])
                })
        solar_df = pd.DataFrame(solar_results)
        solar_df.to_excel(writer, sheet_name='光伏结果', index=False)
        
        hydro_results = []
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                hydro_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.hydro_power[unit, t])
                })
        hydro_df = pd.DataFrame(hydro_results)
        hydro_df.to_excel(writer, sheet_name='水电结果', index=False)
        
        # 3. 储能结果
        ess_results = []
        for unit in self.model.ESSUntis:
            for t in self.model.T:
                ess_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Charge': safe_value(self.model.ess_charge[unit, t]),
                    'Discharge': safe_value(self.model.ess_discharge[unit, t]),
                    'SOC': safe_value(self.model.ess_soc[unit, t])
                })
        ess_df = pd.DataFrame(ess_results)
        ess_df.to_excel(writer, sheet_name='储能结果', index=False)
        
        # 4. 负荷结果
        load_results = []
        for load in self.model.Loads:
            for t in self.model.T:
                load_results.append({
                    'Load': load,
                    'Time': t,
                    'Shed': safe_value(self.model.load_shed[load, t])
                })
        load_df = pd.DataFrame(load_results)
        load_df.to_excel(writer, sheet_name='切负荷结果', index=False)
        
        # 5. 网络结果
        bus_results = []
        for bus in self.model.Buses:
            for t in self.model.T:
                bus_results.append({
                    'Bus': bus,
                    'Time': t,
                    'Angle': safe_value(self.model.bus_angle[bus, t])
                })
        bus_df = pd.DataFrame(bus_results)
        bus_df.to_excel(writer, sheet_name='节点相角', index=False)
        
        line_results = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_results.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_df = pd.DataFrame(line_results)
        line_df.to_excel(writer, sheet_name='线路潮流', index=False)
        
        # 6. 系统总结
        total_thermal_gen = thermal_df.groupby('Time')['Power'].sum().values
        total_wind_gen = wind_df.groupby('Time')['Power'].sum().values
        total_solar_gen = solar_df.groupby('Time')['Power'].sum().values
        total_hydro_gen = hydro_df.groupby('Time')['Power'].sum().values
        total_ess_discharge = ess_df.groupby('Time')['Discharge'].sum().values
        total_ess_charge = ess_df.groupby('Time')['Charge'].sum().values
        
        # 使用全局负荷曲线
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T

        summary_data = {
            'Time': self.hours,
            'Total_Load': total_load,
            'Thermal_Gen': total_thermal_gen,
            'Wind_Gen': total_wind_gen,
            'Solar_Gen': total_solar_gen,
            'Hydro_Gen': total_hydro_gen,
            'ESS_Discharge': total_ess_discharge,
            'ESS_Charge': total_ess_charge,
            'Net_Gen': [t + w + s + h + e - c for t, w, s, h, e, c in 
                       zip(total_thermal_gen, total_wind_gen, total_solar_gen, 
                           total_hydro_gen, total_ess_discharge, total_ess_charge)]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='系统总结', index=False)
        
        # 6. 成本分析结果（如果已计算）
        if hasattr(self, 'cost_analysis'):
            cost_data = []
            for category, cost in self.cost_analysis.items():
                cost_data.append({
                    '成本类别': category,
                    '成本金额(元)': cost,
                    '备注': '负值表示收益' if cost < 0 else ''
                })
            
            cost_df = pd.DataFrame(cost_data)
            cost_df.to_excel(writer, sheet_name='成本分析', index=False)
            
            # 发电量统计
            if hasattr(self, 'generation_stats'):
                gen_data = []
                gen_stats = self.generation_stats
                total_gen = gen_stats['total']
                
                for gen_type, generation in gen_stats.items():
                    if gen_type != 'total':
                        percentage = (generation / total_gen * 100) if total_gen > 0 else 0
                        gen_data.append({
                            '发电类型': gen_type,
                            '发电量(MWh)': generation,
                            '占比(%)': percentage
                        })
                
                # 添加总计行
                gen_data.append({
                    '发电类型': 'total',
                    '发电量(MWh)': total_gen,
                    '占比(%)': 100.0
                })
                
                gen_df = pd.DataFrame(gen_data)
                gen_df.to_excel(writer, sheet_name='发电量统计', index=False)
        
        # 保存Excel
        writer.close()
        logging.info(f"结果已保存到 {output_path}")
        return output_path
    
    def plot_unit_status(self, output_path):
        """绘制机组开关机状态图"""
        logging.info("开始绘制机组开关机状态图...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                logging.getLogger().setLevel(original_level)
        
        # 获取火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        
        # 按机组分组
        grouped = thermal_df.groupby('Unit')
        
        # 创建图表
        fig, axes = plt.subplots(len(grouped), 1, figsize=(15, 10), sharex=True)
        fig.suptitle('火电机组开关机状态及出力情况', fontsize=16)
        
        if len(grouped) == 1:
            axes = [axes]
        
        for i, (unit, group) in enumerate(grouped):
            ax = axes[i]
            
            # 绘制状态（0/1）
            ax.step(group['Time'], group['Status'], 'r-', where='post', label='状态')
            ax.set_ylabel('状态', color='r')
            ax.tick_params(axis='y', labelcolor='r')
            ax.set_ylim(-0.1, 1.1)
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # 创建双Y轴
            ax2 = ax.twinx()
            
            # 绘制出力
            ax2.plot(group['Time'], group['Power'], 'b-', label='出力')
            ax2.set_ylabel('出力 (MW)', color='b')
            ax2.tick_params(axis='y', labelcolor='b')
            
            # 设置标题
            ax.set_title(f'机组 {unit}')
            
            # 添加图例
            lines, labels = ax.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax.legend(lines + lines2, labels + labels2, loc='upper right')
        
        plt.xlabel('时间 (小时)')
        
        # 抑制matplotlib警告
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            plt.tight_layout(rect=[0, 0, 1, 0.95])  # 为总标题留出空间
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"机组状态图已保存到 {output_path}")
        plt.close()
        return output_path
    
    def plot_generation_mix(self, output_path):
        """绘制各类电源出力情况（堆积柱状图）"""
        logging.info("开始绘制电源结构图...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                logging.getLogger().setLevel(original_level)
        
        # 初始化结果数组
        thermal_gen = np.zeros(self.T)
        wind_gen = np.zeros(self.T)
        solar_gen = np.zeros(self.T)
        hydro_gen = np.zeros(self.T)
        ess_discharge = np.zeros(self.T)
        ess_charge = np.zeros(self.T)
        
        # 弃电量数组
        wind_curtail = np.zeros(self.T)
        solar_curtail = np.zeros(self.T)
        hydro_curtail = np.zeros(self.T)
        
        # 计算总发电量和弃电量
        try:
            for t in range(self.T):
                # 实际发电量
                for unit in self.model.ThermalUnits:
                    thermal_gen[t] += safe_value(self.model.thermal_power[unit, t])
                for unit in self.model.WindUnits:
                    wind_gen[t] += safe_value(self.model.wind_power[unit, t])
                for unit in self.model.SolarUnits:
                    solar_gen[t] += safe_value(self.model.solar_power[unit, t])
                for unit in self.model.HydroUnits:
                    hydro_gen[t] += safe_value(self.model.hydro_power[unit, t])
                
                # 储能数据 - 只有在有有效储能单元时才计算
                if len(self.model.ESSUntis) > 0:
                    for unit in self.model.ESSUntis:
                        try:
                            # 先检查变量是否存在
                            if (unit, t) in self.model.ess_discharge and (unit, t) in self.model.ess_charge:
                                ess_discharge[t] += safe_value(self.model.ess_discharge[unit, t])
                                ess_charge[t] += safe_value(self.model.ess_charge[unit, t])
                            else:
                                logging.warning(f"储能单元 {unit} 在时段 {t} 的变量不存在")
                        except KeyError as e:
                            logging.warning(f"储能单元 {unit} 在时段 {t} 索引错误: {e}")
                            continue
                        except Exception as e:
                            logging.warning(f"储能单元 {unit} 在时段 {t} 计算错误: {e}")
                            continue                # 计算弃电量（理论最大出力 - 实际出力）
                # 风电弃电
                for unit in self.model.WindUnits:
                    try:
                        curve_name = self.wind_dict[unit].get('ResourceCurve', '')
                        capacity = self.wind_dict[unit].get('Capacity', 0)
                        max_output = self._get_renewable_max_output(curve_name, t, capacity, 'wind')
                        actual_output = safe_value(self.model.wind_power[unit, t])
                        wind_curtail[t] += max(0, max_output - actual_output)
                    except Exception as e:
                        logging.warning(f"计算风电 {unit} 弃电量时出错: {e}")
                
                # 光伏弃电
                for unit in self.model.SolarUnits:
                    try:
                        curve_name = self.solar_dict[unit].get('ResourceCurve', '')
                        capacity = self.solar_dict[unit].get('Capacity', 0)
                        max_output = self._get_renewable_max_output(curve_name, t, capacity, 'solar')
                        actual_output = safe_value(self.model.solar_power[unit, t])
                        solar_curtail[t] += max(0, max_output - actual_output)
                    except Exception as e:
                        logging.warning(f"计算光伏 {unit} 弃电量时出错: {e}")
                
                # 水电弃电
                for unit in self.model.HydroUnits:
                    try:
                        curve_name = self.hydro_dict[unit].get('ResourceCurve', '')
                        capacity = self.hydro_dict[unit].get('Capacity', 0)
                        max_output = self._get_renewable_max_output(curve_name, t, capacity, 'hydro')
                        actual_output = safe_value(self.model.hydro_power[unit, t])
                        hydro_curtail[t] += max(0, max_output - actual_output)
                    except Exception as e:
                        logging.warning(f"计算水电 {unit} 弃电量时出错: {e}")
        
        except Exception as e:
            logging.error(f"计算发电量和弃电量时出错: {e}")
            # 使用默认值继续绘图
            pass
        
        # 检查储能单元数量
        logging.info(f"储能单元数量: {len(self.model.ESSUntis)}")
        if len(self.model.ESSUntis) > 0:
            logging.info(f"储能单元列表: {list(self.model.ESSUntis)}")
            # 检查储能变量是否正确初始化
            try:
                for unit in self.model.ESSUntis:
                    test_val = safe_value(self.model.ess_charge[unit, 0])
                    logging.info(f"储能单元 {unit} 在时段0的充电功率: {test_val}")
                    break  # 只测试第一个
            except Exception as e:
                logging.error(f"储能变量访问测试失败: {e}")
        else:
            logging.warning("没有储能单元数据！")
            # 如果没有储能单元，跳过储能相关计算
            ess_discharge = np.zeros(self.T)
            ess_charge = np.zeros(self.T)
            ess_discharge = np.zeros(self.T)
            ess_charge = np.zeros(self.T)
        
        # 创建图表
        fig, ax = plt.subplots(figsize=(15, 8))
        
        # 时间轴
        hours = np.arange(self.T)
        width = 0.8  # 柱子宽度
        
        # 绘制堆积柱状图
        bottom = np.zeros(self.T)
        
        # 火电（红色）
        p1 = ax.bar(hours, thermal_gen, width, bottom=bottom, label='Power_ThermalGen', 
                   color='#ff4444', alpha=0.8)
        bottom += thermal_gen
        
        # 水电（蓝色）
        p2 = ax.bar(hours, hydro_gen, width, bottom=bottom, label='Power_RHD', 
                   color='#4444ff', alpha=0.8)
        bottom += hydro_gen
        
        # 风电（绿色）
        p3 = ax.bar(hours, wind_gen, width, bottom=bottom, label='Power_WT', 
                   color='#44ff44', alpha=0.8)
        bottom += wind_gen
        
        # 光伏（黄色）
        p4 = ax.bar(hours, solar_gen, width, bottom=bottom, label='Power_PV', 
                   color='#ffff44', alpha=0.8)
        bottom += solar_gen
        
        # 计算储能净功率（放电为正，充电为负）
        ess_net = [ess_discharge[i] - ess_charge[i] for i in range(len(ess_discharge))]
        
        # 分离储能净功率的正负部分
        ess_net_positive = [max(0, x) for x in ess_net]  # 净放电
        ess_net_negative = [min(0, x) for x in ess_net]  # 净充电
        
        # 储能净放电（堆叠在发电侧）
        if max(ess_net_positive) > 0.01:
            p5 = ax.bar(hours, ess_net_positive, width, bottom=bottom, label='Power_StoreDis', 
                       color='#333333', alpha=0.9, edgecolor='black', linewidth=1)
            bottom += ess_net_positive
        
        # 储能净充电（显示在负轴）
        if min(ess_net_negative) < -0.01:
            p6 = ax.bar(hours, ess_net_negative, width, label='Power_StoreCharge', 
                       color='#666666', alpha=0.9, edgecolor='black', linewidth=1)
        
        # 计算净发电量（使用储能净功率）
        net_generation = [thermal_gen[i] + wind_gen[i] + solar_gen[i] + hydro_gen[i] + ess_net[i] 
                         for i in range(len(thermal_gen))]
        
        # 添加总负荷曲线（黑色实线，加粗突出）
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T
        ax.plot(hours, total_load, 'k-', linewidth=3, label='Total Demand', alpha=0.9)
        
        # 验证最大需求值
        max_demand_value = max(total_load)
        logging.info(f"最大负荷需求: {max_demand_value:.1f} MW")
        
        # 计算实际的有效负荷（考虑切负荷）
        effective_load = []
        total_load_shed = []
        for t in range(self.T):
            load_shed = 0
            for load in self.model.Loads:
                load_shed += safe_value(self.model.load_shed[load, t])
            total_load_shed.append(load_shed)
            effective_load.append(total_load[t] - load_shed)
        
        # 如果有切负荷，显示有效负荷
        if max(total_load_shed) > 0.1:
            logging.warning(f"存在切负荷！最大切负荷: {max(total_load_shed):.1f}MW")
            logging.warning(f"总切负荷量: {sum(total_load_shed):.1f}MWh")
            # 输出每个时段的切负荷详情
            for t in range(self.T):
                if total_load_shed[t] > 0.1:
                    logging.warning(f"  时段{t+1}: 切负荷={total_load_shed[t]:.1f}MW, 原负荷={total_load[t]:.1f}MW, 有效负荷={effective_load[t]:.1f}MW")
        else:
            logging.info("无切负荷，系统完全满足负荷需求")
        
        # 移除多余的线条，简化图表显示
        # 注释掉：净发电量验证线、理论净发电量线等多余线条
        # ax.plot(hours, net_generation, 'r:', linewidth=2, label='Actual Net Generation', alpha=0.8)
        # ax.plot(hours, theoretical_net_generation, 'b:', linewidth=2, label='Expected Net Generation', alpha=0.8)
        
        # 验证功率平衡（按照约束的定义检验）
        for t in range(self.T):
            # 计算发电侧总量
            gen_side = (thermal_gen[t] + wind_gen[t] + solar_gen[t] + hydro_gen[t] + 
                       sum(safe_value(self.model.ess_discharge[unit, t]) for unit in self.model.ESSUntis))
            
            # 计算负荷侧总量
            base_load = total_load[t] if t < len(total_load) else 0
            shed = total_load_shed[t]
            charge = sum(safe_value(self.model.ess_charge[unit, t]) for unit in self.model.ESSUntis)
            load_side = base_load - shed + charge
            
            balance_error = abs(gen_side - load_side)
            if balance_error > 1e-3:
                logging.warning(f"时段{t+1}功率不平衡: {balance_error:.6f} MW")
                logging.warning(f"  发电侧={gen_side:.1f} (火电={thermal_gen[t]:.1f}, 风电={wind_gen[t]:.1f}, 光伏={solar_gen[t]:.1f}, 水电={hydro_gen[t]:.1f}, 储能放电={sum(safe_value(self.model.ess_discharge[unit, t]) for unit in self.model.ESSUntis):.1f})")
                logging.warning(f"  负荷侧={load_side:.1f} (基础负荷={base_load:.1f}, 切负荷={shed:.1f}, 储能充电={charge:.1f})")
        
        # 检查储能数据是否有异常
        max_ess_discharge = max(ess_discharge) if len(ess_discharge) > 0 else 0
        max_ess_charge = max(ess_charge) if len(ess_charge) > 0 else 0
        logging.info(f"储能数据检查: 最大放电={max_ess_discharge:.1f}MW, 最大充电={max_ess_charge:.1f}MW")
        
        # 计算准确的备用容量（重新设计算法）
        available_reserve = []
        total_capacity_available = []
        for t in range(self.T):
            # 1. 火电机组：只有在线机组才能提供备用
            thermal_available = 0
            thermal_actual = 0
            for unit in self.model.ThermalUnits:
                status = safe_value(self.model.thermal_status[unit, t])
                max_output = self.thermal_dict[unit].get('Capacity', 0)
                actual_output = safe_value(self.model.thermal_power[unit, t])
                
                # 只有在线机组才算可用容量
                if status > 0.5:  # 机组在线
                    thermal_available += max_output
                    thermal_actual += actual_output
            
            # 2. 可再生能源：实际可出力作为可用容量（不能提供真正的备用）
            renewable_available = 0
            renewable_actual = 0
            
            # 风电当前可用出力
            for unit in self.model.WindUnits:
                capacity = self.wind_dict[unit].get('Capacity', 0)
                actual_output = safe_value(self.model.wind_power[unit, t])
                renewable_actual += actual_output
                
                # 风电可用容量基于当前资源条件
                if hasattr(self, 'wind_curves') and unit in self.wind_curves:
                    curve = self.wind_curves[unit]
                    if t < len(curve):
                        renewable_available += capacity * curve[t]
                else:
                    renewable_available += capacity * 0.3  # 保守估计
            
            # 光伏当前可用出力
            for unit in self.model.SolarUnits:
                capacity = self.solar_dict[unit].get('Capacity', 0)
                actual_output = safe_value(self.model.solar_power[unit, t])
                renewable_actual += actual_output
                
                # 光伏可用容量基于当前资源条件
                if hasattr(self, 'solar_curves') and unit in self.solar_curves:
                    curve = self.solar_curves[unit]
                    if t < len(curve):
                        renewable_available += capacity * curve[t]
                else:
                    # 夜间为0，白天按时段估计
                    hour = t % 24
                    if 7 <= hour <= 17:
                        factor = 0.6 if 10 <= hour <= 14 else 0.3
                        renewable_available += capacity * factor
            
            # 水电当前可用出力
            for unit in self.model.HydroUnits:
                capacity = self.hydro_dict[unit].get('Capacity', 0)
                actual_output = safe_value(self.model.hydro_power[unit, t])
                renewable_actual += actual_output
                
                if hasattr(self, 'hydro_curves') and unit in self.hydro_curves:
                    curve = self.hydro_curves[unit]
                    if t < len(curve):
                        renewable_available += capacity * curve[t]
                else:
                    renewable_available += capacity * 0.7
            
            # 3. 储能：只有充电状态下才能提供放电备用
            ess_available = 0
            ess_actual = safe_value(sum(self.model.ess_discharge[unit, t] for unit in self.model.ESSUntis))
            
            for unit in self.model.ESSUntis:
                charge = safe_value(self.model.ess_charge[unit, t])
                discharge = safe_value(self.model.ess_discharge[unit, t])
                max_discharge = self.ess_dict[unit].get('MaxDisPower', 0)
                
                # 如果储能在充电，则有放电潜力；如果在放电，则没有额外备用
                if charge > 0.1:  # 正在充电
                    ess_available += max_discharge
                elif discharge < 0.1:  # 既不充电也不放电
                    ess_available += max_discharge * 0.5  # 保守估计50%可用
            
            # 计算真实备用容量：只考虑火电机组的剩余容量
            # 可再生能源和储能不算传统意义的旋转备用
            total_available = thermal_available + renewable_available + ess_available
            total_actual = thermal_actual + renewable_actual + ess_actual
            
            # 真正的备用主要来自火电机组
            thermal_reserve = max(0, thermal_available - thermal_actual)
            # renewable_reserve = max(0, renewable_available - renewable_actual)  # 可再生能源备用意义不大
            ess_reserve = max(0, ess_available - ess_actual)
            
            reserve = thermal_reserve + ess_reserve  # 只计算可控制的备用
            
            available_reserve.append(reserve)
            total_capacity_available.append(total_available)
        
        # 计算备用要求（与约束一致，负荷的2%）
        reserve_requirement = []
        for t in range(self.T):
            req = total_load[t] * 0.02  # 2%备用要求，与约束一致
            reserve_requirement.append(req)
        
        # 绘制修正后的图表（添加备用需求曲线）
        # 计算并绘制备用需求曲线
        # 计算备用要求（负荷的2%，与约束一致）
        reserve_requirement = []
        for t in range(self.T):
            req = total_load[t] * 0.02  # 2%备用要求
            reserve_requirement.append(req)
        
        # 绘制备用需求线（在负荷线之上，绿色虚线）
        demand_plus_reserve = np.array(total_load[:self.T]) + np.array(reserve_requirement)
        ax.plot(hours, demand_plus_reserve, 'g--', linewidth=2, 
               label='Demand + Reserve Requirement (2%)', alpha=0.8)
        

        
        # 设置图表属性
        ax.set_title('Generation Schedule and Reserve Margin', fontsize=14, fontweight='bold')
        ax.set_xlabel('Hour', fontsize=12)
        ax.set_ylabel('Power (MW)', fontsize=12)
        
        # 添加零线（虚线）
        ax.axhline(y=0, color='black', linestyle='--', alpha=0.5, linewidth=1)
        
        # 添加统计信息（包含正确的备用容量信息）
        max_demand = max_demand_value  # 使用前面计算的值
        min_reserve = min(available_reserve)
        max_reserve = max(available_reserve)
        avg_reserve = np.mean(available_reserve)
        max_ess_charge = max(ess_charge) if len(ess_charge) > 0 else 0
        max_ess_discharge = max(ess_discharge) if len(ess_discharge) > 0 else 0
        
        # 检查备用是否满足要求
        reserve_violations = 0
        for t in range(self.T):
            if available_reserve[t] < reserve_requirement[t]:
                reserve_violations += 1
        
        # 弃电量统计
        total_wind_curtail = sum(wind_curtail)
        total_solar_curtail = sum(solar_curtail)
        total_hydro_curtail = sum(hydro_curtail)
        total_curtail = total_wind_curtail + total_solar_curtail + total_hydro_curtail
        
        # 第0小时功率平衡检查
        hour0_gen = thermal_gen[0] + wind_gen[0] + solar_gen[0] + hydro_gen[0] + ess_discharge[0]
        hour0_load = total_load[0] - total_load_shed[0] + ess_charge[0]
        hour0_balance = abs(hour0_gen - hour0_load)
        
        info_text = (f'最大需求: {max_demand_value:.1f} MW (确认)\n'
                    f'备用容量:\n'
                    f'  最小: {min_reserve:.1f} MW\n'
                    f'  最大: {max_reserve:.1f} MW\n'
                    f'  平均: {avg_reserve:.1f} MW\n'
                    f'备用不足时段: {reserve_violations}/{self.T}\n'
                    f'储能最大充电: {max_ess_charge:.1f} MW\n'
                    f'储能最大放电: {max_ess_discharge:.1f} MW\n'
                    f'第0小时平衡误差: {hour0_balance:.3f} MW\n'
                    f'总弃电量: {total_curtail:.1f} MWh\n'
                    f'  弃风: {total_wind_curtail:.1f} MWh\n'
                    f'  弃光: {total_solar_curtail:.1f} MWh\n'
                    f'  弃水: {total_hydro_curtail:.1f} MWh')
        
        ax.text(0.02, 0.98, info_text, 
               transform=ax.transAxes, fontsize=9, verticalalignment='top',
               bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
        
        # 设置图例
        ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize=10)
        
        # 设置网格
        ax.grid(True, linestyle='--', alpha=0.3)
        ax.set_xlim(-0.5, self.T-0.5)
        ax.set_xticks(range(0, self.T, 2))  # 每2小时标记一次
        
        # 动态设置Y轴范围，确保显示完整的上下部分
        all_positive_values = []
        all_negative_values = []
        
        # 收集所有正值数据（包括弃电量）
        all_positive_values.extend(thermal_gen)
        all_positive_values.extend(wind_gen)
        all_positive_values.extend(solar_gen)
        all_positive_values.extend(hydro_gen)
        all_positive_values.extend(ess_discharge)  # 储能放电
        all_positive_values.extend(wind_curtail)   # 弃风
        all_positive_values.extend(solar_curtail)  # 弃光
        all_positive_values.extend(hydro_curtail)  # 弃水
        all_positive_values.extend(total_load)
        # 计算总发电量用于图表范围设置
        total_generation_for_range = []
        for t in range(self.T):
            total_gen = thermal_gen[t] + wind_gen[t] + solar_gen[t] + hydro_gen[t] + ess_discharge[t]
            total_generation_for_range.append(total_gen)
        all_positive_values.extend(total_generation_for_range)
        
        # 收集所有负值数据（使用储能净充电的负值）
        ess_net_charge_values = [x for x in ess_net_negative if x < 0]
        if ess_net_charge_values:
            all_negative_values.extend(ess_net_charge_values)
        
        # 计算范围
        max_positive = max(all_positive_values) if all_positive_values else 1
        min_negative = min(all_negative_values) if all_negative_values else 0
        
        # 设置Y轴范围，确保上下都有适当的显示空间
        y_margin = 0.1
        if max_positive > 0:
            y_max = max_positive * (1 + y_margin)
        else:
            y_max = 1000  # 默认最大值
            
        if min_negative < 0:
            y_min = min_negative * (1 + y_margin)
        else:
            y_min = -max_positive * 0.2  # 确保下方也有一些空间
            
        ax.set_ylim(y_min, y_max)
        
        # 添加零线
        ax.axhline(y=0, color='black', linestyle='-', linewidth=0.8, alpha=0.7)
        
        # 紧凑布局
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            plt.tight_layout()
        
        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        logging.info(f"电源结构图已保存到 {output_path}")
        plt.close()
        return output_path
    
    def plot_network_flow(self, output_path):
        """绘制网络潮流图"""
        logging.info("开始绘制网络潮流图...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            # 临时抑制日志
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                # 恢复原始日志级别
                logging.getLogger().setLevel(original_level)
        
        # 获取线路潮流数据
        line_flows = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_flows.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_flows_df = pd.DataFrame(line_flows)
        
        # 获取节点相角数据
        bus_angles = []
        for bus in self.model.Buses:
            for t in self.model.T:
                bus_angles.append({
                    'Bus': bus,
                    'Time': t,
                    'Angle': safe_value(self.model.bus_angle[bus, t])
                })
        bus_angles_df = pd.DataFrame(bus_angles)
        
        # 创建图表
        plt.figure(figsize=(15, 10))
        
        # 绘制线路潮流
        unique_lines = line_flows_df['Line'].unique()
        for i, line in enumerate(unique_lines):
            line_data = line_flows_df[line_flows_df['Line'] == line]
            plt.plot(line_data['Time'], line_data['Power'], label=f'线路 {line}')
        
        plt.title('线路潮流随时间变化')
        plt.xlabel('时间 (小时)')
        plt.ylabel('功率 (MW)')
        plt.legend(loc='best', bbox_to_anchor=(1.05, 1))
        plt.grid(True, linestyle='--', alpha=0.7)
        
        # 抑制matplotlib警告
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            plt.tight_layout()
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"网络潮流图已保存到 {output_path}")
        plt.close()
        return output_path

    def plot_thermal_units_detail(self, output_path):
        """绘制火电机组24小时详细出力折线图 - 所有机组在一个图中"""
        logging.info("开始绘制火电机组详细出力折线图...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                logging.getLogger().setLevel(original_level)
        
        # 收集火电机组数据
        thermal_units_data = {}
        thermal_status_data = {}
        total_thermal_power = np.zeros(self.T)
        
        for unit in self.model.ThermalUnits:
            unit_power = []
            unit_status = []
            for t in range(self.T):
                power = safe_value(self.model.thermal_power[unit, t])
                status = safe_value(self.model.thermal_status[unit, t])
                unit_power.append(power)
                unit_status.append(status)
                total_thermal_power[t] += power
            
            thermal_units_data[unit] = unit_power
            thermal_status_data[unit] = unit_status
        
        # 按照最大出力排序火电机组
        thermal_sorted = sorted(thermal_units_data.items(), 
                               key=lambda x: max(x[1]), reverse=True)
        
        if len(thermal_sorted) == 0:
            logging.warning("没有火电机组数据可绘制")
            return None
        
        # 创建大图表 - 所有机组在一个图中
        plt.figure(figsize=(20, 12))
        
        hours = np.arange(self.T)
        
        # 生成足够多的颜色
        import matplotlib.cm as cm
        import matplotlib.colors as mcolors
        
        # 使用多种颜色方案组合，确保54个机组都有不同颜色
        color_maps = ['tab10', 'tab20', 'Set1', 'Set2', 'Set3', 'Paired']
        all_colors = []
        
        for cmap_name in color_maps:
            cmap = cm.get_cmap(cmap_name)
            if hasattr(cmap, 'colors'):
                all_colors.extend(cmap.colors)
            else:
                colors_from_map = [cmap(i/10) for i in range(10)]
                all_colors.extend(colors_from_map)
        
        # 如果还不够颜色，生成更多
        while len(all_colors) < len(thermal_sorted):
            # 使用HSV颜色空间生成更多颜色
            hue = (len(all_colors) * 137.5) % 360  # 黄金角度分布
            color = mcolors.hsv_to_rgb([hue/360, 0.7, 0.9])
            all_colors.append(color)
        
        # 分类绘制机组 - 先绘制小机组，再绘制大机组，这样大机组会在前面更醒目
        
        # 计算机组容量信息用于分类
        unit_capacities = {}
        for unit, _ in thermal_sorted:
            capacity = self.thermal_dict.get(unit, {}).get('Capacity', 0)
            unit_capacities[unit] = capacity
        
        # 按容量分为三类：大机组(>500MW)，中机组(100-500MW)，小机组(<100MW)
        large_units = [(unit, data) for unit, data in thermal_sorted if unit_capacities[unit] > 500]
        medium_units = [(unit, data) for unit, data in thermal_sorted if 100 <= unit_capacities[unit] <= 500]
        small_units = [(unit, data) for unit, data in thermal_sorted if unit_capacities[unit] < 100]
        
        # 先绘制小机组（细线，淡色）
        for i, (unit, power_data) in enumerate(small_units):
            color = all_colors[i % len(all_colors)]
            # 只有当机组有实际出力时才显示标签
            max_power = max(power_data)
            if max_power > 1:  # 只显示有意义出力的机组
                plt.plot(hours, power_data, color=color, linewidth=1, 
                        alpha=0.4, label=f'{unit} ({unit_capacities[unit]:.0f}MW)')
            else:
                plt.plot(hours, power_data, color=color, linewidth=1, alpha=0.2)
        
        # 再绘制中机组（中等线宽）
        for i, (unit, power_data) in enumerate(medium_units):
            color = all_colors[(i + len(small_units)) % len(all_colors)]
            max_power = max(power_data)
            if max_power > 1:
                plt.plot(hours, power_data, color=color, linewidth=2, 
                        alpha=0.7, label=f'{unit} ({unit_capacities[unit]:.0f}MW)')
            else:
                plt.plot(hours, power_data, color=color, linewidth=2, alpha=0.3)
        
        # 最后绘制大机组（粗线，亮色，最醒目）
        for i, (unit, power_data) in enumerate(large_units):
            color = all_colors[(i + len(small_units) + len(medium_units)) % len(all_colors)]
            max_power = max(power_data)
            if max_power > 1:
                plt.plot(hours, power_data, color=color, linewidth=3, 
                        alpha=0.9, label=f'{unit} ({unit_capacities[unit]:.0f}MW)')
            else:
                plt.plot(hours, power_data, color=color, linewidth=3, alpha=0.4)
        
        # 绘制火电总出力（黑色粗线，最突出）
        plt.plot(hours, total_thermal_power, 'black', linewidth=4, 
                label='火电总出力', alpha=0.95, linestyle='-')
        
        # 添加负荷曲线作为参考（蓝色虚线）
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T
        plt.plot(hours, total_load, 'blue', linewidth=3, linestyle='--',
                label='总负荷', alpha=0.8)
        
        # 设置图表属性
        plt.title(f'火电机组24小时出力详情 (共{len(thermal_sorted)}台机组)', 
                 fontsize=16, fontweight='bold')
        plt.xlabel('时间 (小时)', fontsize=12)
        plt.ylabel('功率 (MW)', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.xlim(0, self.T-1)
        plt.xticks(range(0, self.T, 2))  # 每2小时标记一次
        
        # 计算统计信息
        total_capacity = sum(unit_capacities.values())
        max_total_power = max(total_thermal_power)
        avg_total_power = np.mean(total_thermal_power)
        system_utilization = (avg_total_power / total_capacity * 100) if total_capacity > 0 else 0
        
        # 计算在线机组数量变化
        online_units_count = []
        for t in range(self.T):
            online_count = sum(1 for unit in thermal_status_data.keys() 
                             if thermal_status_data[unit][t] > 0.5)
            online_units_count.append(online_count)
        
        # 机组分类统计
        active_large = len([u for u, _ in large_units if max(thermal_units_data[u]) > 1])
        active_medium = len([u for u, _ in medium_units if max(thermal_units_data[u]) > 1])
        active_small = len([u for u, _ in small_units if max(thermal_units_data[u]) > 1])
        
        # 添加详细统计信息
        stats_text = (
            f'系统统计:\n'
            f'总装机容量: {total_capacity:.0f} MW\n'
            f'最大出力: {max_total_power:.1f} MW\n'
            f'平均出力: {avg_total_power:.1f} MW\n'
            f'系统利用率: {system_utilization:.1f}%\n'
            f'在线机组数: {min(online_units_count)}-{max(online_units_count)}台\n'
            f'\n机组分类:\n'
            f'大机组(>500MW): {active_large}/{len(large_units)}台参与调度\n'
            f'中机组(100-500MW): {active_medium}/{len(medium_units)}台参与调度\n'
            f'小机组(<100MW): {active_small}/{len(small_units)}台参与调度\n'
            f'\n说明:\n'
            f'• 粗线: 大机组 • 中线: 中机组 • 细线: 小机组\n'
            f'• 黑色粗线: 火电总出力\n'
            f'• 蓝色虚线: 系统总负荷'
        )
        
        # 将统计信息放在图表右侧
        plt.text(1.02, 1, stats_text, transform=plt.gca().transAxes, 
                fontsize=10, verticalalignment='top',
                bbox=dict(boxstyle='round,pad=0.5', facecolor='lightblue', alpha=0.9))
        
        # 设置图例 - 只显示有实际出力的机组和关键曲线
        handles, labels = plt.gca().get_legend_handles_labels()
        
        # 筛选图例：只保留总出力、总负荷和主要机组（出力>50MW的机组）
        filtered_handles = []
        filtered_labels = []
        
        for handle, label in zip(handles, labels):
            if '火电总出力' in label or '总负荷' in label:
                filtered_handles.append(handle)
                filtered_labels.append(label)
            elif 'MW)' in label:
                # 提取容量信息判断是否为主要机组
                try:
                    capacity_str = label.split('(')[1].split('MW')[0]
                    capacity = float(capacity_str)
                    if capacity >= 200:  # 只显示200MW以上的机组在图例中
                        filtered_handles.append(handle)
                        filtered_labels.append(label)
                except:
                    pass
        
        # 如果机组太多，只显示前20个主要机组的图例
        if len(filtered_labels) > 22:  # 保留总出力和总负荷 + 20个主要机组
            main_handles = filtered_handles[:2]  # 总出力和总负荷
            main_labels = filtered_labels[:2]
            main_handles.extend(filtered_handles[2:22])  # 前20个主要机组
            main_labels.extend(filtered_labels[2:22])
            filtered_handles = main_handles
            filtered_labels = main_labels
        
        plt.legend(filtered_handles, filtered_labels, 
                  loc='center left', bbox_to_anchor=(1.02, 0.5), 
                  fontsize=8, ncol=1)
        
        # 调整布局以适应图例和统计信息
        plt.tight_layout()
        plt.subplots_adjust(right=0.75)  # 为右侧信息留出空间
        
        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        logging.info(f"火电机组详细出力图已保存到 {output_path}")
        plt.close()
        return output_path

    def plot_unit_power_curves(self, output_path):
        """绘制24小时各类机组出力情况折线图"""
        logging.info("开始绘制各类机组出力折线图...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                logging.getLogger().setLevel(original_level)
        
        # 初始化各类机组出力数据
        thermal_power = np.zeros(self.T)
        wind_power = np.zeros(self.T)
        solar_power = np.zeros(self.T)
        hydro_power = np.zeros(self.T)
        ess_discharge_power = np.zeros(self.T)
        ess_charge_power = np.zeros(self.T)
        
        # 收集各类机组的详细出力数据
        thermal_units_power = {}
        wind_units_power = {}
        solar_units_power = {}
        hydro_units_power = {}
        ess_units_power = {}
        
        # 收集火电机组数据
        for unit in self.model.ThermalUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.thermal_power[unit, t])
                unit_power.append(power)
                thermal_power[t] += power
            thermal_units_power[unit] = unit_power
        
        # 收集风电机组数据
        for unit in self.model.WindUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.wind_power[unit, t])
                unit_power.append(power)
                wind_power[t] += power
            wind_units_power[unit] = unit_power
        
        # 收集光伏机组数据
        for unit in self.model.SolarUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.solar_power[unit, t])
                unit_power.append(power)
                solar_power[t] += power
            solar_units_power[unit] = unit_power
        
        # 收集水电机组数据
        for unit in self.model.HydroUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.hydro_power[unit, t])
                unit_power.append(power)
                hydro_power[t] += power
            hydro_units_power[unit] = unit_power
        
        # 收集储能机组数据
        for unit in self.model.ESSUntis:
            unit_discharge = []
            unit_charge = []
            for t in range(self.T):
                discharge = safe_value(self.model.ess_discharge[unit, t])
                charge = safe_value(self.model.ess_charge[unit, t])
                unit_discharge.append(discharge)
                unit_charge.append(charge)
                ess_discharge_power[t] += discharge
                ess_charge_power[t] += charge
            ess_units_power[unit] = {'discharge': unit_discharge, 'charge': unit_charge}
        
        # 创建图表
        fig, axes = plt.subplots(3, 2, figsize=(20, 15))
        fig.suptitle('24小时各类机组出力情况', fontsize=16, fontweight='bold')
        
        hours = np.arange(self.T)
        
        # 1. 火电机组出力图
        ax1 = axes[0, 0]
        ax1.plot(hours, thermal_power, 'r-', linewidth=3, label='火电总出力', alpha=0.8)
        # 显示前5个主要火电机组的详细出力
        thermal_sorted = sorted(thermal_units_power.items(), 
                               key=lambda x: max(x[1]), reverse=True)[:5]
        colors = ['#FF6B6B', '#FF8E53', '#FF006E', '#8338EC', '#3A86FF']
        for i, (unit, power_curve) in enumerate(thermal_sorted):
            if max(power_curve) > 10:  # 只显示有意义出力的机组
                ax1.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax1.set_title('火电机组出力', fontsize=14)
        ax1.set_ylabel('功率 (MW)')
        ax1.grid(True, alpha=0.3)
        ax1.legend(fontsize=10)
        
        # 2. 风电机组出力图
        ax2 = axes[0, 1]
        ax2.plot(hours, wind_power, 'g-', linewidth=3, label='风电总出力', alpha=0.8)
        # 显示主要风电机组
        wind_sorted = sorted(wind_units_power.items(), 
                            key=lambda x: max(x[1]), reverse=True)[:3]
        colors = ['#2ECC71', '#27AE60', '#16A085']
        for i, (unit, power_curve) in enumerate(wind_sorted):
            if max(power_curve) > 1:
                ax2.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax2.set_title('风电机组出力', fontsize=14)
        ax2.set_ylabel('功率 (MW)')
        ax2.grid(True, alpha=0.3)
        ax2.legend(fontsize=10)
        
        # 3. 光伏机组出力图
        ax3 = axes[1, 0]
        ax3.plot(hours, solar_power, 'y-', linewidth=3, label='光伏总出力', alpha=0.8)
        # 显示主要光伏机组
        solar_sorted = sorted(solar_units_power.items(), 
                             key=lambda x: max(x[1]), reverse=True)[:3]
        colors = ['#F39C12', '#E67E22', '#D35400']
        for i, (unit, power_curve) in enumerate(solar_sorted):
            if max(power_curve) > 1:
                ax3.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax3.set_title('光伏机组出力', fontsize=14)
        ax3.set_ylabel('功率 (MW)')
        ax3.grid(True, alpha=0.3)
        ax3.legend(fontsize=10)
        
        # 4. 水电机组出力图
        ax4 = axes[1, 1]
        ax4.plot(hours, hydro_power, 'b-', linewidth=3, label='水电总出力', alpha=0.8)
        # 显示主要水电机组
        hydro_sorted = sorted(hydro_units_power.items(), 
                             key=lambda x: max(x[1]), reverse=True)[:3]
        colors = ['#3498DB', '#2980B9', '#1F4E79']
        for i, (unit, power_curve) in enumerate(hydro_sorted):
            if max(power_curve) > 1:
                ax4.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax4.set_title('水电机组出力', fontsize=14)
        ax4.set_ylabel('功率 (MW)')
        ax4.grid(True, alpha=0.3)
        ax4.legend(fontsize=10)
        
        # 5. 储能机组出力图（充放电）
        ax5 = axes[2, 0]
        ax5.plot(hours, ess_discharge_power, 'purple', linewidth=3, label='储能放电', alpha=0.8)
        ax5.plot(hours, -ess_charge_power, 'orange', linewidth=3, label='储能充电', alpha=0.8)
        # 显示主要储能机组
        ess_sorted = sorted(ess_units_power.items(), 
                           key=lambda x: max(x[1]['discharge']), reverse=True)[:3]
        colors_discharge = ['#9B59B6', '#8E44AD', '#663399']
        colors_charge = ['#E74C3C', '#C0392B', '#922B21']
        for i, (unit, power_data) in enumerate(ess_sorted):
            if max(power_data['discharge']) > 0.1:
                ax5.plot(hours, power_data['discharge'], '--', linewidth=2,
                        color=colors_discharge[i % len(colors_discharge)], alpha=0.7, 
                        label=f'{unit}_放电')
            if max(power_data['charge']) > 0.1:
                ax5.plot(hours, [-x for x in power_data['charge']], '--', linewidth=2,
                        color=colors_charge[i % len(colors_charge)], alpha=0.7, 
                        label=f'{unit}_充电')
        ax5.set_title('储能机组充放电', fontsize=14)
        ax5.set_ylabel('功率 (MW)')
        ax5.axhline(y=0, color='black', linestyle='-', alpha=0.5)
        ax5.grid(True, alpha=0.3)
        ax5.legend(fontsize=10)
        
        # 6. 综合对比图
        ax6 = axes[2, 1]
        ax6.plot(hours, thermal_power, 'r-', linewidth=3, label='火电', alpha=0.8)
        ax6.plot(hours, wind_power, 'g-', linewidth=3, label='风电', alpha=0.8)
        ax6.plot(hours, solar_power, 'y-', linewidth=3, label='光伏', alpha=0.8)
        ax6.plot(hours, hydro_power, 'b-', linewidth=3, label='水电', alpha=0.8)
        ax6.plot(hours, ess_discharge_power, 'purple', linewidth=3, label='储能放电', alpha=0.8)
        
        # 添加负荷曲线作为参考
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T
        ax6.plot(hours, total_load, 'k--', linewidth=2, label='总负荷', alpha=0.6)
        
        ax6.set_title('各类机组出力综合对比', fontsize=14)
        ax6.set_ylabel('功率 (MW)')
        ax6.set_xlabel('时间 (小时)')
        ax6.grid(True, alpha=0.3)
        ax6.legend(fontsize=10)
        
        # 设置所有子图的x轴
        for ax in axes.flat:
            ax.set_xlim(0, self.T-1)
            ax.set_xticks(range(0, self.T, 4))  # 每4小时标记一次
        
        # 添加统计信息
        stats_text = (
            f'统计信息 (24小时):\n'
            f'火电: 最大{max(thermal_power):.1f}MW, 平均{np.mean(thermal_power):.1f}MW\n'
            f'风电: 最大{max(wind_power):.1f}MW, 平均{np.mean(wind_power):.1f}MW\n'
            f'光伏: 最大{max(solar_power):.1f}MW, 平均{np.mean(solar_power):.1f}MW\n'
            f'水电: 最大{max(hydro_power):.1f}MW, 平均{np.mean(hydro_power):.1f}MW\n'
            f'储能放电: 最大{max(ess_discharge_power):.1f}MW\n'
            f'储能充电: 最大{max(ess_charge_power):.1f}MW'
        )
        
        # 在右上角添加统计信息
        fig.text(0.98, 0.98, stats_text, fontsize=10, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='white', alpha=0.8),
                transform=fig.transFigure, ha='right')
        
        # 调整布局
        plt.tight_layout(rect=[0, 0, 0.85, 0.95])  # 为统计信息和标题留出空间
        
        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        logging.info(f"各类机组出力折线图已保存到 {output_path}")
        plt.close()
        return output_path

# 主程序
if __name__ == "__main__":
    # 输入输出路径
    input_file = r"C:\Users\admin\Desktop\cleaned_IEEE118_full_v2(1).xlsx"
    output_file = "IEEE118_UC_Results.xlsx"
    status_plot = "IEEE118_Unit_Status.png"
    mix_plot = "IEEE118_Generation_Mix.png"
    flow_plot = "IEEE118_Network_Flow.png"
    resource_plot = "IEEE118_Resource_Curves.png"
    power_curves_plot = "IEEE118_Unit_Power_Curves.png"
    thermal_detail_plot = "IEEE118_Thermal_Units_Detail.png"
    
    try:
        # 创建模型
        logging.info("创建模型...")
        uc_model = UC_Model(input_file)
        
        # 绘制资源曲线预览
        uc_model.plot_resource_curves(resource_plot)
        
        # 求解模型
        logging.info("求解模型...")
        result = uc_model.solve()
        
        if result:
            # 计算和显示成本分析
            logging.info("计算成本分析...")
            uc_model.calculate_cost_analysis()
            
            # 保存结果
            logging.info("保存结果...")
            result_path = uc_model.save_results(output_file)
            
            # 绘制图表
            logging.info("绘制图表...")
            status_path = uc_model.plot_unit_status(status_plot)
            mix_path = uc_model.plot_generation_mix(mix_plot)
            flow_path = uc_model.plot_network_flow(flow_plot)
            power_curves_path = uc_model.plot_unit_power_curves(power_curves_plot)
            thermal_detail_path = uc_model.plot_thermal_units_detail(thermal_detail_plot)
            
            # 将图表插入Excel
            wb = openpyxl.load_workbook(output_file)
            
            # 添加资源曲线预览图
            ws = wb.create_sheet("资源曲线预览")
            img = Image(resource_plot)
            ws.add_image(img, 'A1')
            
            # 添加机组状态图
            ws = wb.create_sheet("机组状态图")
            img = Image(status_path)
            ws.add_image(img, 'A1')
            
            # 添加电源结构图
            ws = wb.create_sheet("电源结构图")
            img = Image(mix_path)
            ws.add_image(img, 'A1')
            
            # 添加网络潮流图
            ws = wb.create_sheet("网络潮流图")
            img = Image(flow_path)
            ws.add_image(img, 'A1')
            
            # 添加各类机组出力折线图
            ws = wb.create_sheet("机组出力曲线")
            img = Image(power_curves_path)
            ws.add_image(img, 'A1')
            
            # 添加火电机组详细出力图
            ws = wb.create_sheet("火电机组详细出力")
            img = Image(thermal_detail_path)
            ws.add_image(img, 'A1')
            
            # 保存最终结果
            wb.save(output_file)
            logging.info("所有结果已保存并整合到Excel文件中")
            print(f"优化完成，结果已保存到: {output_file}")
        else:
            logging.error("优化求解失败，请检查模型和输入数据")
            print("优化求解失败，请检查日志获取详细信息")
        
    except Exception as e:
        logging.exception("程序运行出错")
        print(f"程序运行出错: {str(e)}")
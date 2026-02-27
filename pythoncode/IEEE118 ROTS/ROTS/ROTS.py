import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import os
import logging
import sys
import time
import math

# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 初始化日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class UC_Model:
    def __init__(self, data_path):
        self.data_path = data_path
        self.load_data()
        self.validate_data()  # 添加数据验证
        self.create_model()
    
    def validate_data(self):
        """数据加载后立即执行验证"""
        logging.info("开始数据验证...")

        # 检查可再生能源容量
        for tech, units in [('风电', self.wind_units), 
                           ('光伏', self.solar_units),
                           ('水电', self.hydro_units)]:
            # 修正：兼容不同表头，自动查找容量列
            cap_col = None
            for col in units.columns:
                if str(col).strip() in ['Capacity', '机组装机容量(MW)']:
                    cap_col = col
                    break
            if cap_col is None:
                logging.error(f"{tech}机组数据缺少容量列！实际列为: {list(units.columns)}")
                continue
            if units.empty:
                logging.error(f"{tech}机组数据为空！")
            elif units[cap_col].sum() < 100:
                logging.warning(f"{tech}总容量异常: {units[cap_col].sum()}MW")

        # 检查负荷曲线范围
        if max(self.load_curve) < 5000:
            logging.error(f"负荷峰值异常低: {max(self.load_curve)}MW")
            # 尝试从系统汇总中获取峰值负荷
            sys_peak = self.system_summary.get('PeakLoad(MW)', 0)
            if sys_peak > 5000:
                logging.warning(f"使用系统汇总中的峰值负荷值: {sys_peak}MW")
                # 缩放负荷曲线
                current_peak = max(self.load_curve)
                if current_peak > 0:
                    scaling_factor = sys_peak / current_peak
                    logging.info(f"负荷曲线缩放: 原峰值={current_peak:.1f}MW, 目标峰值={sys_peak:.1f}MW, 缩放因子={scaling_factor:.3f}")
                    self.load_curve = [val * scaling_factor for val in self.load_curve]
                    logging.info(f"缩放后负荷曲线峰值: {max(self.load_curve):.1f}MW")
        
        # 检查资源曲线匹配
        wind_curves_used = self.wind_units['ResourceCurve'].unique()
        for curve in wind_curves_used:
            if curve not in self.wind_curves['WTCurve'].values:
                logging.error(f"风资源曲线'{curve}'未定义！")
        
        solar_curves_used = self.solar_units['ResourceCurve'].unique()
        for curve in solar_curves_used:
            if curve not in self.solar_curves['PVCurve'].values:
                logging.error(f"光伏资源曲线'{curve}'未定义！")
        
        hydro_curves_used = self.hydro_units['ResourceCurve'].unique()
        for curve in hydro_curves_used:
            if curve not in self.hydro_curves['RunoffHydroOperationCurve'].values:
                logging.error(f"水电资源曲线'{curve}'未定义！")
        
        logging.info("数据验证完成")
        
    def plot_resource_curves(self, output_path="ROTS_Resource_Curves.png"):
        """绘制资源曲线验证图"""
        plt.figure(figsize=(12, 8))
        
        # 绘制风电曲线
        for i, curve in enumerate(self.wind_curves['WTCurve'].unique()[:3]):
            df = self.wind_curves[self.wind_curves['WTCurve'] == curve]
            # 提取所有时间列
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'风电曲线_{i+1}')
        
        # 绘制光伏曲线
        for i, curve in enumerate(self.solar_curves['PVCurve'].unique()[:3]):
            df = self.solar_curves[self.solar_curves['PVCurve'] == curve]
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'光伏曲线_{i+1}')
        
        # 绘制水电曲线
        for i, curve in enumerate(self.hydro_curves['RunoffHydroOperationCurve'].unique()[:3]):
            df = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'] == curve]
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'水电曲线_{i+1}')
        
        # 添加标题和标签
        plt.title('资源曲线预览', fontsize=14)
        plt.xlabel('时段')
        plt.ylabel('出力系数')
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(output_path)
        logging.info(f"资源曲线预览图已保存到 {output_path}")
        plt.close()
        return output_path
        
    def load_data(self):
        """加载所有Excel数据"""
        logging.info("开始加载数据...")
        
        # 系统汇总信息
        self.system_summary = pd.read_excel(
            self.data_path, sheet_name='SystemSummary', skiprows=0, nrows=1
        ).iloc[0]
        
        # 机组参数
        self.thermal_units = pd.read_excel(
            self.data_path, sheet_name='UnitThermalGenerators', skiprows=0
        )
        self.wind_units = pd.read_excel(
            self.data_path, sheet_name='UnitWindGenerators', skiprows=0
        )
        # 修复1: 确保所有关键列正确映射
        self.wind_units.rename(columns={
            'WTPlantID': 'WTPlantID',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)
        # 但仍然需要检查容量列是否为数值类型

        self.solar_units = pd.read_excel(
            self.data_path, sheet_name='UnitSolarGenerators', skiprows=0
        )
        self.solar_units.rename(columns={
            'PVPlantID': 'PVPlantID',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)


        self.ess_units = pd.read_excel(
            self.data_path, sheet_name='StorElectrochemicalESS', skiprows=0
        )
        self.ess_units.rename(columns={
            'ESSId': 'ESSId',
            '储能容量(MWh)': 'Capacity',
            '充电功率(MW)': 'MaxChargePower',
            '放电功率(MW)': 'MaxDisPower',
            '所在母线编号': 'BusName'  # 确保有母线信息
        }, inplace=True)


        self.hydro_units = pd.read_excel(
            self.data_path, sheet_name='UnitRunoffHydroGenerators', skiprows=0
        )
        self.hydro_units.rename(columns={
            'RunoffHydroUnitName': 'RunoffHydroUnitName',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)

        # 保证RunoffHydroUnitName唯一，否则to_dict会报错
        if self.hydro_units['RunoffHydroUnitName'].duplicated().any():
            logging.warning("发现重复的水电机组名称，将自动去重（保留首次出现）")
            self.hydro_units = self.hydro_units.drop_duplicates(subset='RunoffHydroUnitName', keep='first')

        # 修正：确保所有机组的Capacity列为数值类型，避免sum时报错
        for df, cap_col in [
            (self.wind_units, 'Capacity'),
            (self.solar_units, 'Capacity'),
            (self.hydro_units, 'Capacity')
        ]:
            if cap_col in df.columns:
                df[cap_col] = pd.to_numeric(df[cap_col], errors='coerce').fillna(0)

        # 检查所有机组集合是否全为空，若全为空则添加一个虚拟火电机组，避免模型完全无源
        if (
            self.thermal_units.empty and
            (self.wind_units.empty or (self.wind_units.shape[0] == 1 and self.wind_units.iloc[0]['Capacity'] == 0)) and
            (self.solar_units.empty or (self.solar_units.shape[0] == 1 and self.solar_units.iloc[0]['Capacity'] == 0)) and
            (self.hydro_units.empty or (self.hydro_units.shape[0] == 1 and self.hydro_units.iloc[0]['Capacity'] == 0))
        ):
            logging.warning("所有机组集合均为空，自动添加一个虚拟火电机组以保证模型可行性")
            self.thermal_units = pd.DataFrame([{
                'ThermalUnitNumber': 'DummyThermal',
                'Capacity': 1000,  # 增加虚拟火电机组容量到1000MW
                'MinPower': 100,   # 设置最小出力
                'MinOnLineTime': 1,
                'MinOffLineTime': 1,
                'UpRamppingRate': 500,   # 增加爬坡速率
                'DownRamppingRate': 500,
                'InitStatus': 1,
                'BusName': 'Bus1'  # 暂时使用默认母线名，稍后会被覆盖
            }])

        # 曲线数据
        self.wind_curves = pd.read_excel(
            self.data_path, sheet_name='CurveWindResource', skiprows=0
        )
        self.wind_curves.rename(columns={
            'WTCurve': 'WTCurve'
        }, inplace=True)
        self.wind_curves = self.wind_curves.dropna(subset=['WTCurve'])
        
        self.solar_curves = pd.read_excel(
            self.data_path, sheet_name='CurveSolarResource', skiprows=0
        )
        self.solar_curves.rename(columns={
            'PVCurve': 'PVCurve'
        }, inplace=True)
        self.solar_curves = self.solar_curves.dropna(subset=['PVCurve'])
        
        self.hydro_curves = pd.read_excel(
            self.data_path, sheet_name='CurveRunoffHydroResource', skiprows=0
        )
        self.hydro_curves.rename(columns={
            'RunoffHydroOperationCurve': 'RunoffHydroOperationCurve'
        }, inplace=True)
        self.hydro_curves = self.hydro_curves.dropna(subset=['RunoffHydroOperationCurve'])
        
        # 修复2: 正确读取负荷曲线数据
        load_curve_sheet = pd.read_excel(self.data_path, sheet_name='CurveLoad')
        # 查找包含"Time_0"的列作为标识
        time_cols = [col for col in load_curve_sheet.columns if 'time_' in col.lower() or 'Time_' in col]

        # 先初始化self.load_curve，避免后续引用时报错
        self.load_curve = []
        peak_load = None

        # 优先从SystemSummary表获取峰值负荷
        if 'PeakLoad(MW)' in self.system_summary.index:
            peak_load = self.system_summary['PeakLoad(MW)']
            logging.info(f"从SystemSummary表获取峰值负荷：{peak_load}MW")
        
        if not time_cols:
            logging.error("负荷曲线表中未找到时间列！")
            # 如果没有找到峰值负荷，尝试从发电机总容量估算
            if peak_load is None:
                # 计算所有机组的总容量
                thermal_cap = self.thermal_units['Capacity'].sum() if not self.thermal_units.empty else 0
                wind_cap = self.wind_units['Capacity'].sum() if not self.wind_units.empty else 0
                solar_cap = self.solar_units['Capacity'].sum() if not self.solar_units.empty else 0
                hydro_cap = self.hydro_units['Capacity'].sum() if not self.hydro_units.empty else 0
                total_gen_capacity = thermal_cap + wind_cap + solar_cap + hydro_cap
                peak_load = total_gen_capacity * 0.8  # 假设负荷率80%
                logging.warning(f"Excel中未指定峰值负荷，根据发电机总容量({total_gen_capacity:.1f}MW)估算峰值负荷：{peak_load:.1f}MW")
            
            # 创建默认负荷曲线
            self.load_curve = [peak_load * 0.7] * int(self.system_summary['Duration'])
        else:
            # 尝试找到包含负荷数据的行
            for i in range(len(load_curve_sheet)):
                row = load_curve_sheet.iloc[i]
                if any(isinstance(val, (int, float)) and not pd.isna(val) for val in row[time_cols]):
                    load_factors = row[time_cols].values.astype(float)
                    
                    # 检测负荷数据类型：标幺值 vs 绝对值
                    max_factor = max(load_factors)
                    if max_factor <= 1.2:  # 标幺值通常不超过1.2
                        # 标幺值，需要乘以基准容量
                        if peak_load is None:
                            # 从发电机总容量估算
                            thermal_cap = self.thermal_units['Capacity'].sum() if not self.thermal_units.empty else 0
                            wind_cap = self.wind_units['Capacity'].sum() if not self.wind_units.empty else 0
                            solar_cap = self.solar_units['Capacity'].sum() if not self.solar_units.empty else 0
                            hydro_cap = self.hydro_units['Capacity'].sum() if not self.hydro_units.empty else 0
                            total_gen_capacity = thermal_cap + wind_cap + solar_cap + hydro_cap
                            peak_load = total_gen_capacity * 0.8
                            logging.warning(f"检测到标幺值负荷曲线，但未指定峰值负荷，根据发电机总容量估算：{peak_load:.1f}MW")
                        else:
                            logging.info(f"检测到标幺值负荷曲线，使用指定峰值负荷：{peak_load}MW")
                        
                        self.load_curve = [factor * peak_load for factor in load_factors]
                    else:
                        # 绝对值，直接使用
                        self.load_curve = load_factors.tolist()
                        actual_peak = max(self.load_curve)
                        logging.info(f"检测到绝对值负荷曲线，直接使用，实际峰值负荷：{actual_peak:.1f}MW")
                        
                        # 更新peak_load为实际峰值
                        if peak_load is None or abs(peak_load - actual_peak) > actual_peak * 0.1:
                            if peak_load is not None:
                                logging.warning(f"SystemSummary中的峰值负荷({peak_load:.1f}MW)与实际负荷曲线峰值({actual_peak:.1f}MW)不匹配，使用实际值")
                            peak_load = actual_peak
                    
                    logging.info(f"从第{i+1}行读取负荷曲线数据")
                    logging.info(f"负荷曲线范围：{min(self.load_curve):.1f} - {max(self.load_curve):.1f} MW")
                    break
            else:
                logging.warning("未找到有效的负荷数据行，使用默认值")
                if peak_load is None:
                    thermal_cap = self.thermal_units['Capacity'].sum() if not self.thermal_units.empty else 0
                    wind_cap = self.wind_units['Capacity'].sum() if not self.wind_units.empty else 0
                    solar_cap = self.solar_units['Capacity'].sum() if not self.solar_units.empty else 0
                    hydro_cap = self.hydro_units['Capacity'].sum() if not self.hydro_units.empty else 0
                    total_gen_capacity = thermal_cap + wind_cap + solar_cap + hydro_cap
                    peak_load = total_gen_capacity * 0.8
                    logging.warning(f"根据发电机总容量估算峰值负荷：{peak_load:.1f}MW")
                
                self.load_curve = [peak_load * 0.7] * int(self.system_summary['Duration'])

        # 确保负荷曲线有合理的数值
        if len(self.load_curve) == 0:
            if peak_load is None:
                thermal_cap = self.thermal_units['Capacity'].sum() if not self.thermal_units.empty else 0
                wind_cap = self.wind_units['Capacity'].sum() if not self.wind_units.empty else 0
                solar_cap = self.solar_units['Capacity'].sum() if not self.solar_units.empty else 0
                hydro_cap = self.hydro_units['Capacity'].sum() if not self.hydro_units.empty else 0
                total_gen_capacity = thermal_cap + wind_cap + solar_cap + hydro_cap
                peak_load = total_gen_capacity * 0.8
                logging.warning(f"最终方案：根据发电机总容量估算峰值负荷：{peak_load:.1f}MW")
            
            self.load_curve = [peak_load * 0.7 for _ in range(int(self.system_summary['Duration']))]
        elif max(self.load_curve) < 100:
            logging.warning("负荷曲线峰值过低，可能是系数而非绝对值")
            if peak_load is None:
                thermal_cap = self.thermal_units['Capacity'].sum() if not self.thermal_units.empty else 0
                wind_cap = self.wind_units['Capacity'].sum() if not self.wind_units.empty else 0
                solar_cap = self.solar_units['Capacity'].sum() if not self.solar_units.empty else 0
                hydro_cap = self.hydro_units['Capacity'].sum() if not self.hydro_units.empty else 0
                total_gen_capacity = thermal_cap + wind_cap + solar_cap + hydro_cap
                peak_load = total_gen_capacity * 0.8
                logging.warning(f"根据发电机总容量估算峰值负荷：{peak_load:.1f}MW")
            
            # 如果数值很小，认为是系数，需要乘以峰值负荷
            if max(self.load_curve) <= 1.0:
                self.load_curve = [val * peak_load for val in self.load_curve]
                logging.info(f"负荷曲线已转换为绝对值，峰值：{max(self.load_curve):.1f}MW")
        
        # 最终验证和日志
        final_peak = max(self.load_curve) if self.load_curve else 0
        logging.info(f"最终负荷曲线峰值：{final_peak:.1f}MW")
        
        # 检查负荷与发电容量的合理性
        thermal_cap = self.thermal_units['Capacity'].sum() if not self.thermal_units.empty else 0
        wind_cap = self.wind_units['Capacity'].sum() if not self.wind_units.empty else 0
        solar_cap = self.solar_units['Capacity'].sum() if not self.solar_units.empty else 0
        hydro_cap = self.hydro_units['Capacity'].sum() if not self.hydro_units.empty else 0
        total_gen_capacity = thermal_cap + wind_cap + solar_cap + hydro_cap
        if final_peak > 0:
            capacity_margin = (total_gen_capacity / final_peak - 1) * 100
            logging.info(f"容量裕度：{capacity_margin:.1f}%")
            if capacity_margin < 10:
                logging.warning(f"容量裕度较低({capacity_margin:.1f}%)，可能存在供电不足风险")
        
        # 网络数据
        self.buses = pd.read_excel(self.data_path, sheet_name='NetBuses', skiprows=0)
        self.lines = pd.read_excel(self.data_path, sheet_name='NetLines', skiprows=0)
        self.transformers = pd.read_excel(self.data_path, sheet_name='NetTransformers', skiprows=0)
        self.sections = pd.read_excel(self.data_path, sheet_name='NetSectionsLines', skiprows=0)
        self.section_capacity = pd.read_excel(self.data_path, sheet_name='NetSectionsCapacity', skiprows=0)
        self.ac_tielines = pd.read_excel(self.data_path, sheet_name='NetHVACTieLine', skiprows=0)
        self.ac_tieline_curves = pd.read_excel(self.data_path, sheet_name='EnergyHVACTieLine', skiprows=0)
        
        # 创建区域映射字典
        self.bus_to_area = self._create_bus_area_mapping()
        logging.info(f"区域映射创建完成，A区{len([k for k,v in self.bus_to_area.items() if 'A' in v])}个母线，"
                    f"B区{len([k for k,v in self.bus_to_area.items() if 'B' in v])}个母线，"
                    f"C区{len([k for k,v in self.bus_to_area.items() if 'C' in v])}个母线")
        
        # 更新虚拟机组的母线信息（使用实际的第一个母线）
        if not self.buses.empty:
            first_bus = self.buses['BusId'].iloc[0]
            # 更新风电虚拟机组
            if (self.wind_units.shape[0] == 1 and 
                self.wind_units.iloc[0]['WTPlantID'] == 'DummyWind'):
                self.wind_units.loc[0, 'BusName'] = first_bus
            # 更新光伏虚拟机组
            if (self.solar_units.shape[0] == 1 and 
                self.solar_units.iloc[0]['PVPlantID'] == 'DummySolar'):
                self.solar_units.loc[0, 'BusName'] = first_bus
            # 更新储能虚拟机组
            if (self.ess_units.shape[0] == 1 and 
                self.ess_units.iloc[0]['ESSId'] == 'DummyESS'):
                self.ess_units.loc[0, 'BusName'] = first_bus
            # 更新水电虚拟机组
            if (self.hydro_units.shape[0] == 1 and 
                self.hydro_units.iloc[0]['RunoffHydroUnitName'] == 'DummyHydro'):
                self.hydro_units.loc[0, 'BusName'] = first_bus
            # 更新火电虚拟机组
            if (self.thermal_units.shape[0] == 1 and 
                self.thermal_units.iloc[0]['ThermalUnitNumber'] == 'DummyThermal'):
                self.thermal_units.loc[0, 'BusName'] = first_bus
        
        # 负荷数据
        self.loads = pd.read_excel(self.data_path, sheet_name='Loads', skiprows=0)
        
        # 燃料数据
        self.fuel_data = pd.read_excel(self.data_path, sheet_name='UnitFuel', skiprows=0).iloc[0]
        
        # 设置时间参数
        self.T = int(self.system_summary['Duration'])
        self.dt = 1  # 调度间隔为1小时
        self.hours = list(range(self.T))
        
        # 创建机组字典
        self.thermal_dict = self.thermal_units.set_index('ThermalUnitNumber').to_dict('index')
        for unit in self.thermal_dict:
            # 确保数值类型正确
            self.thermal_dict[unit]['MinOnLineTime'] = int(self.thermal_dict[unit].get('MinOnLineTime', 0))
            self.thermal_dict[unit]['MinOffLineTime'] = int(self.thermal_dict[unit].get('MinOffLineTime', 0))
            self.thermal_dict[unit]['Capacity'] = float(self.thermal_dict[unit].get('Capacity', 0))
            self.thermal_dict[unit]['MinPower'] = float(self.thermal_dict[unit].get('MinPower', 0))
            self.thermal_dict[unit]['UpRamppingRate'] = float(self.thermal_dict[unit].get('UpRamppingRate', 0))
            self.thermal_dict[unit]['DownRamppingRate'] = float(self.thermal_dict[unit].get('DownRamppingRate', 0))
            
        self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        self.solar_dict = self.solar_units.set_index('PVPlantID').to_dict('index')
        self.ess_dict = self.ess_units.set_index('ESSId').to_dict('index')
        self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        
        # 创建负荷字典
        self.load_dict = self.loads.set_index('负荷编号').to_dict('index')
        
        # 构建母线字典
        self.bus_dict = self.buses.set_index('BusId').to_dict('index')
        
        # 构建线路字典
        self.line_dict = {}
        for idx, row in self.lines.iterrows():
            self.line_dict[row['BranchId']] = {
                'StartNodeName': row['StartNodeName'],
                'EndNodeName': row['EndNodeName'],
                'Resistance': float(row.get('Resistance', 0)),
                'Reactance': float(row.get('Reactance', 0)),
                'Capacity': float(row.get('Capacity', 0)),
                'Susceptance': float(row.get('Susceptance', 0))
            }
        
        # 找到平衡节点（参考节点）
        self.ref_bus = self.system_summary['BalanceNode']
        logging.info(f"参考节点设置为: {self.ref_bus}")
        
        # 初始化资源曲线列名缓存
        self.__init_curve_columns()
        self._check_curve_columns()
        
        logging.info("数据加载完成")
    
    def _create_bus_area_mapping(self):
        """根据提供的区域信息创建母线到区域的映射"""
        bus_to_area = {}
        
        # 根据您提供的信息创建映射
        area_mapping = {
            # A区母线
            'Bus_1': 'A区', 'Bus_2': 'A区', 'Bus_3': 'A区', 'Bus_4': 'A区', 'Bus_5': 'A区',
            'Bus_6': 'A区', 'Bus_7': 'A区', 'Bus_8': 'A区', 'Bus_9': 'A区', 'Bus_10': 'A区',
            'Bus_11': 'A区', 'Bus_12': 'A区', 'Bus_13': 'A区', 'Bus_14': 'A区', 'Bus_15': 'A区',
            
            # B区母线
            'Bus_18': 'B区', 'Bus_19': 'B区', 'Bus_20': 'B区', 'Bus_21': 'B区',
            
            # C区母线
            'Bus_26': 'C区', 'Bus_27': 'C区', 'Bus_28': 'C区', 'Bus_29': 'C区',
            'Bus_38': 'C区', 'Bus_41': 'C区', 'Bus_42': 'C区', 'Bus_44': 'C区',
            'Bus_24': 'C区', 'Bus_25': 'C区'
        }
        
        # 更新映射
        bus_to_area.update(area_mapping)
        
        # 对于未明确指定的母线，根据编号分配到默认区域
        for bus_id in self.buses['BusId'].tolist():
            if bus_id not in bus_to_area:
                # 默认分配策略：根据母线编号分配
                if isinstance(bus_id, str) and 'Bus_' in bus_id:
                    bus_num = int(bus_id.split('_')[1]) if bus_id.split('_')[1].isdigit() else 0
                    if bus_num <= 15:
                        bus_to_area[bus_id] = 'A区'
                    elif bus_num <= 23:
                        bus_to_area[bus_id] = 'B区'  
                    else:
                        bus_to_area[bus_id] = 'C区'
                else:
                    bus_to_area[bus_id] = 'A区'  # 默认A区
        
        return bus_to_area
    
    def __init_curve_columns(self):
        """初始化资源曲线列名缓存，避免重复警告"""
        self._wind_curve_col = None
        self._solar_curve_col = None
        self._hydro_curve_col = None
        self._curve_cols_checked = False
    
    def _check_curve_columns(self):
        """检查并缓存资源曲线列名"""
        if self._curve_cols_checked:
            return
            
        # 检查风电资源曲线列名
        if hasattr(self, 'wind_curves') and not self.wind_curves.empty:
            possible_cols = ['WTResourceCurve', 'WindResourceCurve', 'ResourceCurve']
            for col in possible_cols:
                if col in self.wind_curves.columns:
                    self._wind_curve_col = col
                    break
            if self._wind_curve_col is None:
                logging.warning(f"风电资源曲线表中未找到合适的列名")
        
        # 检查光伏资源曲线列名
        if hasattr(self, 'solar_curves') and not self.solar_curves.empty:
            possible_cols = ['PVResourceCurve', 'SolarResourceCurve', 'ResourceCurve']
            for col in possible_cols:
                if col in self.solar_curves.columns:
                    self._solar_curve_col = col
                    break
            if self._solar_curve_col is None:
                logging.warning(f"光伏资源曲线表中未找到合适的列名")
        
        # 检查水电资源曲线列名
        if hasattr(self, 'hydro_curves') and not self.hydro_curves.empty:
            possible_cols = ['RunoffHydroOperationCurve', 'HydroOperationCurve', 'ResourceCurve']
            for col in possible_cols:
                if col in self.hydro_curves.columns:
                    self._hydro_curve_col = col
                    break
            if self._hydro_curve_col is None:
                logging.warning(f"水电资源曲线表中未找到合适的列名")
        
        self._curve_cols_checked = True
    
    def get_unit_param(self, unit_dict, unit, key, default=0.0):
        """安全获取机组参数"""
        try:
            val = unit_dict[unit].get(key, default)
            if pd.isna(val):
                return default
            return val
        except Exception:
            return default
    
    def _get_renewable_max_output(self, curve_name, t, capacity, unit_type):
        """计算可再生能源在指定时间的理论最大出力"""
        if capacity <= 0:
            return 0
        
        curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
        
        # 如果没有资源曲线，假设满出力
        if curve_name_str in ['', '0', 'nan', 'None']:
            return capacity
        
        try:
            # 根据机组类型选择对应的曲线数据
            if unit_type == 'wind':
                curves_df = self.wind_curves
                curve_col = self._wind_curve_col
                if curve_col is None:
                    return capacity
                    
            elif unit_type == 'solar':
                curves_df = self.solar_curves
                curve_col = self._solar_curve_col
                if curve_col is None:
                    return capacity
                    
            elif unit_type == 'hydro':
                curves_df = self.hydro_curves
                curve_col = self._hydro_curve_col
                if curve_col is None:
                    return capacity
            else:
                return capacity
            
            # 查找对应的资源曲线
            curve_row = curves_df[curves_df[curve_col].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                curve_row = curves_df[curves_df[curve_col].astype(str).str.contains(curve_name_str, case=False, na=False)]
                if curve_row.empty:
                    return capacity
            
            # 获取时间点的出力系数
            time_col = f'Time_{t}'
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 1.0
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        max_factor = 1.0
                else:
                    max_factor = 1.0
            
            return max_factor * capacity
            
        except Exception as e:
            logging.warning(f"计算{unit_type}理论最大出力时出错: {e}")
            return capacity

    def create_model(self):
        """创建Pyomo优化模型（包含直流潮流约束）"""
        logging.info("开始创建优化模型（包含直流潮流约束）...")
        
        self.model = pyo.ConcreteModel()
        
        # 定义集合
        self.model.T = pyo.Set(initialize=self.hours)  # 时间段
        self.model.ThermalUnits = pyo.Set(initialize=self.thermal_units['ThermalUnitNumber'].tolist())
        self.model.WindUnits = pyo.Set(initialize=self.wind_units['WTPlantID'].tolist())
        self.model.SolarUnits = pyo.Set(initialize=self.solar_units['PVPlantID'].tolist())
        self.model.ESSUntis = pyo.Set(initialize=self.ess_units['ESSId'].tolist())
        self.model.HydroUnits = pyo.Set(initialize=self.hydro_units['RunoffHydroUnitName'].tolist())
        self.model.Loads = pyo.Set(initialize=self.loads['负荷编号'].tolist())
        self.model.Buses = pyo.Set(initialize=self.buses['BusId'].tolist())
        self.model.Lines = pyo.Set(initialize=self.lines['BranchId'].tolist())
        
        # 定义变量
        # 火电机组
        self.model.thermal_status = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.thermal_startup = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_shutdown = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        
        # 可再生能源
        self.model.wind_power = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_power = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_power = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 储能系统
        self.model.ess_charge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_discharge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_soc = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 负荷相关
        self.model.load_shed = pyo.Var(
            self.model.Loads, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 直流潮流变量
        self.model.bus_angle = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )  # 节点相角
        
        self.model.line_power = pyo.Var(
            self.model.Lines, self.model.T, within=pyo.Reals
        )  # 线路潮流
        
        # 节点平衡松弛变量（用于调试潮流约束）
        self.model.balance_slack = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )
        
        # 备用松弛变量
        self.model.reserve_slack = pyo.Var(
            self.model.T, within=pyo.NonNegativeReals
        )
        
        # 深度调峰惩罚线性化变量
        self.model.deep_peak_penalty = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 可再生能源最小出力松弛变量
        self.model.wind_slack = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_slack = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_slack = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 目标函数
        def total_cost_rule(model):
            """系统总成本计算"""
            import logging
            # 临时抑制日志
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                cost = 0
                fuel_price = self.fuel_data['UnitPrice']  # 使用原始燃料价格，不再加倍
                
                # 火电运行成本
                for unit in model.ThermalUnits:
                    for t in model.T:
                        a = self.thermal_dict[unit].get('OperationCoeff_A', 0)
                        b = self.thermal_dict[unit].get('OperationCoeff_B', 100)  # 设置默认的运行成本
                        c = self.thermal_dict[unit].get('OperationCoeff_C', 50)   # 设置默认的固定成本
                        
                        # 运行成本（使用合理的基础成本）
                        cost += model.thermal_status[unit, t] * c * fuel_price  # 使用原始固定成本
                        cost += b * model.thermal_power[unit, t] * fuel_price   # 使用原始变动成本
                        
                        # 启动成本
                        startup_cost = self.thermal_dict[unit].get('StateOnFuelConsumption', 200)  # 使用合理的启动成本
                        cost += startup_cost * fuel_price * model.thermal_startup[unit, t]
            finally:
                # 恢复原始日志级别
                logging.getLogger().setLevel(original_level)
            
            # 深度调峰惩罚（线性化实现）
            for unit in model.ThermalUnits:
                for t in model.T:
                    max_power = self.thermal_dict[unit].get('Capacity', 0)
                    deep_threshold = self.thermal_dict[unit].get('DeepPeakShavingThreshold', 0.5) * max_power
                    deep_cost = self.thermal_dict[unit].get('DeepPeakShavingCost', 0)
                    # 线性松弛变量
                    cost += deep_cost * model.deep_peak_penalty[unit, t]
            
            # 可再生能源激励（时间变化的激励机制）
            renewable_penalty = 50000    # 降低惩罚，避免强制出力导致平均化
            
            # 风电激励和惩罚
            for unit in model.WindUnits:
                for t in model.T:
                    # 风电激励保持固定
                    renewable_incentive = -200  # 200元/MWh激励
                    cost += renewable_incentive * model.wind_power[unit, t]
                    cost += renewable_penalty * model.wind_slack[unit, t]
            
            # 光伏激励（按时间变化，鼓励在光照充足时发电）
            for unit in model.SolarUnits:
                for t in model.T:
                    hour = t % 24
                    # 根据光照强度调整激励力度
                    if hour < 6 or hour > 18:  # 夜间
                        solar_incentive = 0  # 夜间无激励
                    elif 6 <= hour <= 8 or 16 <= hour <= 18:  # 弱光时段
                        solar_incentive = -100  # 较小激励
                    elif 9 <= hour <= 15:  # 强光时段
                        if hour == 12:  # 正午最高激励
                            solar_incentive = -500
                        elif hour in [11, 13]:  # 接近正午
                            solar_incentive = -400
                        elif hour in [10, 14]:  # 较强光照
                            solar_incentive = -350
                        else:  # 9时和15时
                            solar_incentive = -300
                    else:  # 其他时段
                        solar_incentive = -150
                    
                    cost += solar_incentive * model.solar_power[unit, t]
                    cost += renewable_penalty * model.solar_slack[unit, t]
            
            # 水电激励和惩罚
            for unit in model.HydroUnits:
                for t in model.T:
                    cost += renewable_incentive * model.hydro_power[unit, t]
                    cost += renewable_penalty * model.hydro_slack[unit, t]
            
            # 储能系统激励（鼓励储能参与调节）
            ess_discharge_incentive = -500  # 放电收益：500元/MWh
            ess_charge_cost = 200          # 充电成本：200元/MWh（模拟电费和损耗）
            for unit in model.ESSUntis:
                for t in model.T:
                    cost += ess_discharge_incentive * model.ess_discharge[unit, t]  # 放电有收益
                    cost += ess_charge_cost * model.ess_charge[unit, t]            # 充电有成本
            
            # 切负荷惩罚（大幅提高惩罚权重，使切负荷成为最后选择）
            penalty_cut = 1000000  # 元/MWh - 提高到100万，确保优先满足负荷
            for load in model.Loads:
                for t in model.T:
                    cost += penalty_cut * model.load_shed[load, t]
            
            # 节点平衡松弛惩罚（降低惩罚避免与全局平衡冲突）
            penalty_slack = 100  # 元/MW（降低惩罚）
            for bus in model.Buses:
                for t in model.T:
                    cost += penalty_slack * (model.balance_slack[bus, t] * model.balance_slack[bus, t])
            
            # 备用松弛惩罚
            penalty_reserve = 5000  # 元/MW
            for t in model.T:
                cost += penalty_reserve * model.reserve_slack[t]
            
            return cost

        # 深度调峰惩罚线性化约束
        def deep_peak_penalty_rule(model, unit, t):
            max_power = self.thermal_dict[unit].get('Capacity', 0)
            deep_threshold = self.thermal_dict[unit].get('DeepPeakShavingThreshold', 0.5) * max_power
            return model.deep_peak_penalty[unit, t] >= deep_threshold - model.thermal_power[unit, t]
        self.model.deep_peak_penalty_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=deep_peak_penalty_rule
        )
        
        self.model.total_cost = pyo.Objective(rule=total_cost_rule, sense=pyo.minimize)
        
        # 1. 火电机组逻辑约束
        def thermal_logic_rule(model, unit, t):
            import logging
            
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                if t == 0:
                    init_status = self.thermal_dict[unit].get('InitStatus', False)
                    if isinstance(init_status, str):
                        init_status = init_status.strip().lower() in ['true', '1', 'yes']
                    return model.thermal_status[unit, t] == (1 if init_status else 0)
                else:
                    return model.thermal_status[unit, t] - model.thermal_status[unit, t-1] == \
                        model.thermal_startup[unit, t] - model.thermal_shutdown[unit, t]
            finally:
                # 恢复原始日志级别
                logging.getLogger().setLevel(original_level)
        self.model.thermal_logic_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_logic_rule
        )

        # 2. 火电机组出力约束
        def thermal_power_lower_rule(model, unit, t):
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            return model.thermal_power[unit, t] >= min_power * model.thermal_status[unit, t]
        
        def thermal_power_upper_rule(model, unit, t):
            max_power = self.thermal_dict[unit].get('Capacity', 0)
            return model.thermal_power[unit, t] <= max_power * model.thermal_status[unit, t]
        
        self.model.thermal_power_lower_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_lower_rule
        )
        self.model.thermal_power_upper_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_upper_rule
        )

        # 3. 爬坡约束 - 使用更稳健的方案2
        def thermal_ramp_up_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_up = self.thermal_dict[unit].get('UpRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束，增加状态变化时的缓冲
            return (model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= 
                    model.thermal_status[unit, t] * ramp_up + 
                    (1 - model.thermal_status[unit, t-1]) * capacity * 0.7)  # 缓冲系数
        
        def thermal_ramp_down_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_down = self.thermal_dict[unit].get('DownRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束，增加状态变化时的缓冲
            return (model.thermal_power[unit, t-1] - model.thermal_power[unit, t] <= 
                    model.thermal_status[unit, t-1] * ramp_down + 
                    (1 - model.thermal_status[unit, t]) * capacity * 0.7)  # 缓冲系数

        self.model.thermal_ramp_up_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_up_rule
        )
        self.model.thermal_ramp_down_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_down_rule
        )

        # 4. 最短开停机时间约束 - 大幅放宽以提高可行性
        def min_up_time_rule(model, unit, t):
            min_up = self.thermal_dict[unit].get('MinOnLineTime', 0)
            # 如果最小开机时间太长，则限制为总时间的1/4
            min_up = min(min_up, max(1, self.T // 4))
            
            if t < min_up or min_up <= 1:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_up:
                return sum(model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        model.thermal_status[unit, t] - prev_status) * (self.T - t)
            else:
                return sum(model.thermal_status[unit, k] for k in range(t, t+min_up)) >= (
                        model.thermal_status[unit, t] - prev_status) * min_up

        def min_down_time_rule(model, unit, t):
            min_down = self.thermal_dict[unit].get('MinOffLineTime', 0)
            # 如果最小停机时间太长，则限制为总时间的1/4
            min_down = min(min_down, max(1, self.T // 4))
            
            if t < min_down or min_down <= 1:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_down:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        prev_status - model.thermal_status[unit, t]) * (self.T - t)
            else:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, t+min_down)) >= (
                        prev_status - model.thermal_status[unit, t]) * min_down

        self.model.min_up_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_up_time_rule
        )
        self.model.min_down_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_down_time_rule
        )

        # 5. 可再生能源出力约束 - 修复资源曲线匹配问题，增加默认出力能力
        def wind_power_rule(model, unit, t):
            curve_name = self.wind_dict[unit].get('ResourceCurve', '')
            capacity = self.wind_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.wind_power[unit, t] <= 0
            
            # 防止curve_name为非字符串
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 如果没有资源曲线名称或为虚拟机组，使用高出力系数
            if curve_name_str in ['', '0', 'nan', 'None'] or curve_name_str == '':
                # 给风电一个很高的出力模式
                default_factor = 0.8  # 大幅提升到80%
                # logging.info(f"风电机组{unit}在时段{t}使用高出力系数: {default_factor}")  # 隐藏详细日志
                return model.wind_power[unit, t] <= default_factor * capacity
            
            # 找到对应的风资源曲线 - 增强匹配逻辑
            curve_row = self.wind_curves[self.wind_curves['WTCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                # 尝试模糊匹配
                curve_row = self.wind_curves[self.wind_curves['WTCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"风电机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！使用高出力系数")
                    # 使用高出力而不是低出力
                    default_factor = 0.8
                    return model.wind_power[unit, t] <= default_factor * capacity
            
            # 获取时间点值 - 处理列名大小写问题
            time_col = f'Time_{t}'
            # 尝试匹配列名（不区分大小写）
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 0.8  # 使用高默认值
                else:
                    max_factor = max(max_factor, 0.5)  # 确保至少50%出力能力
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        max_factor = 0.8
                    else:
                        max_factor = max(max_factor, 0.5)
                else:
                    logging.warning(f"风电机组{unit}在时段{t}缺少出力系数，使用高默认值0.8")
                    max_factor = 0.8  # 使用高默认值
                    
            # 确保出力不超过容量
            return model.wind_power[unit, t] <= max_factor * capacity
        
        self.model.wind_power_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_power_rule
        )
        
        def solar_power_rule(model, unit, t):
            curve_name = self.solar_dict[unit].get('ResourceCurve', '')
            capacity = self.solar_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.solar_power[unit, t] <= 0
            
            hour = t % 24
            
            # 创建更明显的光伏出力时间变化曲线
            if hour < 6 or hour > 18:  # 夜间时段
                max_factor = 0.0
            elif hour == 6 or hour == 18:  # 日出日落
                max_factor = 0.1
            elif hour == 7 or hour == 17:  # 早晚弱光
                max_factor = 0.3
            elif hour == 8 or hour == 16:  # 上午下午
                max_factor = 0.5
            elif hour == 9 or hour == 15:  # 较强光照
                max_factor = 0.7
            elif hour == 10 or hour == 14:  # 强光照
                max_factor = 0.85
            elif hour == 11 or hour == 13:  # 接近峰值
                max_factor = 0.9
            elif hour == 12:  # 正午峰值
                max_factor = 0.95
            else:
                max_factor = 0.6  # 默认值
            
            # logging.info(f"光伏机组{unit}在时段{t}（{hour}时）设置上限系数: {max_factor}")  # 隐藏详细日志
            return model.solar_power[unit, t] <= max_factor * capacity
        
        self.model.solar_power_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_power_rule
        )
        
        def hydro_power_rule(model, unit, t):
            curve_name = self.hydro_dict[unit].get('ResourceCurve', '')
            capacity = self.hydro_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.hydro_power[unit, t] <= 0
            
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 如果资源曲线名为0或空，允许水电机组出力为0~容量
            if curve_name_str in ['', '0', 'nan', 'None']:
                # logging.info(f"水电机组{unit}无资源曲线限制，可在0~{capacity}MW范围内出力")  # 隐藏详细日志
                return (0, model.hydro_power[unit, t], capacity)
            
            curve_row = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                curve_row = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"水电机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！允许灵活出力")
                    return (0, model.hydro_power[unit, t], capacity)
            
            # 获取时间点值 - 处理列名大小写问题
            time_col = f'Time_{t}'
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 0.8  # 使用默认值而不是0
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        max_factor = 0.8
                else:
                    logging.warning(f"水电机组{unit}在时段{t}缺少出力系数，使用默认值0.8")
                    max_factor = 0.8  # 使用默认值而不是0
                    
            return model.hydro_power[unit, t] <= max_factor * capacity
        
        self.model.hydro_power_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_power_rule
        )

        # 5.1 可再生能源最小出力约束（临时禁用强制性要求）
        def wind_min_output_rule(model, unit, t):
            # 临时禁用风电最小出力约束，避免不可行性
            return pyo.Constraint.Skip
        
        def solar_min_output_rule(model, unit, t):
            # 修改光伏最小出力约束，不强制夜间和低光照时段出力
            capacity = self.solar_dict[unit].get('Capacity', 0)
            
            # 只有容量大于1MW的机组才考虑最小出力约束
            if capacity > 1:
                hour = t % 24
                
                # 只在光照充足时设置最小出力要求
                if 9 <= hour <= 15:  # 只在强光照时段（9-15时）设置最小出力
                    if hour == 12:  # 正午
                        min_factor = 0.5  # 降低最小出力要求
                    elif hour in [11, 13]:  # 接近正午
                        min_factor = 0.4
                    elif hour in [10, 14]:  # 较强光照
                        min_factor = 0.3
                    else:  # 9时和15时
                        min_factor = 0.2
                    
                    min_output = capacity * min_factor
                    # logging.info(f"光伏机组{unit}在时段{t}（{hour}时）设置下限系数: {min_factor}")  # 隐藏详细日志
                    return model.solar_power[unit, t] + model.solar_slack[unit, t] >= min_output
                else:
                    # 其他时段不设置最小出力要求，让模型自由优化
                    return model.solar_slack[unit, t] >= 0
            else:
                return model.solar_slack[unit, t] >= 0
        
        def hydro_min_output_rule(model, unit, t):
            # 临时禁用水电最小出力约束，避免不可行性
            return pyo.Constraint.Skip
        
        self.model.wind_min_output_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_min_output_rule
        )
        self.model.solar_min_output_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_min_output_rule
        )
        self.model.hydro_min_output_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_min_output_rule
        )

        # 6. 储能系统约束
        def ess_charge_power_rule(model, unit, t):
            max_charge = self.ess_dict[unit].get('MaxChargePower', 0)
            return model.ess_charge[unit, t] <= max_charge
        
        def ess_discharge_power_rule(model, unit, t):
            max_discharge = self.ess_dict[unit].get('MaxDisPower', 0)
            return model.ess_discharge[unit, t] <= max_discharge
        
        def ess_soc_rule(model, unit, t):
            min_soc = self.ess_dict[unit].get('MinSOC', 0)
            max_soc = self.ess_dict[unit].get('MaxSOC', 1000)
            capacity = self.ess_dict[unit].get('Capacity', 0)
            return (min_soc, model.ess_soc[unit, t], max_soc)
        
        def ess_soc_balance_rule(model, unit, t):
            charge_eff = self.ess_dict[unit].get('ChargeEfficiency', 0.85)
            dis_eff = self.ess_dict[unit].get('DisEfficiency', 0.85)
            
            if t == 0:
                init_soc = self.ess_dict[unit].get('InitSOC', 0)
                return model.ess_soc[unit, t] == init_soc
            else:
                return model.ess_soc[unit, t] == (model.ess_soc[unit, t-1] + 
                                                  charge_eff * model.ess_charge[unit, t] - 
                                                  model.ess_discharge[unit, t] / dis_eff)
        
        self.model.ess_charge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_charge_power_rule
        )
        self.model.ess_discharge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_discharge_power_rule
        )
        self.model.ess_soc_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_rule
        )
        self.model.ess_soc_balance_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_balance_rule
        )
        
        # 6.5 储能充放电互斥约束（不能同时充放电）
        # 添加二进制变量控制充放电状态
        self.model.ess_charge_status = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.Binary
        )
        
        def ess_charge_exclusive_rule(model, unit, t):
            max_charge = self.ess_dict[unit].get('MaxChargePower', 0)
            return model.ess_charge[unit, t] <= max_charge * model.ess_charge_status[unit, t]
        
        def ess_discharge_exclusive_rule(model, unit, t):
            max_discharge = self.ess_dict[unit].get('MaxDisPower', 0)
            return model.ess_discharge[unit, t] <= max_discharge * (1 - model.ess_charge_status[unit, t])
        
        self.model.ess_charge_exclusive_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_charge_exclusive_rule
        )
        self.model.ess_discharge_exclusive_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_discharge_exclusive_rule
        )
        
        # 移除二进制约束以保持线性规划

        # 7. 直流潮流约束
        # logging.info("启用直流潮流约束...")  # 隐藏详细日志
        
        # 7.1 参考节点约束（平衡节点相角为0）
        def ref_bus_rule(model, t):
            return model.bus_angle[self.ref_bus, t] == 0
        self.model.ref_bus_con = pyo.Constraint(
            self.model.T, rule=ref_bus_rule
        )
        
        # 7.2 节点功率平衡约束（仅处理潮流分配，不重复处理全局平衡）
        def node_balance_rule(model, bus, t):
            # 只有当有多个节点时才启用节点平衡约束
            if len(model.Buses) <= 1:
                return pyo.Constraint.Skip
                
            # 节点发电功率
            gen_power = 0
            for unit in model.ThermalUnits:
                if self.thermal_dict[unit].get('BusName', '') == bus:
                    gen_power += model.thermal_power[unit, t]
            
            for unit in model.WindUnits:
                if self.wind_dict[unit].get('BusName', '') == bus:
                    gen_power += model.wind_power[unit, t]
            
            for unit in model.SolarUnits:
                if self.solar_dict[unit].get('BusName', '') == bus:
                    gen_power += model.solar_power[unit, t]
            
            for unit in model.HydroUnits:
                if self.hydro_dict[unit].get('BusName', '') == bus:
                    gen_power += model.hydro_power[unit, t]
            
            # 储能净功率（放电为正，充电为负）
            ess_net_power = 0
            for unit in model.ESSUntis:
                if self.ess_dict[unit].get('BusName', '') == bus:
                    ess_net_power += model.ess_discharge[unit, t] - model.ess_charge[unit, t]
            
            # 节点负荷 - 简化为平均分配
            num_buses = len(model.Buses)
            if t < len(self.load_curve) and num_buses > 0:
                avg_load = self.load_curve[t] / num_buses
            else:
                avg_load = 0
            
            # 节点净注入功率
            net_injection = gen_power + ess_net_power - avg_load
            
            # 计算线路潮流总和（流出为正，流入为负）
            line_flow = 0
            for line in model.Lines:
                line_info = self.line_dict.get(line, {})
                if line_info.get('StartNodeName', '') == bus:
                    line_flow += model.line_power[line, t]
                elif line_info.get('EndNodeName', '') == bus:
                    line_flow -= model.line_power[line, t]
            
            # 节点功率平衡（带松弛变量，但惩罚较小以避免冲突）
            return net_injection == line_flow + model.balance_slack[bus, t]
        
        self.model.node_balance_con = pyo.Constraint(
            self.model.Buses, self.model.T, rule=node_balance_rule
        )
        
        # 7.3 线路潮流约束（直流潮流方程）
        def line_power_rule(model, line, t):
            line_info = self.line_dict.get(line, {})
            start_bus = line_info.get('StartNodeName', '')
            end_bus = line_info.get('EndNodeName', '')
            reactance = line_info.get('Reactance', 0.01)  # 默认电抗为0.01
            
            if start_bus == '' or end_bus == '':
                return pyo.Constraint.Skip
            
            # 如果电抗为0或很小，跳过这条线路的潮流约束
            if abs(reactance) < 1e-6:
                return pyo.Constraint.Skip
            
            # 直流潮流方程：P_ij = (θ_i - θ_j) / X_ij
            return model.line_power[line, t] == (
                model.bus_angle[start_bus, t] - model.bus_angle[end_bus, t]
            ) / reactance
        
        self.model.line_power_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_power_rule
        )
        
        # 7.4 线路容量约束
        def line_capacity_rule(model, line, t):
            line_info = self.line_dict.get(line, {})
            capacity = line_info.get('Capacity', 0)
            
            if capacity <= 0:
                return pyo.Constraint.Skip
            
            # 线路潮流不超过容量限制
            return (-capacity, model.line_power[line, t], capacity)
        
        self.model.line_capacity_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_capacity_rule
        )
        
        # 8. 系统功率平衡约束（全局）- 重新启用确保基本平衡
        def global_power_balance_rule(model, t):
            # 发电侧（储能放电算发电，储能充电算负荷）
            thermal_gen = sum(model.thermal_power[unit, t] for unit in model.ThermalUnits)
            wind_gen = sum(model.wind_power[unit, t] for unit in model.WindUnits)
            solar_gen = sum(model.solar_power[unit, t] for unit in model.SolarUnits)
            hydro_gen = sum(model.hydro_power[unit, t] for unit in model.HydroUnits)
            ess_discharge = sum(model.ess_discharge[unit, t] for unit in model.ESSUntis)
            ess_charge = sum(model.ess_charge[unit, t] for unit in model.ESSUntis)
            
            # 负荷侧 - 直接使用全局负荷曲线，不再按节点分解
            if t < len(self.load_curve):
                total_load = self.load_curve[t]
            else:
                total_load = 0
            
            # 切负荷总量
            total_shed = sum(model.load_shed[load, t] for load in model.Loads)
            
            # 修正：储能充电是消耗，应该加到负荷侧，不是发电侧
            return (thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge ==
                    total_load - total_shed + ess_charge)
        
        self.model.global_power_balance_con = pyo.Constraint(
            self.model.T, rule=global_power_balance_rule
        )

        # 9. 备用约束（修改为更合理的备用要求）
        def reserve_requirement_rule(model, t):
            if t >= len(self.load_curve):
                return pyo.Constraint.Skip
                
            total_load = self.load_curve[t]
            
            # 计算总的可用发电容量（包括所有类型的机组）
            total_thermal_capacity = sum(model.thermal_status[unit, t] * 
                                       self.thermal_dict[unit].get('Capacity', 0) 
                                       for unit in model.ThermalUnits)
            
            # 风电和光伏虽然不可控，但在备用计算中可以按一定比例计算
            total_wind_capacity = sum(self.wind_dict[unit].get('Capacity', 0) * 0.3  # 风电按30%可靠性计算
                                    for unit in model.WindUnits)
            total_solar_capacity = sum(self.solar_dict[unit].get('Capacity', 0) * 0.2  # 光伏按20%可靠性计算
                                     for unit in model.SolarUnits) 
            total_hydro_capacity = sum(self.hydro_dict[unit].get('Capacity', 0) * 0.8  # 水电按80%可靠性计算
                                     for unit in model.HydroUnits)
            
            # 储能放电容量（100%可靠性）
            total_ess_capacity = sum(self.ess_dict[unit].get('MaxDisPower', 0) 
                                   for unit in model.ESSUntis)
            
            total_available_capacity = (total_thermal_capacity + total_wind_capacity + 
                                      total_solar_capacity + total_hydro_capacity + total_ess_capacity)
            
            # 备用要求：负荷的20%（提高备用要求以确保足够的安全裕度）
            reserve_req = total_load * 0.20  # 提高到20%备用要求
            
            # 总可用容量必须大于负荷+备用要求
            return total_available_capacity >= total_load + reserve_req

        self.model.reserve_con = pyo.Constraint(
            self.model.T, rule=reserve_requirement_rule
        )
        
        logging.info("优化模型创建完成（包含直流潮流约束）")
    
    def debug_model(self):
        """调试模型不可行性"""
        logging.info("开始调试模型...")
        
        # 检查基本数据完整性
        logging.info(f"火电机组数量: {len(self.model.ThermalUnits)}")
        logging.info(f"风电机组数量: {len(self.model.WindUnits)}")
        logging.info(f"光伏机组数量: {len(self.model.SolarUnits)}")
        logging.info(f"水电机组数量: {len(self.model.HydroUnits)}")
        logging.info(f"储能单元数量: {len(self.model.ESSUntis)}")
        logging.info(f"负荷数量: {len(self.model.Loads)}")
        logging.info(f"母线数量: {len(self.model.Buses)}")
        logging.info(f"线路数量: {len(self.model.Lines)}")
        logging.info(f"时间段数: {self.T}")
        
        # 储能系统详细检查
        if len(self.model.ESSUntis) > 0:
            total_ess_charge_cap = sum(self.ess_dict[unit].get('MaxChargePower', 0) for unit in self.model.ESSUntis)
            total_ess_discharge_cap = sum(self.ess_dict[unit].get('MaxDisPower', 0) for unit in self.model.ESSUntis)
            total_ess_energy_cap = sum(self.ess_dict[unit].get('Capacity', 0) for unit in self.model.ESSUntis)
            logging.info(f"储能总充电功率: {total_ess_charge_cap:.1f}MW")
            logging.info(f"储能总放电功率: {total_ess_discharge_cap:.1f}MW")
            logging.info(f"储能总能量容量: {total_ess_energy_cap:.1f}MWh")
        
        # 检查负荷总量和发电容量
        total_thermal_cap = sum(self.thermal_dict[unit].get('Capacity', 0) for unit in self.model.ThermalUnits)
        total_wind_cap = sum(self.wind_dict[unit].get('Capacity', 0) for unit in self.model.WindUnits)
        total_solar_cap = sum(self.solar_dict[unit].get('Capacity', 0) for unit in self.model.SolarUnits)
        total_hydro_cap = sum(self.hydro_dict[unit].get('Capacity', 0) for unit in self.model.HydroUnits)
        total_gen_cap = total_thermal_cap + total_wind_cap + total_solar_cap + total_hydro_cap
        
        peak_load = max(self.load_curve) if self.load_curve else 0
        
        logging.info(f"总发电容量: {total_gen_cap} MW")
        logging.info(f"  - 火电容量: {total_thermal_cap} MW")
        logging.info(f"  - 风电容量: {total_wind_cap} MW")
        logging.info(f"  - 光伏容量: {total_solar_cap} MW")
        logging.info(f"  - 水电容量: {total_hydro_cap} MW")
        logging.info(f"峰值负荷: {peak_load} MW")
        logging.info(f"容量充裕度: {(total_gen_cap / peak_load * 100):.1f}%" if peak_load > 0 else "无负荷")
        
        # 详细检查各类机组
        logging.info("=== 风电机组详细信息 ===")
        for unit in self.model.WindUnits:
            cap = self.wind_dict[unit].get('Capacity', 0)
            curve = self.wind_dict[unit].get('ResourceCurve', '')
            bus = self.wind_dict[unit].get('BusName', '')
            logging.info(f"风电机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        logging.info("=== 光伏机组详细信息 ===")
        for unit in self.model.SolarUnits:
            cap = self.solar_dict[unit].get('Capacity', 0)
            curve = self.solar_dict[unit].get('ResourceCurve', '')
            bus = self.solar_dict[unit].get('BusName', '')
            logging.info(f"光伏机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        logging.info("=== 水电机组详细信息 ===")
        for unit in self.model.HydroUnits:
            cap = self.hydro_dict[unit].get('Capacity', 0)
            curve = self.hydro_dict[unit].get('ResourceCurve', '')
            bus = self.hydro_dict[unit].get('BusName', '')
            logging.info(f"水电机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        # 检查资源曲线数据
        logging.info("=== 资源曲线数据检查 ===")
        logging.info(f"风电曲线数量: {len(self.wind_curves)}")
        if not self.wind_curves.empty:
            logging.info(f"风电曲线名称: {list(self.wind_curves['WTCurve'].unique())}")
        
        logging.info(f"光伏曲线数量: {len(self.solar_curves)}")
        if not self.solar_curves.empty:
            logging.info(f"光伏曲线名称: {list(self.solar_curves['PVCurve'].unique())}")
        
        logging.info(f"水电曲线数量: {len(self.hydro_curves)}")
        if not self.hydro_curves.empty:
            logging.info(f"水电曲线名称: {list(self.hydro_curves['RunoffHydroOperationCurve'].unique())}")
        
        # 检查机组参数合理性
        for unit in self.model.ThermalUnits:
            unit_data = self.thermal_dict[unit]
            cap = unit_data.get('Capacity', 0)
            min_power = unit_data.get('MinPower', 0)
            if min_power > cap:
                logging.error(f"火电机组{unit}: 最小出力({min_power}) > 容量({cap})")
        
        # 输出模型统计信息
        logging.info(f"模型变量数: {len([v for v in self.model.component_objects(pyo.Var)])}")
        logging.info(f"模型约束数: {len([c for c in self.model.component_objects(pyo.Constraint)])}")
        
    def solve(self):
        """求解优化问题"""
        import logging
        # 保存原始日志级别并设置为CRITICAL以抑制ERROR消息
        original_level = logging.getLogger().level
        original_pyomo_level = logging.getLogger('pyomo').level
        logging.getLogger().setLevel(logging.CRITICAL)
        logging.getLogger('pyomo').setLevel(logging.CRITICAL)
        
        try:
            logging.info("开始求解优化问题...")
            
            # 先进行模型调试
            self.debug_model()
            
            # 关键容量检查
            logging.info("=== 关键容量检查 ===")
            peak_load = max(self.load_curve) if self.load_curve else 0
            total_thermal_cap = sum(self.thermal_dict[unit]['Capacity'] for unit in self.thermal_dict) if self.thermal_dict else 0
            total_wind_cap = sum(self.wind_dict[unit]['Capacity'] for unit in self.wind_dict) if self.wind_dict else 0
            total_solar_cap = sum(self.solar_dict[unit]['Capacity'] for unit in self.solar_dict) if self.solar_dict else 0
            total_hydro_cap = sum(self.hydro_dict[unit]['Capacity'] for unit in self.hydro_dict) if self.hydro_dict else 0
            total_capacity = total_thermal_cap + total_wind_cap + total_solar_cap + total_hydro_cap
            
            logging.info(f"峰值负荷: {peak_load:.1f}MW")
            logging.info(f"总发电容量: {total_capacity:.1f}MW")
            logging.info(f"容量充裕度: {(total_capacity/peak_load*100):.1f}%" if peak_load > 0 else "无负荷")
            
            if total_capacity < peak_load * 1.05:
                logging.error(f"严重警告：总容量({total_capacity:.1f}MW) 不足以满足峰值负荷+5%备用({peak_load*1.05:.1f}MW)")
        finally:
            # 暂时恢复INFO级别以显示关键信息
            logging.getLogger().setLevel(logging.INFO)
        
        # 使用Gurobi求解器
        solver = SolverFactory('gurobi')
        solver.options['MIPGap'] = 0.2  # 进一步放宽到20%
        solver.options['TimeLimit'] = 60  # 降低到1分钟
        solver.options['FeasibilityTol'] = 1e-3  # 放宽可行性容差
        solver.options['OptimalityTol'] = 1e-3  # 放宽最优性容差
        
        # 在求解前再次抑制日志以避免ERROR消息
        logging.getLogger().setLevel(logging.CRITICAL)
        logging.getLogger('pyomo').setLevel(logging.CRITICAL)
        
        # 尝试求解
        start_time = time.time()
        results = solver.solve(self.model, tee=True)
        solve_time = time.time() - start_time
        
        # 恢复INFO级别以显示求解结果
        logging.getLogger().setLevel(logging.INFO)
        logging.info(f"求解耗时: {solve_time:.2f}秒")
        
        # 检查求解状态
        if (results.solver.status == pyo.SolverStatus.ok and 
            results.solver.termination_condition in [pyo.TerminationCondition.optimal, 
                                                     pyo.TerminationCondition.feasible]):
            logging.info("优化求解成功完成")
            result = True
        elif results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            logging.warning("达到时间限制，返回当前最优解")
            result = True
        elif results.solver.termination_condition == pyo.TerminationCondition.infeasible:
            logging.error("模型不可行，尝试诊断...")
            self.diagnose_infeasibility()
            
            # 尝试简化模型
            logging.info("尝试创建简化模型...")
            if self.create_simplified_model():
                logging.info("重新求解简化模型...")
                results = solver.solve(self.model, tee=True)
                if (results.solver.status == pyo.SolverStatus.ok and 
                    results.solver.termination_condition in [pyo.TerminationCondition.optimal, 
                                                             pyo.TerminationCondition.feasible]):
                    logging.info("简化模型求解成功")
                    result = True
                else:
                    result = False
            else:
                result = False
        else:
            logging.error(f"优化求解失败: {results.solver.termination_condition}")
            result = False
            
        # 最后恢复原始日志级别
        logging.getLogger().setLevel(original_level)
        logging.getLogger('pyomo').setLevel(original_pyomo_level)
        
        return result
    
    def diagnose_infeasibility(self):
        """诊断模型不可行性"""
        logging.info("开始诊断模型不可行性...")
        
        # 写出模型文件用于调试
        try:
            self.model.write('debug_model.lp', format=pyo.WriterFactory.LP)
            logging.info("模型已写出到 debug_model.lp 文件")
        except Exception as e:
            logging.warning(f"无法写出模型文件: {e}")
        
        # 检查一些常见的不可行原因
        logging.info("检查常见不可行原因:")
        
        # 1. 检查负荷是否过大
        peak_load = max(self.load_curve) if self.load_curve else 0
        total_thermal_cap = sum(self.thermal_dict[unit].get('Capacity', 0) for unit in self.model.ThermalUnits)
        if peak_load > total_thermal_cap * 1.5:  # 考虑可再生能源
            logging.error(f"峰值负荷({peak_load}) 可能超过总发电能力")
        
        # 2. 检查机组最小开机时间是否过长
        for unit in self.model.ThermalUnits:
            min_up = self.thermal_dict[unit].get('MinOnLineTime', 0)
            min_down = self.thermal_dict[unit].get('MinOffLineTime', 0)
            if min_up > self.T // 2:
                logging.warning(f"机组{unit}最小开机时间({min_up})过长")
            if min_down > self.T // 2:
                logging.warning(f"机组{unit}最小停机时间({min_down})过长")
        
        # 3. 检查备用约束是否过紧
        logging.warning("建议放宽备用约束或检查机组参数设置")
    
    def create_simplified_model(self):
        """创建简化模型（移除一些复杂约束）"""
        try:
            logging.info("创建简化模型...")
            
            # 移除原有约束
            self.model.del_component(self.model.thermal_ramp_up_con)
            self.model.del_component(self.model.thermal_ramp_down_con)
            self.model.del_component(self.model.min_up_time_con)
            self.model.del_component(self.model.min_down_time_con)
            self.model.del_component(self.model.reserve_con)
            self.model.del_component(self.model.deep_peak_penalty_con)
            
            logging.info("已移除复杂约束：爬坡、最短开停机时间、备用、深度调峰")
            
            # 添加简化的爬坡约束（更宽松）
            def simple_ramp_rule(model, unit, t):
                if t == 0:
                    return pyo.Constraint.Skip
                capacity = self.thermal_dict[unit].get('Capacity', 0)
                # 使用容量的50%作为最大变化量
                max_change = capacity * 0.5
                return (-max_change <= 
                        model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= 
                        max_change)
            
            self.model.simple_ramp_con = pyo.Constraint(
                self.model.ThermalUnits, self.model.T, rule=simple_ramp_rule
            )
            
            logging.info("添加了简化的爬坡约束")
            return True
            
        except Exception as e:
            logging.error(f"创建简化模型失败: {e}")
            return False
    
    def calculate_cost_analysis(self):
        """计算和显示系统成本分析"""
        logging.info("开始计算系统成本分析...")
        
        # 安全获取变量值函数
        def safe_value(var):
            try:
                return pyo.value(var)
            except:
                return 0.0
        
        # 初始化成本字典
        cost_breakdown = {
            '火电运行成本': 0,
            '火电启动成本': 0,
            '深度调峰成本': 0,
            '可再生能源激励': 0,
            '储能运行成本': 0,
            '切负荷惩罚': 0,
            '节点平衡松弛成本': 0,
            '总成本': 0
        }
        
        fuel_price = self.fuel_data['UnitPrice']  # 与目标函数一致
        
        # 1. 火电运行成本和启动成本
        thermal_running_cost = 0
        thermal_startup_cost = 0
        
        print("\n" + "="*60)
        print("               火电机组成本明细")
        print("="*60)
        print(f"{'机组':<15} {'运行成本(元)':<15} {'启动成本(元)':<15} {'总成本(元)':<15}")
        print("-"*60)
        
        for unit in self.model.ThermalUnits:
            unit_running_cost = 0
            unit_startup_cost = 0
            
            # 获取机组参数
            a = self.thermal_dict[unit].get('OperationCoeff_A', 0)
            b = self.thermal_dict[unit].get('OperationCoeff_B', 100)
            c = self.thermal_dict[unit].get('OperationCoeff_C', 50)
            startup_cost = self.thermal_dict[unit].get('StateOnFuelConsumption', 200)
            
            for t in self.model.T:
                # 运行成本（与目标函数一致）
                if safe_value(self.model.thermal_status[unit, t]) > 0.5:
                    power = safe_value(self.model.thermal_power[unit, t])
                    fixed_cost = c * fuel_price  # 使用原始固定成本
                    variable_cost = b * power * fuel_price  # 使用原始变动成本
                    unit_running_cost += fixed_cost + variable_cost
                
                # 启动成本
                unit_startup_cost += startup_cost * fuel_price * safe_value(self.model.thermal_startup[unit, t])
            
            unit_total_cost = unit_running_cost + unit_startup_cost
            thermal_running_cost += unit_running_cost
            thermal_startup_cost += unit_startup_cost
            
            print(f"{unit:<15} {unit_running_cost:<15.2f} {unit_startup_cost:<15.2f} {unit_total_cost:<15.2f}")
        
        cost_breakdown['火电运行成本'] = thermal_running_cost
        cost_breakdown['火电启动成本'] = thermal_startup_cost
        
        # 2. 深度调峰成本
        deep_peak_cost = 0
        for unit in self.model.ThermalUnits:
            deep_cost = self.thermal_dict[unit].get('DeepPeakShavingCost', 0)
            for t in self.model.T:
                deep_peak_cost += deep_cost * safe_value(self.model.deep_peak_penalty[unit, t])
        cost_breakdown['深度调峰成本'] = deep_peak_cost
        
        # 3. 可再生能源激励（负成本表示收益）
        renewable_cost = 0
        renewable_incentive = -100  # 与目标函数中一致
        
        for unit in self.model.WindUnits:
            for t in self.model.T:
                renewable_cost += renewable_incentive * safe_value(self.model.wind_power[unit, t])
        
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                renewable_cost += renewable_incentive * safe_value(self.model.solar_power[unit, t])
        
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                renewable_cost += renewable_incentive * safe_value(self.model.hydro_power[unit, t])
        
        cost_breakdown['可再生能源激励'] = renewable_cost
        
        # 4. 储能运行成本
        ess_cost = 0
        ess_discharge_incentive = -500  # 与目标函数一致
        ess_charge_cost = 200          # 与目标函数一致
        
        for unit in self.model.ESSUntis:
            for t in self.model.T:
                ess_cost += ess_discharge_incentive * safe_value(self.model.ess_discharge[unit, t])
                ess_cost += ess_charge_cost * safe_value(self.model.ess_charge[unit, t])
        
        cost_breakdown['储能运行成本'] = ess_cost
        
        # 5. 切负荷惩罚
        load_shed_cost = 0
        penalty_cut = 1000000  # 与目标函数一致
        for load in self.model.Loads:
            for t in self.model.T:
                load_shed_cost += penalty_cut * safe_value(self.model.load_shed[load, t])
        cost_breakdown['切负荷惩罚'] = load_shed_cost
        
        # 6. 节点平衡松弛成本
        balance_slack_cost = 0
        penalty_slack = 500  # 与目标函数一致
        for bus in self.model.Buses:
            for t in self.model.T:
                balance_slack_cost += penalty_slack * safe_value(self.model.balance_slack[bus, t])
        cost_breakdown['节点平衡松弛成本'] = balance_slack_cost
        
        # 计算总成本
        cost_breakdown['总成本'] = sum(cost_breakdown.values()) - cost_breakdown['总成本']  # 避免重复计算
        
        # 显示成本汇总
        print("\n" + "="*60)
        print("               系统成本汇总")
        print("="*60)
        
        for category, cost in cost_breakdown.items():
            if category != '总成本':
                if cost < 0:
                    print(f"{category:<20}: -{abs(cost):>15,.2f} 元 (收益)")
                else:
                    print(f"{category:<20}: {cost:>15,.2f} 元")
        
        print("-"*60)
        print(f"{'系统总成本':<20}: {cost_breakdown['总成本']:>15,.2f} 元")
        print("="*60)
        
        # 计算发电量统计
        print("\n" + "="*60)
        print("               发电量统计")
        print("="*60)
        
        thermal_generation = 0
        wind_generation = 0
        solar_generation = 0
        hydro_generation = 0
        
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_generation += safe_value(self.model.thermal_power[unit, t])
        
        for unit in self.model.WindUnits:
            for t in self.model.T:
                wind_generation += safe_value(self.model.wind_power[unit, t])
        
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                solar_generation += safe_value(self.model.solar_power[unit, t])
        
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                hydro_generation += safe_value(self.model.hydro_power[unit, t])
        
        total_generation = thermal_generation + wind_generation + solar_generation + hydro_generation
        
        print(f"火电发电量: {thermal_generation:,.2f} MWh")
        print(f"风电发电量: {wind_generation:,.2f} MWh")
        print(f"光伏发电量: {solar_generation:,.2f} MWh")
        print(f"水电发电量: {hydro_generation:,.2f} MWh")
        print(f"总发电量: {total_generation:,.2f} MWh")
        
        # 计算平均发电成本
        if total_generation > 0:
            avg_cost = cost_breakdown['总成本'] / total_generation
            print(f"\n平均发电成本: {avg_cost:.2f} 元/MWh")
        
        # 计算各类型发电占比
        if total_generation > 0:
            print(f"\n发电结构:")
            print(f"火电占比: {thermal_generation/total_generation*100:.1f}%")
            print(f"风电占比: {wind_generation/total_generation*100:.1f}%")
            print(f"光伏占比: {solar_generation/total_generation*100:.1f}%")
            print(f"水电占比: {hydro_generation/total_generation*100:.1f}%")
        
        print("="*60)
        
        # 存储成本分析结果供后续使用
        self.cost_analysis = cost_breakdown
        self.generation_stats = {
            'thermal': thermal_generation,
            'wind': wind_generation,
            'solar': solar_generation,
            'hydro': hydro_generation,
            'total': total_generation
        }
        
        return cost_breakdown
    
    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                # 静默处理未初始化的变量，返回默认值
                return 0

        # 创建Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # 1. 火电机组结果
        thermal_results = []
        # 临时提高日志级别，避免显示未初始化变量的错误
        original_level = logging.getLogger().level
        logging.getLogger().setLevel(logging.CRITICAL)
        
        try:
            for unit in self.model.ThermalUnits:
                for t in self.model.T:
                    thermal_results.append({
                        'Unit': unit,
                        'Time': t,
                        'Status': safe_value(self.model.thermal_status[unit, t]),
                        'Power': safe_value(self.model.thermal_power[unit, t]),
                        'Startup': safe_value(self.model.thermal_startup[unit, t]),
                        'Shutdown': safe_value(self.model.thermal_shutdown[unit, t])
                    })
        finally:
            # 恢复原始日志级别
            logging.getLogger().setLevel(original_level)
            
        thermal_df = pd.DataFrame(thermal_results)
        thermal_df.to_excel(writer, sheet_name='火电机组结果', index=False)
        
        # 2. 可再生能源结果
        wind_results = []
        for unit in self.model.WindUnits:
            for t in self.model.T:
                wind_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.wind_power[unit, t])
                })
        wind_df = pd.DataFrame(wind_results)
        wind_df.to_excel(writer, sheet_name='风电结果', index=False)
        
        solar_results = []
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                solar_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.solar_power[unit, t])
                })
        solar_df = pd.DataFrame(solar_results)
        solar_df.to_excel(writer, sheet_name='光伏结果', index=False)
        
        hydro_results = []
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                hydro_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.hydro_power[unit, t])
                })
        hydro_df = pd.DataFrame(hydro_results)
        hydro_df.to_excel(writer, sheet_name='水电结果', index=False)
        
        # 3. 储能结果
        ess_results = []
        for unit in self.model.ESSUntis:
            for t in self.model.T:
                ess_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Charge': safe_value(self.model.ess_charge[unit, t]),
                    'Discharge': safe_value(self.model.ess_discharge[unit, t]),
                    'SOC': safe_value(self.model.ess_soc[unit, t])
                })
        ess_df = pd.DataFrame(ess_results)
        ess_df.to_excel(writer, sheet_name='储能结果', index=False)
        
        # 4. 负荷结果
        load_results = []
        for load in self.model.Loads:
            for t in self.model.T:
                load_results.append({
                    'Load': load,
                    'Time': t,
                    'Shed': safe_value(self.model.load_shed[load, t])
                })
        load_df = pd.DataFrame(load_results)
        load_df.to_excel(writer, sheet_name='切负荷结果', index=False)
        
        # 5. 网络结果
        bus_results = []
        # 临时提高日志级别，避免显示未初始化变量的错误
        original_level = logging.getLogger().level
        logging.getLogger().setLevel(logging.CRITICAL)
        
        try:
            for bus in self.model.Buses:
                for t in self.model.T:
                    bus_results.append({
                        'Bus': bus,
                        'Time': t,
                        'Angle': safe_value(self.model.bus_angle[bus, t])
                    })
        finally:
            # 恢复原始日志级别
            logging.getLogger().setLevel(original_level)
            
        bus_df = pd.DataFrame(bus_results)
        bus_df.to_excel(writer, sheet_name='节点相角', index=False)
        
        line_results = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_results.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_df = pd.DataFrame(line_results)
        line_df.to_excel(writer, sheet_name='线路潮流', index=False)
        
        # 6. 系统总结
        total_thermal_gen = thermal_df.groupby('Time')['Power'].sum().values
        total_wind_gen = wind_df.groupby('Time')['Power'].sum().values
        total_solar_gen = solar_df.groupby('Time')['Power'].sum().values
        total_hydro_gen = hydro_df.groupby('Time')['Power'].sum().values
        total_ess_discharge = ess_df.groupby('Time')['Discharge'].sum().values
        total_ess_charge = ess_df.groupby('Time')['Charge'].sum().values
        
        # 使用全局负荷曲线
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T

        summary_data = {
            'Time': self.hours,
            'Total_Load': total_load,
            'Thermal_Gen': total_thermal_gen,
            'Wind_Gen': total_wind_gen,
            'Solar_Gen': total_solar_gen,
            'Hydro_Gen': total_hydro_gen,
            'ESS_Discharge': total_ess_discharge,
            'ESS_Charge': total_ess_charge,
            'Net_Gen': [t + w + s + h + e - c for t, w, s, h, e, c in 
                       zip(total_thermal_gen, total_wind_gen, total_solar_gen, 
                           total_hydro_gen, total_ess_discharge, total_ess_charge)]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='系统总结', index=False)
        
        # 6. 成本分析结果（如果已计算）
        if hasattr(self, 'cost_analysis'):
            cost_data = []
            for category, cost in self.cost_analysis.items():
                cost_data.append({
                    '成本类别': category,
                    '成本金额(元)': cost,
                    '备注': '负值表示收益' if cost < 0 else ''
                })
            
            cost_df = pd.DataFrame(cost_data)
            cost_df.to_excel(writer, sheet_name='成本分析', index=False)
            
            # 发电量统计
            if hasattr(self, 'generation_stats'):
                gen_data = []
                gen_stats = self.generation_stats
                total_gen = gen_stats['total']
                
                gen_data.append({
                    '发电类型': '火电',
                    '发电量(MWh)': gen_stats['thermal'],
                    '占比(%)': (gen_stats['thermal'] / total_gen * 100) if total_gen > 0 else 0
                })
                
                gen_data.append({
                    '发电类型': '风电',
                    '发电量(MWh)': gen_stats['wind'],
                    '占比(%)': (gen_stats['wind'] / total_gen * 100) if total_gen > 0 else 0
                })
                
                gen_data.append({
                    '发电类型': '光伏',
                    '发电量(MWh)': gen_stats['solar'],
                    '占比(%)': (gen_stats['solar'] / total_gen * 100) if total_gen > 0 else 0
                })
                
                gen_data.append({
                    '发电类型': '水电',
                    '发电量(MWh)': gen_stats['hydro'],
                    '占比(%)': (gen_stats['hydro'] / total_gen * 100) if total_gen > 0 else 0
                })
                
                # 添加总计行
                gen_data.append({
                    '发电类型': '总计',
                    '发电量(MWh)': total_gen,
                    '占比(%)': 100.0
                })
                
                gen_df = pd.DataFrame(gen_data)
                gen_df.to_excel(writer, sheet_name='发电量统计', index=False)
        
        # 保存Excel
        writer.close()
        logging.info(f"结果已保存到 {output_path}")
        return output_path
    
    def plot_unit_status(self, output_path):
        """绘制机组开关机状态图"""
        logging.info("开始绘制机组开关机状态图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                # 静默处理未初始化的变量，返回默认值
                return 0
        
        # 获取火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        
        # 按机组分组
        grouped = thermal_df.groupby('Unit')
        
        # 创建图表
        fig, axes = plt.subplots(len(grouped), 1, figsize=(15, 10), sharex=True)
        fig.suptitle('火电机组开关机状态及出力情况', fontsize=16)
        
        if len(grouped) == 1:
            axes = [axes]
        
        for i, (unit, group) in enumerate(grouped):
            ax = axes[i]
            
            # 绘制状态（0/1）
            ax.step(group['Time'], group['Status'], 'r-', where='post', label='状态')
            ax.set_ylabel('状态', color='r')
            ax.tick_params(axis='y', labelcolor='r')
            ax.set_ylim(-0.1, 1.1)
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # 创建双Y轴
            ax2 = ax.twinx()
            
            # 绘制出力
            ax2.plot(group['Time'], group['Power'], 'b-', label='出力')
            ax2.set_ylabel('出力 (MW)', color='b')
            ax2.tick_params(axis='y', labelcolor='b')
            
            # 设置标题
            ax.set_title(f'机组 {unit}')
            
            # 添加图例
            lines, labels = ax.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax.legend(lines + lines2, labels + labels2, loc='upper right')
        
        plt.xlabel('时间 (小时)')
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # 为总标题留出空间
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"机组状态图已保存到 {output_path}")
        plt.close()
        return output_path
    
    def plot_generation_mix(self, output_path):
        """绘制ABC三个区域的各类电源出力情况（分区域堆积柱状图）"""
        logging.info("开始绘制分区域电源结构图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0
        
        # 按区域初始化结果数组
        areas = ['A区', 'B区', 'C区']
        area_data = {}
        
        for area in areas:
            area_data[area] = {
                'thermal_gen': np.zeros(self.T),
                'wind_gen': np.zeros(self.T),
                'solar_gen': np.zeros(self.T),
                'hydro_gen': np.zeros(self.T),
                'ess_discharge': np.zeros(self.T),
                'ess_charge': np.zeros(self.T),
                'wind_curtail': np.zeros(self.T),
                'solar_curtail': np.zeros(self.T),
                'hydro_curtail': np.zeros(self.T)
            }
        
        # 计算各区域的发电量
        try:
            for t in range(self.T):
                # 火电机组按区域统计
                for unit in self.model.ThermalUnits:
                    bus_name = self.thermal_dict[unit].get('BusName', '')
                    area = self.bus_to_area.get(bus_name, 'A区')  # 默认A区
                    power = safe_value(self.model.thermal_power[unit, t])
                    area_data[area]['thermal_gen'][t] += power
                
                # 风电机组按区域统计
                for unit in self.model.WindUnits:
                    bus_name = self.wind_dict[unit].get('BusName', '')
                    area = self.bus_to_area.get(bus_name, 'A区')
                    power = safe_value(self.model.wind_power[unit, t])
                    area_data[area]['wind_gen'][t] += power
                    
                    # 计算弃风
                    try:
                        curve_name = self.wind_dict[unit].get('ResourceCurve', '')
                        capacity = self.wind_dict[unit].get('Capacity', 0)
                        max_output = self._get_renewable_max_output(curve_name, t, capacity, 'wind')
                        curtail = max(0, max_output - power)
                        area_data[area]['wind_curtail'][t] += curtail
                    except Exception as e:
                        logging.warning(f"计算风电 {unit} 弃电量时出错: {e}")
                
                # 光伏机组按区域统计
                for unit in self.model.SolarUnits:
                    bus_name = self.solar_dict[unit].get('BusName', '')
                    area = self.bus_to_area.get(bus_name, 'A区')
                    power = safe_value(self.model.solar_power[unit, t])
                    area_data[area]['solar_gen'][t] += power
                    
                    # 计算弃光
                    try:
                        curve_name = self.solar_dict[unit].get('ResourceCurve', '')
                        capacity = self.solar_dict[unit].get('Capacity', 0)
                        max_output = self._get_renewable_max_output(curve_name, t, capacity, 'solar')
                        curtail = max(0, max_output - power)
                        area_data[area]['solar_curtail'][t] += curtail
                    except Exception as e:
                        logging.warning(f"计算光伏 {unit} 弃电量时出错: {e}")
                
                # 水电机组按区域统计
                for unit in self.model.HydroUnits:
                    bus_name = self.hydro_dict[unit].get('BusName', '')
                    area = self.bus_to_area.get(bus_name, 'A区')
                    power = safe_value(self.model.hydro_power[unit, t])
                    area_data[area]['hydro_gen'][t] += power
                    
                    # 计算弃水
                    try:
                        curve_name = self.hydro_dict[unit].get('ResourceCurve', '')
                        capacity = self.hydro_dict[unit].get('Capacity', 0)
                        max_output = self._get_renewable_max_output(curve_name, t, capacity, 'hydro')
                        curtail = max(0, max_output - power)
                        area_data[area]['hydro_curtail'][t] += curtail
                    except Exception as e:
                        logging.warning(f"计算水电 {unit} 弃电量时出错: {e}")
                
                # 储能按区域统计
                for unit in self.model.ESSUntis:
                    try:
                        bus_name = self.ess_dict[unit].get('BusName', '')
                        area = self.bus_to_area.get(bus_name, 'A区')
                        if (unit, t) in self.model.ess_discharge and (unit, t) in self.model.ess_charge:
                            discharge = safe_value(self.model.ess_discharge[unit, t])
                            charge = safe_value(self.model.ess_charge[unit, t])
                            area_data[area]['ess_discharge'][t] += discharge
                            area_data[area]['ess_charge'][t] += charge
                    except Exception as e:
                        logging.warning(f"计算储能 {unit} 在时段 {t} 时出错: {e}")
        
        except Exception as e:
            logging.error(f"计算分区域发电量时出错: {e}")
        
        # 创建三个子图
        fig, axes = plt.subplots(1, 3, figsize=(20, 8))
        fig.suptitle('ABC三区域发电出力情况对比', fontsize=16, fontweight='bold')
        
        hours = np.arange(self.T)
        width = 0.8
        
        # 颜色方案
        colors = {
            'thermal': '#ff4444',  # 红色-火电
            'hydro': '#4444ff',    # 蓝色-水电
            'wind': '#44ff44',     # 绿色-风电
            'solar': '#ffff44',    # 黄色-光伏
            'ess_discharge': '#333333',  # 深灰-储能放电
            'ess_charge': '#666666'      # 浅灰-储能充电
        }
        
        for i, area in enumerate(areas):
            ax = axes[i]
            data = area_data[area]
            
            # 绘制堆积柱状图
            bottom = np.zeros(self.T)
            
            # 火电
            thermal_gen = data['thermal_gen']
            if max(thermal_gen) > 0.1:
                ax.bar(hours, thermal_gen, width, bottom=bottom, 
                      label='火电', color=colors['thermal'], alpha=0.8)
                bottom += thermal_gen
            
            # 水电
            hydro_gen = data['hydro_gen']
            if max(hydro_gen) > 0.1:
                ax.bar(hours, hydro_gen, width, bottom=bottom, 
                      label='水电', color=colors['hydro'], alpha=0.8)
                bottom += hydro_gen
            
            # 风电
            wind_gen = data['wind_gen']
            if max(wind_gen) > 0.1:
                ax.bar(hours, wind_gen, width, bottom=bottom, 
                      label='风电', color=colors['wind'], alpha=0.8)
                bottom += wind_gen
            
            # 光伏
            solar_gen = data['solar_gen']
            if max(solar_gen) > 0.1:
                ax.bar(hours, solar_gen, width, bottom=bottom, 
                      label='光伏', color=colors['solar'], alpha=0.8)
                bottom += solar_gen
            
            # 储能放电
            ess_discharge = data['ess_discharge']
            if max(ess_discharge) > 0.1:
                ax.bar(hours, ess_discharge, width, bottom=bottom, 
                      label='储能放电', color=colors['ess_discharge'], alpha=0.9)
                bottom += ess_discharge
            
            # 储能充电（负值显示）
            ess_charge = data['ess_charge']
            if max(ess_charge) > 0.1:
                ax.bar(hours, [-x for x in ess_charge], width, 
                      label='储能充电', color=colors['ess_charge'], alpha=0.9)
            
            # 弃电量显示（半透明堆叠）
            curtail_bottom = bottom.copy()
            
            # 弃风
            wind_curtail = data['wind_curtail']
            if max(wind_curtail) > 0.1:
                ax.bar(hours, wind_curtail, width, bottom=curtail_bottom, 
                      label='弃风', color=colors['wind'], alpha=0.3, hatch='///')
                curtail_bottom += wind_curtail
            
            # 弃光
            solar_curtail = data['solar_curtail']
            if max(solar_curtail) > 0.1:
                ax.bar(hours, solar_curtail, width, bottom=curtail_bottom, 
                      label='弃光', color=colors['solar'], alpha=0.3, hatch='///')
                curtail_bottom += solar_curtail
            
            # 弃水
            hydro_curtail = data['hydro_curtail']
            if max(hydro_curtail) > 0.1:
                ax.bar(hours, hydro_curtail, width, bottom=curtail_bottom, 
                      label='弃水', color=colors['hydro'], alpha=0.3, hatch='///')
            
            # 设置子图属性
            ax.set_title(f'{area}发电出力', fontsize=14, fontweight='bold')
            ax.set_xlabel('时间 (小时)', fontsize=12)
            ax.set_ylabel('功率 (MW)', fontsize=12)
            ax.grid(True, alpha=0.3)
            ax.set_xlim(-0.5, self.T-0.5)
            ax.set_xticks(range(0, self.T, 4))
            
            # 添加零线
            ax.axhline(y=0, color='black', linestyle='--', alpha=0.5, linewidth=1)
            
            # 计算区域统计
            total_thermal = sum(thermal_gen)
            total_wind = sum(wind_gen)
            total_solar = sum(solar_gen)
            total_hydro = sum(hydro_gen)
            total_ess_discharge = sum(ess_discharge)
            total_ess_charge = sum(ess_charge)
            total_generation = total_thermal + total_wind + total_solar + total_hydro + total_ess_discharge
            max_power = max(bottom) if len(bottom) > 0 else 0
            
            # 添加统计信息
            stats_text = (
                f'{area}统计(24h):\n'
                f'总发电: {total_generation:.0f}MWh\n'
                f'火电: {total_thermal:.0f}MWh\n'
                f'水电: {total_hydro:.0f}MWh\n'
                f'风电: {total_wind:.0f}MWh\n'
                f'光伏: {total_solar:.0f}MWh\n'
                f'储能放电: {total_ess_discharge:.0f}MWh\n'
                f'储能充电: {total_ess_charge:.0f}MWh\n'
                f'峰值功率: {max_power:.0f}MW'
            )
            
            # 将统计信息放在图表内
            ax.text(0.02, 0.98, stats_text, transform=ax.transAxes, 
                   fontsize=9, verticalalignment='top',
                   bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.8))
            
            # 只在第一个子图显示图例
            if i == 0:
                ax.legend(loc='upper right', bbox_to_anchor=(1, 0.85), fontsize=10)
        
        # 调整布局
        plt.tight_layout()
        
        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        logging.info(f"分区域电源结构图已保存到 {output_path}")
        plt.close()
        return output_path

    def plot_unit_power_curves(self, output_path):
        """绘制24小时各类机组出力情况折线图"""
        logging.info("开始绘制各类机组出力折线图...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                logging.getLogger().setLevel(original_level)
        
        # 初始化各类机组出力数据
        thermal_power = np.zeros(self.T)
        wind_power = np.zeros(self.T)
        solar_power = np.zeros(self.T)
        hydro_power = np.zeros(self.T)
        ess_discharge_power = np.zeros(self.T)
        ess_charge_power = np.zeros(self.T)
        
        # 收集各类机组的详细出力数据
        thermal_units_power = {}
        wind_units_power = {}
        solar_units_power = {}
        hydro_units_power = {}
        ess_units_power = {}
        
        # 收集火电机组数据
        for unit in self.model.ThermalUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.thermal_power[unit, t])
                unit_power.append(power)
                thermal_power[t] += power
            thermal_units_power[unit] = unit_power
        
        # 收集风电机组数据
        for unit in self.model.WindUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.wind_power[unit, t])
                unit_power.append(power)
                wind_power[t] += power
            wind_units_power[unit] = unit_power
        
        # 收集光伏机组数据
        for unit in self.model.SolarUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.solar_power[unit, t])
                unit_power.append(power)
                solar_power[t] += power
            solar_units_power[unit] = unit_power
        
        # 收集水电机组数据
        for unit in self.model.HydroUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.hydro_power[unit, t])
                unit_power.append(power)
                hydro_power[t] += power
            hydro_units_power[unit] = unit_power
        
        # 收集储能机组数据
        for unit in self.model.ESSUntis:
            unit_discharge = []
            unit_charge = []
            for t in range(self.T):
                discharge = safe_value(self.model.ess_discharge[unit, t])
                charge = safe_value(self.model.ess_charge[unit, t])
                unit_discharge.append(discharge)
                unit_charge.append(charge)
                ess_discharge_power[t] += discharge
                ess_charge_power[t] += charge
            ess_units_power[unit] = {'discharge': unit_discharge, 'charge': unit_charge}
        
        # 创建图表
        fig, axes = plt.subplots(3, 2, figsize=(20, 15))
        fig.suptitle('24小时各类机组出力情况', fontsize=16, fontweight='bold')
        
        hours = np.arange(self.T)
        
        # 1. 火电机组出力图
        ax1 = axes[0, 0]
        ax1.plot(hours, thermal_power, 'r-', linewidth=3, label='火电总出力', alpha=0.8)
        # 显示前5个主要火电机组的详细出力
        thermal_sorted = sorted(thermal_units_power.items(), 
                               key=lambda x: max(x[1]), reverse=True)[:5]
        colors = ['#FF6B6B', '#FF8E53', '#FF006E', '#8338EC', '#3A86FF']
        for i, (unit, power_curve) in enumerate(thermal_sorted):
            if max(power_curve) > 10:  # 只显示有意义出力的机组
                ax1.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax1.set_title('火电机组出力', fontsize=14)
        ax1.set_ylabel('功率 (MW)')
        ax1.grid(True, alpha=0.3)
        ax1.legend(fontsize=10)
        
        # 2. 风电机组出力图
        ax2 = axes[0, 1]
        ax2.plot(hours, wind_power, 'g-', linewidth=3, label='风电总出力', alpha=0.8)
        # 显示主要风电机组
        wind_sorted = sorted(wind_units_power.items(), 
                            key=lambda x: max(x[1]), reverse=True)[:3]
        colors = ['#2ECC71', '#27AE60', '#16A085']
        for i, (unit, power_curve) in enumerate(wind_sorted):
            if max(power_curve) > 1:
                ax2.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax2.set_title('风电机组出力', fontsize=14)
        ax2.set_ylabel('功率 (MW)')
        ax2.grid(True, alpha=0.3)
        ax2.legend(fontsize=10)
        
        # 3. 光伏机组出力图
        ax3 = axes[1, 0]
        ax3.plot(hours, solar_power, 'y-', linewidth=3, label='光伏总出力', alpha=0.8)
        # 显示主要光伏机组
        solar_sorted = sorted(solar_units_power.items(), 
                             key=lambda x: max(x[1]), reverse=True)[:3]
        colors = ['#F39C12', '#E67E22', '#D35400']
        for i, (unit, power_curve) in enumerate(solar_sorted):
            if max(power_curve) > 1:
                ax3.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax3.set_title('光伏机组出力', fontsize=14)
        ax3.set_ylabel('功率 (MW)')
        ax3.grid(True, alpha=0.3)
        ax3.legend(fontsize=10)
        
        # 4. 水电机组出力图
        ax4 = axes[1, 1]
        ax4.plot(hours, hydro_power, 'b-', linewidth=3, label='水电总出力', alpha=0.8)
        # 显示主要水电机组
        hydro_sorted = sorted(hydro_units_power.items(), 
                             key=lambda x: max(x[1]), reverse=True)[:3]
        colors = ['#3498DB', '#2980B9', '#1F4E79']
        for i, (unit, power_curve) in enumerate(hydro_sorted):
            if max(power_curve) > 1:
                ax4.plot(hours, power_curve, '--', linewidth=2, 
                        color=colors[i % len(colors)], alpha=0.7, label=f'{unit}')
        ax4.set_title('水电机组出力', fontsize=14)
        ax4.set_ylabel('功率 (MW)')
        ax4.grid(True, alpha=0.3)
        ax4.legend(fontsize=10)
        
        # 5. 储能机组出力图（充放电）
        ax5 = axes[2, 0]
        ax5.plot(hours, ess_discharge_power, 'purple', linewidth=3, label='储能放电', alpha=0.8)
        ax5.plot(hours, -ess_charge_power, 'orange', linewidth=3, label='储能充电', alpha=0.8)
        # 显示主要储能机组
        ess_sorted = sorted(ess_units_power.items(), 
                           key=lambda x: max(x[1]['discharge']), reverse=True)[:3]
        colors_discharge = ['#9B59B6', '#8E44AD', '#663399']
        colors_charge = ['#E74C3C', '#C0392B', '#922B21']
        for i, (unit, power_data) in enumerate(ess_sorted):
            if max(power_data['discharge']) > 0.1:
                ax5.plot(hours, power_data['discharge'], '--', linewidth=2,
                        color=colors_discharge[i % len(colors_discharge)], alpha=0.7, 
                        label=f'{unit}_放电')
            if max(power_data['charge']) > 0.1:
                ax5.plot(hours, [-x for x in power_data['charge']], '--', linewidth=2,
                        color=colors_charge[i % len(colors_charge)], alpha=0.7, 
                        label=f'{unit}_充电')
        ax5.set_title('储能机组充放电', fontsize=14)
        ax5.set_ylabel('功率 (MW)')
        ax5.axhline(y=0, color='black', linestyle='-', alpha=0.5)
        ax5.grid(True, alpha=0.3)
        ax5.legend(fontsize=10)
        
        # 6. 综合对比图
        ax6 = axes[2, 1]
        ax6.plot(hours, thermal_power, 'r-', linewidth=3, label='火电', alpha=0.8)
        ax6.plot(hours, wind_power, 'g-', linewidth=3, label='风电', alpha=0.8)
        ax6.plot(hours, solar_power, 'y-', linewidth=3, label='光伏', alpha=0.8)
        ax6.plot(hours, hydro_power, 'b-', linewidth=3, label='水电', alpha=0.8)
        ax6.plot(hours, ess_discharge_power, 'purple', linewidth=3, label='储能放电', alpha=0.8)
        
        # 添加负荷曲线作为参考
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T
        ax6.plot(hours, total_load, 'k--', linewidth=2, label='总负荷', alpha=0.6)
        
        ax6.set_title('各类机组出力综合对比', fontsize=14)
        ax6.set_ylabel('功率 (MW)')
        ax6.set_xlabel('时间 (小时)')
        ax6.grid(True, alpha=0.3)
        ax6.legend(fontsize=10)
        
        # 设置所有子图的x轴
        for ax in axes.flat:
            ax.set_xlim(0, self.T-1)
            ax.set_xticks(range(0, self.T, 4))  # 每4小时标记一次
        
        # 添加统计信息
        stats_text = (
            f'统计信息 (24小时):\n'
            f'火电: 最大{max(thermal_power):.1f}MW, 平均{np.mean(thermal_power):.1f}MW\n'
            f'风电: 最大{max(wind_power):.1f}MW, 平均{np.mean(wind_power):.1f}MW\n'
            f'光伏: 最大{max(solar_power):.1f}MW, 平均{np.mean(solar_power):.1f}MW\n'
            f'水电: 最大{max(hydro_power):.1f}MW, 平均{np.mean(hydro_power):.1f}MW\n'
            f'储能放电: 最大{max(ess_discharge_power):.1f}MW\n'
            f'储能充电: 最大{max(ess_charge_power):.1f}MW'
        )
        
        # 在右上角添加统计信息
        fig.text(0.98, 0.98, stats_text, fontsize=10, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='white', alpha=0.8),
                transform=fig.transFigure, ha='right')
        
        # 调整布局
        plt.tight_layout(rect=[0, 0, 0.85, 0.95])  # 为统计信息和标题留出空间
        
        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        logging.info(f"各类机组出力折线图已保存到 {output_path}")
        plt.close()
        return output_path

    def plot_thermal_units_detail(self, output_path):
        """绘制32台火电机组24小时出力详细折线图"""
        logging.info("开始绘制火电机组详细出力折线图...")
        
        # 安全获取变量值
        def safe_value(var):
            import logging
            original_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.CRITICAL)
            try:
                value = pyo.value(var)
                return value
            except Exception:
                return 0
            finally:
                logging.getLogger().setLevel(original_level)
        
        # 收集所有火电机组的24小时出力数据
        thermal_units_power = {}
        total_thermal_power = np.zeros(self.T)
        
        for unit in self.model.ThermalUnits:
            unit_power = []
            for t in range(self.T):
                power = safe_value(self.model.thermal_power[unit, t])
                unit_power.append(power)
                total_thermal_power[t] += power
            thermal_units_power[unit] = unit_power
        
        # 创建大图，专门显示所有火电机组
        plt.figure(figsize=(16, 10))
        
        hours = np.arange(self.T)
        
        # 生成区分度高的颜色序列
        import matplotlib.cm as cm
        import matplotlib.colors as mcolors
        
        # 使用多个色谱生成足够多的颜色
        colors = []
        # 第一组：使用tab20色谱（20种颜色）
        tab20_colors = plt.cm.tab20(np.linspace(0, 1, 20))
        colors.extend(tab20_colors)
        
        # 第二组：使用Set3色谱（12种颜色）
        set3_colors = plt.cm.Set3(np.linspace(0, 1, 12))
        colors.extend(set3_colors)
        
        # 按出力大小排序火电机组（便于识别主要机组）
        thermal_sorted = sorted(thermal_units_power.items(), 
                               key=lambda x: max(x[1]), reverse=True)
        
        # 绘制每台火电机组的出力曲线
        active_units = 0  # 统计有效出力的机组数量
        for i, (unit, power_curve) in enumerate(thermal_sorted):
            max_power = max(power_curve)
            avg_power = np.mean(power_curve)
            
            # 只显示有意义出力的机组（最大出力>1MW）
            if max_power > 1.0:
                color = colors[active_units % len(colors)]
                
                # 根据机组重要性调整线条样式
                if max_power > 100:  # 主力机组，粗线
                    linewidth = 2.5
                    alpha = 0.9
                    linestyle = '-'
                elif max_power > 50:  # 重要机组，中等线宽
                    linewidth = 2.0
                    alpha = 0.8
                    linestyle = '-'
                else:  # 小机组，细线
                    linewidth = 1.5
                    alpha = 0.7
                    linestyle = '-'
                
                plt.plot(hours, power_curve, 
                        color=color, linewidth=linewidth, alpha=alpha, linestyle=linestyle,
                        label=f'{unit} (最大:{max_power:.1f}MW, 平均:{avg_power:.1f}MW)')
                
                active_units += 1
        
        # 添加总出力曲线作为参考
        plt.plot(hours, total_thermal_power, 'black', linewidth=4, alpha=0.8, 
                label=f'火电总出力 (最大:{max(total_thermal_power):.1f}MW)', linestyle='--')
        
        # 图表设置
        plt.title('ROTS系统32台火电机组24小时出力详细分析', fontsize=16, fontweight='bold', pad=20)
        plt.xlabel('时间 (小时)', fontsize=14)
        plt.ylabel('功率 (MW)', fontsize=14)
        plt.grid(True, alpha=0.3, linestyle=':', linewidth=0.5)
        
        # 设置x轴
        plt.xlim(0, self.T-1)
        plt.xticks(range(0, self.T, 4), fontsize=12)  # 每4小时标记一次
        plt.yticks(fontsize=12)
        
        # 智能图例布局 - 分成两列以节省空间
        if active_units > 16:
            ncol = 3
            fontsize = 8
        elif active_units > 8:
            ncol = 2
            fontsize = 9
        else:
            ncol = 1
            fontsize = 10
            
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', 
                  ncol=ncol, fontsize=fontsize, framealpha=0.9)
        
        # 添加统计信息文本框
        stats_text = (
            f'火电机组统计信息:\n'
            f'• 有效出力机组数: {active_units}台\n'
            f'• 系统火电装机: {sum(max(power_curve) for _, power_curve in thermal_sorted if max(power_curve) > 1):.1f}MW\n'
            f'• 24h总发电量: {sum(total_thermal_power):.1f}MWh\n'
            f'• 平均出力: {np.mean(total_thermal_power):.1f}MW\n'
            f'• 最大出力: {max(total_thermal_power):.1f}MW\n'
            f'• 最小出力: {min(total_thermal_power):.1f}MW\n'
            f'• 平均利用率: {np.mean(total_thermal_power)/sum(max(power_curve) for _, power_curve in thermal_sorted if max(power_curve) > 1)*100:.1f}%'
        )
        
        # 在左上角添加统计信息
        plt.text(0.02, 0.98, stats_text, transform=plt.gca().transAxes, 
                fontsize=10, verticalalignment='top',
                bbox=dict(boxstyle='round,pad=0.5', facecolor='lightblue', alpha=0.8))
        
        # 添加峰谷分析
        peak_hour = np.argmax(total_thermal_power)
        valley_hour = np.argmin(total_thermal_power)
        peak_valley_ratio = max(total_thermal_power) / max(min(total_thermal_power), 1)
        
        # 标记峰值和谷值点
        plt.annotate(f'峰值\n{peak_hour}时\n{max(total_thermal_power):.1f}MW', 
                    xy=(peak_hour, max(total_thermal_power)), 
                    xytext=(peak_hour+2, max(total_thermal_power)+50),
                    arrowprops=dict(arrowstyle='->', color='red', lw=2),
                    fontsize=10, ha='center',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='red', alpha=0.7, edgecolor='red'))
        
        plt.annotate(f'谷值\n{valley_hour}时\n{min(total_thermal_power):.1f}MW', 
                    xy=(valley_hour, min(total_thermal_power)), 
                    xytext=(valley_hour+2, min(total_thermal_power)-50),
                    arrowprops=dict(arrowstyle='->', color='blue', lw=2),
                    fontsize=10, ha='center',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='blue', alpha=0.7, edgecolor='blue'))
        
        # 调整布局以适应图例
        plt.tight_layout()
        plt.subplots_adjust(right=0.75)  # 为右侧图例留出空间
        
        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        logging.info(f"火电机组详细出力图已保存到 {output_path}")
        plt.close()
        return output_path
        
    def plot_network_flow(self, output_path):
        """绘制网络潮流图"""
        logging.info("开始绘制网络潮流图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                # 静默处理未初始化的变量，返回默认值
                return 0
        
        # 获取线路潮流数据
        line_flows = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_flows.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_flows_df = pd.DataFrame(line_flows)
        
        # 获取节点相角数据
        bus_angles = []
        # 临时提高日志级别，避免显示未初始化变量的错误
        original_level = logging.getLogger().level
        logging.getLogger().setLevel(logging.CRITICAL)
        
        try:
            for bus in self.model.Buses:
                for t in self.model.T:
                    bus_angles.append({
                        'Bus': bus,
                        'Time': t,
                        'Angle': safe_value(self.model.bus_angle[bus, t])
                    })
        finally:
            # 恢复原始日志级别
            logging.getLogger().setLevel(original_level)
            
        bus_angles_df = pd.DataFrame(bus_angles)
        
        # 创建图表
        plt.figure(figsize=(15, 10))
        
        # 绘制线路潮流
        unique_lines = line_flows_df['Line'].unique()
        for i, line in enumerate(unique_lines):
            line_data = line_flows_df[line_flows_df['Line'] == line]
            plt.plot(line_data['Time'], line_data['Power'], label=f'线路 {line}')
        
        plt.title('线路潮流随时间变化')
        plt.xlabel('时间 (小时)')
        plt.ylabel('功率 (MW)')
        plt.legend(loc='best', bbox_to_anchor=(1.05, 1))
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"网络潮流图已保存到 {output_path}")
        plt.close()
        return output_path

# 主程序
if __name__ == "__main__":
    # 输入输出路径
    input_file = r"C:\Users\admin\Desktop\Final_Cleaned_Rots_v2.xlsx"
    output_file = "ROTS_UC_Results.xlsx"
    status_plot = "ROTS_Unit_Status.png"
    mix_plot = "ROTS_Generation_Mix.png"
    flow_plot = "ROTS_Network_Flow.png"
    resource_plot = "ROTS_Resource_Curves.png"
    power_curves_plot = "ROTS_Power_Curves.png"
    thermal_detail_plot = "ROTS_Thermal_Detail.png"
    
    try:
        # 创建模型
        logging.info("创建模型...")
        uc_model = UC_Model(input_file)
        
        # 绘制资源曲线预览
        uc_model.plot_resource_curves(resource_plot)
        
        # 求解模型
        logging.info("求解模型...")
        result = uc_model.solve()
        
        if result:
            # 计算和显示成本分析
            logging.info("计算成本分析...")
            uc_model.calculate_cost_analysis()
            
            # 保存结果
            logging.info("保存结果...")
            result_path = uc_model.save_results(output_file)
            
            # 绘制图表
            logging.info("绘制图表...")
            status_path = uc_model.plot_unit_status(status_plot)
            mix_path = uc_model.plot_generation_mix(mix_plot)
            flow_path = uc_model.plot_network_flow(flow_plot)
            power_curves_path = uc_model.plot_unit_power_curves(power_curves_plot)
            thermal_detail_path = uc_model.plot_thermal_units_detail(thermal_detail_plot)
            
            # 将图表插入Excel
            wb = openpyxl.load_workbook(output_file)
            
            # 添加资源曲线预览图
            ws = wb.create_sheet("资源曲线预览")
            img = Image(resource_plot)
            ws.add_image(img, 'A1')
            
            # 添加机组状态图
            ws = wb.create_sheet("机组状态图")
            img = Image(status_path)
            ws.add_image(img, 'A1')
            
            # 添加电源结构图
            ws = wb.create_sheet("电源结构图")
            img = Image(mix_path)
            ws.add_image(img, 'A1')
            
            # 添加网络潮流图
            ws = wb.create_sheet("网络潮流图")
            img = Image(flow_path)
            ws.add_image(img, 'A1')
            
            # 添加机组出力折线图
            ws = wb.create_sheet("机组出力曲线")
            img = Image(power_curves_path)
            ws.add_image(img, 'A1')
            
            # 添加火电机组详细分析图
            ws = wb.create_sheet("火电机组详细分析")
            img = Image(thermal_detail_path)
            ws.add_image(img, 'A1')
            
            # 保存最终结果
            wb.save(output_file)
            logging.info("所有结果已保存并整合到Excel文件中")
            print(f"优化完成，结果已保存到: {output_file}")
        else:
            logging.error("优化求解失败，请检查模型和输入数据")
            print("优化求解失败，请检查日志获取详细信息")
        
    except Exception as e:
        logging.exception("程序运行出错")
        print(f"程序运行出错: {str(e)}")
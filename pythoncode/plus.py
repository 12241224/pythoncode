import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import os
import logging
import sys
import time
import math

# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 初始化日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class UC_Model:
    def __init__(self, data_path):
        self.data_path = data_path
        self.load_data()
        self.create_model()
        
    def load_data(self):
        """加载所有Excel数据"""
        logging.info("开始加载数据...")
        
        # 系统汇总信息
        self.system_summary = pd.read_excel(
            self.data_path, sheet_name='SystemSummary', skiprows=0, nrows=1
        ).iloc[0]
        
        # 机组参数
        self.thermal_units = pd.read_excel(
            self.data_path, sheet_name='UnitThermalGenerators', skiprows=0
        )
        self.wind_units = pd.read_excel(
            self.data_path, sheet_name='UnitWindGenerators', skiprows=0
        )
        self.wind_units.rename(columns={
            'WTPlantID': 'WTPlantID',
            '机组装机容量(MW)': 'Capacity',
            '风电资源曲线名称': 'ResourceCurve'
        }, inplace=True)
        
        self.solar_units = pd.read_excel(
            self.data_path, sheet_name='UnitSolarGenerators', skiprows=0
        )
        self.solar_units.rename(columns={
            'PVPlantID': 'PVPlantID',
            '机组装机容量(MW)': 'Capacity',
            '光伏资源曲线名称': 'ResourceCurve'
        }, inplace=True)
        
        self.ess_units = pd.read_excel(
            self.data_path, sheet_name='StorElectrochemicalESS', skiprows=0
        )
        self.ess_units.rename(columns={
            'ESSId': 'ESSId',
            '储能容量(MWh)': 'Capacity',
            '充电功率(MW)': 'MaxChargePower',
            '放电功率(MW)': 'MaxDisPower'
        }, inplace=True)
        
        self.hydro_units = pd.read_excel(
            self.data_path, sheet_name='UnitRunoffHydroGenerators', skiprows=0
        )
        self.hydro_units.rename(columns={
            '径流式水电机组名称': 'RunoffHydroUnitName',
            '机组装机容量(MW)': 'Capacity',
            '径流式水电资源曲线名称': 'ResourceCurve'
        }, inplace=True)
        # 保证RunoffHydroUnitName唯一，否则to_dict会报错
        if self.hydro_units['RunoffHydroUnitName'].duplicated().any():
            logging.warning("发现重复的水电机组名称，将自动去重（保留首次出现）")
            self.hydro_units = self.hydro_units.drop_duplicates(subset='RunoffHydroUnitName', keep='first')
        self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        
        # 曲线数据
        self.wind_curves = pd.read_excel(
            self.data_path, sheet_name='CurveWindResource', skiprows=0
        )
        self.wind_curves.rename(columns={
            'WTCurve': 'WTCurve'
        }, inplace=True)
        self.wind_curves = self.wind_curves.dropna(subset=['WTCurve'])
        
        self.solar_curves = pd.read_excel(
            self.data_path, sheet_name='CurveSolarResource', skiprows=0
        )
        self.solar_curves.rename(columns={
            'PVCurve': 'PVCurve'
        }, inplace=True)
        self.solar_curves = self.solar_curves.dropna(subset=['PVCurve'])
        
        self.hydro_curves = pd.read_excel(
            self.data_path, sheet_name='CurveRunoffHydroResource', skiprows=0
        )
        self.hydro_curves.rename(columns={
            'RunoffHydroOperationCurve': 'RunoffHydroOperationCurve'
        }, inplace=True)
        self.hydro_curves = self.hydro_curves.dropna(subset=['RunoffHydroOperationCurve'])
        
        self.load_curve = pd.read_excel(
            self.data_path, sheet_name='CurveLoad', skiprows=0
        ).iloc[0]
        
        # 网络数据
        self.buses = pd.read_excel(self.data_path, sheet_name='NetBuses', skiprows=0)
        self.lines = pd.read_excel(self.data_path, sheet_name='NetLines', skiprows=0)
        self.transformers = pd.read_excel(self.data_path, sheet_name='NetTransformers', skiprows=0)
        self.sections = pd.read_excel(self.data_path, sheet_name='NetSectionsLines', skiprows=0)
        self.section_capacity = pd.read_excel(self.data_path, sheet_name='NetSectionsCapacity', skiprows=0)
        self.ac_tielines = pd.read_excel(self.data_path, sheet_name='NetHVACTieLine', skiprows=0)
        self.ac_tieline_curves = pd.read_excel(self.data_path, sheet_name='EnergyHVACTieLine', skiprows=0)
        
        # 负荷数据
        self.loads = pd.read_excel(self.data_path, sheet_name='Loads', skiprows=0)
        
        # 燃料数据
        self.fuel_data = pd.read_excel(self.data_path, sheet_name='UnitFuel', skiprows=0).iloc[0]
        
        # 设置时间参数
        self.T = int(self.system_summary['Duration'])
        self.dt = 1  # 调度间隔为1小时
        self.hours = list(range(self.T))
        
        # 创建机组字典
        self.thermal_dict = self.thermal_units.set_index('ThermalUnitNumber').to_dict('index')
        for unit in self.thermal_dict:
            # 确保数值类型正确
            self.thermal_dict[unit]['MinOnLineTime'] = int(self.thermal_dict[unit].get('MinOnLineTime', 0))
            self.thermal_dict[unit]['MinOffLineTime'] = int(self.thermal_dict[unit].get('MinOffLineTime', 0))
            self.thermal_dict[unit]['Capacity'] = float(self.thermal_dict[unit].get('Capacity', 0))
            self.thermal_dict[unit]['MinPower'] = float(self.thermal_dict[unit].get('MinPower', 0))
            self.thermal_dict[unit]['UpRamppingRate'] = float(self.thermal_dict[unit].get('UpRamppingRate', 0))
            self.thermal_dict[unit]['DownRamppingRate'] = float(self.thermal_dict[unit].get('DownRamppingRate', 0))
            
        self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        self.solar_dict = self.solar_units.set_index('PVPlantID').to_dict('index')
        self.ess_dict = self.ess_units.set_index('ESSId').to_dict('index')
        self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        
        # 构建母线字典
        self.bus_dict = self.buses.set_index('BusId').to_dict('index')
        
        # 构建线路字典
        self.line_dict = {}
        for idx, row in self.lines.iterrows():
            self.line_dict[row['BranchId']] = {
                'StartNodeName': row['StartNodeName'],
                'EndNodeName': row['EndNodeName'],
                'Resistance': float(row.get('Resistance', 0)),
                'Reactance': float(row.get('Reactance', 0)),
                'Capacity': float(row.get('Capacity', 0)),
                'Susceptance': float(row.get('Susceptance', 0))
            }
        
        # 找到平衡节点（参考节点）
        self.ref_bus = self.system_summary['BalanceNode']
        logging.info(f"参考节点设置为: {self.ref_bus}")
        
        logging.info("数据加载完成")
    
    def get_unit_param(self, unit_dict, unit, key, default=0.0):
        """安全获取机组参数"""
        try:
            val = unit_dict[unit].get(key, default)
            if pd.isna(val):
                return default
            return val
        except Exception:
            return default

    def create_model(self):
        """创建Pyomo优化模型（包含直流潮流约束）"""
        logging.info("开始创建优化模型（包含直流潮流约束）...")
        
        self.model = pyo.ConcreteModel()
        
        # 定义集合
        self.model.T = pyo.Set(initialize=self.hours)  # 时间段
        self.model.ThermalUnits = pyo.Set(initialize=self.thermal_units['ThermalUnitNumber'].tolist())
        self.model.WindUnits = pyo.Set(initialize=self.wind_units['WTPlantID'].tolist())
        self.model.SolarUnits = pyo.Set(initialize=self.solar_units['PVPlantID'].tolist())
        self.model.ESSUntis = pyo.Set(initialize=self.ess_units['ESSId'].tolist())
        self.model.HydroUnits = pyo.Set(initialize=self.hydro_units['RunoffHydroUnitName'].tolist())
        self.model.Loads = pyo.Set(initialize=self.loads['负荷编号'].tolist())
        self.model.Buses = pyo.Set(initialize=self.buses['BusId'].tolist())
        self.model.Lines = pyo.Set(initialize=self.lines['BranchId'].tolist())
        
        # 定义变量
        # 火电机组
        self.model.thermal_status = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.thermal_startup = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_shutdown = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        
        # 可再生能源
        self.model.wind_power = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_power = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_power = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 储能系统
        self.model.ess_charge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_discharge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_soc = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 负荷相关
        self.model.load_shed = pyo.Var(
            self.model.Loads, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 直流潮流变量
        self.model.bus_angle = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )  # 节点相角
        
        self.model.line_power = pyo.Var(
            self.model.Lines, self.model.T, within=pyo.Reals
        )  # 线路潮流
        
        # 目标函数
        def total_cost_rule(model):
            """系统总成本计算"""
            cost = 0
            fuel_price = self.fuel_data['UnitPrice']
            
            # 火电运行成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    a = self.thermal_dict[unit].get('OperationCoeff_A', 0)
                    b = self.thermal_dict[unit].get('OperationCoeff_B', 0)
                    c = self.thermal_dict[unit].get('OperationCoeff_C', 0)
                    
                    # 运行成本
                    cost += model.thermal_status[unit, t] * c * fuel_price
                    cost += b * model.thermal_power[unit, t] * fuel_price
                    
                    # 启动成本
                    startup_cost = self.thermal_dict[unit].get('StateOnFuelConsumption', 0)
                    cost += startup_cost * fuel_price * model.thermal_startup[unit, t]
            
            # 深度调峰惩罚
            self.model.deep_peak_penalty = pyo.Var(
                model.ThermalUnits, model.T, within=pyo.NonNegativeReals
            )
            def deep_peak_penalty_rule(model, unit, t):
                max_power = self.thermal_dict[unit].get('Capacity', 0)
                deep_threshold = self.thermal_dict[unit].get('DeepPeakShavingThreshold', 0.5) * max_power
                deep_cost = self.thermal_dict[unit].get('DeepPeakShavingCost', 0)
                
                # 深度调峰惩罚
                return model.deep_peak_penalty[unit, t] >= deep_cost * (
                    deep_threshold - model.thermal_power[unit, t]
                )
            
            model.deep_peak_penalty_con = pyo.Constraint(
                model.ThermalUnits, model.T, rule=deep_peak_penalty_rule
            )
            
            # 将深度调峰惩罚加入总成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    cost += model.deep_peak_penalty[unit, t]
            
            # 切负荷惩罚
            penalty_cut = 100000  # 元/MWh
            for load in model.Loads:
                for t in model.T:
                    cost += penalty_cut * model.load_shed[load, t]
            
            return cost
        
        self.model.total_cost = pyo.Objective(rule=total_cost_rule, sense=pyo.minimize)
        
        # 1. 火电机组逻辑约束
        def thermal_logic_rule(model, unit, t):
            if t == 0:
                init_status = self.thermal_dict[unit].get('InitStatus', False)
                if isinstance(init_status, str):
                    init_status = init_status.strip().lower() in ['true', '1', 'yes']
                return model.thermal_status[unit, t] == (1 if init_status else 0)
            else:
                return model.thermal_status[unit, t] - model.thermal_status[unit, t-1] == \
                    model.thermal_startup[unit, t] - model.thermal_shutdown[unit, t]
        self.model.thermal_logic_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_logic_rule
        )

        # 2. 火电机组出力约束
        def thermal_power_lower_rule(model, unit, t):
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            return model.thermal_power[unit, t] >= min_power * model.thermal_status[unit, t]
        
        def thermal_power_upper_rule(model, unit, t):
            max_power = self.thermal_dict[unit].get('Capacity', 0)
            return model.thermal_power[unit, t] <= max_power * model.thermal_status[unit, t]
        
        self.model.thermal_power_lower_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_lower_rule
        )
        self.model.thermal_power_upper_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_upper_rule
        )

        # 3. 爬坡约束 - 使用更稳健的方案2
        def thermal_ramp_up_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_up = self.thermal_dict[unit].get('UpRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束
            return (model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= 
                    model.thermal_status[unit, t] * ramp_up + 
                    (1 - model.thermal_status[unit, t-1]) * capacity)
        
        def thermal_ramp_down_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_down = self.thermal_dict[unit].get('DownRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束
            return (model.thermal_power[unit, t-1] - model.thermal_power[unit, t] <= 
                    model.thermal_status[unit, t-1] * ramp_down + 
                    (1 - model.thermal_status[unit, t]) * capacity)

        self.model.thermal_ramp_up_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_up_rule
        )
        self.model.thermal_ramp_down_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_down_rule
        )

        # 4. 最短开停机时间约束 - 使用更稳健的方案
        def min_up_time_rule(model, unit, t):
            min_up = self.thermal_dict[unit].get('MinOnLineTime', 0)
            if t < min_up:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_up:
                return sum(model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        model.thermal_status[unit, t] - prev_status) * (self.T - t)
            else:
                return sum(model.thermal_status[unit, k] for k in range(t, t+min_up)) >= (
                        model.thermal_status[unit, t] - prev_status) * min_up

        def min_down_time_rule(model, unit, t):
            min_down = self.thermal_dict[unit].get('MinOffLineTime', 0)
            if t < min_down:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_down:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        prev_status - model.thermal_status[unit, t]) * (self.T - t)
            else:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, t+min_down)) >= (
                        prev_status - model.thermal_status[unit, t]) * min_down

        self.model.min_up_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_up_time_rule
        )
        self.model.min_down_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_down_time_rule
        )

        # 5. 可再生能源出力约束
        def wind_power_rule(model, unit, t):
            curve_name = self.wind_dict[unit].get('ResourceCurve', '')
            capacity = self.wind_dict[unit].get('Capacity', 0)
            
            # 找到对应的风资源曲线
            curve_row = self.wind_curves[self.wind_curves['WTCurve'] == curve_name]
            if curve_row.empty:
                return model.wind_power[unit, t] <= 0
            
            # 获取时间点值
            time_col = f'Time_{t}'
            if time_col in curve_row.columns:
                max_factor = curve_row[time_col].values[0]
            else:
                max_factor = 0
                
            return model.wind_power[unit, t] <= max_factor * capacity
        
        self.model.wind_power_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_power_rule
        )
        
        def solar_power_rule(model, unit, t):
            curve_name = self.solar_dict[unit].get('ResourceCurve', '')
            capacity = self.solar_dict[unit].get('Capacity', 0)
            
            # 找到对应的光伏资源曲线
            curve_row = self.solar_curves[self.solar_curves['PVCurve'] == curve_name]
            if curve_row.empty:
                return model.solar_power[unit, t] <= 0
            
            # 获取时间点值
            time_col = f'Time_{t}'
            if time_col in curve_row.columns:
                max_factor = curve_row[time_col].values[0]
            else:
                max_factor = 0
                
            return model.solar_power[unit, t] <= max_factor * capacity
        
        self.model.solar_power_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_power_rule
        )
        
        def hydro_power_rule(model, unit, t):
            curve_name = self.hydro_dict[unit].get('ResourceCurve', '')
            capacity = self.hydro_dict[unit].get('Capacity', 0)
            
            # 找到对应的水电资源曲线
            curve_row = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'] == curve_name]
            if curve_row.empty:
                return model.hydro_power[unit, t] <= 0
            
            # 获取时间点值
            time_col = f'Time_{t}'
            if time_col in curve_row.columns:
                max_factor = curve_row[time_col].values[0]
            else:
                max_factor = 0
                
            return model.hydro_power[unit, t] <= max_factor * capacity
        
        self.model.hydro_power_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_power_rule
        )

        # 6. 储能系统约束
        def ess_charge_power_rule(model, unit, t):
            max_charge = self.ess_dict[unit].get('MaxChargePower', 0)
            return model.ess_charge[unit, t] <= max_charge
        
        def ess_discharge_power_rule(model, unit, t):
            max_discharge = self.ess_dict[unit].get('MaxDisPower', 0)
            return model.ess_discharge[unit, t] <= max_discharge
        
        def ess_soc_rule(model, unit, t):
            min_soc = self.ess_dict[unit].get('MinSOC', 0)
            max_soc = self.ess_dict[unit].get('MaxSOC', 1000)
            capacity = self.ess_dict[unit].get('Capacity', 0)
            return (min_soc, model.ess_soc[unit, t], max_soc)
        
        def ess_soc_balance_rule(model, unit, t):
            charge_eff = self.ess_dict[unit].get('ChargeEfficiency', 0.85)
            dis_eff = self.ess_dict[unit].get('DisEfficiency', 0.85)
            
            if t == 0:
                init_soc = self.ess_dict[unit].get('InitSOC', 0)
                return model.ess_soc[unit, t] == init_soc
            else:
                return model.ess_soc[unit, t] == (model.ess_soc[unit, t-1] + 
                                                  charge_eff * model.ess_charge[unit, t] - 
                                                  model.ess_discharge[unit, t] / dis_eff)
        
        self.model.ess_charge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_charge_power_rule
        )
        self.model.ess_discharge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_discharge_power_rule
        )
        self.model.ess_soc_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_rule
        )
        self.model.ess_soc_balance_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_balance_rule
        )

        # 7. 直流潮流约束
        logging.info("添加直流潮流约束...")
        
        # 7.1 参考节点约束（平衡节点相角为0）
        def ref_bus_rule(model, t):
            return model.bus_angle[self.ref_bus, t] == 0
        self.model.ref_bus_con = pyo.Constraint(
            self.model.T, rule=ref_bus_rule
        )
        
        # 7.2 节点功率平衡约束
        def node_balance_rule(model, bus, t):
            # 节点发电量（包括所有连接到该节点的发电单元）
            gen_power = 0
            # 火电
            for unit in model.ThermalUnits:
                if self.thermal_dict[unit].get('BusName', '') == bus:
                    gen_power += model.thermal_power[unit, t]
            # 风电
            for unit in model.WindUnits:
                if self.wind_dict[unit].get('BusName', '') == bus:
                    gen_power += model.wind_power[unit, t]
            # 光伏
            for unit in model.SolarUnits:
                if self.solar_dict[unit].get('BusName', '') == bus:
                    gen_power += model.solar_power[unit, t]
            # 水电
            for unit in model.HydroUnits:
                if self.hydro_dict[unit].get('BusName', '') == bus:
                    gen_power += model.hydro_power[unit, t]
            # 储能放电
            for unit in model.ESSUntis:
                if self.ess_dict[unit].get('BusName', '') == bus:
                    gen_power += model.ess_discharge[unit, t]
            # 储能充电
            ess_charge = 0
            for unit in model.ESSUntis:
                if self.ess_dict[unit].get('BusName', '') == bus:
                    ess_charge += model.ess_charge[unit, t]
            
            # 节点负荷（包括切负荷）
            load_power = 0
            load_shed = 0
            for load in model.Loads:
                load_info = self.loads[self.loads['负荷编号'] == load]
                if not load_info.empty and load_info['所在母线编号'].values[0] == bus:
                    coef = load_info['有功比例系数'].values[0]
                    time_col = f'Time_{t}'
                    if time_col in self.load_curve:
                        curve_val = self.load_curve[time_col]
                    else:
                        curve_val = 0
                    # 保证为float
                    try:
                        coef = float(coef)
                    except Exception:
                        coef = 0.0
                    try:
                        curve_val = float(curve_val)
                    except Exception:
                        curve_val = 0.0
                    load_power += coef * curve_val
                    load_shed += model.load_shed[load, t]
            
            # 节点净注入功率
            net_injection = gen_power - ess_charge - (load_power - load_shed)
            
            # 计算线路潮流总和（流出为正，流入为负）
            line_flow = 0
            for line in model.Lines:
                line_info = self.line_dict.get(line, {})
                if line_info.get('StartNodeName', '') == bus:
                    line_flow += model.line_power[line, t]
                elif line_info.get('EndNodeName', '') == bus:
                    line_flow -= model.line_power[line, t]
            
            # 节点功率平衡
            return net_injection == line_flow
        
        self.model.node_balance_con = pyo.Constraint(
            self.model.Buses, self.model.T, rule=node_balance_rule
        )
        
        # 7.3 线路潮流约束（直流潮流方程）
        def line_power_rule(model, line, t):
            line_info = self.line_dict.get(line, {})
            start_bus = line_info.get('StartNodeName', '')
            end_bus = line_info.get('EndNodeName', '')
            reactance = line_info.get('Reactance', 0.001)  # 避免除零错误
            
            # 线路潮流与相角差的关系
            return model.line_power[line, t] == (
                model.bus_angle[start_bus, t] - model.bus_angle[end_bus, t]
            ) / reactance
        
        self.model.line_power_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_power_rule
        )
        
        # 7.4 线路容量约束
        def line_capacity_rule(model, line, t):
            capacity = self.line_dict.get(line, {}).get('Capacity', 10000)  # 默认大容量
            return (-capacity, model.line_power[line, t], capacity)
        
        self.model.line_capacity_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_capacity_rule
        )
        
        # 8. 系统功率平衡约束（全局）
        def global_power_balance_rule(model, t):
            # 发电侧
            thermal_gen = sum(model.thermal_power[unit, t] for unit in model.ThermalUnits)
            wind_gen = sum(model.wind_power[unit, t] for unit in model.WindUnits)
            solar_gen = sum(model.solar_power[unit, t] for unit in model.SolarUnits)
            hydro_gen = sum(model.hydro_power[unit, t] for unit in model.HydroUnits)
            ess_discharge = sum(model.ess_discharge[unit, t] for unit in model.ESSUntis)
            ess_charge = sum(model.ess_charge[unit, t] for unit in model.ESSUntis)
            
            # 负荷侧
            total_load = 0
            total_shed = 0
            for load in model.Loads:
                load_info = self.loads[self.loads['负荷编号'] == load]
                if not load_info.empty:
                    coef = load_info['有功比例系数'].values[0]
                else:
                    coef = 0

                time_col = f'Time_{t}'
                if time_col in self.load_curve:
                    curve_val = self.load_curve[time_col]
                else:
                    curve_val = 0

                try:
                    coef = float(coef)
                except Exception:
                    coef = 0.0
                try:
                    curve_val = float(curve_val)
                except Exception:
                    curve_val = 0.0

                total_load += coef * curve_val
                total_shed += model.load_shed[load, t]
            
            return (thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge ==
                    total_load - total_shed + ess_charge)

        self.model.global_power_balance_con = pyo.Constraint(
            self.model.T, rule=global_power_balance_rule
        )

        # 9. 备用约束
        def reserve_requirement_rule(model, t):
            # 计算总负荷
            total_load = 0
            for load in model.Loads:
                load_info = self.loads[self.loads['负荷编号'] == load]
                if not load_info.empty:
                    coef = load_info['有功比例系数'].values[0]
                else:
                    coef = 0

                time_col = f'Time_{t}'
                if time_col in self.load_curve:
                    curve_val = self.load_curve[time_col]
                else:
                    curve_val = 0

                try:
                    coef = float(coef)
                except Exception:
                    coef = 0.0
                try:
                    curve_val = float(curve_val)
                except Exception:
                    curve_val = 0.0

                total_load += coef * curve_val

            # 计算可用备用
            available_reserve = 0
            for unit in model.ThermalUnits:
                max_power = self.thermal_dict[unit].get('Capacity', 0)
                min_power = self.thermal_dict[unit].get('MinPower', 0)
                available_reserve += (model.thermal_status[unit, t] * max_power -
                                     model.thermal_power[unit, t])

            # 10%的备用要求
            reserve_req = 0.1 * total_load
            return available_reserve >= reserve_req

        self.model.reserve_con = pyo.Constraint(
            self.model.T, rule=reserve_requirement_rule
        )
        
        logging.info("优化模型创建完成（包含直流潮流约束）")
    
    def solve(self):
        """求解优化问题"""
        logging.info("开始求解优化问题...")
        
        # 使用Gurobi求解器
        solver = SolverFactory('gurobi')
        solver.options['MIPGap'] = 0.02  # 设置MIP gap为2%
        solver.options['TimeLimit'] = 600  # 设置10分钟时间限制
        
        # 尝试求解
        start_time = time.time()
        results = solver.solve(self.model, tee=True)
        solve_time = time.time() - start_time
        logging.info(f"求解耗时: {solve_time:.2f}秒")
        
        # 检查求解状态
        if (results.solver.status == pyo.SolverStatus.ok and 
            results.solver.termination_condition in [pyo.TerminationCondition.optimal, 
                                                     pyo.TerminationCondition.feasible]):
            logging.info("优化求解成功完成")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            logging.warning("达到时间限制，返回当前最优解")
            return True
        else:
            logging.error(f"优化求解失败: {results.solver.termination_condition}")
            return False
    
    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0

        # 创建Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # 1. 火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t]),
                    'Startup': safe_value(self.model.thermal_startup[unit, t]),
                    'Shutdown': safe_value(self.model.thermal_shutdown[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        thermal_df.to_excel(writer, sheet_name='火电机组结果', index=False)
        
        # 2. 可再生能源结果
        wind_results = []
        for unit in self.model.WindUnits:
            for t in self.model.T:
                wind_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.wind_power[unit, t])
                })
        wind_df = pd.DataFrame(wind_results)
        wind_df.to_excel(writer, sheet_name='风电结果', index=False)
        
        solar_results = []
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                solar_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.solar_power[unit, t])
                })
        solar_df = pd.DataFrame(solar_results)
        solar_df.to_excel(writer, sheet_name='光伏结果', index=False)
        
        hydro_results = []
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                hydro_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.hydro_power[unit, t])
                })
        hydro_df = pd.DataFrame(hydro_results)
        hydro_df.to_excel(writer, sheet_name='水电结果', index=False)
        
        # 3. 储能结果
        ess_results = []
        for unit in self.model.ESSUntis:
            for t in self.model.T:
                ess_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Charge': safe_value(self.model.ess_charge[unit, t]),
                    'Discharge': safe_value(self.model.ess_discharge[unit, t]),
                    'SOC': safe_value(self.model.ess_soc[unit, t])
                })
        ess_df = pd.DataFrame(ess_results)
        ess_df.to_excel(writer, sheet_name='储能结果', index=False)
        
        # 4. 负荷结果
        load_results = []
        for load in self.model.Loads:
            for t in self.model.T:
                load_results.append({
                    'Load': load,
                    'Time': t,
                    'Shed': safe_value(self.model.load_shed[load, t])
                })
        load_df = pd.DataFrame(load_results)
        load_df.to_excel(writer, sheet_name='切负荷结果', index=False)
        
        # 5. 网络结果
        bus_results = []
        for bus in self.model.Buses:
            for t in self.model.T:
                bus_results.append({
                    'Bus': bus,
                    'Time': t,
                    'Angle': safe_value(self.model.bus_angle[bus, t])
                })
        bus_df = pd.DataFrame(bus_results)
        bus_df.to_excel(writer, sheet_name='节点相角', index=False)
        
        line_results = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_results.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_df = pd.DataFrame(line_results)
        line_df.to_excel(writer, sheet_name='线路潮流', index=False)
        
        # 6. 系统总结
        total_thermal_gen = thermal_df.groupby('Time')['Power'].sum().values
        total_wind_gen = wind_df.groupby('Time')['Power'].sum().values
        total_solar_gen = solar_df.groupby('Time')['Power'].sum().values
        total_hydro_gen = hydro_df.groupby('Time')['Power'].sum().values
        total_ess_discharge = ess_df.groupby('Time')['Discharge'].sum().values
        total_ess_charge = ess_df.groupby('Time')['Charge'].sum().values
        
        total_load = []
        for t in self.hours:
            time_col = f'Time_{t}'
            if time_col in self.load_curve:
                curve_val = self.load_curve[time_col]
            else:
                curve_val = 0
            total_load.append(curve_val)

        # 修正：只有total_load长度与self.T一致时才绘制
        if len(total_load) == self.T:
            plt.plot(range(self.T), total_load, 'k--', linewidth=2, label='总负荷')
        
        summary_data = {
            'Time': self.hours,
            'Total_Load': total_load,
            'Thermal_Gen': total_thermal_gen,
            'Wind_Gen': total_wind_gen,
            'Solar_Gen': total_solar_gen,
            'Hydro_Gen': total_hydro_gen,
            'ESS_Discharge': total_ess_discharge,
            'ESS_Charge': total_ess_charge,
            'Net_Gen': [t + w + s + h + e - c for t, w, s, h, e, c in 
                       zip(total_thermal_gen, total_wind_gen, total_solar_gen, 
                           total_hydro_gen, total_ess_discharge, total_ess_charge)]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='系统总结', index=False)
        
        # 保存Excel
        writer.close()
        logging.info(f"结果已保存到 {output_path}")
        return output_path
    
    def plot_unit_status(self, output_path):
        """绘制机组开关机状态图"""
        logging.info("开始绘制机组开关机状态图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0
        
        # 获取火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        
        # 按机组分组
        grouped = thermal_df.groupby('Unit')
        
        # 创建图表
        fig, axes = plt.subplots(len(grouped), 1, figsize=(15, 10), sharex=True)
        fig.suptitle('火电机组开关机状态及出力情况', fontsize=16)
        
        if len(grouped) == 1:
            axes = [axes]
        
        for i, (unit, group) in enumerate(grouped):
            ax = axes[i]
            
            # 绘制状态（0/1）
            ax.step(group['Time'], group['Status'], 'r-', where='post', label='状态')
            ax.set_ylabel('状态', color='r')
            ax.tick_params(axis='y', labelcolor='r')
            ax.set_ylim(-0.1, 1.1)
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # 创建双Y轴
            ax2 = ax.twinx()
            
            # 绘制出力
            ax2.plot(group['Time'], group['Power'], 'b-', label='出力')
            ax2.set_ylabel('出力 (MW)', color='b')
            ax2.tick_params(axis='y', labelcolor='b')
            
            # 设置标题
            ax.set_title(f'机组 {unit}')
            
            # 添加图例
            lines, labels = ax.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax.legend(lines + lines2, labels + labels2, loc='upper right')
        
        plt.xlabel('时间 (小时)')
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # 为总标题留出空间
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"机组状态图已保存到 {output_path}")
        plt.close()
        return output_path
    
    def plot_generation_mix(self, output_path):
        """绘制各类电源出力情况"""
        logging.info("开始绘制电源结构图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0
        
        # 初始化结果数组
        thermal_gen = np.zeros(self.T)
        wind_gen = np.zeros(self.T)
        solar_gen = np.zeros(self.T)
        hydro_gen = np.zeros(self.T)
        ess_discharge = np.zeros(self.T)
        
        # 计算总发电量
        for t in range(self.T):
            for unit in self.model.ThermalUnits:
                thermal_gen[t] += safe_value(self.model.thermal_power[unit, t])
            for unit in self.model.WindUnits:
                wind_gen[t] += safe_value(self.model.wind_power[unit, t])
            for unit in self.model.SolarUnits:
                solar_gen[t] += safe_value(self.model.solar_power[unit, t])
            for unit in self.model.HydroUnits:
                hydro_gen[t] += safe_value(self.model.hydro_power[unit, t])
            for unit in self.model.ESSUntis:
                ess_discharge[t] += safe_value(self.model.ess_discharge[unit, t])
        
        # 创建图表
        plt.figure(figsize=(12, 6))
        
        # 堆叠面积图
        plt.stackplot(range(self.T), thermal_gen, wind_gen, solar_gen, hydro_gen, ess_discharge,
                      labels=['火电', '风电', '光伏', '水电', '储能放电'])
        
        # 添加总负荷曲线
        total_load = []
        for t in self.hours:
            time_col = f'Time_{t}'
            if time_col in self.load_curve:
                curve_val = self.load_curve[time_col]
            else:
                curve_val = 0
            total_load.append(curve_val)

        # 修正：只有total_load长度与self.T一致时才绘制
        if len(total_load) == self.T:
            plt.plot(range(self.T), total_load, 'k--', linewidth=2, label='总负荷')
        
        # 设置图表属性
        plt.title('各类电源出力及负荷曲线')
        plt.xlabel('时间 (小时)')
        plt.ylabel('功率 (MW)')
        plt.legend(loc='upper left')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.xlim(0, self.T-1)
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"电源结构图已保存到 {output_path}")
        plt.close()
        return output_path
    
    def plot_network_flow(self, output_path):
        """绘制网络潮流图"""
        logging.info("开始绘制网络潮流图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0
        
        # 获取线路潮流数据
        line_flows = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_flows.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_flows_df = pd.DataFrame(line_flows)
        
        # 获取节点相角数据
        bus_angles = []
        for bus in self.model.Buses:
            for t in self.model.T:
                bus_angles.append({
                    'Bus': bus,
                    'Time': t,
                    'Angle': safe_value(self.model.bus_angle[bus, t])
                })
        bus_angles_df = pd.DataFrame(bus_angles)
        
        # 创建图表
        plt.figure(figsize=(15, 10))
        
        # 绘制线路潮流
        unique_lines = line_flows_df['Line'].unique()
        for i, line in enumerate(unique_lines):
            line_data = line_flows_df[line_flows_df['Line'] == line]
            plt.plot(line_data['Time'], line_data['Power'], label=f'线路 {line}')
        
        plt.title('线路潮流随时间变化')
        plt.xlabel('时间 (小时)')
        plt.ylabel('功率 (MW)')
        plt.legend(loc='best', bbox_to_anchor=(1.05, 1))
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"网络潮流图已保存到 {output_path}")
        plt.close()
        return output_path

# 主程序
if __name__ == "__main__":
    # 输入输出路径
    input_file = r"C:\Users\admin\Desktop\Final_Cleaned_Rots_v2.xlsx"
    output_file = "UC_Results.xlsx"
    status_plot = "Unit_Status.png"
    mix_plot = "Generation_Mix.png"
    flow_plot = "Network_Flow.png"
    
    try:
        # 创建模型
        logging.info("创建模型...")
        uc_model = UC_Model(input_file)
        
        # 求解模型
        logging.info("求解模型...")
        result = uc_model.solve()
        
        if result:
            # 保存结果
            logging.info("保存结果...")
            result_path = uc_model.save_results(output_file)
            
            # 绘制图表
            logging.info("绘制图表...")
            status_path = uc_model.plot_unit_status(status_plot)
            mix_path = uc_model.plot_generation_mix(mix_plot)
            flow_path = uc_model.plot_network_flow(flow_plot)
            
            # 将图表插入Excel
            wb = openpyxl.load_workbook(output_file)
            
            # 添加机组状态图
            ws = wb.create_sheet("机组状态图")
            img = Image(status_path)
            ws.add_image(img, 'A1')
            
            # 添加电源结构图
            ws = wb.create_sheet("电源结构图")
            img = Image(mix_path)
            ws.add_image(img, 'A1')
            
            # 添加网络潮流图
            ws = wb.create_sheet("网络潮流图")
            img = Image(flow_path)
            ws.add_image(img, 'A1')
            
            # 保存最终结果
            wb.save(output_file)
            logging.info("所有结果已保存并整合到Excel文件中")
            print(f"优化完成，结果已保存到: {output_file}")
        else:
            logging.error("优化求解失败，请检查模型和输入数据")
            print("优化求解失败，请检查日志获取详细信息")
        
    except Exception as e:
        logging.exception("程序运行出错")
        print(f"程序运行出错: {str(e)}")
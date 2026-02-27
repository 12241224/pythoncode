import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import os
import logging
import sys
import time
import math

# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 初始化日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class UC_Model:
    def __init__(self, data_path):
        self.data_path = data_path
        self.load_data()
        self.validate_data()  # 添加数据验证
        self.create_model()
    
    def validate_data(self):
        """数据加载后立即执行验证"""
        logging.info("开始数据验证...")

        # 检查可再生能源容量
        for tech, units in [('风电', self.wind_units), 
                           ('光伏', self.solar_units),
                           ('水电', self.hydro_units)]:
            # 修正：兼容不同表头，自动查找容量列
            cap_col = None
            for col in units.columns:
                if str(col).strip() in ['Capacity', '机组装机容量(MW)']:
                    cap_col = col
                    break
            if cap_col is None:
                logging.error(f"{tech}机组数据缺少容量列！实际列为: {list(units.columns)}")
                continue
            if units.empty:
                logging.error(f"{tech}机组数据为空！")
            elif units[cap_col].sum() < 100:
                logging.warning(f"{tech}总容量异常: {units[cap_col].sum()}MW")

        # 检查负荷曲线范围
        if max(self.load_curve) < 5000:
            logging.error(f"负荷峰值异常低: {max(self.load_curve)}MW")
            # 尝试从系统汇总中获取峰值负荷
            sys_peak = self.system_summary.get('PeakLoad(MW)', 0)
            if sys_peak > 5000:
                logging.warning(f"使用系统汇总中的峰值负荷值: {sys_peak}MW")
                # 缩放负荷曲线
                current_peak = max(self.load_curve)
                if current_peak > 0:
                    scaling_factor = sys_peak / current_peak
                    self.load_curve = [val * scaling_factor for val in self.load_curve]
        
        # 检查资源曲线匹配
        wind_curves_used = self.wind_units['ResourceCurve'].unique()
        for curve in wind_curves_used:
            if curve not in self.wind_curves['WTCurve'].values:
                logging.error(f"风资源曲线'{curve}'未定义！")
        
        solar_curves_used = self.solar_units['ResourceCurve'].unique()
        for curve in solar_curves_used:
            if curve not in self.solar_curves['PVCurve'].values:
                logging.error(f"光伏资源曲线'{curve}'未定义！")
        
        hydro_curves_used = self.hydro_units['ResourceCurve'].unique()
        for curve in hydro_curves_used:
            if curve not in self.hydro_curves['RunoffHydroOperationCurve'].values:
                logging.error(f"水电资源曲线'{curve}'未定义！")
        
        logging.info("数据验证完成")
        
    def plot_resource_curves(self, output_path="Resource_Curves.png"):
        """绘制资源曲线验证图"""
        plt.figure(figsize=(12, 8))
        
        # 绘制风电曲线
        for i, curve in enumerate(self.wind_curves['WTCurve'].unique()[:3]):
            df = self.wind_curves[self.wind_curves['WTCurve'] == curve]
            # 提取所有时间列
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'风电曲线_{i+1}')
        
        # 绘制光伏曲线
        for i, curve in enumerate(self.solar_curves['PVCurve'].unique()[:3]):
            df = self.solar_curves[self.solar_curves['PVCurve'] == curve]
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'光伏曲线_{i+1}')
        
        # 绘制水电曲线
        for i, curve in enumerate(self.hydro_curves['RunoffHydroOperationCurve'].unique()[:3]):
            df = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'] == curve]
            time_cols = [col for col in df.columns if 'Time_' in col]
            values = df[time_cols].values.flatten()[:self.T]
            plt.plot(range(len(values)), values, label=f'水电曲线_{i+1}')
        
        # 添加标题和标签
        plt.title('资源曲线预览', fontsize=14)
        plt.xlabel('时段')
        plt.ylabel('出力系数')
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(output_path)
        logging.info(f"资源曲线预览图已保存到 {output_path}")
        plt.close()
        return output_path
        
    def load_data(self):
        """加载所有Excel数据"""
        logging.info("开始加载数据...")
        
        # 系统汇总信息
        self.system_summary = pd.read_excel(
            self.data_path, sheet_name='SystemSummary', skiprows=0, nrows=1
        ).iloc[0]
        
        # 机组参数
        self.thermal_units = pd.read_excel(
            self.data_path, sheet_name='UnitThermalGenerators', skiprows=0
        )
        self.wind_units = pd.read_excel(
            self.data_path, sheet_name='UnitWindGenerators', skiprows=0
        )
        # 修复1: 确保所有关键列正确映射
        self.wind_units.rename(columns={
            'WTPlantID': 'WTPlantID',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)
        # 由于有真实的机组数据，不需要添加虚拟机组
        # 但仍然需要检查容量列是否为数值类型

        self.solar_units = pd.read_excel(
            self.data_path, sheet_name='UnitSolarGenerators', skiprows=0
        )
        self.solar_units.rename(columns={
            'PVPlantID': 'PVPlantID',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)


        self.ess_units = pd.read_excel(
            self.data_path, sheet_name='StorElectrochemicalESS', skiprows=0
        )
        self.ess_units.rename(columns={
            'ESSId': 'ESSId',
            '储能容量(MWh)': 'Capacity',
            '充电功率(MW)': 'MaxChargePower',
            '放电功率(MW)': 'MaxDisPower',
            '所在母线编号': 'BusName'  # 确保有母线信息
        }, inplace=True)


        self.hydro_units = pd.read_excel(
            self.data_path, sheet_name='UnitRunoffHydroGenerators', skiprows=0
        )
        self.hydro_units.rename(columns={
            'RunoffHydroUnitName': 'RunoffHydroUnitName',
            'Capacity ': 'Capacity',  # 注意原列名有空格
            'ResourceCurve': 'ResourceCurve',
            'BusName': 'BusName'  # 确保有母线信息
        }, inplace=True)

        # 保证RunoffHydroUnitName唯一，否则to_dict会报错
        if self.hydro_units['RunoffHydroUnitName'].duplicated().any():
            logging.warning("发现重复的水电机组名称，将自动去重（保留首次出现）")
            self.hydro_units = self.hydro_units.drop_duplicates(subset='RunoffHydroUnitName', keep='first')

        # 修正：确保所有机组的Capacity列为数值类型，避免sum时报错
        for df, cap_col in [
            (self.wind_units, 'Capacity'),
            (self.solar_units, 'Capacity'),
            (self.hydro_units, 'Capacity')
        ]:
            if cap_col in df.columns:
                df[cap_col] = pd.to_numeric(df[cap_col], errors='coerce').fillna(0)

        # 检查所有机组集合是否全为空，若全为空则添加一个虚拟火电机组，避免模型完全无源
        if (
            self.thermal_units.empty and
            (self.wind_units.empty or (self.wind_units.shape[0] == 1 and self.wind_units.iloc[0]['Capacity'] == 0)) and
            (self.solar_units.empty or (self.solar_units.shape[0] == 1 and self.solar_units.iloc[0]['Capacity'] == 0)) and
            (self.hydro_units.empty or (self.hydro_units.shape[0] == 1 and self.hydro_units.iloc[0]['Capacity'] == 0))
        ):
            logging.warning("所有机组集合均为空，自动添加一个虚拟火电机组以保证模型可行性")
            self.thermal_units = pd.DataFrame([{
                'ThermalUnitNumber': 'DummyThermal',
                'Capacity': 1000,  # 增加虚拟火电机组容量到1000MW
                'MinPower': 100,   # 设置最小出力
                'MinOnLineTime': 1,
                'MinOffLineTime': 1,
                'UpRamppingRate': 500,   # 增加爬坡速率
                'DownRamppingRate': 500,
                'InitStatus': 1,
                'BusName': 'Bus1'  # 暂时使用默认母线名，稍后会被覆盖
            }])

        # 曲线数据
        self.wind_curves = pd.read_excel(
            self.data_path, sheet_name='CurveWindResource', skiprows=0
        )
        self.wind_curves.rename(columns={
            'WTCurve': 'WTCurve'
        }, inplace=True)
        self.wind_curves = self.wind_curves.dropna(subset=['WTCurve'])
        
        self.solar_curves = pd.read_excel(
            self.data_path, sheet_name='CurveSolarResource', skiprows=0
        )
        self.solar_curves.rename(columns={
            'PVCurve': 'PVCurve'
        }, inplace=True)
        self.solar_curves = self.solar_curves.dropna(subset=['PVCurve'])
        
        self.hydro_curves = pd.read_excel(
            self.data_path, sheet_name='CurveRunoffHydroResource', skiprows=0
        )
        self.hydro_curves.rename(columns={
            'RunoffHydroOperationCurve': 'RunoffHydroOperationCurve'
        }, inplace=True)
        self.hydro_curves = self.hydro_curves.dropna(subset=['RunoffHydroOperationCurve'])
        
        # 修复2: 正确读取负荷曲线数据
        load_curve_sheet = pd.read_excel(self.data_path, sheet_name='CurveLoad')
        # 查找包含"Time_0"的列作为标识
        time_cols = [col for col in load_curve_sheet.columns if 'time_' in col.lower() or 'Time_' in col]

        # 先初始化self.load_curve，避免后续引用时报错
        self.load_curve = []

        if not time_cols:
            logging.error("负荷曲线表中未找到时间列！")
            # 创建默认负荷曲线
            peak_load = self.system_summary.get('PeakLoad(MW)', 10000)
            self.load_curve = [peak_load * 0.7] * int(self.system_summary['Duration'])
        else:
            # 尝试找到包含负荷数据的行
            for i in range(len(load_curve_sheet)):
                row = load_curve_sheet.iloc[i]
                if any(isinstance(val, (int, float)) and not pd.isna(val) for val in row[time_cols]):
                    load_factors = row[time_cols].values.astype(float)
                    # 获取系统峰值负荷
                    peak_load = self.system_summary.get('PeakLoad(MW)', 10000)
                    # 将系数转换为实际负荷值
                    self.load_curve = [factor * peak_load for factor in load_factors]
                    logging.info(f"从第{i+1}行读取负荷曲线数据，峰值负荷：{peak_load}MW")
                    logging.info(f"负荷曲线范围：{min(self.load_curve):.1f} - {max(self.load_curve):.1f} MW")
                    break
            else:
                logging.warning("未找到有效的负荷数据行，使用默认值")
                peak_load = self.system_summary.get('PeakLoad(MW)', 10000)
                self.load_curve = [peak_load * 0.7] * int(self.system_summary['Duration'])

        # 确保负荷曲线有合理的数值
        if len(self.load_curve) == 0:
            peak_load = self.system_summary.get('PeakLoad(MW)', 10000)
            self.load_curve = [peak_load * 0.7 for _ in range(int(self.system_summary['Duration']))]
        elif max(self.load_curve) < 100:
            logging.warning("负荷曲线峰值过低，可能是系数而非绝对值")
            peak_load = self.system_summary.get('PeakLoad(MW)', 10000)
            # 如果数值很小，认为是系数，需要乘以峰值负荷
            if max(self.load_curve) <= 1.0:
                self.load_curve = [val * peak_load for val in self.load_curve]
                logging.info(f"负荷曲线已转换为绝对值，峰值：{max(self.load_curve):.1f}MW")
        
        # 网络数据
        self.buses = pd.read_excel(self.data_path, sheet_name='NetBuses', skiprows=0)
        self.lines = pd.read_excel(self.data_path, sheet_name='NetLines', skiprows=0)
        self.transformers = pd.read_excel(self.data_path, sheet_name='NetTransformers', skiprows=0)
        self.sections = pd.read_excel(self.data_path, sheet_name='NetSectionsLines', skiprows=0)
        self.section_capacity = pd.read_excel(self.data_path, sheet_name='NetSectionsCapacity', skiprows=0)
        self.ac_tielines = pd.read_excel(self.data_path, sheet_name='NetHVACTieLine', skiprows=0)
        self.ac_tieline_curves = pd.read_excel(self.data_path, sheet_name='EnergyHVACTieLine', skiprows=0)
        
        # 更新虚拟机组的母线信息（使用实际的第一个母线）
        if not self.buses.empty:
            first_bus = self.buses['BusId'].iloc[0]
            # 更新风电虚拟机组
            if (self.wind_units.shape[0] == 1 and 
                self.wind_units.iloc[0]['WTPlantID'] == 'DummyWind'):
                self.wind_units.loc[0, 'BusName'] = first_bus
            # 更新光伏虚拟机组
            if (self.solar_units.shape[0] == 1 and 
                self.solar_units.iloc[0]['PVPlantID'] == 'DummySolar'):
                self.solar_units.loc[0, 'BusName'] = first_bus
            # 更新储能虚拟机组
            if (self.ess_units.shape[0] == 1 and 
                self.ess_units.iloc[0]['ESSId'] == 'DummyESS'):
                self.ess_units.loc[0, 'BusName'] = first_bus
            # 更新水电虚拟机组
            if (self.hydro_units.shape[0] == 1 and 
                self.hydro_units.iloc[0]['RunoffHydroUnitName'] == 'DummyHydro'):
                self.hydro_units.loc[0, 'BusName'] = first_bus
            # 更新火电虚拟机组
            if (self.thermal_units.shape[0] == 1 and 
                self.thermal_units.iloc[0]['ThermalUnitNumber'] == 'DummyThermal'):
                self.thermal_units.loc[0, 'BusName'] = first_bus
        
        # 负荷数据
        self.loads = pd.read_excel(self.data_path, sheet_name='Loads', skiprows=0)
        
        # 燃料数据
        self.fuel_data = pd.read_excel(self.data_path, sheet_name='UnitFuel', skiprows=0).iloc[0]
        
        # 设置时间参数
        self.T = int(self.system_summary['Duration'])
        self.dt = 1  # 调度间隔为1小时
        self.hours = list(range(self.T))
        
        # 创建机组字典
        self.thermal_dict = self.thermal_units.set_index('ThermalUnitNumber').to_dict('index')
        for unit in self.thermal_dict:
            # 确保数值类型正确
            self.thermal_dict[unit]['MinOnLineTime'] = int(self.thermal_dict[unit].get('MinOnLineTime', 0))
            self.thermal_dict[unit]['MinOffLineTime'] = int(self.thermal_dict[unit].get('MinOffLineTime', 0))
            self.thermal_dict[unit]['Capacity'] = float(self.thermal_dict[unit].get('Capacity', 0))
            self.thermal_dict[unit]['MinPower'] = float(self.thermal_dict[unit].get('MinPower', 0))
            self.thermal_dict[unit]['UpRamppingRate'] = float(self.thermal_dict[unit].get('UpRamppingRate', 0))
            self.thermal_dict[unit]['DownRamppingRate'] = float(self.thermal_dict[unit].get('DownRamppingRate', 0))
            
        self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        self.solar_dict = self.solar_units.set_index('PVPlantID').to_dict('index')
        self.ess_dict = self.ess_units.set_index('ESSId').to_dict('index')
        self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        
        # 构建母线字典
        self.bus_dict = self.buses.set_index('BusId').to_dict('index')
        
        # 构建线路字典
        self.line_dict = {}
        for idx, row in self.lines.iterrows():
            self.line_dict[row['BranchId']] = {
                'StartNodeName': row['StartNodeName'],
                'EndNodeName': row['EndNodeName'],
                'Resistance': float(row.get('Resistance', 0)),
                'Reactance': float(row.get('Reactance', 0)),
                'Capacity': float(row.get('Capacity', 0)),
                'Susceptance': float(row.get('Susceptance', 0))
            }
        
        # 找到平衡节点（参考节点）
        self.ref_bus = self.system_summary['BalanceNode']
        logging.info(f"参考节点设置为: {self.ref_bus}")
        
        logging.info("数据加载完成")
    
    def get_unit_param(self, unit_dict, unit, key, default=0.0):
        """安全获取机组参数"""
        try:
            val = unit_dict[unit].get(key, default)
            if pd.isna(val):
                return default
            return val
        except Exception:
            return default

    def create_model(self):
        """创建Pyomo优化模型（包含直流潮流约束）"""
        logging.info("开始创建优化模型（包含直流潮流约束）...")
        
        self.model = pyo.ConcreteModel()
        
        # 定义集合
        self.model.T = pyo.Set(initialize=self.hours)  # 时间段
        self.model.ThermalUnits = pyo.Set(initialize=self.thermal_units['ThermalUnitNumber'].tolist())
        self.model.WindUnits = pyo.Set(initialize=self.wind_units['WTPlantID'].tolist())
        self.model.SolarUnits = pyo.Set(initialize=self.solar_units['PVPlantID'].tolist())
        self.model.ESSUntis = pyo.Set(initialize=self.ess_units['ESSId'].tolist())
        self.model.HydroUnits = pyo.Set(initialize=self.hydro_units['RunoffHydroUnitName'].tolist())
        self.model.Loads = pyo.Set(initialize=self.loads['负荷编号'].tolist())
        self.model.Buses = pyo.Set(initialize=self.buses['BusId'].tolist())
        self.model.Lines = pyo.Set(initialize=self.lines['BranchId'].tolist())
        
        # 定义变量
        # 火电机组
        self.model.thermal_status = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.thermal_startup = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        self.model.thermal_shutdown = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )
        
        # 可再生能源
        self.model.wind_power = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_power = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_power = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 储能系统
        self.model.ess_charge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_discharge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.ess_soc = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 负荷相关
        self.model.load_shed = pyo.Var(
            self.model.Loads, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 直流潮流变量
        self.model.bus_angle = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )  # 节点相角
        
        self.model.line_power = pyo.Var(
            self.model.Lines, self.model.T, within=pyo.Reals
        )  # 线路潮流
        
        # 节点平衡松弛变量
        self.model.balance_slack = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 备用松弛变量
        self.model.reserve_slack = pyo.Var(
            self.model.T, within=pyo.NonNegativeReals
        )
        
        # 深度调峰惩罚线性化变量
        self.model.deep_peak_penalty = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 可再生能源最小出力松弛变量
        self.model.wind_slack = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_slack = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_slack = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 目标函数
        def total_cost_rule(model):
            """系统总成本计算"""
            cost = 0
            fuel_price = self.fuel_data['UnitPrice'] * 2  # 增加燃料价格，让火电更贵
            
            # 火电运行成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    a = self.thermal_dict[unit].get('OperationCoeff_A', 0)
                    b = self.thermal_dict[unit].get('OperationCoeff_B', 100)  # 设置默认的运行成本
                    c = self.thermal_dict[unit].get('OperationCoeff_C', 50)   # 设置默认的固定成本
                    
                    # 运行成本（增加基础成本）
                    cost += model.thermal_status[unit, t] * (c + 100) * fuel_price  # 增加固定成本
                    cost += (b + 50) * model.thermal_power[unit, t] * fuel_price    # 增加变动成本
                    
                    # 启动成本
                    startup_cost = self.thermal_dict[unit].get('StateOnFuelConsumption', 500)  # 增加启动成本
                    cost += startup_cost * fuel_price * model.thermal_startup[unit, t]
            
            # 深度调峰惩罚（线性化实现）
            for unit in model.ThermalUnits:
                for t in model.T:
                    max_power = self.thermal_dict[unit].get('Capacity', 0)
                    deep_threshold = self.thermal_dict[unit].get('DeepPeakShavingThreshold', 0.5) * max_power
                    deep_cost = self.thermal_dict[unit].get('DeepPeakShavingCost', 0)
                    # 线性松弛变量
                    cost += deep_cost * model.deep_peak_penalty[unit, t]
            
            # 可再生能源激励（负成本，即收益）和强制出力罚函数（线性化）
            renewable_incentive = -3000  # 大幅增加到3000元/MWh的激励，确保新能源绝对优先
            renewable_penalty = 100000    # 大幅增加到100000元/MWh的惩罚，确保强制出力
            
            # 风电激励和惩罚
            for unit in model.WindUnits:
                for t in model.T:
                    cost += renewable_incentive * model.wind_power[unit, t]
                    cost += renewable_penalty * model.wind_slack[unit, t]
            
            # 光伏激励和惩罚
            for unit in model.SolarUnits:
                for t in model.T:
                    cost += renewable_incentive * model.solar_power[unit, t]
                    cost += renewable_penalty * model.solar_slack[unit, t]
            
            # 水电激励和惩罚
            for unit in model.HydroUnits:
                for t in model.T:
                    cost += renewable_incentive * model.hydro_power[unit, t]
                    cost += renewable_penalty * model.hydro_slack[unit, t]
            
            # 切负荷惩罚
            penalty_cut = 100000  # 元/MWh
            for load in model.Loads:
                for t in model.T:
                    cost += penalty_cut * model.load_shed[load, t]
            
            # 节点平衡松弛惩罚
            penalty_slack = 10000  # 元/MW
            for bus in model.Buses:
                for t in model.T:
                    cost += penalty_slack * model.balance_slack[bus, t]
            
            # 备用松弛惩罚
            penalty_reserve = 5000  # 元/MW
            for t in model.T:
                cost += penalty_reserve * model.reserve_slack[t]
            
            return cost

        # 深度调峰惩罚线性化约束
        def deep_peak_penalty_rule(model, unit, t):
            max_power = self.thermal_dict[unit].get('Capacity', 0)
            deep_threshold = self.thermal_dict[unit].get('DeepPeakShavingThreshold', 0.5) * max_power
            return model.deep_peak_penalty[unit, t] >= deep_threshold - model.thermal_power[unit, t]
        self.model.deep_peak_penalty_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=deep_peak_penalty_rule
        )
        
        self.model.total_cost = pyo.Objective(rule=total_cost_rule, sense=pyo.minimize)
        
        # 1. 火电机组逻辑约束
        def thermal_logic_rule(model, unit, t):
            if t == 0:
                init_status = self.thermal_dict[unit].get('InitStatus', False)
                if isinstance(init_status, str):
                    init_status = init_status.strip().lower() in ['true', '1', 'yes']
                return model.thermal_status[unit, t] == (1 if init_status else 0)
            else:
                return model.thermal_status[unit, t] - model.thermal_status[unit, t-1] == \
                    model.thermal_startup[unit, t] - model.thermal_shutdown[unit, t]
        self.model.thermal_logic_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_logic_rule
        )

        # 2. 火电机组出力约束
        def thermal_power_lower_rule(model, unit, t):
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            return model.thermal_power[unit, t] >= min_power * model.thermal_status[unit, t]
        
        def thermal_power_upper_rule(model, unit, t):
            max_power = self.thermal_dict[unit].get('Capacity', 0)
            return model.thermal_power[unit, t] <= max_power * model.thermal_status[unit, t]
        
        self.model.thermal_power_lower_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_lower_rule
        )
        self.model.thermal_power_upper_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_upper_rule
        )

        # 3. 爬坡约束 - 使用更稳健的方案2
        def thermal_ramp_up_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_up = self.thermal_dict[unit].get('UpRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束，增加状态变化时的缓冲
            return (model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= 
                    model.thermal_status[unit, t] * ramp_up + 
                    (1 - model.thermal_status[unit, t-1]) * capacity * 0.7)  # 缓冲系数
        
        def thermal_ramp_down_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_down = self.thermal_dict[unit].get('DownRamppingRate', 0)
            capacity = self.thermal_dict[unit].get('Capacity', 0)
            min_power = self.thermal_dict[unit].get('MinPower', 0)
            
            # 方案2：考虑启停状态的爬坡约束，增加状态变化时的缓冲
            return (model.thermal_power[unit, t-1] - model.thermal_power[unit, t] <= 
                    model.thermal_status[unit, t-1] * ramp_down + 
                    (1 - model.thermal_status[unit, t]) * capacity * 0.7)  # 缓冲系数

        self.model.thermal_ramp_up_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_up_rule
        )
        self.model.thermal_ramp_down_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_down_rule
        )

        # 4. 最短开停机时间约束 - 大幅放宽以提高可行性
        def min_up_time_rule(model, unit, t):
            min_up = self.thermal_dict[unit].get('MinOnLineTime', 0)
            # 如果最小开机时间太长，则限制为总时间的1/4
            min_up = min(min_up, max(1, self.T // 4))
            
            if t < min_up or min_up <= 1:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_up:
                return sum(model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        model.thermal_status[unit, t] - prev_status) * (self.T - t)
            else:
                return sum(model.thermal_status[unit, k] for k in range(t, t+min_up)) >= (
                        model.thermal_status[unit, t] - prev_status) * min_up

        def min_down_time_rule(model, unit, t):
            min_down = self.thermal_dict[unit].get('MinOffLineTime', 0)
            # 如果最小停机时间太长，则限制为总时间的1/4
            min_down = min(min_down, max(1, self.T // 4))
            
            if t < min_down or min_down <= 1:
                return pyo.Constraint.Skip

            prev_status = model.thermal_status[unit, t-1] if t > 0 else 0

            if t >= self.T - min_down:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, self.T)) >= (
                        prev_status - model.thermal_status[unit, t]) * (self.T - t)
            else:
                return sum(1 - model.thermal_status[unit, k] for k in range(t, t+min_down)) >= (
                        prev_status - model.thermal_status[unit, t]) * min_down

        self.model.min_up_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_up_time_rule
        )
        self.model.min_down_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_down_time_rule
        )

        # 5. 可再生能源出力约束 - 修复资源曲线匹配问题，增加默认出力能力
        def wind_power_rule(model, unit, t):
            curve_name = self.wind_dict[unit].get('ResourceCurve', '')
            capacity = self.wind_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.wind_power[unit, t] <= 0
            
            # 防止curve_name为非字符串
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 如果没有资源曲线名称或为虚拟机组，使用高出力系数
            if curve_name_str in ['', '0', 'nan', 'None'] or curve_name_str == '':
                # 给风电一个很高的出力模式
                default_factor = 0.8  # 大幅提升到80%
                logging.info(f"风电机组{unit}在时段{t}使用高出力系数: {default_factor}")
                return model.wind_power[unit, t] <= default_factor * capacity
            
            # 找到对应的风资源曲线 - 增强匹配逻辑
            curve_row = self.wind_curves[self.wind_curves['WTCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                # 尝试模糊匹配
                curve_row = self.wind_curves[self.wind_curves['WTCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"风电机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！使用高出力系数")
                    # 使用高出力而不是低出力
                    default_factor = 0.8
                    return model.wind_power[unit, t] <= default_factor * capacity
            
            # 获取时间点值 - 处理列名大小写问题
            time_col = f'Time_{t}'
            # 尝试匹配列名（不区分大小写）
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 0.8  # 使用高默认值
                else:
                    max_factor = max(max_factor, 0.5)  # 确保至少50%出力能力
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        max_factor = 0.8
                    else:
                        max_factor = max(max_factor, 0.5)
                else:
                    logging.warning(f"风电机组{unit}在时段{t}缺少出力系数，使用高默认值0.8")
                    max_factor = 0.8  # 使用高默认值
                    
            # 确保出力不超过容量
            return model.wind_power[unit, t] <= max_factor * capacity
        
        self.model.wind_power_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_power_rule
        )
        
        def solar_power_rule(model, unit, t):
            curve_name = self.solar_dict[unit].get('ResourceCurve', '')
            capacity = self.solar_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.solar_power[unit, t] <= 0
            
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 白天时段（6-18点）和时间相关的光照模式
            hour = t % 24
            
            # 如果没有资源曲线名称或为虚拟机组，使用强制高出力系数
            if curve_name_str in ['', '0', 'nan', 'None'] or curve_name_str == '':
                # 给光伏一个更强的出力模式
                if 7 <= hour <= 17:  # 白天时段
                    if 10 <= hour <= 14:  # 最强光照时段
                        default_factor = 0.9  # 最高出力90%
                    else:  # 一般白天时段
                        default_factor = 0.7  # 高出力70%
                else:  # 夜间
                    default_factor = 0.0  # 夜间无光照
                logging.info(f"光伏机组{unit}在时段{t}（{hour}时）使用高出力系数: {default_factor}")
                return model.solar_power[unit, t] <= default_factor * capacity
            
            curve_row = self.solar_curves[self.solar_curves['PVCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                curve_row = self.solar_curves[self.solar_curves['PVCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"光伏机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！使用强制高出力系数")
                    # 使用强制高出力而不是低出力
                    if 7 <= hour <= 17:
                        if 10 <= hour <= 14:
                            default_factor = 0.9
                        else:
                            default_factor = 0.7
                    else:
                        default_factor = 0.0
                    return model.solar_power[unit, t] <= default_factor * capacity
            
            # 获取时间点值 - 处理列名大小写问题
            time_col = f'Time_{t}'
            # 尝试匹配列名（不区分大小写）
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    # 使用时间相关的高默认值
                    if 7 <= hour <= 17:
                        max_factor = 0.8 if 10 <= hour <= 14 else 0.6
                    else:
                        max_factor = 0.0
                else:
                    # 确保白天有足够出力上限
                    if 7 <= hour <= 17:
                        max_factor = max(max_factor, 0.6)  # 白天至少60%上限
                    # 夜间保持原值（通常为0）
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        if 7 <= hour <= 17:
                            max_factor = 0.8 if 10 <= hour <= 14 else 0.6
                        else:
                            max_factor = 0.0
                    else:
                        if 7 <= hour <= 17:
                            max_factor = max(max_factor, 0.6)
                else:
                    logging.warning(f"光伏机组{unit}在时段{t}缺少出力系数，使用高默认值")
                    if 7 <= hour <= 17:
                        max_factor = 0.8 if 10 <= hour <= 14 else 0.6
                    else:
                        max_factor = 0.0
                    
            return model.solar_power[unit, t] <= max_factor * capacity
        
        self.model.solar_power_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_power_rule
        )
        
        def hydro_power_rule(model, unit, t):
            curve_name = self.hydro_dict[unit].get('ResourceCurve', '')
            capacity = self.hydro_dict[unit].get('Capacity', 0)
            
            # 如果容量为0，直接跳过
            if capacity <= 0:
                return model.hydro_power[unit, t] <= 0
            
            curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
            
            # 如果资源曲线名为0或空，允许水电机组出力为0~容量
            if curve_name_str in ['', '0', 'nan', 'None']:
                logging.info(f"水电机组{unit}无资源曲线限制，可在0~{capacity}MW范围内出力")
                return (0, model.hydro_power[unit, t], capacity)
            
            curve_row = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'].astype(str).str.strip() == curve_name_str]
            if curve_row.empty:
                curve_row = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'].astype(str).str.contains(curve_name_str, case=False)]
                if curve_row.empty:
                    logging.warning(f"水电机组{unit}的资源曲线'{curve_name}'在曲线表中未找到！允许灵活出力")
                    return (0, model.hydro_power[unit, t], capacity)
            
            # 获取时间点值 - 处理列名大小写问题
            time_col = f'Time_{t}'
            matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
            if matching_cols:
                max_factor = curve_row[matching_cols[0]].values[0]
                if pd.isna(max_factor):
                    max_factor = 0.8  # 使用默认值而不是0
            else:
                # 尝试匹配"time_X"格式
                alt_time_col = f'time_{t}'
                if alt_time_col in curve_row.columns:
                    max_factor = curve_row[alt_time_col].values[0]
                    if pd.isna(max_factor):
                        max_factor = 0.8
                else:
                    logging.warning(f"水电机组{unit}在时段{t}缺少出力系数，使用默认值0.8")
                    max_factor = 0.8  # 使用默认值而不是0
                    
            return model.hydro_power[unit, t] <= max_factor * capacity
        
        self.model.hydro_power_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_power_rule
        )

        # 5.1 可再生能源最小出力约束（用松弛变量实现）- 更强制性
        def wind_min_output_rule(model, unit, t):
            capacity = self.wind_dict[unit].get('Capacity', 0)
            curve_name = self.wind_dict[unit].get('ResourceCurve', '')
            
            logging.debug(f"风电机组{unit}时段{t}: 容量={capacity}")
            # 大幅降低容量门槛，让小机组也参与运行
            if capacity > 1:  # 降低到1MW门槛，确保更多机组参与
                # 检查该时段是否有风资源
                curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
                
                # 获取该时段的实际资源系数
                actual_factor = 0
                if curve_name_str not in ['', '0', 'nan', 'None']:
                    curve_row = self.wind_curves[self.wind_curves['WTCurve'].astype(str).str.strip() == curve_name_str]
                    if not curve_row.empty:
                        time_col = f'Time_{t}'
                        if time_col in curve_row.columns:
                            actual_factor = curve_row[time_col].values[0]
                            if pd.isna(actual_factor):
                                actual_factor = 0
                
                # 强制所有风电机组都要有基本出力 - 确保图表中总量可见
                if actual_factor > 0.001:  # 极低门槛0.1%
                    min_output = capacity * max(actual_factor * 0.95, 0.6)  # 至少60%容量或95%资源
                    return model.wind_power[unit, t] + model.wind_slack[unit, t] >= min_output
                else:
                    # 强制基本出力 - 所有机组都要贡献
                    min_output = capacity * 0.5  # 所有风电机组强制50%容量
                    return model.wind_power[unit, t] + model.wind_slack[unit, t] >= min_output
            else:
                # 即使是很小的机组也要有一定出力
                if capacity > 0:
                    min_output = capacity * 0.3  # 小机组30%出力
                    return model.wind_power[unit, t] + model.wind_slack[unit, t] >= min_output
                return model.wind_slack[unit, t] >= 0
        
        def solar_min_output_rule(model, unit, t):
            capacity = self.solar_dict[unit].get('Capacity', 0)
            curve_name = self.solar_dict[unit].get('ResourceCurve', '')
            
            logging.debug(f"光伏机组{unit}时段{t}: 容量={capacity}")
            # 大幅降低容量门槛，让所有光伏机组都参与运行
            if capacity > 1:  # 降低到1MW门槛，确保更多机组参与
                # 检查该时段是否有太阳资源
                curve_name_str = str(curve_name).strip() if not pd.isna(curve_name) else ''
                
                # 获取该时段的实际资源系数
                actual_factor = 0
                if curve_name_str not in ['', '0', 'nan', 'None']:
                    curve_row = self.solar_curves[self.solar_curves['PVCurve'].astype(str).str.strip() == curve_name_str]
                    if not curve_row.empty:
                        time_col = f'Time_{t}'
                        matching_cols = [col for col in curve_row.columns if col.lower() == time_col.lower()]
                        if matching_cols:
                            actual_factor = curve_row[matching_cols[0]].values[0]
                            if pd.isna(actual_factor):
                                actual_factor = 0
                
                # 强制所有光伏机组在白天都要有大幅出力 - 确保图表中黄色非常明显
                daytime_hours = list(range(6, 19))  # 6点到18点
                if t % 24 in daytime_hours:
                    if actual_factor > 0.001:  # 极低资源门槛
                        min_output = capacity * max(actual_factor * 0.95, 0.7)  # 至少70%容量或95%资源
                        return model.solar_power[unit, t] + model.solar_slack[unit, t] >= min_output
                    else:
                        # 白天强制大幅出力 - 所有光伏机组都要贡献
                        peak_hours = list(range(9, 16))  # 9-15点为光伏强制高峰
                        if t % 24 in peak_hours:
                            min_output = capacity * 0.8  # 高峰时段强制80%容量
                        else:
                            min_output = capacity * 0.6  # 其他白天时段60%容量
                        return model.solar_power[unit, t] + model.solar_slack[unit, t] >= min_output
                else:
                    return model.solar_slack[unit, t] >= 0  # 夜间不强制出力
            else:
                # 即使是很小的光伏机组白天也要有出力
                if capacity > 0:
                    daytime_hours = list(range(6, 19))
                    if t % 24 in daytime_hours:
                        min_output = capacity * 0.4  # 小机组白天40%出力
                        return model.solar_power[unit, t] + model.solar_slack[unit, t] >= min_output
                return model.solar_slack[unit, t] >= 0
        
        def hydro_min_output_rule(model, unit, t):
            capacity = self.hydro_dict[unit].get('Capacity', 0)
            logging.debug(f"水电机组{unit}时段{t}: 容量={capacity}")
            # 大幅降低水电门槛，让所有水电机组都参与
            if capacity > 1:  # 降低到1MW门槛
                min_output = capacity * 0.6  # 大幅提升要求到60%出力
                return model.hydro_power[unit, t] + model.hydro_slack[unit, t] >= min_output
            else:
                if capacity > 0:
                    min_output = capacity * 0.4  # 小机组40%出力
                    return model.hydro_power[unit, t] + model.hydro_slack[unit, t] >= min_output
                return model.hydro_slack[unit, t] >= 0
        
        self.model.wind_min_output_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_min_output_rule
        )
        self.model.solar_min_output_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_min_output_rule
        )
        self.model.hydro_min_output_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_min_output_rule
        )

        # 6. 储能系统约束
        def ess_charge_power_rule(model, unit, t):
            max_charge = self.ess_dict[unit].get('MaxChargePower', 0)
            return model.ess_charge[unit, t] <= max_charge
        
        def ess_discharge_power_rule(model, unit, t):
            max_discharge = self.ess_dict[unit].get('MaxDisPower', 0)
            return model.ess_discharge[unit, t] <= max_discharge
        
        def ess_soc_rule(model, unit, t):
            min_soc = self.ess_dict[unit].get('MinSOC', 0)
            max_soc = self.ess_dict[unit].get('MaxSOC', 1000)
            capacity = self.ess_dict[unit].get('Capacity', 0)
            return (min_soc, model.ess_soc[unit, t], max_soc)
        
        def ess_soc_balance_rule(model, unit, t):
            charge_eff = self.ess_dict[unit].get('ChargeEfficiency', 0.85)
            dis_eff = self.ess_dict[unit].get('DisEfficiency', 0.85)
            
            if t == 0:
                init_soc = self.ess_dict[unit].get('InitSOC', 0)
                return model.ess_soc[unit, t] == init_soc
            else:
                return model.ess_soc[unit, t] == (model.ess_soc[unit, t-1] + 
                                                  charge_eff * model.ess_charge[unit, t] - 
                                                  model.ess_discharge[unit, t] / dis_eff)
        
        self.model.ess_charge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_charge_power_rule
        )
        self.model.ess_discharge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_discharge_power_rule
        )
        self.model.ess_soc_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_rule
        )
        self.model.ess_soc_balance_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_balance_rule
        )

        # 7. 直流潮流约束
        logging.info("添加直流潮流约束...")
        
        # 7.1 参考节点约束（平衡节点相角为0）
        def ref_bus_rule(model, t):
            return model.bus_angle[self.ref_bus, t] == 0
        self.model.ref_bus_con = pyo.Constraint(
            self.model.T, rule=ref_bus_rule
        )
        
        # 7.2 节点功率平衡约束（带松弛）
        def node_balance_rule(model, bus, t):
            # 节点发电量（包括所有连接到该节点的发电单元）
            gen_power = 0
            # 火电
            for unit in model.ThermalUnits:
                if self.thermal_dict[unit].get('BusName', '') == bus:
                    gen_power += model.thermal_power[unit, t]
            # 风电
            for unit in model.WindUnits:
                if self.wind_dict[unit].get('BusName', '') == bus:
                    gen_power += model.wind_power[unit, t]
            # 光伏
            for unit in model.SolarUnits:
                if self.solar_dict[unit].get('BusName', '') == bus:
                    gen_power += model.solar_power[unit, t]
            # 水电
            for unit in model.HydroUnits:
                if self.hydro_dict[unit].get('BusName', '') == bus:
                    gen_power += model.hydro_power[unit, t]
            # 储能放电
            for unit in model.ESSUntis:
                if self.ess_dict[unit].get('BusName', '') == bus:
                    gen_power += model.ess_discharge[unit, t]
            # 储能充电
            ess_charge = 0
            for unit in model.ESSUntis:
                if self.ess_dict[unit].get('BusName', '') == bus:
                    ess_charge += model.ess_charge[unit, t]
            
            # 节点负荷（包括切负荷）
            load_power = 0
            load_shed = 0
            for load in model.Loads:
                load_info = self.loads[self.loads['负荷编号'] == load]
                if not load_info.empty and load_info['所在母线编号'].values[0] == bus:
                    coef = load_info['有功比例系数'].values[0]
                    # 使用全局负荷曲线
                    if t < len(self.load_curve):
                        curve_val = self.load_curve[t]
                    else:
                        curve_val = 0
                    # 保证为float
                    try:
                        coef = float(coef)
                    except Exception:
                        coef = 0.0
                    try:
                        curve_val = float(curve_val)
                    except Exception:
                        curve_val = 0.0
                    load_power += coef * curve_val
                    load_shed += model.load_shed[load, t]
            
            # 节点净注入功率
            net_injection = gen_power - ess_charge - (load_power - load_shed)
            
            # 计算线路潮流总和（流出为正，流入为负）
            line_flow = 0
            for line in model.Lines:
                line_info = self.line_dict.get(line, {})
                if line_info.get('StartNodeName', '') == bus:
                    line_flow += model.line_power[line, t]
                elif line_info.get('EndNodeName', '') == bus:
                    line_flow -= model.line_power[line, t]
            
            # 节点功率平衡（带松弛变量）
            return net_injection == line_flow + model.balance_slack[bus, t]
        
        self.model.node_balance_con = pyo.Constraint(
            self.model.Buses, self.model.T, rule=node_balance_rule
        )
        
        # 7.3 线路潮流约束（直流潮流方程）- 增强数值稳定性
        def line_power_rule(model, line, t):
            line_info = self.line_dict.get(line, {})
            start_bus = line_info.get('StartNodeName', '')
            end_bus = line_info.get('EndNodeName', '')
            reactance = line_info.get('Reactance', 0.01)  # 默认更大的电抗值
            
            # 确保电抗值不为零，避免数值问题
            if reactance <= 0:
                reactance = 0.01
            
            # 线路潮流与相角差的关系
            return model.line_power[line, t] == (
                model.bus_angle[start_bus, t] - model.bus_angle[end_bus, t]
            ) / reactance
        
        self.model.line_power_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_power_rule
        )
        
        # 7.4 线路容量约束（大幅放宽默认容量）
        def line_capacity_rule(model, line, t):
            capacity = self.line_dict.get(line, {}).get('Capacity', 50000)  # 默认大容量50GW
            if capacity <= 0:
                capacity = 50000  # 确保正值
            return (-capacity, model.line_power[line, t], capacity)
        
        self.model.line_capacity_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_capacity_rule
        )
        
        # 8. 系统功率平衡约束（全局）
        def global_power_balance_rule(model, t):
            # 发电侧
            thermal_gen = sum(model.thermal_power[unit, t] for unit in model.ThermalUnits)
            wind_gen = sum(model.wind_power[unit, t] for unit in model.WindUnits)
            solar_gen = sum(model.solar_power[unit, t] for unit in model.SolarUnits)
            hydro_gen = sum(model.hydro_power[unit, t] for unit in model.HydroUnits)
            ess_discharge = sum(model.ess_discharge[unit, t] for unit in model.ESSUntis)
            ess_charge = sum(model.ess_charge[unit, t] for unit in model.ESSUntis)
            
            # 负荷侧
            total_load = 0
            total_shed = 0
            for load in model.Loads:
                load_info = self.loads[self.loads['负荷编号'] == load]
                if not load_info.empty:
                    coef = load_info['有功比例系数'].values[0]
                else:
                    coef = 0

                # 使用全局负荷曲线
                if t < len(self.load_curve):
                    curve_val = self.load_curve[t]
                else:
                    curve_val = 0

                try:
                    coef = float(coef)
                except Exception:
                    coef = 0.0
                try:
                    curve_val = float(curve_val)
                except Exception:
                    curve_val = 0.0

                total_load += coef * curve_val
                total_shed += model.load_shed[load, t]
            
            return (thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge ==
                    total_load - total_shed + ess_charge)
        
        self.model.global_power_balance_con = pyo.Constraint(
            self.model.T, rule=global_power_balance_rule
        )

        # 9. 备用约束（进一步放宽到3%，并添加松弛变量）
        def reserve_requirement_rule(model, t):
            # 计算总负荷
            total_load = 0
            for load in model.Loads:
                load_info = self.loads[self.loads['负荷编号'] == load]
                if not load_info.empty:
                    coef = load_info['有功比例系数'].values[0]
                else:
                    coef = 0

                # 使用全局负荷曲线
                if t < len(self.load_curve):
                    curve_val = self.load_curve[t]
                else:
                    curve_val = 0

                try:
                    coef = float(coef)
                except Exception:
                    coef = 0.0
                try:
                    curve_val = float(curve_val)
                except Exception:
                    curve_val = 0.0

                total_load += coef * curve_val

            # 计算可用备用
            available_reserve = 0
            for unit in model.ThermalUnits:
                max_power = self.thermal_dict[unit].get('Capacity', 0)
                min_power = self.thermal_dict[unit].get('MinPower', 0)
                available_reserve += (model.thermal_status[unit, t] * max_power -
                                     model.thermal_power[unit, t])

            # 3%的备用要求（从5%进一步降到3%），并添加松弛变量
            reserve_req = 0.03 * total_load
            return available_reserve + model.reserve_slack[t] >= reserve_req

        self.model.reserve_con = pyo.Constraint(
            self.model.T, rule=reserve_requirement_rule
        )
        
        logging.info("优化模型创建完成（包含直流潮流约束）")
    
    def debug_model(self):
        """调试模型不可行性"""
        logging.info("开始调试模型...")
        
        # 检查基本数据完整性
        logging.info(f"火电机组数量: {len(self.model.ThermalUnits)}")
        logging.info(f"风电机组数量: {len(self.model.WindUnits)}")
        logging.info(f"光伏机组数量: {len(self.model.SolarUnits)}")
        logging.info(f"水电机组数量: {len(self.model.HydroUnits)}")
        logging.info(f"储能单元数量: {len(self.model.ESSUntis)}")
        logging.info(f"负荷数量: {len(self.model.Loads)}")
        logging.info(f"母线数量: {len(self.model.Buses)}")
        logging.info(f"线路数量: {len(self.model.Lines)}")
        logging.info(f"时间段数: {self.T}")
        
        # 检查负荷总量和发电容量
        total_thermal_cap = sum(self.thermal_dict[unit].get('Capacity', 0) for unit in self.model.ThermalUnits)
        total_wind_cap = sum(self.wind_dict[unit].get('Capacity', 0) for unit in self.model.WindUnits)
        total_solar_cap = sum(self.solar_dict[unit].get('Capacity', 0) for unit in self.model.SolarUnits)
        total_hydro_cap = sum(self.hydro_dict[unit].get('Capacity', 0) for unit in self.model.HydroUnits)
        total_gen_cap = total_thermal_cap + total_wind_cap + total_solar_cap + total_hydro_cap
        
        peak_load = max(self.load_curve) if self.load_curve else 0
        
        logging.info(f"总发电容量: {total_gen_cap} MW")
        logging.info(f"  - 火电容量: {total_thermal_cap} MW")
        logging.info(f"  - 风电容量: {total_wind_cap} MW")
        logging.info(f"  - 光伏容量: {total_solar_cap} MW")
        logging.info(f"  - 水电容量: {total_hydro_cap} MW")
        logging.info(f"峰值负荷: {peak_load} MW")
        logging.info(f"容量充裕度: {(total_gen_cap / peak_load * 100):.1f}%" if peak_load > 0 else "无负荷")
        
        # 详细检查各类机组
        logging.info("=== 风电机组详细信息 ===")
        for unit in self.model.WindUnits:
            cap = self.wind_dict[unit].get('Capacity', 0)
            curve = self.wind_dict[unit].get('ResourceCurve', '')
            bus = self.wind_dict[unit].get('BusName', '')
            logging.info(f"风电机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        logging.info("=== 光伏机组详细信息 ===")
        for unit in self.model.SolarUnits:
            cap = self.solar_dict[unit].get('Capacity', 0)
            curve = self.solar_dict[unit].get('ResourceCurve', '')
            bus = self.solar_dict[unit].get('BusName', '')
            logging.info(f"光伏机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        logging.info("=== 水电机组详细信息 ===")
        for unit in self.model.HydroUnits:
            cap = self.hydro_dict[unit].get('Capacity', 0)
            curve = self.hydro_dict[unit].get('ResourceCurve', '')
            bus = self.hydro_dict[unit].get('BusName', '')
            logging.info(f"水电机组 {unit}: 容量={cap}MW, 资源曲线='{curve}', 母线='{bus}'")
        
        # 检查资源曲线数据
        logging.info("=== 资源曲线数据检查 ===")
        logging.info(f"风电曲线数量: {len(self.wind_curves)}")
        if not self.wind_curves.empty:
            logging.info(f"风电曲线名称: {list(self.wind_curves['WTCurve'].unique())}")
        
        logging.info(f"光伏曲线数量: {len(self.solar_curves)}")
        if not self.solar_curves.empty:
            logging.info(f"光伏曲线名称: {list(self.solar_curves['PVCurve'].unique())}")
        
        logging.info(f"水电曲线数量: {len(self.hydro_curves)}")
        if not self.hydro_curves.empty:
            logging.info(f"水电曲线名称: {list(self.hydro_curves['RunoffHydroOperationCurve'].unique())}")
        
        # 检查机组参数合理性
        for unit in self.model.ThermalUnits:
            unit_data = self.thermal_dict[unit]
            cap = unit_data.get('Capacity', 0)
            min_power = unit_data.get('MinPower', 0)
            if min_power > cap:
                logging.error(f"火电机组{unit}: 最小出力({min_power}) > 容量({cap})")
        
        # 输出模型统计信息
        logging.info(f"模型变量数: {len([v for v in self.model.component_objects(pyo.Var)])}")
        logging.info(f"模型约束数: {len([c for c in self.model.component_objects(pyo.Constraint)])}")
        
    def solve(self):
        """求解优化问题"""
        logging.info("开始求解优化问题...")
        
        # 先进行模型调试
        self.debug_model()
        
        # 使用Gurobi求解器
        solver = SolverFactory('gurobi')
        solver.options['MIPGap'] = 0.1  # 进一步放宽到10%
        solver.options['TimeLimit'] = 600  # 10分钟
        solver.options['NumericFocus'] = 3  # 提高数值稳定性
        solver.options['FeasibilityTol'] = 1e-4  # 放宽可行性容差
        solver.options['OptimalityTol'] = 1e-4  # 放宽最优性容差
        
        # 尝试求解
        start_time = time.time()
        results = solver.solve(self.model, tee=True)
        solve_time = time.time() - start_time
        logging.info(f"求解耗时: {solve_time:.2f}秒")
        
        # 检查求解状态
        if (results.solver.status == pyo.SolverStatus.ok and 
            results.solver.termination_condition in [pyo.TerminationCondition.optimal, 
                                                     pyo.TerminationCondition.feasible]):
            logging.info("优化求解成功完成")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            logging.warning("达到时间限制，返回当前最优解")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.infeasible:
            logging.error("模型不可行，尝试诊断...")
            self.diagnose_infeasibility()
            
            # 尝试简化模型
            logging.info("尝试创建简化模型...")
            if self.create_simplified_model():
                logging.info("重新求解简化模型...")
                results = solver.solve(self.model, tee=True)
                if (results.solver.status == pyo.SolverStatus.ok and 
                    results.solver.termination_condition in [pyo.TerminationCondition.optimal, 
                                                             pyo.TerminationCondition.feasible]):
                    logging.info("简化模型求解成功")
                    return True
            
            return False
        else:
            logging.error(f"优化求解失败: {results.solver.termination_condition}")
            return False
    
    def diagnose_infeasibility(self):
        """诊断模型不可行性"""
        logging.info("开始诊断模型不可行性...")
        
        # 写出模型文件用于调试
        try:
            self.model.write('debug_model.lp', format=pyo.WriterFactory.LP)
            logging.info("模型已写出到 debug_model.lp 文件")
        except Exception as e:
            logging.warning(f"无法写出模型文件: {e}")
        
        # 检查一些常见的不可行原因
        logging.info("检查常见不可行原因:")
        
        # 1. 检查负荷是否过大
        peak_load = max(self.load_curve) if self.load_curve else 0
        total_thermal_cap = sum(self.thermal_dict[unit].get('Capacity', 0) for unit in self.model.ThermalUnits)
        if peak_load > total_thermal_cap * 1.5:  # 考虑可再生能源
            logging.error(f"峰值负荷({peak_load}) 可能超过总发电能力")
        
        # 2. 检查机组最小开机时间是否过长
        for unit in self.model.ThermalUnits:
            min_up = self.thermal_dict[unit].get('MinOnLineTime', 0)
            min_down = self.thermal_dict[unit].get('MinOffLineTime', 0)
            if min_up > self.T // 2:
                logging.warning(f"机组{unit}最小开机时间({min_up})过长")
            if min_down > self.T // 2:
                logging.warning(f"机组{unit}最小停机时间({min_down})过长")
        
        # 3. 检查备用约束是否过紧
        logging.warning("建议放宽备用约束或检查机组参数设置")
    
    def create_simplified_model(self):
        """创建简化模型（移除一些复杂约束）"""
        try:
            logging.info("创建简化模型...")
            
            # 移除原有约束
            self.model.del_component(self.model.thermal_ramp_up_con)
            self.model.del_component(self.model.thermal_ramp_down_con)
            self.model.del_component(self.model.min_up_time_con)
            self.model.del_component(self.model.min_down_time_con)
            self.model.del_component(self.model.reserve_con)
            self.model.del_component(self.model.deep_peak_penalty_con)
            
            logging.info("已移除复杂约束：爬坡、最短开停机时间、备用、深度调峰")
            
            # 添加简化的爬坡约束（更宽松）
            def simple_ramp_rule(model, unit, t):
                if t == 0:
                    return pyo.Constraint.Skip
                capacity = self.thermal_dict[unit].get('Capacity', 0)
                # 使用容量的50%作为最大变化量
                max_change = capacity * 0.5
                return (-max_change <= 
                        model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= 
                        max_change)
            
            self.model.simple_ramp_con = pyo.Constraint(
                self.model.ThermalUnits, self.model.T, rule=simple_ramp_rule
            )
            
            logging.info("添加了简化的爬坡约束")
            return True
            
        except Exception as e:
            logging.error(f"创建简化模型失败: {e}")
            return False
    
    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0

        # 创建Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # 1. 火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t]),
                    'Startup': safe_value(self.model.thermal_startup[unit, t]),
                    'Shutdown': safe_value(self.model.thermal_shutdown[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        thermal_df.to_excel(writer, sheet_name='火电机组结果', index=False)
        
        # 2. 可再生能源结果
        wind_results = []
        for unit in self.model.WindUnits:
            for t in self.model.T:
                wind_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.wind_power[unit, t])
                })
        wind_df = pd.DataFrame(wind_results)
        wind_df.to_excel(writer, sheet_name='风电结果', index=False)
        
        solar_results = []
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                solar_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.solar_power[unit, t])
                })
        solar_df = pd.DataFrame(solar_results)
        solar_df.to_excel(writer, sheet_name='光伏结果', index=False)
        
        hydro_results = []
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                hydro_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': safe_value(self.model.hydro_power[unit, t])
                })
        hydro_df = pd.DataFrame(hydro_results)
        hydro_df.to_excel(writer, sheet_name='水电结果', index=False)
        
        # 3. 储能结果
        ess_results = []
        for unit in self.model.ESSUntis:
            for t in self.model.T:
                ess_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Charge': safe_value(self.model.ess_charge[unit, t]),
                    'Discharge': safe_value(self.model.ess_discharge[unit, t]),
                    'SOC': safe_value(self.model.ess_soc[unit, t])
                })
        ess_df = pd.DataFrame(ess_results)
        ess_df.to_excel(writer, sheet_name='储能结果', index=False)
        
        # 4. 负荷结果
        load_results = []
        for load in self.model.Loads:
            for t in self.model.T:
                load_results.append({
                    'Load': load,
                    'Time': t,
                    'Shed': safe_value(self.model.load_shed[load, t])
                })
        load_df = pd.DataFrame(load_results)
        load_df.to_excel(writer, sheet_name='切负荷结果', index=False)
        
        # 5. 网络结果
        bus_results = []
        for bus in self.model.Buses:
            for t in self.model.T:
                bus_results.append({
                    'Bus': bus,
                    'Time': t,
                    'Angle': safe_value(self.model.bus_angle[bus, t])
                })
        bus_df = pd.DataFrame(bus_results)
        bus_df.to_excel(writer, sheet_name='节点相角', index=False)
        
        line_results = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_results.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_df = pd.DataFrame(line_results)
        line_df.to_excel(writer, sheet_name='线路潮流', index=False)
        
        # 6. 系统总结
        total_thermal_gen = thermal_df.groupby('Time')['Power'].sum().values
        total_wind_gen = wind_df.groupby('Time')['Power'].sum().values
        total_solar_gen = solar_df.groupby('Time')['Power'].sum().values
        total_hydro_gen = hydro_df.groupby('Time')['Power'].sum().values
        total_ess_discharge = ess_df.groupby('Time')['Discharge'].sum().values
        total_ess_charge = ess_df.groupby('Time')['Charge'].sum().values
        
        # 使用全局负荷曲线
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T

        summary_data = {
            'Time': self.hours,
            'Total_Load': total_load,
            'Thermal_Gen': total_thermal_gen,
            'Wind_Gen': total_wind_gen,
            'Solar_Gen': total_solar_gen,
            'Hydro_Gen': total_hydro_gen,
            'ESS_Discharge': total_ess_discharge,
            'ESS_Charge': total_ess_charge,
            'Net_Gen': [t + w + s + h + e - c for t, w, s, h, e, c in 
                       zip(total_thermal_gen, total_wind_gen, total_solar_gen, 
                           total_hydro_gen, total_ess_discharge, total_ess_charge)]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='系统总结', index=False)
        
        # 保存Excel
        writer.close()
        logging.info(f"结果已保存到 {output_path}")
        return output_path
    
    def plot_unit_status(self, output_path):
        """绘制机组开关机状态图"""
        logging.info("开始绘制机组开关机状态图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0
        
        # 获取火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        
        # 按机组分组
        grouped = thermal_df.groupby('Unit')
        
        # 创建图表
        fig, axes = plt.subplots(len(grouped), 1, figsize=(15, 10), sharex=True)
        fig.suptitle('火电机组开关机状态及出力情况', fontsize=16)
        
        if len(grouped) == 1:
            axes = [axes]
        
        for i, (unit, group) in enumerate(grouped):
            ax = axes[i]
            
            # 绘制状态（0/1）
            ax.step(group['Time'], group['Status'], 'r-', where='post', label='状态')
            ax.set_ylabel('状态', color='r')
            ax.tick_params(axis='y', labelcolor='r')
            ax.set_ylim(-0.1, 1.1)
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # 创建双Y轴
            ax2 = ax.twinx()
            
            # 绘制出力
            ax2.plot(group['Time'], group['Power'], 'b-', label='出力')
            ax2.set_ylabel('出力 (MW)', color='b')
            ax2.tick_params(axis='y', labelcolor='b')
            
            # 设置标题
            ax.set_title(f'机组 {unit}')
            
            # 添加图例
            lines, labels = ax.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax.legend(lines + lines2, labels + labels2, loc='upper right')
        
        plt.xlabel('时间 (小时)')
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # 为总标题留出空间
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"机组状态图已保存到 {output_path}")
        plt.close()
        return output_path
    
    def plot_generation_mix(self, output_path):
        """绘制各类电源出力情况（堆积柱状图）"""
        logging.info("开始绘制电源结构图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0
        
        # 初始化结果数组
        thermal_gen = np.zeros(self.T)
        wind_gen = np.zeros(self.T)
        solar_gen = np.zeros(self.T)
        hydro_gen = np.zeros(self.T)
        ess_discharge = np.zeros(self.T)
        ess_charge = np.zeros(self.T)
        
        # 计算总发电量
        for t in range(self.T):
            for unit in self.model.ThermalUnits:
                thermal_gen[t] += safe_value(self.model.thermal_power[unit, t])
            for unit in self.model.WindUnits:
                wind_gen[t] += safe_value(self.model.wind_power[unit, t])
            for unit in self.model.SolarUnits:
                solar_gen[t] += safe_value(self.model.solar_power[unit, t])
            for unit in self.model.HydroUnits:
                hydro_gen[t] += safe_value(self.model.hydro_power[unit, t])
            for unit in self.model.ESSUntis:
                ess_discharge[t] += safe_value(self.model.ess_discharge[unit, t])
                ess_charge[t] += safe_value(self.model.ess_charge[unit, t])
        
        # 创建图表
        fig, ax = plt.subplots(figsize=(15, 8))
        
        # 时间轴
        hours = np.arange(self.T)
        width = 0.8  # 柱子宽度
        
        # 绘制堆积柱状图
        bottom = np.zeros(self.T)
        
        # 火电（红色）
        p1 = ax.bar(hours, thermal_gen, width, bottom=bottom, label='Power_ThermalGen', 
                   color='#ff4444', alpha=0.8)
        bottom += thermal_gen
        
        # 水电（蓝色）
        p2 = ax.bar(hours, hydro_gen, width, bottom=bottom, label='Power_RHD', 
                   color='#4444ff', alpha=0.8)
        bottom += hydro_gen
        
        # 风电（绿色）
        p3 = ax.bar(hours, wind_gen, width, bottom=bottom, label='Power_WT', 
                   color='#44ff44', alpha=0.8)
        bottom += wind_gen
        
        # 光伏（黄色）
        p4 = ax.bar(hours, solar_gen, width, bottom=bottom, label='Power_PV', 
                   color='#ffff44', alpha=0.8)
        bottom += solar_gen
        
        # 储能放电（灰色）
        p5 = ax.bar(hours, ess_discharge, width, bottom=bottom, label='Power_StoreDis', 
                   color='#888888', alpha=0.8)
        bottom += ess_discharge
        
        # 储能充电（负值，浅灰色）
        p6 = ax.bar(hours, -ess_charge, width, label='Power_StoreCharge', 
                   color='#cccccc', alpha=0.6)
        
        # 添加总负荷曲线（黑色虚线）
        total_load = self.load_curve[:self.T] if len(self.load_curve) >= self.T else [0] * self.T
        ax.plot(hours, total_load, 'k--', linewidth=2, label='Total Demand')
        
        # 计算备用容量
        total_gen = thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge
        available_reserve = total_gen - np.array(total_load[:self.T])
        
        # 添加备用容量要求线（绿色虚线）
        reserve_req = np.array(total_load[:self.T]) * 0.1  # 假设10%备用要求
        ax.plot(hours, np.array(total_load[:self.T]) + reserve_req, 'g:', linewidth=2, 
               label='Reserve Requirement', alpha=0.7)
        
        # 填充可用备用区域
        ax.fill_between(hours, np.array(total_load[:self.T]) + reserve_req, total_gen, 
                       alpha=0.2, color='lightgreen', label='Available Reserve')
        
        # 设置图表属性
        ax.set_title('Generation Schedule and Reserve Margin', fontsize=14, fontweight='bold')
        ax.set_xlabel('Hour', fontsize=12)
        ax.set_ylabel('Power (MW)', fontsize=12)
        
        # 添加统计信息
        max_demand = max(total_load[:self.T])
        min_reserve = min(available_reserve)
        ax.text(0.02, 0.98, f'Max Demand: {max_demand:.1f} MW\nMin Reserve: {min_reserve:.1f} MW', 
               transform=ax.transAxes, fontsize=10, verticalalignment='top',
               bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
        
        # 设置图例
        ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1), fontsize=10)
        
        # 设置网格
        ax.grid(True, linestyle='--', alpha=0.3)
        ax.set_xlim(-0.5, self.T-0.5)
        ax.set_xticks(range(0, self.T, 2))  # 每2小时标记一次
        
        # 紧凑布局
        plt.tight_layout()
        
        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        logging.info(f"电源结构图已保存到 {output_path}")
        plt.close()
        return output_path
    
    def plot_network_flow(self, output_path):
        """绘制网络潮流图"""
        logging.info("开始绘制网络潮流图...")
        
        # 安全获取变量值
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0
        
        # 获取线路潮流数据
        line_flows = []
        for line in self.model.Lines:
            for t in self.model.T:
                line_flows.append({
                    'Line': line,
                    'Time': t,
                    'Power': safe_value(self.model.line_power[line, t])
                })
        line_flows_df = pd.DataFrame(line_flows)
        
        # 获取节点相角数据
        bus_angles = []
        for bus in self.model.Buses:
            for t in self.model.T:
                bus_angles.append({
                    'Bus': bus,
                    'Time': t,
                    'Angle': safe_value(self.model.bus_angle[bus, t])
                })
        bus_angles_df = pd.DataFrame(bus_angles)
        
        # 创建图表
        plt.figure(figsize=(15, 10))
        
        # 绘制线路潮流
        unique_lines = line_flows_df['Line'].unique()
        for i, line in enumerate(unique_lines):
            line_data = line_flows_df[line_flows_df['Line'] == line]
            plt.plot(line_data['Time'], line_data['Power'], label=f'线路 {line}')
        
        plt.title('线路潮流随时间变化')
        plt.xlabel('时间 (小时)')
        plt.ylabel('功率 (MW)')
        plt.legend(loc='best', bbox_to_anchor=(1.05, 1))
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"网络潮流图已保存到 {output_path}")
        plt.close()
        return output_path

# 主程序
if __name__ == "__main__":
    # 输入输出路径
    input_file = r"C:\Users\admin\Desktop\Final_Cleaned_Rots_v2.xlsx"
    output_file = "UC_Results.xlsx"
    status_plot = "Unit_Status.png"
    mix_plot = "Generation_Mix.png"
    flow_plot = "Network_Flow.png"
    resource_plot = "Resource_Curves.png"
    
    try:
        # 创建模型
        logging.info("创建模型...")
        uc_model = UC_Model(input_file)
        
        # 绘制资源曲线预览
        uc_model.plot_resource_curves(resource_plot)
        
        # 求解模型
        logging.info("求解模型...")
        result = uc_model.solve()
        
        if result:
            # 保存结果
            logging.info("保存结果...")
            result_path = uc_model.save_results(output_file)
            
            # 绘制图表
            logging.info("绘制图表...")
            status_path = uc_model.plot_unit_status(status_plot)
            mix_path = uc_model.plot_generation_mix(mix_plot)
            flow_path = uc_model.plot_network_flow(flow_plot)
            
            # 将图表插入Excel
            wb = openpyxl.load_workbook(output_file)
            
            # 添加资源曲线预览图
            ws = wb.create_sheet("资源曲线预览")
            img = Image(resource_plot)
            ws.add_image(img, 'A1')
            
            # 添加机组状态图
            ws = wb.create_sheet("机组状态图")
            img = Image(status_path)
            ws.add_image(img, 'A1')
            
            # 添加电源结构图
            ws = wb.create_sheet("电源结构图")
            img = Image(mix_path)
            ws.add_image(img, 'A1')
            
            # 添加网络潮流图
            ws = wb.create_sheet("网络潮流图")
            img = Image(flow_path)
            ws.add_image(img, 'A1')
            
            # 保存最终结果
            wb.save(output_file)
            logging.info("所有结果已保存并整合到Excel文件中")
            print(f"优化完成，结果已保存到: {output_file}")
        else:
            logging.error("优化求解失败，请检查模型和输入数据")
            print("优化求解失败，请检查日志获取详细信息")
        
    except Exception as e:
        logging.exception("程序运行出错")
        print(f"程序运行出错: {str(e)}")
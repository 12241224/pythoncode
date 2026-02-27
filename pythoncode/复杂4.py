import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import os
import logging
import pyomo.environ as pyo
import sys


# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 初始化日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class UC_Model:
    def __init__(self, data_path):
        self.data_path = data_path
        self.load_data()
        self.create_model()
    def check_all_df_for_illegal(df, name):
        for col in df.columns:
            if df[col].dtype == object:
                if df[col].str.contains(r'[a-zA-Z]|nan|inf|\+nan|x\d+', na=False).any():
                    print(f"{name} {col} 存在非法字符串")
            if df[col].isnull().any():
                print(f"{name} {col} 存在 nan")
            if np.isinf(df[col]).any():
                print(f"{name} {col} 存在 inf")
    def get_load_curve_val(self, t):
        """获取指定时间点 t 的负荷曲线值"""
        if f'Time_{t}' in self.load_curve:
            val = self.load_curve[f'Time_{t}']
            if isinstance(val, (pd.Series, np.ndarray)):  # 如果返回的是Series或ndarray类型
                logging.debug(f"load_curve_val (sequence) for Time_{t}: {val}")
                return float(val.iloc[0]) if isinstance(val, pd.Series) else float(val[0])  # 强制转换为单一数值
            return float(val)  # 如果是标量，则直接转换为float
        else:
            logging.warning(f"时间点 {t} 不在负荷曲线中，返回默认值 0")
            return 0  # 如果没有找到对应时间点，返回默认值 0
        
    def load_data(self):
        """加载所有Excel数据"""
        logging.info("开始加载数据...")
        
        # 系统汇总信息
        self.system_summary = pd.read_excel(
            self.data_path, sheet_name='SystemSummary', skiprows=0, nrows=1
        ).iloc[0]
        
        # 机组参数
        self.thermal_units = pd.read_excel(
            self.data_path, sheet_name='UnitThermalGenerators', skiprows=0
        )
        self.wind_units = pd.read_excel(
            self.data_path, sheet_name='UnitWindGenerators', skiprows=0
        )
        # 检查wind_units的实际列名
        print("wind_units columns:", self.wind_units.columns.tolist())
        # 列名自动修正
        self.wind_units.rename(columns={
            '#风电机组编号': 'WTPlantID',
            '机组装机容量(MW)': 'Capacity ',
            '风电资源曲线名称': 'ResourceCurve'
        }, inplace=True)
        # 检查风电机组数据是否包含 'Capacity ' 列
        if 'Capacity ' not in self.wind_units.columns:
            logging.error("风电机组数据中没有 'Capacity ' 列，请检查数据源或列名。")
            raise ValueError("风电机组数据中没有 'Capacity '列")
        else:
            self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        
        self.solar_units = pd.read_excel(
                    self.data_path, sheet_name='UnitSolarGenerators', skiprows=0
                )
        # 检查solar_units的实际列名
        print("solar_units columns:", self.solar_units.columns.tolist())
        # 列名自动修正
        self.solar_units.rename(columns={
            '#光伏机组编号': 'PVPlantID',
            '机组装机容量(MW)': 'Capacity',
            '光伏资源曲线名称': 'ResourceCurve'
        }, inplace=True)
        
        self.ess_units = pd.read_excel(
            self.data_path, sheet_name='StorElectrochemicalESS', skiprows=0
        )
        # 检查ess_units的实际列名
        print("ess_units columns:", self.ess_units.columns.tolist())
        # 列名自动修正
        self.ess_units.rename(columns={
            '电化学储能系统编号': 'ESSId',
            '储能容量(MWh)': 'Capacity',
            '机组装机容量(MW)': 'Capacity',
            # 新增：兼容充放电功率字段
            '充电功率(MW)': 'MaxChargePower',
            '放电功率(MW)': 'MaxDisPower'
        }, inplace=True)
        
        self.hydro_units = pd.read_excel(
            self.data_path, sheet_name='UnitRunoffHydroGenerators', skiprows=0
        )
        # 检查hydro_units的实际列名
        print("hydro_units columns:", self.hydro_units.columns.tolist())
        # 列名自动修正
        self.hydro_units.rename(columns={
            '径流式水电机组名称': 'RunoffHydroUnitName',
            '机组装机容量(MW)': 'Capacity',
            # 注意：实际列名为 '径流式 水电资源曲线名称'，有空格
            '径流式 水电资源曲线名称': 'ResourceCurve',
            # 兼容无空格的情况
            '径流式水电资源曲线名称': 'ResourceCurve'
        }, inplace=True)
        
        # 曲线数据
        self.wind_curves = pd.read_excel(
            self.data_path, sheet_name='CurveWindResource', skiprows=0
        )
        # 检查wind_curves的实际列名
        print("wind_curves columns:", self.wind_curves.columns.tolist())
        # 列名自动修正
        self.wind_curves.rename(columns={
            '#风电资源曲线名称/风电资源曲线': 'WTCurve',
            # 如有其他可能的列名，也可补充
        }, inplace=True)
        # 修正：去除wind_curves中的无效行（如全空或含nan的行）
        self.wind_curves = self.wind_curves.dropna(subset=['WTCurve'])

        self.solar_curves = pd.read_excel(
            self.data_path, sheet_name='CurveSolarResource', skiprows=0
        )
        # 检查solar_curves的实际列名
        print("solar_curves columns:", self.solar_curves.columns.tolist())
        # 列名自动修正
        self.solar_curves.rename(columns={
            '#光伏资源曲线名称/光伏资源曲线': 'PVCurve',
            # 如有其他可能的列名，也可补充
        }, inplace=True)
        # 修正：去除solar_curves中的无效行（如全空或含nan的行）
        self.solar_curves = self.solar_curves.dropna(subset=['PVCurve'])

        self.hydro_curves = pd.read_excel(
            self.data_path, sheet_name='CurveRunoffHydroResource', skiprows=0
        )
        # 检查hydro_curves的实际列名
        print("hydro_curves columns:", self.hydro_curves.columns.tolist())
        # 列名自动修正
        self.hydro_curves.rename(columns={
            '#径流式水电资源曲线名称/径流式水电资源曲线': 'RunoffHydroOperationCurve',
            # 兼容无#和无斜杠的情况
            '径流式水电资源曲线名称': 'RunoffHydroOperationCurve',
            '径流式 水电资源曲线名称': 'RunoffHydroOperationCurve'
        }, inplace=True)
        # 修正：去除hydro_curves中的无效行（如全空或含nan的行）
        self.hydro_curves = self.hydro_curves.dropna(subset=['RunoffHydroOperationCurve'])
        # 新增：去除所有曲线表中的全为nan或非数字的时间列
        def drop_all_nan_timecols(df, prefix="Time_"):
            time_cols = [col for col in df.columns if col.startswith(prefix)]
            for col in time_cols:
                # 尝试转换为float
                df[col] = pd.to_numeric(df[col], errors='coerce')
                # 如果全部为nan或全部为0，删除该列（允许负值，防止负风/光/水出力导致nan）
                if df[col].isnull().all() or (df[col].notnull().all() and (df[col] == 0).all()):
                    df.drop(columns=[col], inplace=True)
            return df
        self.wind_curves = drop_all_nan_timecols(self.wind_curves, "Time_")
        self.solar_curves = drop_all_nan_timecols(self.solar_curves, "Time_")
        self.hydro_curves = drop_all_nan_timecols(self.hydro_curves, "Time_")
        
        self.load_curve = pd.read_excel(
            self.data_path, sheet_name='CurveLoad', skiprows=0
        ).iloc[0]
        
        
        # 网络数据
        self.buses = pd.read_excel(self.data_path, sheet_name='NetBuses', skiprows=0)
        # 检查buses的实际列名
        print("buses columns:", self.buses.columns.tolist())
        # 列名自动修正
        self.buses.rename(columns={
            '母线编号': 'BusId',
            # 如有其他可能的列名，也可补充
        }, inplace=True)
        self.lines = pd.read_excel(self.data_path, sheet_name='NetLines', skiprows=0)
        # 检查lines的实际列名
        print("lines columns:", self.lines.columns.tolist())
        # 列名自动修正
        self.lines.rename(columns={
            '输电线路编号': 'BranchId',
            # 如有其他可能的列名，也可补充
        }, inplace=True)
        self.transformers = pd.read_excel(self.data_path, sheet_name='NetTransformers', skiprows=0)

        self.sections = pd.read_excel(self.data_path, sheet_name='NetSectionsLines', skiprows=0)
        # 检查sections的实际列名
        print("sections columns:", self.sections.columns.tolist())
        # 列名自动修正
        self.sections.rename(columns={
            '断面编号': 'NetSectionId',
            # 如有其他可能的列名，也可
        }, inplace=True)
        self.section_capacity = pd.read_excel(self.data_path, sheet_name='NetSectionsCapacity', skiprows=0)
        # 检查section_capacity的实际列名
        print("section_capacity columns:", self.section_capacity.columns.tolist())
        # 列名自动修正
        self.section_capacity.rename(columns={
            '#断面编号': 'NetSectionId',
            # 如有其他可能的列名，也可补充
        }, inplace=True)
        self.ac_tielines = pd.read_excel(self.data_path, sheet_name='NetHVACTieLine', skiprows=0)
        self.ac_tieline_curves = pd.read_excel(self.data_path, sheet_name='EnergyHVACTieLine', skiprows=0)
        
        # 负荷数据
        self.loads = pd.read_excel(self.data_path, sheet_name='Loads', skiprows=0)
        print("loads columns:", self.loads.columns.tolist())
        # 燃料数据
        self.fuel_data = pd.read_excel(self.data_path, sheet_name='UnitFuel', skiprows=0).iloc[0]
        
        # 设置时间参数
        self.T = int(self.system_summary['Duration'])
        self.dt = 1  # 调度间隔为1小时
        self.hours = list(range(self.T))
        
        
        # 创建机组字典
        self.thermal_dict = self.thermal_units.set_index('ThermalUnitNumber').to_dict('index')
        # 确保时间相关参数是整数，并处理可能的NaN值
        for unit in self.thermal_dict:
            # 处理最小运行时间
            min_on_time = self.thermal_dict[unit]['MinOnLineTime']
            if pd.isna(min_on_time):
                self.thermal_dict[unit]['MinOnLineTime'] = 0
            else:
                self.thermal_dict[unit]['MinOnLineTime'] = int(min_on_time)
                
            # 处理最小停机时间
            min_off_time = self.thermal_dict[unit]['MinOffLineTime']
            if pd.isna(min_off_time):
                self.thermal_dict[unit]['MinOffLineTime'] = 0
            else:
                self.thermal_dict[unit]['MinOffLineTime'] = int(min_off_time)
            
        self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        self.solar_dict = self.solar_units.set_index('PVPlantID').to_dict('index')
        self.ess_dict = self.ess_units.set_index('ESSId').to_dict('index')
        if self.hydro_units['RunoffHydroUnitName'].duplicated().any():
            logging.warning("发现重复的水电机组名称，将通过去重后处理。")
            self.hydro_units = self.hydro_units.drop_duplicates(subset='RunoffHydroUnitName')
            self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        else:
            self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
                # 检查所有曲线表和负荷曲线是否有 nan 或非法值
        def check_curve_nan(df, name):
            for col in df.columns:
                if col.startswith("Time_"):
                    if df[col].isnull().any():
                        raise ValueError(f"{name} 存在 nan，请检查 {col} 列！")
                    if not np.issubdtype(df[col].dtype, np.number):
                        raise ValueError(f"{name} {col} 列存在非数字，请检查！")
        check_curve_nan(self.wind_curves, "风电曲线")
        check_curve_nan(self.solar_curves, "光伏曲线")
        check_curve_nan(self.hydro_curves, "水电曲线")
        # 检查负荷曲线
        for k, v in self.load_curve.items():
            # 跳过明显的表头或非时间列（如 'LoadCurve', '负荷曲线1' 等）
            if not str(k).startswith("Time_"):
                continue
            try:
                v2 = float(v)
                if np.isnan(v2):
                    raise ValueError(f"负荷曲线 {k} 存在 nan，请检查 CurveLoad 表！")
            except Exception:
                raise ValueError(f"负荷曲线 {k} 存在非法值，请检查 CurveLoad 表！")


    
    def get_unit_param_as_float(self, unit_dict, unit, key, default=0.0):
        """安全获取机组参数（float）"""
        try:
            val = unit_dict[unit][key]
            return float(val) if not pd.isna(val) else default
        except Exception:
            return default

    def get_unit_param_as_int(self, unit_dict, unit, key, default=0):
        """安全获取机组参数（int）"""
        try:
            val = unit_dict[unit][key]
            return int(val) if not pd.isna(val) else default
        except Exception:
            return default

    def create_model(self):
        """创建Pyomo优化模型"""
        logging.info("开始创建优化模型...")
        
        self.model = pyo.ConcreteModel()
        
        # 定义集合
        self.model.T = pyo.Set(initialize=self.hours)  # 时间段
        self.model.ThermalUnits = pyo.Set(initialize=self.thermal_units['ThermalUnitNumber'].tolist())
        self.model.WindUnits = pyo.Set(initialize=self.wind_units['WTPlantID'].tolist())
        self.model.SolarUnits = pyo.Set(initialize=self.solar_units['PVPlantID'].tolist())
        self.model.ESSUntis = pyo.Set(initialize=self.ess_units['ESSId'].tolist())
        self.model.HydroUnits = pyo.Set(initialize=self.hydro_units['RunoffHydroUnitName'].tolist())
        self.model.Loads = pyo.Set(initialize=self.loads['负荷编号'].tolist())
        self.model.Buses = pyo.Set(initialize=self.buses['BusId'].tolist())
        self.model.Lines = pyo.Set(initialize=self.lines['BranchId'].tolist())
        self.model.Sections = pyo.Set(initialize=self.section_capacity['NetSectionId'].tolist())
      

        # 定义变量
        # 火电机组
        self.model.thermal_status = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 机组状态 (0/1)
        self.model.thermal_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )  # 机组出力
        self.model.thermal_startup = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 启动状态
        self.model.thermal_shutdown = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 停机状态
        
        # 可再生能源
        self.model.wind_power = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_power = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_power = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 储能系统
        self.model.ess_charge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 充电功率
        self.model.ess_discharge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 放电功率
        self.model.ess_soc = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 储能状态
        
        # 负荷相关
        self.model.load_shed = pyo.Var(
            self.model.Loads, self.model.T, within=pyo.NonNegativeReals
        )  # 切负荷量
        
        # 网络相关
        self.model.bus_angle = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )  # 节点相角
        self.model.line_power = pyo.Var(
            self.model.Lines, self.model.T, within=pyo.Reals
        )  # 线路功率
        
        # 目标函数
        def total_cost_rule(model):
            """系统总成本计算"""
            cost = 0
            
            # 火电运行成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    # 二次成本函数线性化近似
                    a = self.thermal_dict[unit]['OperationCoeff_A']
                    b = self.thermal_dict[unit]['OperationCoeff_B']
                    c = self.thermal_dict[unit]['OperationCoeff_C']
                    min_power = self.thermal_dict[unit]['MinPower']
                    max_power = self.thermal_dict[unit]['Capacity']
                    
                    # 分段线性化近似二次函数
                    cost += model.thermal_status[unit, t] * c
                    cost += b * model.thermal_power[unit, t]
            
            model.deep_peak_penalty = pyo.Var(
            model.ThermalUnits, model.T, within=pyo.NonNegativeReals
            )        
            
            def deep_peak_penalty_rule(model, unit, t):
                max_power = self.thermal_dict[unit]['Capacity']
                deep_threshold = self.thermal_dict[unit]['DeepPeakShavingThreshold'] * max_power
                deep_cost = self.thermal_dict[unit]['DeepPeakShavingCost']
                
                # 使用 big-M 方法处理条件约束
                M = max_power  # 一个足够大的数
                # 这里简化处理，实际应添加适当的约束来计算惩罚
                
                # 惩罚成本项
                return model.deep_peak_penalty[unit, t] >= deep_cost * (
                    deep_threshold - model.thermal_power[unit, t]
                )
            
            model.deep_peak_penalty_con = pyo.Constraint(
                model.ThermalUnits, model.T, rule=deep_peak_penalty_rule
            )
            
            # 将深度调峰惩罚加入总成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    cost += model.deep_peak_penalty[unit, t]
            
            # 启停成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    startup_cost = self.thermal_dict[unit]['StateOnFuelConsumption']
                    cost += startup_cost * model.thermal_startup[unit, t]
                    
                    # 停机成本通常为0，这里省略
            
            # 切负荷惩罚
            penalty_cut = 10000  # 元/MWh
            for load in model.Loads:
                for t in model.T:
                    cost += penalty_cut * model.load_shed[load, t]
            
            # 备用不足惩罚
            penalty_reserve = 5000  # 元/MWh
            # 这里简化处理，实际应计算备用不足量
            
            return cost
        
        self.model.total_cost = pyo.Objective(rule=total_cost_rule, sense=pyo.minimize)
        
        # ========== 关键调试：输出模型约束和变量数量，辅助定位 infeasible ==========
        logging.info(f"模型变量总数: {len(list(self.model.component_data_objects(pyo.Var)))}")
        logging.info(f"模型约束总数: {len(list(self.model.component_data_objects(pyo.Constraint)))}")
        # 可选：输出所有约束名
        for cname, c in self.model.component_map(pyo.Constraint, active=True).items():
            logging.info(f"约束: {cname}, 条数: {len(list(c.values()))}")

        # ========== 可选：保存模型LP文件，人工用文本编辑器检查 ==========
        try:
            self.model.write('debug_model.lp', io_options={'symbolic_solver_labels': True})
            logging.info("已导出 debug_model.lp，可用文本编辑器检查约束和变量")
        except Exception as e:
            logging.warning(f"导出LP文件失败: {e}")

        # 1. 火电机组逻辑约束
        def thermal_logic_rule(model, unit, t):
            if t == 0:
                init_status = self.thermal_dict[unit].get('InitStatus', False)
                if isinstance(init_status, str):
                    init_status = init_status.strip().lower() in ['true', '1', 'yes']
                return model.thermal_status[unit, t] == (1 if init_status else 0)
            else:
                return model.thermal_status[unit, t] - model.thermal_status[unit, t-1] == \
                    model.thermal_startup[unit, t] - model.thermal_shutdown[unit, t]
        self.model.thermal_logic_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_logic_rule
        )

        # 2. 火电机组出力约束
        def thermal_power_lower_rule(model, unit, t):
            min_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'MinPower', 0.0)
            return model.thermal_power[unit, t] >= min_power * model.thermal_status[unit, t]
        def thermal_power_upper_rule(model, unit, t):
            max_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'Capacity', 0.0)
            return model.thermal_power[unit, t] <= max_power * model.thermal_status[unit, t]
        self.model.thermal_power_lower_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_lower_rule
        )
        self.model.thermal_power_upper_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_upper_rule
        )

        # 3. 深度调峰约束
        self.model.thermal_deep_power = pyo.Var(self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals)
        def deep_power_rule(model, unit, t):
            threshold = self.get_unit_param_as_float(self.thermal_dict, unit, 'DeepPeakShavingThreshold', 0.0)
            min_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'MinPower', 0.0)
            if threshold > 0 and min_power > 0:
                deep_threshold = min_power * threshold
                return model.thermal_deep_power[unit, t] >= deep_threshold - model.thermal_power[unit, t]
            else:
                return model.thermal_deep_power[unit, t] == 0
        self.model.deep_power_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=deep_power_rule
        )

        # 4. 爬坡约束
        def thermal_ramp_up_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_up = self.get_unit_param_as_float(self.thermal_dict, unit, 'UpRamppingRate', 0.0)
            capacity = self.get_unit_param_as_float(self.thermal_dict, unit, 'Capacity', 0.0)
            return model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= ramp_up + model.thermal_startup[unit, t] * capacity
        self.model.thermal_ramp_up_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_up_rule
        )
        def thermal_ramp_down_rule(model, unit, t):
            if t == 0:
                return pyo.Constraint.Skip
            ramp_down = self.get_unit_param_as_float(self.thermal_dict, unit, 'DownRamppingRate', 0.0)
            capacity = self.get_unit_param_as_float(self.thermal_dict, unit, 'Capacity', 0.0)
            return model.thermal_power[unit, t-1] - model.thermal_power[unit, t] <= ramp_down + model.thermal_shutdown[unit, t] * capacity
        self.model.thermal_ramp_down_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_down_rule
        )

        # 5. 最短开停机时间约束
        def min_up_time_rule(model, unit, t):
            min_up = self.get_unit_param_as_int(self.thermal_dict, unit, 'MinOnLineTime', 0)
            if t < min_up:
                return pyo.Constraint.Skip
            min_up = int(min_up)
            start_time = int(max(0, t - min_up + 1))
            end_time = int(t + 1)
            return sum(model.thermal_shutdown[unit, k] for k in range(start_time, end_time)) <= 1 - model.thermal_status[unit, t]
        self.model.min_up_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_up_time_rule
        )
        def min_down_time_rule(model, unit, t):
            min_down = self.get_unit_param_as_int(self.thermal_dict, unit, 'MinOffLineTime', 0)
            if t < min_down:
                return pyo.Constraint.Skip
            min_down = int(min_down)
            start_time = int(max(0, t - min_down + 1))
            end_time = int(t + 1)
            return sum(model.thermal_startup[unit, k] for k in range(start_time, end_time)) <= model.thermal_status[unit, t]
        self.model.min_down_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_down_time_rule
        )

        # 6. 可再生能源出力约束
        def wind_power_rule(model, unit, t):
            curve_name = self.wind_dict[unit]['ResourceCurve']
            capacity = self.get_unit_param_as_float(self.wind_dict, unit, 'Capacity ', 0.0)
            wind_curve = self.wind_curves[self.wind_curves['WTCurve'] == curve_name].iloc[0]
            max_power = wind_curve[f'Time_{t}'] * capacity
            return model.wind_power[unit, t] <= max_power
        self.model.wind_power_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_power_rule
        )
        def solar_power_rule(model, unit, t):
            curve_name = self.solar_dict[unit]['ResourceCurve']
            capacity = self.get_unit_param_as_float(self.solar_dict, unit, 'Capacity', 0.0)
            solar_curve = self.solar_curves[self.solar_curves['PVCurve'] == curve_name].iloc[0]
            # 兼容不同时间列名：优先Time_{t}，否则尝试str(t)，否则取第一个时间列
            time_col = f'Time_{t}'
            if time_col in solar_curve:
                curve_val = solar_curve[time_col]
            elif str(t) in solar_curve:
                curve_val = solar_curve[str(t)]
            else:
                # 取第一个以Time_开头的列
                time_cols = [col for col in solar_curve.index if str(col).startswith('Time_')]
                if time_cols:
                    curve_val = solar_curve[time_cols[0]]
                else:
                    curve_val = 0.0
            max_power = curve_val * capacity
            return model.solar_power[unit, t] <= max_power
        self.model.solar_power_con = pyo.Constraint(
            self.model.SolarUnits, self.model.T, rule=solar_power_rule
        )
        def hydro_power_rule(model, unit, t):
            curve_name = self.hydro_dict[unit]['ResourceCurve']
            capacity = self.get_unit_param_as_float(self.hydro_dict, unit, 'Capacity', 0.0)
            # 先判断曲线是否存在，避免iloc[0]越界
            hydro_curve_df = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'] == curve_name]
            if not hydro_curve_df.empty:
                hydro_curve = hydro_curve_df.iloc[0]
                # 兼容不同时间列名：优先Time_{t}，否则尝试str(t)，否则取第一个时间列
                time_col = f'Time_{t}'
                if time_col in hydro_curve:
                    curve_val = hydro_curve[time_col]
                elif str(t) in hydro_curve:
                    curve_val = hydro_curve[str(t)]
                else:
                    time_cols = [col for col in hydro_curve.index if str(col).startswith('Time_')]
                    if time_cols:
                        curve_val = hydro_curve[time_cols[0]]
                    else:
                        curve_val = 0.0
            else:
                # 没有找到对应曲线，最大出力为0
                curve_val = 0.0
            max_power = curve_val * capacity
            return model.hydro_power[unit, t] <= max_power
        self.model.hydro_power_con = pyo.Constraint(
            self.model.HydroUnits, self.model.T, rule=hydro_power_rule
        )

        # 7. 储能系统约束
        def ess_charge_power_rule(model, unit, t):
            max_charge = self.get_unit_param_as_float(self.ess_dict, unit, 'MaxChargePower', 0.0)
            return model.ess_charge[unit, t] <= max_charge
        self.model.ess_charge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_charge_power_rule
        )
        def ess_discharge_power_rule(model, unit, t):
            max_discharge = self.get_unit_param_as_float(self.ess_dict, unit, 'MaxDisPower', 0.0)
            return model.ess_discharge[unit, t] <= max_discharge
        self.model.ess_discharge_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_discharge_power_rule
        )
        def ess_soc_rule(model, unit, t):
            min_soc = self.get_unit_param_as_float(self.ess_dict, unit, 'MinSOC', 0.0)
            max_soc = self.get_unit_param_as_float(self.ess_dict, unit, 'MaxSOC', 1.0)
            capacity = self.get_unit_param_as_float(self.ess_dict, unit, 'Capacity', 0.0)
            return (min_soc * capacity, model.ess_soc[unit, t], max_soc * capacity)
        self.model.ess_soc_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_rule
        )
        def ess_soc_balance_rule(model, unit, t):
            eff = self.get_unit_param_as_float(self.ess_dict, unit, 'Eff', 1.0)
            if t == 0:
                return model.ess_soc[unit, t] == 0
            return model.ess_soc[unit, t] == model.ess_soc[unit, t-1] + eff * model.ess_charge[unit, t] - model.ess_discharge[unit, t] / eff
        self.model.ess_soc_balance_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_balance_rule
        )

        # 8. 系统功率平衡约束
        def power_balance_rule(model, t):
            thermal_gen = sum(model.thermal_power[unit, t] for unit in model.ThermalUnits)
            wind_gen = sum(model.wind_power[unit, t] for unit in model.WindUnits)
            solar_gen = sum(model.solar_power[unit, t] for unit in model.SolarUnits)
            hydro_gen = sum(model.hydro_power[unit, t] for unit in model.HydroUnits)
            ess_discharge = sum(model.ess_discharge[unit, t] for unit in model.ESSUntis)
            ess_charge = sum(model.ess_charge[unit, t] for unit in model.ESSUntis)
            total_load = 0
            # 自动兼容负荷编号字段
            load_id_col = 'LoadId' if 'LoadId' in self.loads.columns else self.loads.columns[0]
            coef_col = 'ActivePowerCoef' if 'ActivePowerCoef' in self.loads.columns else self.loads.columns[1]
            for load in model.Loads:
                # 获取该负荷的有功比例系数
                mask = self.loads[load_id_col].astype(str) == str(load)
                if mask.any():
                    active_coef = self.loads.loc[mask, coef_col].values[0]
                else:
                    active_coef = 0
                # 确保active_coef为float
                try:
                    active_coef = float(active_coef)
                except (ValueError, TypeError):
                    active_coef = 0.0
                time_key = f'Time_{t}'
                if time_key in self.load_curve:
                    load_curve_val = self.load_curve[time_key]
                else:
                    try:
                        load_curve_val = self.load_curve[t]
                    except KeyError:
                        load_curve_val = 0.0
                try:
                    load_curve_val = float(load_curve_val)
                except (ValueError, TypeError):
                    load_curve_val = 0.0
                total_load += active_coef * load_curve_val
            return thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge == total_load - sum(model.load_shed[load, t] for load in model.Loads) + ess_charge
        self.model.power_balance_con = pyo.Constraint(
            self.model.T, rule=power_balance_rule
        )

        # 9. 备用约束
        def reserve_requirement_rule(model, t):
            reserve_req = 0.1
            total_load = 0
            # 自动兼容负荷编号字段
            load_id_col = 'LoadId' if 'LoadId' in self.loads.columns else self.loads.columns[0]
            coef_col = 'ActivePowerCoef' if 'ActivePowerCoef' in self.loads.columns else self.loads.columns[1]
            for load in model.Loads:
                mask = self.loads[load_id_col].astype(str) == str(load)
                if mask.any():
                    active_coef = self.loads.loc[mask, coef_col].values[0]
                else:
                    active_coef = 0
                try:
                    active_coef = float(active_coef)
                except (ValueError, TypeError):
                    active_coef = 0.0
                time_key = f'Time_{t}'
                if time_key in self.load_curve:
                    load_curve_val = self.load_curve[time_key]
                else:
                    try:
                        load_curve_val = self.load_curve[t]
                    except KeyError:
                        load_curve_val = 0.0
                try:
                    load_curve_val = float(load_curve_val)
                except (ValueError, TypeError):
                    load_curve_val = 0.0
                total_load += active_coef * load_curve_val
            available_reserve = 0
            for unit in model.ThermalUnits:
                max_power = self.get_unit_param_as_float(self.thermal_dict, unit, 'Capacity', 0.0)
                available_reserve += model.thermal_status[unit, t] * max_power - model.thermal_power[unit, t]
            return available_reserve >= reserve_req * total_load
        self.model.reserve_con = pyo.Constraint(
            self.model.T, rule=reserve_requirement_rule
        )

        logging.info("优化模型创建完成")
    
    def solve(self):
        """求解优化问题"""
        logging.info("开始求解优化问题...")
        
        # 使用Gurobi求解器
        solver = SolverFactory('gurobi')
        solver.options['MIPGap'] = 0.01  # 设置MIP gap为1%
        solver.options['TimeLimit'] = 600  # 设置10分钟时间限制
        results = solver.solve(self.model, tee=True)
        
        if (results.solver.status == pyo.SolverStatus.ok and 
            results.solver.termination_condition == pyo.TerminationCondition.optimal):
            logging.info("优化求解成功完成")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            logging.warning("达到时间限制，返回当前最优解")
            return True
        else:
            logging.error("优化求解失败")
            return False

    
    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")

        # 新增：安全获取变量值，未初始化时返回0
        def safe_value(var):
            try:
                v = pyo.value(var)
                # 若为None或nan也返回0
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    return 0
                return v
            except Exception:
                return 0

        # 创建Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')

        # 1. 火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t]),
                    'Startup': safe_value(self.model.thermal_startup[unit, t]),
                    'Shutdown': safe_value(self.model.thermal_shutdown[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        thermal_df.to_excel(writer, sheet_name='火电机组结果', index=False)

        # 2. 可再生能源结果
        wind_results = []
        for unit in self.model.WindUnits:
            for t in self.model.T:
                # wind_power变量的索引可能不全，需判断
                if (unit, t) in self.model.wind_power:
                    power = safe_value(self.model.wind_power[unit, t])
                else:
                    power = 0
                wind_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': power
                })
        wind_df = pd.DataFrame(wind_results)
        wind_df.to_excel(writer, sheet_name='风电结果', index=False)

        # 3. 光伏结果
        solar_results = []
        for unit in self.model.SolarUnits:
            for t in self.model.T:
                if (unit, t) in self.model.solar_power:
                    power = safe_value(self.model.solar_power[unit, t])
                else:
                    power = 0
                solar_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': power
                })
        solar_df = pd.DataFrame(solar_results)
        solar_df.to_excel(writer, sheet_name='光伏结果', index=False)

        # 4. 水电结果
        hydro_results = []
        for unit in self.model.HydroUnits:
            for t in self.model.T:
                if (unit, t) in self.model.hydro_power:
                    power = safe_value(self.model.hydro_power[unit, t])
                else:
                    power = 0
                hydro_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Power': power
                })
        hydro_df = pd.DataFrame(hydro_results)
        hydro_df.to_excel(writer, sheet_name='水电结果', index=False)
                    
        # 5. 储能结果
        ess_results = []
        for unit in self.model.ESSUntis:
            for t in self.model.T:
                ess_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Charge': safe_value(self.model.ess_charge[unit, t]),
                    'Discharge': safe_value(self.model.ess_discharge[unit, t]),
                    'SOC': safe_value(self.model.ess_soc[unit, t])
                })
        ess_df = pd.DataFrame(ess_results)
        ess_df.to_excel(writer, sheet_name='储能结果', index=False)
       
        # 6. 网络结果
        line_results = []
        valid_line_keys = set(self.model.line_power.keys())
        for line in self.model.Lines:
            for t in self.model.T:
                # 只在索引有效时才取值，避免pyomo内部报错
                if (line, t) in valid_line_keys:
                    power = safe_value(self.model.line_power[line, t])
                else:
                    power = 0
                line_results.append({
                    'Line': line,
                    'Time': t,
                    'Power': power
                })
        line_df = pd.DataFrame(line_results)
        line_df.to_excel(writer, sheet_name='线路潮流', index=False)
        
        # 7. 断面结果
        section_results = []
        for section in self.model.Sections:
            section_lines = self.sections[self.sections['NetSectionId'] == section]
            for t in self.model.T:
                total_power = 0
                for _, row in section_lines.iterrows():
                    line = row['BranchId']
                    direction = 1 if row['BranchDirection'] else -1
                    # 修正：先判断(line, t)是否在变量索引中，避免KeyError
                    if (line, t) in self.model.line_power:
                        total_power += safe_value(self.model.line_power[line, t]) * direction
                    else:
                        total_power += 0  # 不存在的索引直接加0
                section_results.append({
                    'Section': section,
                    'Time': t,
                    'Power': total_power
                })
        section_df = pd.DataFrame(section_results)
        section_df.to_excel(writer, sheet_name='断面潮流', index=False)
        
        # 8. 系统总结
        summary_data = {
            '总成本': [safe_value(self.model.total_cost)],
            '总发电量': [thermal_df['Power'].sum() + wind_df['Power'].sum()],
            '总弃风量': [0],  # 计算实际弃风量
            '总切负荷量': [0]  # 计算总切负荷量
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='系统总结', index=False)
        
        # 保存Excel
        writer.close()
        
        logging.info(f"结果已保存到 {output_path}")
        
        # 返回结果文件路径
        return output_path
    
    def plot_unit_status(self, output_path):
        """绘制机组开关机状态图"""
        logging.info("开始绘制机组开关机状态图...")

        # 新增：安全获取变量值，未初始化时返回0
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0

        # 获取火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': safe_value(self.model.thermal_status[unit, t]),
                    'Power': safe_value(self.model.thermal_power[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        
        # 按机组分组
        grouped = thermal_df.groupby('Unit')
        
        # 创建图表
        fig, axes = plt.subplots(len(grouped), 1, figsize=(15, 10), sharex=True)
        fig.suptitle('火电机组开关机状态及出力情况', fontsize=16)
        
        if len(grouped) == 1:
            axes = [axes]
        
        for i, (unit, group) in enumerate(grouped):
            ax = axes[i]
            
            # 绘制状态（0/1）
            ax.step(group['Time'], group['Status'], 'r-', where='post', label='状态')
            ax.set_ylabel('状态', color='r')
            ax.tick_params(axis='y', labelcolor='r')
            ax.set_ylim(-0.1, 1.1)
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # 创建双Y轴
            ax2 = ax.twinx()
            
            # 绘制出力
            ax2.plot(group['Time'], group['Power'], 'b-', label='出力')
            ax2.set_ylabel('出力 (MW)', color='b')
            ax2.tick_params(axis='y', labelcolor='b')
            
            # 设置标题
            ax.set_title(f'机组 {unit}')
            
            # 添加图例
            lines, labels = ax.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax.legend(lines + lines2, labels + labels2, loc='upper right')
        
        plt.xlabel('时间 (小时)')
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # 为总标题留出空间
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"机组状态图已保存到 {output_path}")
        plt.close()
        
        return output_path
    
    def plot_generation_mix(self, output_path):
        """绘制各类电源出力情况"""
        logging.info("开始绘制电源结构图...")

        # 新增：安全获取变量值，未初始化时返回0
        def safe_value(var):
            try:
                return pyo.value(var)
            except Exception:
                return 0

        # 初始化结果数组
        thermal_gen = np.zeros(self.T)
        wind_gen = np.zeros(self.T)
        solar_gen = np.zeros(self.T)
        hydro_gen = np.zeros(self.T)
        ess_discharge = np.zeros(self.T)

        # 计算总发电量
        for t in range(self.T):
            for unit in self.model.ThermalUnits:
                thermal_gen[t] += safe_value(self.model.thermal_power[unit, t])
            for unit in self.model.WindUnits:
                wind_gen[t] += safe_value(self.model.wind_power[unit, t])
            for unit in self.model.SolarUnits:
                solar_gen[t] += safe_value(self.model.solar_power[unit, t])
            for unit in self.model.HydroUnits:
                hydro_gen[t] += safe_value(self.model.hydro_power[unit, t])
            for unit in self.model.ESSUntis:
                ess_discharge[t] += safe_value(self.model.ess_discharge[unit, t])
        
        # 创建图表
        plt.figure(figsize=(12, 6))
        
        # 堆叠面积图
        plt.stackplot(range(self.T), thermal_gen, wind_gen, solar_gen, hydro_gen, ess_discharge,
                      labels=['火电', '风电', '光伏', '水电', '储能放电'])
        
        # 添加总负荷曲线
        total_load = []
        # 负荷编号字段兼容
        load_id_col = 'LoadId' if 'LoadId' in self.loads.columns else self.loads.columns[0]
        coef_col = 'ActivePowerCoef' if 'ActivePowerCoef' in self.loads.columns else self.loads.columns[1]
        # 负荷曲线时间字段兼容（只允许字符串key，避免KeyError: '0'）
        def get_time_col(t):
            if hasattr(self.load_curve, 'index'):
                idx = self.load_curve.index
            else:
                idx = self.load_curve.keys()
            # 优先Time_0、Time_1等
            if f'Time_{t}' in idx:
                return f'Time_{t}'
            # 其次字符串数字
            if str(t) in idx:
                return str(t)
            # 只取第一个可用key
            return list(idx)[0]
        def get_load_coef(load_id):
            mask = self.loads[load_id_col].astype(str) == str(load_id)
            if not mask.any():
                return 0.0
            val = self.loads.loc[mask, coef_col].values[0]
            try:
                return float(val)
            except Exception:
                return 0.0
        def get_load_curve_val(t):
            col = get_time_col(t)
            if col not in self.load_curve:
                logging.error(f"负荷曲线缺少 {col} 列，请检查 CurveLoad 表！")
                raise ValueError(f"负荷曲线缺少 {col} 列，请检查 CurveLoad 表！")
            val = self.load_curve[col]
            try:
                v = float(val)
                if np.isnan(v):
                    logging.error(f"负荷曲线 {col} 数据为 nan，请检查 CurveLoad 表！")
                    raise ValueError(f"负荷曲线 {col} 数据为 nan，请检查 CurveLoad 表！")
                return v
            except Exception:
                logging.error(f"负荷曲线 {col} 数据无法转换为数值，请检查 CurveLoad 表！")
                raise
        for t in range(self.T):
            load = 0
            time_col = get_time_col(t)
            for load_id in self.model.Loads:
                active_coef = get_load_coef(load_id)
                load_curve_val = get_load_curve_val(t)
                load += active_coef * load_curve_val
            total_load.append(load)
        
        plt.plot(range(self.T), total_load, 'k--', linewidth=2, label='总负荷')
        
        # 设置图表属性
        plt.title('各类电源出力及负荷曲线')
        plt.xlabel('时间 (小时)')
        plt.ylabel('功率 (MW)')
        # 检查所有输入的曲线数据是否有nan，如果有则报错并退出
        for arr, name in zip([thermal_gen, wind_gen, solar_gen, hydro_gen, ess_discharge], 
                             ['火电', '风电', '光伏', '水电', '储能放电']):
            if np.isnan(arr).any():
                logging.error(f"{name}出力曲线存在nan，请检查输入数据和曲线表！")
                raise ValueError(f"{name}出力曲线存在nan，请检查输入数据和曲线表！")
        if np.isnan(total_load).any():
            logging.error("总负荷曲线存在nan，请检查负荷曲线数据！")
            raise ValueError("总负荷曲线存在nan，请检查负荷曲线数据！")
        plt.legend(loc='upper left')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.xlim(0, self.T-1)
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"电源结构图已保存到 {output_path}")
        plt.close()
        
        return output_path

# 主程序
if __name__ == "__main__":
    # 输入输出路径
    input_file = r"C:\Users\admin\Desktop\Final_Cleaned_Rots_v2.xlsx"
    output_file = "UC_Results.xlsx"
    status_plot = "Unit_Status.png"
    mix_plot = "Generation_Mix.png"
    
    try:
        # 创建模型
        uc_model = UC_Model(input_file)
        
        # 求解模型
        result = uc_model.solve()
        if result is True:
            # 保存结果
            result_path = uc_model.save_results(output_file)
            
            # 绘制图表
            status_path = uc_model.plot_unit_status(status_plot)
            mix_path = uc_model.plot_generation_mix(mix_plot)
            
            # 将图表插入Excel
            wb = openpyxl.load_workbook(output_file)
            
            # 添加机组状态图
            ws = wb.create_sheet("机组状态图")
            img = Image(status_path)
            ws.add_image(img, 'A1')
            
            # 添加电源结构图
            ws = wb.create_sheet("电源结构图")
            img = Image(mix_path)
            ws.add_image(img, 'A1')
            
            # 添加联络线送电曲线
            ws = wb.create_sheet("联络线送电曲线")

            # 获取联络线数据
            tie_line_results = []
            valid_line_keys = set(uc_model.model.line_power.keys())
            def safe_value(var):
                try:
                    v = pyo.value(var)
                    if v is None or (isinstance(v, float) and np.isnan(v)):
                        return 0
                    return v
                except Exception:
                    return 0
            # 收集所有AC联络线名称
            ac_lines = [line for line in uc_model.model.Lines if "ACLine" in str(line)]
            times = sorted(list(uc_model.model.T))
            # 构建表头
            header = ["Time"] + [str(line) for line in ac_lines]
            ws.append(header)
            # 填充每个时刻每条联络线的功率
            for t in times:
                row = [t]
                for line in ac_lines:
                    if (line, t) in valid_line_keys:
                        power = safe_value(uc_model.model.line_power[line, t])
                    else:
                        power = 0
                    row.append(power)
                ws.append(row)

            # 创建图表
            chart = LineChart()
            chart.title = "联络线送电曲线"
            chart.style = 12
            chart.y_axis.title = '功率 (MW)'
            chart.x_axis.title = '时间 (小时)'

            # 设置数据区域（假设数据从A1开始，A为时间，B~N为各联络线）
            max_row = ws.max_row
            max_col = ws.max_column
            data = Reference(ws, min_col=2, min_row=1, max_row=max_row, max_col=max_col)
            chart.add_data(data, titles_from_data=True)
            cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart.set_categories(cats)
            ws.add_chart(chart, "D5")
        elif result is False:
            logging.error("优化求解失败，请检查模型和输入数据")
            print("优化求解失败，请检查日志获取详细信息")
            sys.exit(1)  # 直接退出主程序，避免后续 wb 未定义报错
        else:
            logging.error("优化求解未返回有效结果")
            print("优化求解未返回有效结果")    
            sys.exit(1)  # 直接退出主程序，避免后续 wb 未定义报错
        # 保存最终结果
        wb.save(output_file)

        logging.info("所有结果已保存并整合到Excel文件中")
        print(f"优化完成，结果已保存到: {output_file}")
        
        
    except Exception as e:
        logging.exception("程序运行出错")
        print(f"程序运行出错: {str(e)}")


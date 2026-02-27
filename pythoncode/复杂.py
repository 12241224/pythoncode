import pandas as pd
import numpy as np
import pyomo.environ as pyo
from pyomo.opt import SolverFactory
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import LineChart, Reference
import os
import logging

# 设置中文显示
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 初始化日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class UC_Model:
    def __init__(self, data_path):
        self.data_path = data_path
        self.load_data()
        self.create_model()
    
    def load_data(self):
        """加载所有Excel数据"""
        logging.info("开始加载数据...")
        
        # 系统汇总信息
        self.system_summary = pd.read_excel(
            self.data_path, sheet_name='SystemSummary', skiprows=2, nrows=1
        ).iloc[0]
        
        # 机组参数
        self.thermal_units = pd.read_excel(
            self.data_path, sheet_name='UnitThermalGenerators', skiprows=2
        )
        self.wind_units = pd.read_excel(
            self.data_path, sheet_name='UnitWindGenerators', skiprows=2
        )
        self.solar_units = pd.read_excel(
            self.data_path, sheet_name='UnitSolarGenerators', skiprows=2
        )
        self.ess_units = pd.read_excel(
            self.data_path, sheet_name='StorElectrochemicalESS', skiprows=2
        )
        self.hydro_units = pd.read_excel(
            self.data_path, sheet_name='UnitRunoffHydroGenerators', skiprows=2
        )
        
        # 曲线数据
        self.wind_curves = pd.read_excel(
            self.data_path, sheet_name='CurveWindResource', skiprows=2
        )
        self.solar_curves = pd.read_excel(
            self.data_path, sheet_name='CurveSolarResource', skiprows=2
        )
        self.hydro_curves = pd.read_excel(
            self.data_path, sheet_name='CurveRunoffHydroResource', skiprows=2
        )
        self.load_curve = pd.read_excel(
            self.data_path, sheet_name='CurveLoad', skiprows=2
        ).iloc[0]
        
        # 网络数据
        self.buses = pd.read_excel(self.data_path, sheet_name='NetBuses', skiprows=2)
        self.lines = pd.read_excel(self.data_path, sheet_name='NetLines', skiprows=2)
        self.transformers = pd.read_excel(self.data_path, sheet_name='NetTransformers', skiprows=2)
        self.sections = pd.read_excel(self.data_path, sheet_name='NetSectionsLines', skiprows=2)
        self.section_capacity = pd.read_excel(self.data_path, sheet_name='NetSectionsCapacity', skiprows=2)
        self.ac_tielines = pd.read_excel(self.data_path, sheet_name='NetHVACTieLine', skiprows=2)
        self.ac_tieline_curves = pd.read_excel(self.data_path, sheet_name='EnergyHVACTieLine', skiprows=2)
        
        # 负荷数据
        self.loads = pd.read_excel(self.data_path, sheet_name='Loads', skiprows=2)
        
        # 燃料数据
        self.fuel_data = pd.read_excel(self.data_path, sheet_name='UnitFuel', skiprows=2).iloc[0]
        
        # 设置时间参数
        self.T = int(self.system_summary['Duration'])
        self.dt = 1  # 调度间隔为1小时
        self.hours = list(range(self.T))
        
        # 创建机组字典
        self.thermal_dict = self.thermal_units.set_index('ThermalUnitNumber').to_dict('index')
        self.wind_dict = self.wind_units.set_index('WTPlantID').to_dict('index')
        self.solar_dict = self.solar_units.set_index('PVPlantID').to_dict('index')
        self.ess_dict = self.ess_units.set_index('ESSId').to_dict('index')
        self.hydro_dict = self.hydro_units.set_index('RunoffHydroUnitName').to_dict('index')
        
        logging.info("数据加载完成")
    
    def create_model(self):
        """创建Pyomo优化模型"""
        logging.info("开始创建优化模型...")
        
        self.model = pyo.ConcreteModel()
        
        # 定义集合
        self.model.T = pyo.Set(initialize=self.hours)  # 时间段
        self.model.ThermalUnits = pyo.Set(initialize=self.thermal_units['ThermalUnitNumber'].tolist())
        self.model.WindUnits = pyo.Set(initialize=self.wind_units['WTPlantID'].tolist())
        self.model.SolarUnits = pyo.Set(initialize=self.solar_units['PVPlantID'].tolist())
        self.model.ESSUntis = pyo.Set(initialize=self.ess_units['ESSId'].tolist())
        self.model.HydroUnits = pyo.Set(initialize=self.hydro_units['RunoffHydroUnitName'].tolist())
        self.model.Loads = pyo.Set(initialize=self.loads['LoadId'].tolist())
        self.model.Buses = pyo.Set(initialize=self.buses['BusId'].tolist())
        self.model.Lines = pyo.Set(initialize=self.lines['BranchId'].tolist())
        self.model.Sections = pyo.Set(initialize=self.section_capacity['NetSectionId'].tolist())
        
        # 定义变量
        # 火电机组
        self.model.thermal_status = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 机组状态 (0/1)
        self.model.thermal_power = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.NonNegativeReals
        )  # 机组出力
        self.model.thermal_startup = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 启动状态
        self.model.thermal_shutdown = pyo.Var(
            self.model.ThermalUnits, self.model.T, within=pyo.Binary
        )  # 停机状态
        
        # 可再生能源
        self.model.wind_power = pyo.Var(
            self.model.WindUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.solar_power = pyo.Var(
            self.model.SolarUnits, self.model.T, within=pyo.NonNegativeReals
        )
        self.model.hydro_power = pyo.Var(
            self.model.HydroUnits, self.model.T, within=pyo.NonNegativeReals
        )
        
        # 储能系统
        self.model.ess_charge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 充电功率
        self.model.ess_discharge = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 放电功率
        self.model.ess_soc = pyo.Var(
            self.model.ESSUntis, self.model.T, within=pyo.NonNegativeReals
        )  # 储能状态
        
        # 负荷相关
        self.model.load_shed = pyo.Var(
            self.model.Loads, self.model.T, within=pyo.NonNegativeReals
        )  # 切负荷量
        
        # 网络相关
        self.model.bus_angle = pyo.Var(
            self.model.Buses, self.model.T, within=pyo.Reals
        )  # 节点相角
        self.model.line_power = pyo.Var(
            self.model.Lines, self.model.T, within=pyo.Reals
        )  # 线路功率
        
        # 目标函数
        def total_cost_rule(model):
            """系统总成本计算"""
            cost = 0
            
            # 火电运行成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    # 二次成本函数线性化近似
                    a = self.thermal_dict[unit]['OperationCoeff_A']
                    b = self.thermal_dict[unit]['OperationCoeff_B']
                    c = self.thermal_dict[unit]['OperationCoeff_C']
                    min_power = self.thermal_dict[unit]['MinPower']
                    max_power = self.thermal_dict[unit]['Capacity']
                    
                    # 分段线性化近似二次函数
                    cost += model.thermal_status[unit, t] * c
                    cost += b * model.thermal_power[unit, t]
                    
                    # 深度调峰成本
                    deep_threshold = self.thermal_dict[unit]['DeepPeakShavingThreshold'] * max_power
                    deep_cost = self.thermal_dict[unit]['DeepPeakShavingCost']
                    if model.thermal_power[unit, t] < deep_threshold:
                        cost += deep_cost * (deep_threshold - model.thermal_power[unit, t])
            
            # 启停成本
            for unit in model.ThermalUnits:
                for t in model.T:
                    startup_cost = self.thermal_dict[unit]['StateOnFuelConsumption']
                    cost += startup_cost * model.thermal_startup[unit, t]
                    
                    # 停机成本通常为0，这里省略
            
            # 切负荷惩罚
            penalty_cut = 10000  # 元/MWh
            for load in model.Loads:
                for t in model.T:
                    cost += penalty_cut * model.load_shed[load, t]
            
            # 备用不足惩罚
            penalty_reserve = 5000  # 元/MWh
            # 这里简化处理，实际应计算备用不足量
            
            return cost
        
        self.model.total_cost = pyo.Objective(rule=total_cost_rule, sense=pyo.minimize)
        
        # 约束条件
        # 1. 火电机组逻辑约束
        def thermal_logic_rule(model, unit, t):
            """机组状态逻辑约束"""
            if t == 0:
                # 初始状态
                init_status = self.thermal_dict[unit]['InitStatus']
                return model.thermal_status[unit, t] == (1 if init_status else 0)
            else:
                # 状态变化约束
                return model.thermal_status[unit, t] - model.thermal_status[unit, t-1] == \
                    model.thermal_startup[unit, t] - model.thermal_shutdown[unit, t]
        self.model.thermal_logic_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_logic_rule
        )
        
        # 2. 火电机组出力约束
        def thermal_power_rule(model, unit, t):
            """机组出力上下限约束"""
            min_power = self.thermal_dict[unit]['MinPower']
            max_power = self.thermal_dict[unit]['Capacity']
            return (
                min_power * model.thermal_status[unit, t],
                model.thermal_power[unit, t],
                max_power * model.thermal_status[unit, t]
            )
        self.model.thermal_power_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_power_rule
        )
        
        # 3. 爬坡约束
        def thermal_ramp_up_rule(model, unit, t):
            """上爬坡约束"""
            if t == 0:
                return pyo.Constraint.Skip
            ramp_up = self.thermal_dict[unit]['UpRamppingRate']
            return model.thermal_power[unit, t] - model.thermal_power[unit, t-1] <= \
                ramp_up + model.thermal_startup[unit, t] * self.thermal_dict[unit]['Capacity']
        self.model.thermal_ramp_up_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_up_rule
        )
        
        def thermal_ramp_down_rule(model, unit, t):
            """下爬坡约束"""
            if t == 0:
                return pyo.Constraint.Skip
            ramp_down = self.thermal_dict[unit]['DownRamppingRate']
            return model.thermal_power[unit, t-1] - model.thermal_power[unit, t] <= \
                ramp_down + model.thermal_shutdown[unit, t] * self.thermal_dict[unit]['Capacity']
        self.model.thermal_ramp_down_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=thermal_ramp_down_rule
        )
        
        # 4. 最短开停机时间约束
        def min_up_time_rule(model, unit, t):
            """最小运行时间约束"""
            min_up = self.thermal_dict[unit]['MinOnLineTime']
            if t < min_up:
                return pyo.Constraint.Skip
            
            # 如果机组在t时刻启动，则从t-min_up+1到t时刻都必须运行
            return sum(model.thermal_shutdown[unit, k] for k in range(t-min_up+1, t+1)) <= \
                1 - model.thermal_status[unit, t]
        self.model.min_up_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_up_time_rule
        )
        
        def min_down_time_rule(model, unit, t):
            """最小停机时间约束"""
            min_down = self.thermal_dict[unit]['MinOffLineTime']
            if t < min_down:
                return pyo.Constraint.Skip
            
            # 如果机组在t时刻停机，则从t-min_down+1到t时刻都必须停机
            return sum(model.thermal_startup[unit, k] for k in range(t-min_down+1, t+1)) <= \
                model.thermal_status[unit, t]
        self.model.min_down_time_con = pyo.Constraint(
            self.model.ThermalUnits, self.model.T, rule=min_down_time_rule
        )
        
        # 5. 可再生能源出力约束
        def wind_power_rule(model, unit, t):
            """风电出力约束"""
            curve_name = self.wind_dict[unit]['ResourceCurve']
            capacity = self.wind_dict[unit]['Capacity']
            
            # 从风电曲线中获取预测值
            wind_curve = self.wind_curves[self.wind_curves['WTCurve'] == curve_name].iloc[0]
            max_power = wind_curve[f'Time_{t}'] * capacity
            
            return model.wind_power[unit, t] <= max_power
        self.model.wind_power_con = pyo.Constraint(
            self.model.WindUnits, self.model.T, rule=wind_power_rule
        )
        
        # 类似地添加光伏和水电约束...
        
        # 6. 储能系统约束
        def ess_power_rule(model, unit, t):
            """储能充放电功率约束"""
            max_charge = self.ess_dict[unit]['MaxChargePower']
            max_discharge = self.ess_dict[unit]['MaxDisPower']
            return (
                model.ess_charge[unit, t] <= max_charge,
                model.ess_discharge[unit, t] <= max_discharge
            )
        self.model.ess_power_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_power_rule
        )
        
        def ess_soc_rule(model, unit, t):
            """储能SOC约束"""
            min_soc = self.ess_dict[unit]['MinSOC']
            max_soc = self.ess_dict[unit]['MaxSOC']
            capacity = self.ess_dict[unit]['Capacity']
            
            return (min_soc * capacity, model.ess_soc[unit, t], max_soc * capacity)
        self.model.ess_soc_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_soc_rule
        )
        
        def ess_balance_rule(model, unit, t):
            """储能能量平衡约束"""
            charge_eff = self.ess_dict[unit]['ChargeEfficiency']
            discharge_eff = self.ess_dict[unit]['DisEfficiency']
            
            if t == 0:
                init_soc = self.ess_dict[unit]['InitSOC']
                return model.ess_soc[unit, t] == init_soc - \
                    model.ess_discharge[unit, t]/discharge_eff + \
                    model.ess_charge[unit, t] * charge_eff
            else:
                return model.ess_soc[unit, t] == model.ess_soc[unit, t-1] - \
                    model.ess_discharge[unit, t]/discharge_eff + \
                    model.ess_charge[unit, t] * charge_eff
        self.model.ess_balance_con = pyo.Constraint(
            self.model.ESSUntis, self.model.T, rule=ess_balance_rule
        )
        
        # 7. 系统功率平衡约束
        def power_balance_rule(model, t):
            """系统功率平衡约束"""
            # 总发电
            thermal_gen = sum(model.thermal_power[unit, t] for unit in model.ThermalUnits)
            wind_gen = sum(model.wind_power[unit, t] for unit in model.WindUnits)
            solar_gen = sum(model.solar_power[unit, t] for unit in model.SolarUnits)
            hydro_gen = sum(model.hydro_power[unit, t] for unit in model.HydroUnits)
            ess_discharge = sum(model.ess_discharge[unit, t] for unit in model.ESSUntis)
            ess_charge = sum(model.ess_charge[unit, t] for unit in model.ESSUntis)
            
            # 总负荷
            total_load = sum(self.loads.loc[self.loads['LoadId'] == load, 'ActivePowerCoef'].values[0] * 
                           self.load_curve[f'Time_{t}'] for load in model.Loads)
            load_shed = sum(model.load_shed[load, t] for load in model.Loads)
            
            # 功率平衡
            return thermal_gen + wind_gen + solar_gen + hydro_gen + ess_discharge == \
                total_load - load_shed + ess_charge
        self.model.power_balance_con = pyo.Constraint(
            self.model.T, rule=power_balance_rule
        )
        
        # 8. 备用约束
        def reserve_requirement_rule(model, t):
            """旋转备用约束"""
            reserve_req = 0.1  # 10%的备用要求
            
            # 总负荷
            total_load = sum(self.loads.loc[self.loads['LoadId'] == load, 'ActivePowerCoef'].values[0] * 
                           self.load_curve[f'Time_{t}'] for load in model.Loads)
            
            # 可用备用
            available_reserve = 0
            for unit in model.ThermalUnits:
                max_power = self.thermal_dict[unit]['Capacity']
                available_reserve += model.thermal_status[unit, t] * max_power - model.thermal_power[unit, t]
            
            return available_reserve >= reserve_req * total_load
        self.model.reserve_con = pyo.Constraint(
            self.model.T, rule=reserve_requirement_rule
        )
        
        # 9. 网络约束（直流潮流）
        # 这里简化处理，实际应建立完整的直流潮流模型
        def line_power_rule(model, line, t):
            """线路功率计算（直流潮流）"""
            line_data = self.lines[self.lines['BranchId'] == line].iloc[0]
            start_bus = line_data['StartNodeName']
            end_bus = line_data['EndNodeName']
            reactance = line_data['Reactance']
            
            # 线路功率 = (θ_start - θ_end) / x
            return model.line_power[line, t] == (
                model.bus_angle[start_bus, t] - model.bus_angle[end_bus, t]
            ) / reactance
        self.model.line_power_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_power_rule
        )
        
        def line_capacity_rule(model, line, t):
            """线路传输容量约束"""
            capacity = self.lines[self.lines['BranchId'] == line]['Capacity'].values[0]
            return (-capacity, model.line_power[line, t], capacity)
        self.model.line_capacity_con = pyo.Constraint(
            self.model.Lines, self.model.T, rule=line_capacity_rule
        )
        
        # 10. 断面约束
        def section_capacity_rule(model, section, t):
            """断面传输容量约束"""
            # 获取断面包含的线路
            section_lines = self.sections[self.sections['NetSectionId'] == section]
            total_power = 0
            
            for _, row in section_lines.iterrows():
                line = row['BranchId']
                direction = 1 if row['BranchDirection'] else -1
                total_power += model.line_power[line, t] * direction
            
            capacity = self.section_capacity[
                self.section_capacity['NetSectionId'] == section
            ]['Capacity'].values[0]
            
            return (-capacity, total_power, capacity)
        self.model.section_capacity_con = pyo.Constraint(
            self.model.Sections, self.model.T, rule=section_capacity_rule
        )
        
        # 设置参考节点相角为0
        balance_bus = self.system_summary['BalanceNode']
        def ref_bus_rule(model, t):
            return model.bus_angle[balance_bus, t] == 0
        self.model.ref_bus_con = pyo.Constraint(
            self.model.T, rule=ref_bus_rule
        )
        
        logging.info("优化模型创建完成")
    
    def solve(self):
        """求解优化问题"""
        logging.info("开始求解优化问题...")
        
        # 使用Gurobi求解器
        solver = SolverFactory('gurobi')
        solver.options['MIPGap'] = 0.01  # 设置MIP gap为1%
        solver.options['TimeLimit'] = 600  # 设置10分钟时间限制
        results = solver.solve(self.model, tee=True)
        
        if (results.solver.status == pyo.SolverStatus.ok and 
            results.solver.termination_condition == pyo.TerminationCondition.optimal):
            logging.info("优化求解成功完成")
            return True
        elif results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            logging.warning("达到时间限制，返回当前最优解")
            return True
        else:
            logging.error("优化求解失败")
            return False
    
    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")
        
        # 创建Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # 省略前面的导入和类定义...

    def save_results(self, output_path):
        """保存结果到Excel"""
        logging.info("开始保存结果到Excel...")
        
        # 创建Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        # 1. 火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    '机组': unit,
                    '时间': t,
                    '状态': pyo.value(self.model.thermal_status[unit, t]),
                    '出力': pyo.value(self.model.thermal_power[unit, t]),
                    '启动': pyo.value(self.model.thermal_startup[unit, t]),
                    '停机': pyo.value(self.model.thermal_shutdown[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        thermal_df.to_excel(writer, sheet_name='火电机组', index=False)
        
        # 2. 风电结果
        wind_results = []
        for unit in self.model.WindUnits:
            curve_name = self.wind_dict[unit]['ResourceCurve']
            capacity = self.wind_dict[unit]['Capacity']
            wind_curve = self.wind_curves[self.wind_curves['WTCurve'] == curve_name].iloc[0]
            
            for t in self.model.T:
                max_power = wind_curve[f'Time_{t}'] * capacity
                actual_power = pyo.value(self.model.wind_power[unit, t])
                curtailment = max(0, max_power - actual_power)  # 弃风量
                
                wind_results.append({
                    '机组': unit,
                    '时间': t,
                    '预测出力': max_power,
                    '实际出力': actual_power,
                    '弃风量': curtailment
                })
        wind_df = pd.DataFrame(wind_results)
        wind_df.to_excel(writer, sheet_name='风电', index=False)
        
        # 3. 光伏结果
        solar_results = []
        for unit in self.model.SolarUnits:
            curve_name = self.solar_dict[unit]['ResourceCurve']
            capacity = self.solar_dict[unit]['Capacity']
            solar_curve = self.solar_curves[self.solar_curves['PVCurve'] == curve_name].iloc[0]
            
            for t in self.model.T:
                max_power = solar_curve[f'Time_{t}'] * capacity
                actual_power = pyo.value(self.model.solar_power[unit, t])
                curtailment = max(0, max_power - actual_power)  # 弃光量
                
                solar_results.append({
                    '机组': unit,
                    '时间': t,
                    '预测出力': max_power,
                    '实际出力': actual_power,
                    '弃光量': curtailment
                })
        solar_df = pd.DataFrame(solar_results)
        solar_df.to_excel(writer, sheet_name='光伏', index=False)
        
        # 4. 水电结果
        hydro_results = []
        for unit in self.model.HydroUnits:
            curve_name = self.hydro_dict[unit]['ResourceCurve']
            capacity = self.hydro_dict[unit]['Capacity']
            hydro_curve = self.hydro_curves[self.hydro_curves['RunoffHydroOperationCurve'] == curve_name].iloc[0]
            
            for t in self.model.T:
                max_power = hydro_curve[f'Time_{t}'] * capacity
                actual_power = pyo.value(self.model.hydro_power[unit, t])
                
                hydro_results.append({
                    '机组': unit,
                    '时间': t,
                    '最大出力': max_power,
                    '实际出力': actual_power
                })
        hydro_df = pd.DataFrame(hydro_results)
        hydro_df.to_excel(writer, sheet_name='水电', index=False)
        
        # 5. 储能结果
        ess_results = []
        for unit in self.model.ESSUntis:
            min_soc = self.ess_dict[unit]['MinSOC']
            max_soc = self.ess_dict[unit]['MaxSOC']
            capacity = self.ess_dict[unit]['Capacity']
            
            for t in self.model.T:
                charge = pyo.value(self.model.ess_charge[unit, t])
                discharge = pyo.value(self.model.ess_discharge[unit, t])
                soc = pyo.value(self.model.ess_soc[unit, t])
                soc_percent = soc / capacity * 100  # SOC百分比
                
                ess_results.append({
                    '储能': unit,
                    '时间': t,
                    '充电功率': charge,
                    '放电功率': discharge,
                    'SOC': soc,
                    'SOC(%)': soc_percent,
                    '状态': '充电' if charge > 0 else ('放电' if discharge > 0 else '空闲')
                })
        ess_df = pd.DataFrame(ess_results)
        ess_df.to_excel(writer, sheet_name='储能', index=False)
        
        # 6. 负荷结果
        load_results = []
        for load in self.model.Loads:
            active_coef = self.loads.loc[self.loads['LoadId'] == load, 'ActivePowerCoef'].values[0]
            bus = self.loads.loc[self.loads['LoadId'] == load, 'BusName'].values[0]
            area = self.loads.loc[self.loads['LoadId'] == load, 'AreaName'].values[0]
            
            for t in self.model.T:
                load_value = active_coef * self.load_curve[f'Time_{t}']
                shed = pyo.value(self.model.load_shed[load, t])
                
                load_results.append({
                    '负荷': load,
                    '母线': bus,
                    '区域': area,
                    '时间': t,
                    '需求': load_value,
                    '切负荷': shed,
                    '满足率': (load_value - shed) / load_value * 100 if load_value > 0 else 100
                })
        load_df = pd.DataFrame(load_results)
        load_df.to_excel(writer, sheet_name='负荷', index=False)
        
        # 7. 网络结果
        line_results = []
        for line in self.model.Lines:
            line_data = self.lines[self.lines['BranchId'] == line].iloc[0]
            start_bus = line_data['StartNodeName']
            end_bus = line_data['EndNodeName']
            capacity = line_data['Capacity']
            
            for t in self.model.T:
                power = pyo.value(self.model.line_power[line, t])
                utilization = abs(power) / capacity * 100  # 利用率
                
                line_results.append({
                    '线路': line,
                    '起点': start_bus,
                    '终点': end_bus,
                    '时间': t,
                    '潮流': power,
                    '容量': capacity,
                    '利用率(%)': utilization
                })
        line_df = pd.DataFrame(line_results)
        line_df.to_excel(writer, sheet_name='线路潮流', index=False)
        
        # 8. 断面结果
        section_results = []
        for section in self.model.Sections:
            section_data = self.section_capacity[self.section_capacity['NetSectionId'] == section].iloc[0]
            section_name = section_data['NetSectionName']
            capacity = section_data['Capacity']
            start_system = section_data['StartSystemName']
            end_system = section_data['EndSystemName']
            
            section_lines = self.sections[self.sections['NetSectionId'] == section]
            for t in self.model.T:
                total_power = 0
                for _, row in section_lines.iterrows():
                    line = row['BranchId']
                    direction = 1 if row['BranchDirection'] else -1
                    total_power += pyo.value(self.model.line_power[line, t]) * direction
                
                utilization = abs(total_power) / capacity * 100
                
                section_results.append({
                    '断面': section,
                    '名称': section_name,
                    '时间': t,
                    '潮流': total_power,
                    '容量': capacity,
                    '利用率(%)': utilization,
                    '送端': start_system,
                    '受端': end_system
                })
        section_df = pd.DataFrame(section_results)
        section_df.to_excel(writer, sheet_name='断面潮流', index=False)
        
        # 9. 联络线结果
        tie_line_results = []
        for line in self.model.Lines:
            if "ACLine" in line:
                line_data = self.ac_tielines[self.ac_tielines['ACTieLineId'] == line].iloc[0]
                send_bus = line_data['SendBusID']
                receive_bus = line_data['ReceiveBusID']
                capacity = line_data['Capacity']
                start_area = line_data['StartAreaName']
                end_area = line_data['EndAreaName']
                
                for t in self.model.T:
                    power = pyo.value(self.model.line_power[line, t])
                    utilization = abs(power) / capacity * 100
                    
                    tie_line_results.append({
                        '联络线': line,
                        '送端母线': send_bus,
                        '受端母线': receive_bus,
                        '时间': t,
                        '潮流': power,
                        '容量': capacity,
                        '利用率(%)': utilization,
                        '送端区域': start_area,
                        '受端区域': end_area
                    })
        tie_line_df = pd.DataFrame(tie_line_results)
        tie_line_df.to_excel(writer, sheet_name='联络线', index=False)
        
        # 10. 系统总结
        total_thermal_gen = sum(pyo.value(self.model.thermal_power[unit, t]) 
                             for unit in self.model.ThermalUnits for t in self.model.T)
        total_wind_gen = sum(pyo.value(self.model.wind_power[unit, t]) 
                           for unit in self.model.WindUnits for t in self.model.T)
        total_solar_gen = sum(pyo.value(self.model.solar_power[unit, t]) 
                            for unit in self.model.SolarUnits for t in self.model.T)
        total_hydro_gen = sum(pyo.value(self.model.hydro_power[unit, t]) 
                            for unit in self.model.HydroUnits for t in self.model.T)
        total_gen = total_thermal_gen + total_wind_gen + total_solar_gen + total_hydro_gen
        
        total_load = sum(self.loads.loc[self.loads['LoadId'] == load, 'ActivePowerCoef'].values[0] * 
                       self.load_curve[f'Time_{t}'] for load in self.model.Loads for t in self.model.T)
        total_shed = sum(pyo.value(self.model.load_shed[load, t]) 
                      for load in self.model.Loads for t in self.model.T)
        
        total_wind_curtail = wind_df['弃风量'].sum()
        total_solar_curtail = solar_df['弃光量'].sum()
        
        summary_data = {
            '指标': ['总成本(元)', '总发电量(MWh)', '总负荷(MWh)', '总切负荷量(MWh)', 
                   '总弃风量(MWh)', '总弃光量(MWh)', '总弃水率(%)', '平均切负荷率(%)'],
            '数值': [
                pyo.value(self.model.total_cost),
                total_gen,
                total_load,
                total_shed,
                total_wind_curtail,
                total_solar_curtail,
                total_wind_curtail / (total_wind_gen + total_wind_curtail) * 100 if (total_wind_gen + total_wind_curtail) > 0 else 0,
                total_shed / total_load * 100 if total_load > 0 else 0
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='系统总结', index=False)
        
        # 保存Excel
        writer.save()
        
        logging.info(f"结果已保存到 {output_path}")
        
        # 返回结果文件路径
        return output_path
         
    def plot_unit_status(self, output_path):
        """绘制机组开关机状态图"""
        logging.info("开始绘制机组开关机状态图...")
        
        # 获取火电机组结果
        thermal_results = []
        for unit in self.model.ThermalUnits:
            for t in self.model.T:
                thermal_results.append({
                    'Unit': unit,
                    'Time': t,
                    'Status': pyo.value(self.model.thermal_status[unit, t]),
                    'Power': pyo.value(self.model.thermal_power[unit, t])
                })
        thermal_df = pd.DataFrame(thermal_results)
        
        # 按机组分组
        grouped = thermal_df.groupby('Unit')
        
        # 创建图表
        fig, axes = plt.subplots(len(grouped), 1, figsize=(15, 10), sharex=True)
        fig.suptitle('火电机组开关机状态及出力情况', fontsize=16)
        
        if len(grouped) == 1:
            axes = [axes]
        
        for i, (unit, group) in enumerate(grouped):
            ax = axes[i]
            
            # 绘制状态（0/1）
            ax.step(group['Time'], group['Status'], 'r-', where='post', label='状态')
            ax.set_ylabel('状态', color='r')
            ax.tick_params(axis='y', labelcolor='r')
            ax.set_ylim(-0.1, 1.1)
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # 创建双Y轴
            ax2 = ax.twinx()
            
            # 绘制出力
            ax2.plot(group['Time'], group['Power'], 'b-', label='出力')
            ax2.set_ylabel('出力 (MW)', color='b')
            ax2.tick_params(axis='y', labelcolor='b')
            
            # 设置标题
            ax.set_title(f'机组 {unit}')
            
            # 添加图例
            lines, labels = ax.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax.legend(lines + lines2, labels + labels2, loc='upper right')
        
        plt.xlabel('时间 (小时)')
        plt.tight_layout(rect=[0, 0, 1, 0.95])  # 为总标题留出空间
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"机组状态图已保存到 {output_path}")
        plt.close()
        
        return output_path
    
    def plot_generation_mix(self, output_path):
        """绘制各类电源出力情况"""
        logging.info("开始绘制电源结构图...")
        
        # 初始化结果数组
        thermal_gen = np.zeros(self.T)
        wind_gen = np.zeros(self.T)
        solar_gen = np.zeros(self.T)
        hydro_gen = np.zeros(self.T)
        ess_discharge = np.zeros(self.T)
        
        # 计算总发电量
        for t in range(self.T):
            for unit in self.model.ThermalUnits:
                thermal_gen[t] += pyo.value(self.model.thermal_power[unit, t])
            for unit in self.model.WindUnits:
                wind_gen[t] += pyo.value(self.model.wind_power[unit, t])
            for unit in self.model.SolarUnits:
                solar_gen[t] += pyo.value(self.model.solar_power[unit, t])
            for unit in self.model.HydroUnits:
                hydro_gen[t] += pyo.value(self.model.hydro_power[unit, t])
            for unit in self.model.ESSUntis:
                ess_discharge[t] += pyo.value(self.model.ess_discharge[unit, t])
        
        # 创建图表
        plt.figure(figsize=(12, 6))
        
        # 堆叠面积图
        plt.stackplot(range(self.T), thermal_gen, wind_gen, solar_gen, hydro_gen, ess_discharge,
                      labels=['火电', '风电', '光伏', '水电', '储能放电'])
        
        # 添加总负荷曲线
        total_load = []
        for t in range(self.T):
            load = 0
            for load_id in self.model.Loads:
                active_coef = self.loads.loc[self.loads['LoadId'] == load_id, 'ActivePowerCoef'].values[0]
                load_curve_val = self.load_curve[f'Time_{t}']
                load += active_coef * load_curve_val
            total_load.append(load)
        
        plt.plot(range(self.T), total_load, 'k--', linewidth=2, label='总负荷')
        
        # 设置图表属性
        plt.title('各类电源出力及负荷曲线')
        plt.xlabel('时间 (小时)')
        plt.ylabel('功率 (MW)')
        plt.legend(loc='upper left')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.xlim(0, self.T-1)
        
        # 保存图表
        plt.savefig(output_path)
        logging.info(f"电源结构图已保存到 {output_path}")
        plt.close()
        
        return output_path

# 主程序
if __name__ == "__main__":
    # 输入输出路径
    input_file = "Rots.xlsx"
    output_file = "UC_Results.xlsx"
    status_plot = "Unit_Status.png"
    mix_plot = "Generation_Mix.png"
    
    try:
        # 创建模型
        uc_model = UC_Model(input_file)
        
        # 求解模型
        if uc_model.solve():
            # 保存结果
            result_path = uc_model.save_results(output_file)
            
            # 绘制图表
            status_path = uc_model.plot_unit_status(status_plot)
            mix_path = uc_model.plot_generation_mix(mix_plot)
            
            # 将图表插入Excel
            wb = openpyxl.load_workbook(output_file)
            
            # 添加机组状态图
            ws = wb.create_sheet("机组状态图")
            img = Image(status_path)
            ws.add_image(img, 'A1')
            
            # 添加电源结构图
            ws = wb.create_sheet("电源结构图")
            img = Image(mix_path)
            ws.add_image(img, 'A1')
            
            # 添加联络线送电曲线
            ws = wb.create_sheet("联络线送电曲线")
            
            # 获取联络线数据
            tie_line_results = []
            for line in uc_model.model.Lines:
                if "ACLine" in line:
                    for t in uc_model.model.T:
                        tie_line_results.append({
                            'Line': line,
                            'Time': t,
                            'Power': pyo.value(uc_model.model.line_power[line, t])
                        })
            tie_line_df = pd.DataFrame(tie_line_results)
            
            # 创建图表
            chart = LineChart()
            chart.title = "联络线送电曲线"
            chart.style = 12
            chart.y_axis.title = '功率 (MW)'
            chart.x_axis.title = '时间 (小时)'
            
            # 添加数据
            data = Reference(ws, min_col=2, min_row=1, max_row=uc_model.T+1, max_col=len(uc_model.ac_tielines)+1)
            chart.add_data(data, titles_from_data=True)
            
            # 设置类别轴
            cats = Reference(ws, min_col=1, min_row=2, max_row=uc_model.T+1)
            chart.set_categories(cats)
            
            # 添加到工作表
            for i, row in enumerate(dataframe_to_rows(tie_line_df, index=False, header=True)):
                ws.append(row)
            
            ws.add_chart(chart, "D5")
            
            # 保存最终结果
            wb.save(output_file)
            
            logging.info("所有结果已保存并整合到Excel文件中")
            print(f"优化完成，结果已保存到: {output_file}")
        else:
            logging.error("优化求解失败，请检查模型和输入数据")
            print("优化求解失败，请检查日志获取详细信息")
    except Exception as e:
        logging.exception("程序运行出错")
        print(f"程序运行出错: {str(e)}")